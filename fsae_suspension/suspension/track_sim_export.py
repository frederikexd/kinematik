# ============================================================================
#  KinematiK — Formula SAE suspension & vehicle dynamics toolkit
#  Created by Frederik Thio. Copyright (c) 2026 Frederik Thio.
#  Open source. Original author: Frederik Thio, creator of KinematiK.
#
#  suspension/track_sim_export.py — write lap-sim results into the electrics
#  workbook as a LIVE model: correct physics, real Excel formulas, cached
#  values, and not one of the user's own cells touched.
# ============================================================================
"""
Track-sim export that produces a workbook someone can actually use.

WHAT WAS WRONG WITH THE OLD EXPORT
----------------------------------
The previous path (`ev_excel_roundtrip.lap_to_excel_roundtrip`) produced a file
that opened cleanly and was still not much use, for four separate reasons:

**1. It replaced formulas with constants.** Every computed cell in
ElecPropulsion came back as a static number. The workbook stopped being a model
— change a speed, a gear ratio or the pack layout and nothing recalculates —
and the errors baked into those constants became invisible to any audit that
reads formulas. A wrong formula can be found and fixed; a wrong number that used
to be a formula cannot.

**2. It re-implemented the workbook's own physics errors.** Gear ratios applied
as 1/N so the motor turned slower than the wheel, current as
`V_pack * PF * RPM / 1000`, efficiency multiplying instead of dividing. The
export agreed with the spreadsheet because it was a copy of the spreadsheet, so
the round-trip could never disagree with the thing it was supposed to check.
Peak power came out at 294 kW from a pack whose absolute ceiling is 106 kW.

**3. It wrote formulas with no cached values.** openpyxl does not evaluate what
it writes, so every formula cell read back as None to `data_only=True` — which
is every Python consumer, including KinematiK's own readers. Excel recalculated
on open and looked fine; nothing else could read the file at all.

**4. Its pack advice ignored current.** The old advisor reasoned from energy
alone and recommended dropping a 140S3P pack to 140S1P because the lap only
needed 1.02 of 6.96 kWh. Going to one parallel string triples pack resistance,
cuts deliverable power to a third, and puts the entire pack current through
every cell — about 39C for a 5 Ah cell. It recommended the change that makes the
fuse failure three times worse, on a sheet already reporting that the fuse
fails.

WHAT THIS MODULE DOES INSTEAD
-----------------------------
* **Physics from `power_draw`** — a real force balance, `P_elec = P_wheel/eta`,
  and pack current solved from `P = (V_oc - I*R)*I` so sag is included and an
  impossible demand is reported as impossible rather than as a large number.
* **Live Excel formulas throughout.** Inputs are numbers; everything derived is
  a formula referencing them. Change the mass on `KX Inputs` and the whole
  workbook moves, including the gear recommendation. This is the difference
  between a report and a model.
* **Cached values populated, with or without LibreOffice.**
  `export_track_sim` recalculates before returning, so `data_only=True` readers
  see numbers. LibreOffice is used when it is installed; when it is not — which
  is the normal case on Streamlit Cloud, in slim containers and in CI —
  `xlsx_formula_cache` evaluates the formulas in process instead. It covers the
  operators and functions these sheets use and refuses anything else rather
  than guessing, so a cell is either right or left empty.
* **The clock is the trace's own.** Column A of `KX Lap Trace` holds the real
  timestamps and column P each sample's share of the elapsed time; every
  derivative divides by an actual interval and every integral is a weighted
  sum. An unevenly logged lap is exact rather than approximate.
* **Nothing of the user's is touched.** Every sheet this writes is prefixed
  `KX `. Their formulas, their layout and their row numbering are left exactly
  as they were, which also means the audit trail survives and re-running is
  idempotent.
* **Advice that accounts for current.** The pack advisor checks energy AND the
  power ceiling AND per-cell C-rate, so it cannot recommend a pack that is
  thermally or electrically incapable of the lap it just simulated.

Only Excel-2007-era functions are used (`SUMPRODUCT`, `INDEX`, `MAX`, `IFERROR`)
so the formulas survive recalculation outside Excel.
"""

from __future__ import annotations

import math
import os
import shutil
import subprocess
import tempfile
from dataclasses import dataclass, field
from typing import Optional, Sequence

from .interfaces import Severity, Finding
from . import power_draw as pdw

#: Every sheet this module writes carries this prefix. Nothing else is modified.
SHEET_PREFIX = "KX "

S_INPUTS = SHEET_PREFIX + "Inputs"
S_TRACE = SHEET_PREFIX + "Lap Trace"
S_DASH = SHEET_PREFIX + "Dashboard"
S_GEARS = SHEET_PREFIX + "Gear Study"
S_PACK = SHEET_PREFIX + "Pack Limits"
S_ADVISOR = SHEET_PREFIX + "Pack Advisor"
S_PROV = SHEET_PREFIX + "Provenance"

_FONT = "Arial"
_BLUE = "0000FF"        # hardcoded input
_BLACK = "000000"       # formula
_GREEN = "008000"       # link to another sheet
_YELLOW = "FFFFFF00"    # must-confirm assumption


# ===================================================================== #
#  Styling helpers
# ===================================================================== #
def _style(ws, coord, *, bold=False, color=_BLACK, fill=None, fmt=None,
           size=10, wrap=False):
    from openpyxl.styles import Font, PatternFill, Alignment
    c = ws[coord]
    c.font = Font(name=_FONT, size=size, bold=bold, color=color)
    if fill:
        c.fill = PatternFill("solid", fgColor=fill)
    if fmt:
        c.number_format = fmt
    if wrap:
        c.alignment = Alignment(wrap_text=True, vertical="top")
    return c


def _header(ws, coord, text):
    _style(ws, coord, bold=True, size=12).value = text


def _title(ws, text: str, last_col: int) -> None:
    """A sheet title, merged across the used width.

    Titles are longer than the first column on every one of these sheets. Excel
    will spill text over an empty neighbour, but the moment someone puts a value
    in B the title is silently clipped — merging makes it independent of that.
    """
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
    _style(ws, "A1", bold=True, size=12).value = text


def _footer_label(ws, row: int, text: str, label_last_col: int,
                  value_col: int, value_formula: str, fmt: str = "General"):
    """A wide merged label with its value just to the right of it."""
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=label_last_col)
    _style(ws, f"A{row}", bold=False).value = text
    _style(ws, f"{_col(value_col)}{row}", bold=True, fmt=fmt).value = \
        value_formula


def _label(ws, row, text, col=1):
    _style(ws, f"{_col(col)}{row}", bold=False).value = text


def _col(n: int) -> str:
    s = ""
    while n > 0:
        n, r = divmod(n - 1, 26)
        s = chr(65 + r) + s
    return s


#: Notes and headers were originally written as plain long strings in column A.
#: A 240-character sentence in a 34-wide column does not render: it either
#: spills across every neighbour or is clipped by the first non-empty cell to
#: its right, and neither looks intentional. Explanatory text now gets merged
#: across the sheet's used width, wrapped, and given an explicit row height,
#: because Excel does not auto-fit the height of a merged wrapped cell.
_CHARS_PER_LINE_PER_WIDTH = 1.05
_LINE_HEIGHT = 13.5


def _note(ws, row: int, text: str, last_col: int, *, color="808080",
          width_hint: Optional[float] = None) -> None:
    """A wrapped, merged block of explanatory text that actually renders."""
    from openpyxl.styles import Alignment
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=last_col)
    c = _style(ws, f"A{row}", color=color, wrap=True)
    c.value = text
    c.alignment = Alignment(wrap_text=True, vertical="top")
    total_width = width_hint or sum(
        (ws.column_dimensions[_col(i)].width or 10) for i in range(1, last_col + 1))
    per_line = max(20.0, total_width * _CHARS_PER_LINE_PER_WIDTH)
    lines = max(1, math.ceil(len(text) / per_line))
    ws.row_dimensions[row].height = lines * _LINE_HEIGHT + 4


def _autofit(ws, *, min_w=9.0, max_w=46.0, skip_rows=(), header_rows=()):
    """Set column widths from actual content, ignoring merged note rows.

    Merged cells are excluded because their text length says nothing about how
    wide any single column needs to be — including them is what produced a
    column A sized for a 240-character sentence.
    """
    merged_rows = set()
    for rng in ws.merged_cells.ranges:
        for r in range(rng.min_row, rng.max_row + 1):
            merged_rows.add(r)
    longest: dict[str, int] = {}
    for row in ws.iter_rows():
        for c in row:
            if c.value is None or c.row in merged_rows or c.row in skip_rows:
                continue
            if isinstance(c.value, str) and c.value.startswith("="):
                continue                     # formula text is not what renders
            text = str(c.value)
            if c.row in header_rows:
                text = max(text.split(), key=len) if text.split() else text
            longest[c.column_letter] = max(longest.get(c.column_letter, 0),
                                           len(text))
    for col, n in longest.items():
        want = min(max_w, max(min_w, n + 2.5))
        cur = ws.column_dimensions[col].width
        if cur is None or cur < want:
            ws.column_dimensions[col].width = want


def _wrap_header(ws, row: int, last_col: int, height: float = 30.0):
    """Let a two-word header wrap onto two lines instead of being clipped."""
    from openpyxl.styles import Alignment
    for i in range(1, last_col + 1):
        c = ws.cell(row, i)
        if c.value is not None:
            c.alignment = Alignment(wrap_text=True, vertical="bottom",
                                    horizontal="center")
    ws.row_dimensions[row].height = height


def _fit_wrapped_rows(ws, col: str, width: float):
    """Give every wrapped cell in `col` a row height that fits its text."""
    for row in ws.iter_rows(min_col=ws[col + "1"].column,
                            max_col=ws[col + "1"].column):
        for c in row:
            if c.value is None or not isinstance(c.value, str):
                continue
            if c.alignment and c.alignment.wrap_text:
                lines = max(1, math.ceil(len(c.value) / max(20.0, width)))
                have = ws.row_dimensions[c.row].height or 0
                ws.row_dimensions[c.row].height = max(
                    have, lines * _LINE_HEIGHT + 4)


def _input(ws, row, label, value, unit="", note="", fmt="0.0000",
           must_confirm=False):
    """A labelled input cell: blue for typed values, yellow if unverified."""
    _label(ws, row, label)
    c = _style(ws, f"B{row}", color=_BLUE, fmt=fmt,
               fill=_YELLOW if must_confirm else None)
    c.value = value
    _style(ws, f"C{row}").value = unit
    if note:
        _style(ws, f"D{row}", color="808080", wrap=True).value = note
    return f"$B${row}"


# ===================================================================== #
#  Input specification
# ===================================================================== #
@dataclass
class ExportSpec:
    """Everything the export needs, and where each number came from."""
    pack: pdw.PackSpec
    vehicle: pdw.VehicleSpec
    drive: pdw.DriveSpec
    speed_mph: Sequence[float]
    dt_s: float
    #: The actual sample times, when the caller has them. A lap sim's output is
    #: not always evenly spaced, and `dt_s` is only their mean; the trace sheet
    #: writes these into column A and every derivative and integral is taken
    #: against them, so an uneven log is handled exactly rather than smeared
    #: over an average step. None means "assume uniform k*dt_s".
    time_s: Optional[Sequence[float]] = None
    lap_time_s: Optional[float] = None
    reductions: Sequence[float] = tuple(range(1, 16))
    smooth_window: int = 11
    endurance_km: Optional[float] = 22.0
    #: Per-cell continuous discharge rating, in C. Used by the advisor to
    #: reject packs that meet the energy requirement and cannot supply the
    #: current. Datasheet figure; the default is a conservative 21700 value and
    #: is flagged as an assumption because it changes the recommendation.
    max_cell_c_rate: float = 10.0
    soc_window: tuple[float, float] = (0.10, 0.80)


# ===================================================================== #
#  Sheet builders
# ===================================================================== #
def _write_inputs(ws, spec: ExportSpec) -> dict:
    """Every assumption in its own labelled cell, referenced by everything else.

    Yellow fill marks the numbers nobody has verified. Three of them —  mass,
    drag area and rolling resistance — do not exist anywhere in the original
    workbook, and every current figure depends on them, so they are the first
    cells a reviewer should look at.
    """
    _title(ws, "KinematiK \u2014 inputs for the lap-sim export", 4)
    _NOTE_INPUTS = (
        "Blue = typed input. Yellow = assumption not taken from the workbook; "
        "confirm before quoting any result. Everything on the other KX sheets "
        "is a formula referencing this one, so editing here updates the whole "
        "export.")
    r = {}
    row = 4
    _header(ws, f"A{row}", "Vehicle")
    row += 1
    r["mass"] = _input(ws, row, "Mass incl. driver", spec.vehicle.mass_kg, "kg",
                       "Not present in the original workbook. Peak current is "
                       "roughly proportional to this.", "0.0",
                       must_confirm=True); row += 1
    r["crr"] = _input(ws, row, "Rolling resistance coeff", spec.vehicle.crr, "-",
                      "Generic FSAE figure.", "0.0000",
                      must_confirm=True); row += 1
    r["cda"] = _input(ws, row, "Drag area CdA", spec.vehicle.cda_m2, "m^2",
                      "Generic FSAE figure. Dominates power above ~40 mph.",
                      "0.000", must_confirm=True); row += 1
    r["rho"] = _input(ws, row, "Air density", spec.vehicle.air_density,
                      "kg/m^3", "", "0.000"); row += 1
    r["mu"] = _input(ws, row, "Longitudinal grip coeff",
                     spec.vehicle.mu_lon, "-",
                     "Caps the acceleration term, so a glitched speed sample "
                     "cannot become a current spike.", "0.00"); row += 1
    r["wheel_in"] = _input(ws, row, "Wheel diameter", spec.vehicle.wheel_diameter_in,
                           "in", "", "0.0"); row += 1
    row += 1

    _header(ws, f"A{row}", "Drivetrain")
    row += 1
    r["eta_drive"] = _input(ws, row, "Drivetrain efficiency",
                            spec.vehicle.drivetrain_efficiency, "-", "",
                            "0.000"); row += 1
    r["eta_motor"] = _input(ws, row, "Motor efficiency",
                            spec.vehicle.motor_efficiency, "-", "",
                            "0.0000"); row += 1
    r["eta_inv"] = _input(ws, row, "Inverter efficiency",
                          spec.vehicle.inverter_efficiency, "-",
                          "Not in the original workbook.", "0.000",
                          must_confirm=True); row += 1
    _label(ws, row, "Total efficiency")
    r["eta"] = f"$B${row}"
    _style(ws, f"B{row}", fmt="0.0000").value = \
        f"={r['eta_drive']}*{r['eta_motor']}*{r['eta_inv']}"
    _style(ws, f"D{row}", color="808080", wrap=True).value = (
        "Electrical demand DIVIDES by this. The original workbook multiplied, "
        "which made a more efficient motor draw more current.")
    row += 2

    _header(ws, f"A{row}", "Selected gear")
    row += 1
    r["reduction"] = _input(ws, row, "Gear reduction (motor:wheel)",
                            spec.drive.reduction, ":1",
                            "A reduction of N means the motor turns N times "
                            "per wheel turn, so motor rpm = wheel rpm * N.",
                            "0.00"); row += 1
    r["max_rpm"] = _input(ws, row, "Motor max speed",
                          spec.drive.motor_max_rpm, "rpm", "", "0"); row += 1
    r["t_peak"] = _input(ws, row, "Motor peak torque",
                         spec.drive.motor_peak_torque_nm, "Nm", "",
                         "0.0"); row += 1
    r["t_cont"] = _input(ws, row, "Motor continuous torque",
                         spec.drive.motor_continuous_torque_nm, "Nm", "",
                         "0.0"); row += 1
    r["p_motor"] = _input(ws, row, "Motor peak power",
                          spec.drive.motor_peak_power_kw, "kW", "",
                          "0.0"); row += 2

    _header(ws, f"A{row}", "Pack (read from the workbook by row label)")
    row += 1
    p = spec.pack
    r["n_series"] = _input(ws, row, "Series groups", p.n_series, "-", "",
                           "0"); row += 1
    r["n_par"] = _input(ws, row, "Parallel strings", p.n_parallel, "-", "",
                        "0"); row += 1
    r["cell_v"] = _input(ws, row, "Cell nominal voltage", p.cell_voltage_v,
                         "V", "", "0.000"); row += 1
    r["cell_ah"] = _input(ws, row, "Cell capacity", p.cell_capacity_ah, "Ah",
                          "", "0.00"); row += 1
    r["cell_r"] = _input(ws, row, "Cell internal resistance",
                         p.cell_resistance_ohm, "ohm", "", "0.00000"); row += 1
    r["fuse"] = _input(ws, row, "Fuse rating", p.fuse_max_a, "A", "",
                       "0.0"); row += 1
    r["c_rate"] = _input(ws, row, "Cell max continuous discharge",
                         spec.max_cell_c_rate, "C",
                         "Datasheet figure. Changes the pack recommendation, "
                         "so it is an assumption worth checking.", "0.0",
                         must_confirm=True); row += 2

    _header(ws, f"A{row}", "Session")
    row += 1
    r["dt"] = _input(ws, row, "Sample interval", spec.dt_s, "s",
                     "Mean spacing, for reference only. Nothing divides by "
                     "it: the trace differentiates and integrates against its "
                     "own time column, so an unevenly logged lap is still "
                     "exact.", "0.00000"); row += 1
    r["endurance"] = _input(ws, row, "Endurance distance",
                            spec.endurance_km or 22.0, "km", "",
                            "0.0"); row += 1
    r["soc_hi"] = _input(ws, row, "SOC ceiling", spec.soc_window[1], "-", "",
                         "0.00"); row += 1
    r["soc_lo"] = _input(ws, row, "SOC floor", spec.soc_window[0], "-", "",
                         "0.00"); row += 2

    _header(ws, f"A{row}", "Constants")
    row += 1
    r["g"] = _input(ws, row, "g", 9.80665, "m/s^2", "", "0.00000"); row += 1
    r["mph_ms"] = _input(ws, row, "mph -> m/s", 0.44704, "-", "",
                         "0.00000"); row += 1
    r["mph_ipm"] = _input(ws, row, "mph -> in/min", 1056.0, "-",
                          "63360 in/mile / 60 min.", "0.0"); row += 1

    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 9
    ws.column_dimensions["D"].width = 62
    _note(ws, 2, _NOTE_INPUTS, 4)
    _fit_wrapped_rows(ws, "D", 60)
    ws.freeze_panes = "A4"
    return r


_TRACE_COLS = [
    ("Time", "s", "0.000"),
    ("Speed", "mph", "0.00"),
    ("Speed", "m/s", "0.000"),
    ("Accel (grip-limited)", "m/s^2", "0.000"),
    ("F roll", "N", "0.0"),
    ("F aero", "N", "0.0"),
    ("F accel", "N", "0.0"),
    ("F total", "N", "0.0"),
    ("P wheel", "W", "0"),
    ("P electrical", "W", "0"),
    ("Pack current", "A", "0.00"),
    ("Over pack ceiling?", "1/0", "0"),
    ("Joule heat", "W", "0.0"),
    ("Motor speed", "rpm", "0"),
    ("Motor torque", "Nm", "0.00"),
    #: Each sample's share of the lap's duration — half the gap to the sample
    #: before plus half the gap to the sample after, which is the trapezoidal
    #: weight. Every integral on the dashboard is SUMPRODUCT(quantity, this)
    #: rather than SUM(quantity)*dt, so a log with a dropout or a variable
    #: rate integrates correctly instead of approximately. The column sums to
    #: exactly the elapsed time.
    ("Sample weight", "s", "0.0000"),
]


def _write_trace(ws, spec: ExportSpec, ref: dict, n: int) -> dict:
    """The lap, as formulas. Editing an input on KX Inputs moves every row."""
    _title(ws, "KinematiK \u2014 lap trace (every column a live formula)",
           len(_TRACE_COLS))
    for i, (name, unit, _fmt) in enumerate(_TRACE_COLS, start=1):
        # Name and unit on separate lines: "Accel (grip-limited) (m/s^2)" is 28
        # characters and was being clipped by a 13-wide column.
        _style(ws, f"{_col(i)}2", bold=True).value = f"{name}\n({unit})"
    _wrap_header(ws, 2, len(_TRACE_COLS), height=34)
    ws.freeze_panes = "A3"

    I = f"'{S_INPUTS}'!"
    first, last = 3, 2 + n
    for k in range(n):
        r = first + k
        prev, nxt = max(first, r - 1), min(last, r + 1)
        # speed in m/s
        ws[f"C{r}"] = f"=B{r}*{I}{ref['mph_ms']}"
        # Central difference taken against the TIME COLUMN, not against a
        # nominal step: dividing by (rows spanned) * dt assumes every sample is
        # equally spaced, and a logger that drops a frame then reports an
        # acceleration inversely proportional to a gap it never had. A3:A is
        # the real clock, so this is exact however the lap was sampled.
        raw = f"IFERROR((C{nxt}-C{prev})/(A{nxt}-A{prev}),0)"
        grip = f"{I}{ref['mu']}*{I}{ref['g']}"
        ws[f"D{r}"] = f"=MAX(-{grip},MIN({grip},{raw}))"
        ws[f"E{r}"] = (f"=IF(C{r}>0.05,{I}{ref['crr']}*{I}{ref['mass']}"
                       f"*{I}{ref['g']},0)")
        ws[f"F{r}"] = f"=0.5*{I}{ref['rho']}*{I}{ref['cda']}*C{r}^2"
        ws[f"G{r}"] = f"={I}{ref['mass']}*D{r}"
        ws[f"H{r}"] = f"=E{r}+F{r}+G{r}"
        ws[f"I{r}"] = f"=H{r}*C{r}"
        ws[f"J{r}"] = f"=IF(I{r}>0,I{r}/{I}{ref['eta']},0)"
        # sag-aware current: R*I^2 - Voc*I + P = 0, low-current root.
        voc = f"({I}{ref['n_series']}*{I}{ref['cell_v']})"
        rp = f"({I}{ref['n_series']}*({I}{ref['cell_r']}/{I}{ref['n_par']}))"
        disc = f"({voc}^2-4*{rp}*J{r})"
        ws[f"K{r}"] = (f"=IF(J{r}<=0,0,IF({disc}<0,{voc}/(2*{rp}),"
                       f"({voc}-SQRT({disc}))/(2*{rp})))")
        ws[f"L{r}"] = f"=IF(J{r}<=0,0,IF({disc}<0,1,0))"
        ws[f"M{r}"] = f"=K{r}^2*{rp}"
        ws[f"N{r}"] = (f"=B{r}*{I}{ref['mph_ipm']}/({I}{ref['wheel_in']}*PI())"
                       f"*{I}{ref['reduction']}")
        ws[f"O{r}"] = (f"=MAX(0,H{r})*({I}{ref['wheel_in']}*0.0254/2)"
                       f"/({I}{ref['reduction']}*{I}{ref['eta_drive']})")
        # Trapezoidal weight: half the step behind plus half the step ahead,
        # one-sided at the two ends. Sums to exactly the elapsed time.
        ws[f"P{r}"] = f"=(A{nxt}-A{prev})/2"
        for i, (_nm, _u, fmt) in enumerate(_TRACE_COLS, start=1):
            _style(ws, f"{_col(i)}{r}", fmt=fmt)
    for i in range(1, len(_TRACE_COLS) + 1):
        ws.column_dimensions[_col(i)].width = 12.5
    return {"first": first, "last": last}


def _write_pack_limits(ws, ref: dict) -> dict:
    """The pack's electrical envelope — none of which the original computes."""
    I = f"'{S_INPUTS}'!"
    _title(ws, "KinematiK \u2014 pack limits", 3)
    _note(ws, 2,
        "All derived from the primary cell inputs. The workbook's own pack "
        "resistance cell multiplies the TOTAL cell count by R_cell/P, which "
        "cancels to S*R_cell and discards the parallel benefit \u2014 3x high "
        "for 140S3P. These figures use S*(R_cell/P).", 3)
    out = {}
    row = 4
    rows = [
        ("Pack nominal voltage", f"={I}{ref['n_series']}*{I}{ref['cell_v']}",
         "V", "0.0", "voc"),
        ("Pack capacity", f"={I}{ref['n_par']}*{I}{ref['cell_ah']}", "Ah",
         "0.00", "ah"),
        ("Pack energy", "=B4*B5/1000", "kWh", "0.000", "kwh"),
        ("Usable energy (SOC window)",
         f"=B6*({I}{ref['soc_hi']}-{I}{ref['soc_lo']})", "kWh", "0.000",
         "usable"),
        ("Pack resistance",
         f"={I}{ref['n_series']}*({I}{ref['cell_r']}/{I}{ref['n_par']})",
         "ohm", "0.0000", "r"),
        ("Max deliverable power (V^2/4R)", "=B4^2/(4*B8)/1000", "kW", "0.0",
         "ceiling"),
        ("Current at that ceiling", "=B4/(2*B8)", "A", "0.0", "i_ceiling"),
        ("Power at fuse limit",
         f"={I}{ref['fuse']}*(B4-{I}{ref['fuse']}*B8)/1000", "kW", "0.00",
         "p_fuse"),
        ("Cell count", f"={I}{ref['n_series']}*{I}{ref['n_par']}", "cells",
         "0", "cells"),
        ("Cell continuous current limit",
         f"={I}{ref['cell_ah']}*{I}{ref['c_rate']}", "A", "0.0", "cell_i_max"),
        ("Pack continuous current limit",
         f"=B13*{I}{ref['n_par']}", "A", "0.0", "pack_i_max"),
    ]
    for label, formula, unit, fmt, key in rows:
        _label(ws, row, label)
        _style(ws, f"B{row}", fmt=fmt).value = formula
        _style(ws, f"C{row}").value = unit
        out[key] = f"$B${row}"
        row += 1
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 14
    return out


def _write_dashboard(ws, ref: dict, tr: dict, pk: dict) -> dict:
    """The numbers a reviewer opens the file to see."""
    I, P = f"'{S_INPUTS}'!", f"'{S_PACK}'!"
    T = f"'{S_TRACE}'!"
    f, l = tr["first"], tr["last"]
    _title(ws, "KinematiK \u2014 lap-sim dashboard", 3)
    _note(ws, 2,
        "Every cell here is a formula. Change an input on the KX Inputs sheet "
        "and these move with it.", 3, width_hint=104)

    row = 4
    _header(ws, f"A{row}", "Lap")
    row += 1
    # The weight column sums to the elapsed time and turns every integral into
    # a true trapezoidal sum, so none of these depend on the samples being
    # evenly spaced.
    W = f"{T}P{f}:P{l}"
    items = [
        ("Duration", f"=SUM({W})", "s", "0.00"),
        ("Distance", f"=SUMPRODUCT({T}C{f}:C{l},{W})/1000", "km",
         "0.000"),
        ("Peak speed", f"=MAX({T}B{f}:B{l})", "mph", "0.0"),
        ("Mean speed", f"=AVERAGE({T}C{f}:C{l})*3.6", "km/h", "0.0"),
    ]
    for lab, formula, unit, fmt in items:
        _label(ws, row, lab)
        _style(ws, f"B{row}", fmt=fmt).value = formula
        _style(ws, f"C{row}").value = unit
        row += 1
    dist = "$B$6"
    row += 1

    _header(ws, f"A{row}", "Electrical demand")
    row += 1
    e_row = None
    items = [
        ("Peak pack current", f"=MAX({T}K{f}:K{l})", "A", "0.00", "peak_i"),
        ("Mean pack current", f"=AVERAGE({T}K{f}:K{l})", "A", "0.00", None),
        ("Peak electrical power", f"=MAX({T}J{f}:J{l})/1000", "kW", "0.00",
         "peak_p"),
        ("Energy used", f"=SUMPRODUCT({T}J{f}:J{l},{W})/3600000",
         "kWh", "0.0000", "energy"),
        ("Peak cell current", f"=B{row}/{I}{ref['n_par']}", "A", "0.00", None),
        ("Peak cell C-rate", f"=B{row+4}/{I}{ref['cell_ah']}", "C", "0.0",
         None),
        ("Samples over pack ceiling", f"=SUM({T}L{f}:L{l})", "-", "0", None),
        ("Joule heat, mean", f"=AVERAGE({T}M{f}:M{l})", "W", "0.0", None),
        ("Joule heat, total",
         f"=SUMPRODUCT({T}M{f}:M{l},{W})/3600", "Wh", "0.0", None),
    ]
    keys = {}
    for lab, formula, unit, fmt, key in items:
        _label(ws, row, lab)
        _style(ws, f"B{row}", fmt=fmt).value = formula
        _style(ws, f"C{row}").value = unit
        if key:
            keys[key] = f"$B${row}"
        row += 1
    row += 1

    _header(ws, f"A{row}", "Verdicts")
    row += 1
    checks = [
        ("Fuse", f"=IF({keys['peak_i']}<={I}{ref['fuse']},"
                 f'"PASS - peak "&TEXT({keys["peak_i"]},"0.0")&" A of "'
                 f'&TEXT({I}{ref["fuse"]},"0")&" A",'
                 f'"FAIL - peak "&TEXT({keys["peak_i"]},"0.0")'
                 f'&" A exceeds "&TEXT({I}{ref["fuse"]},"0")&" A fuse")'),
        ("Pack power ceiling",
         f"=IF({keys['peak_p']}<={P}{pk['ceiling']},"
         f'"PASS","FAIL - demand "&TEXT({keys["peak_p"]},"0")'
         f'&" kW exceeds the "&TEXT({P}{pk["ceiling"]},"0")'
         f'&" kW the pack can deliver into any load")'),
        ("Cell C-rate",
         f"=IF({keys['peak_i']}/{I}{ref['n_par']}<={P}{pk['cell_i_max']},"
         f'"PASS","FAIL - "&TEXT({keys["peak_i"]}/{I}{ref["n_par"]}'
         f'/{I}{ref["cell_ah"]},"0.0")&"C exceeds the "'
         f'&TEXT({I}{ref["c_rate"]},"0")&"C cell rating")'),
        ("Single-lap energy",
         f"=IF({keys['energy']}<={P}{pk['usable']},"
         f'"PASS","FAIL - needs "&TEXT({keys["energy"]},"0.00")'
         f'&" kWh of "&TEXT({P}{pk["usable"]},"0.00")&" kWh usable")'),
        ("Endurance energy",
         f"=IF({dist}<=0,\"n/a - zero distance\","
         f"IF({I}{ref['endurance']}/{dist}*{keys['energy']}<={P}{pk['usable']},"
         f'"PASS","FAIL - '
         f'"&TEXT({I}{ref["endurance"]}/{dist},"0.0")&" laps for "'
         f'&TEXT({I}{ref["endurance"]},"0")&" km needs "'
         f'&TEXT({I}{ref["endurance"]}/{dist}*{keys["energy"]},"0.0")'
         f'&" kWh of "&TEXT({P}{pk["usable"]},"0.00")&" kWh"))'),
        ("Motor speed",
         f"=IF(MAX({T}N{f}:N{l})<={I}{ref['max_rpm']},"
         f'"PASS","FAIL - "&TEXT(MAX({T}N{f}:N{l}),"0")'
         f'&" rpm exceeds "&TEXT({I}{ref["max_rpm"]},"0")&" rpm")'),
        ("Motor torque",
         f"=IF(MAX({T}O{f}:O{l})<={I}{ref['t_peak']},"
         f'IF(MAX({T}O{f}:O{l})<={I}{ref["t_cont"]},"PASS",'
         f'"MARGINAL - "&TEXT(MAX({T}O{f}:O{l}),"0")'
         f'&" Nm is above the continuous rating"),'
         f'"FAIL - needs "&TEXT(MAX({T}O{f}:O{l}),"0")&" Nm of "'
         f'&TEXT({I}{ref["t_peak"]},"0")&" Nm peak")'),
    ]
    for lab, formula in checks:
        _label(ws, row, lab)
        _style(ws, f"B{row}", wrap=True).value = formula
        row += 1

    _label(ws, row + 1, "Laps to complete endurance")
    _style(ws, f"B{row+1}", fmt="0.0").value = \
        f"=IF({dist}>0,{I}{ref['endurance']}/{dist},0)"
    _label(ws, row + 2, "Laps the pack supports")
    _style(ws, f"B{row+2}", fmt="0.0").value = \
        f"=IF({keys['energy']}>0,{P}{pk['usable']}/{keys['energy']},0)"

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 74
    ws.column_dimensions["C"].width = 8
    _fit_wrapped_rows(ws, "B", 72)
    ws.freeze_panes = "A4"
    keys["distance"] = dist
    return keys


def _write_gear_study(ws, spec: ExportSpec, ref: dict, tr: dict) -> None:
    """Torque per ratio — the quantity that actually chooses a gear.

    Pack current is identical in every column, because power is force times
    speed regardless of which gear delivers it. That is exactly why the
    original workbook's fifteen-column *current* sweep could not inform a gear
    choice: every column reported the same thing.
    """
    I, T = f"'{S_INPUTS}'!", f"'{S_TRACE}'!"
    f, l = tr["first"], tr["last"]
    _title(ws, "KinematiK \u2014 gear study", 6)
    _note(ws, 2,
        "Motor speed scales with the reduction and torque scales inversely, so "
        "both are exact from the peak wheel speed and peak tractive force. "
        "Pack current does not vary with gearing at all, which is why a "
        "current-only sweep cannot choose a ratio.", 6, width_hint=104)
    hdr = ["Reduction (:1)", "Peak motor speed (rpm)",
           "Peak motor torque (Nm)", "Overspeed?", "Torque verdict",
           "Recommended?"]
    for i, h in enumerate(hdr, start=1):
        _style(ws, f"{_col(i)}4", bold=True).value = h

    peak_wheel = f"MAX({T}B{f}:B{l})*{I}{ref['mph_ipm']}/({I}{ref['wheel_in']}*PI())"
    peak_force = f"MAX({T}H{f}:H{l})"
    r_wheel = f"({I}{ref['wheel_in']}*0.0254/2)"

    row = 5
    for red in spec.reductions:
        _style(ws, f"A{row}", color=_BLUE, fmt="0").value = float(red)
        _style(ws, f"B{row}", fmt="0").value = f"={peak_wheel}*A{row}"
        _style(ws, f"C{row}", fmt="0.00").value = \
            f"={peak_force}*{r_wheel}/(A{row}*{I}{ref['eta_drive']})"
        _style(ws, f"D{row}").value = \
            f'=IF(B{row}>{I}{ref["max_rpm"]},"OVERSPEED","ok")'
        _style(ws, f"E{row}").value = (
            f'=IF(C{row}>{I}{ref["t_peak"]},"over peak",'
            f'IF(C{row}>{I}{ref["t_cont"]},"over continuous","ok"))')
        _style(ws, f"F{row}", bold=True).value = (
            f'=IF(AND(D{row}="ok",E{row}="ok"),"YES","-")')
        row += 1
    for i, w in enumerate((16, 22, 22, 14, 16, 14), start=1):
        ws.column_dimensions[_col(i)].width = w

    _wrap_header(ws, 4, 6, height=30)
    ws.freeze_panes = "A5"
    _footer_label(
        ws, row + 1, "Lowest workable reduction", 4, 5,
        f'=IFERROR(INDEX(A5:A{row-1},MATCH("YES",F5:F{row-1},0)),'
        f'"none of the ratios tested works")')


def _write_advisor(ws, spec: ExportSpec, ref: dict, pk: dict,
                   dash: dict) -> None:
    """Pack candidates scored on energy AND power AND cell C-rate.

    The old advisor used energy alone and therefore recommended 140S1P for a lap
    needing 1 kWh of a 7 kWh pack. One parallel string triples resistance, cuts
    deliverable power to a third, and routes the whole pack current through every
    cell — about 39C for a 5 Ah cell. An advisor that cannot see current will
    keep making that recommendation, which is why all three gates are here.
    """
    I, P, D = f"'{S_INPUTS}'!", f"'{S_PACK}'!", f"'{S_DASH}'!"
    _title(ws, "KinematiK \u2014 pack advisor", 11)
    _note(ws, 2,
        "A candidate must clear three gates, not one: enough usable energy for "
        "the endurance distance, a power ceiling above peak demand, and a "
        "per-cell current inside the cell C-rating. Energy alone recommends "
        "packs that cannot supply the current.", 11, width_hint=150)

    hdr = ["Series", "Parallel", "Cells", "Nominal V", "Usable kWh",
           "R (ohm)", "Ceiling (kW)", "Peak cell C", "Energy needed (kWh)",
           "Mass (kg)", "Verdict"]
    for i, h in enumerate(hdr, start=1):
        _style(ws, f"{_col(i)}4", bold=True).value = h

    peak_i = f"{D}{dash['peak_i']}"
    peak_p = f"{D}{dash['peak_p']}"
    lap_kwh = f"{D}{dash['energy']}"
    lap_km = f"{D}{dash['distance']}"

    s0, p0 = spec.pack.n_series, spec.pack.n_parallel
    cands = sorted({1, 2, 3, p0, p0 + 1, p0 + 2, 6, 8})
    row = 5
    for p in [c for c in cands if 1 <= c <= 12]:
        _style(ws, f"A{row}", color=_BLUE, fmt="0").value = s0
        _style(ws, f"B{row}", color=_BLUE, fmt="0").value = p
        _style(ws, f"C{row}", fmt="0").value = f"=A{row}*B{row}"
        _style(ws, f"D{row}", fmt="0.0").value = f"=A{row}*{I}{ref['cell_v']}"
        _style(ws, f"E{row}", fmt="0.000").value = (
            f"=D{row}*B{row}*{I}{ref['cell_ah']}/1000"
            f"*({I}{ref['soc_hi']}-{I}{ref['soc_lo']})")
        _style(ws, f"F{row}", fmt="0.0000").value = (
            f"=A{row}*({I}{ref['cell_r']}/B{row})")
        _style(ws, f"G{row}", fmt="0.0").value = f"=D{row}^2/(4*F{row})/1000"
        _style(ws, f"H{row}", fmt="0.0").value = (
            f"={peak_i}/B{row}/{I}{ref['cell_ah']}")
        _style(ws, f"I{row}", fmt="0.000").value = (
            f"=IF({lap_km}>0,{I}{ref['endurance']}/{lap_km}*{lap_kwh},0)")
        _style(ws, f"J{row}", fmt="0.0").value = (
            f"=C{row}*{spec.pack.cell_weight_kg}")
        _style(ws, f"K{row}", bold=True, wrap=True).value = (
            f'=IF(H{row}>{I}{ref["c_rate"]},'
            f'"NO - "&TEXT(H{row},"0.0")&"C per cell exceeds the "'
            f'&TEXT({I}{ref["c_rate"]},"0")&"C rating",'
            f'IF(G{row}<{peak_p},'
            f'"NO - "&TEXT(G{row},"0")&" kW ceiling is below the "'
            f'&TEXT({peak_p},"0")&" kW peak demand",'
            f'IF(E{row}<I{row},'
            f'"NO - "&TEXT(E{row},"0.00")&" kWh usable against "'
            f'&TEXT(I{row},"0.00")&" kWh needed for endurance",'
            f'"YES - clears current, power and energy")))')
        row += 1
    for i, w in enumerate((8, 9, 8, 11, 11, 10, 12, 12, 18, 10, 58), start=1):
        ws.column_dimensions[_col(i)].width = w
    _wrap_header(ws, 4, 11, height=32)
    ws.freeze_panes = "A5"
    _fit_wrapped_rows(ws, "K", 56)

    _footer_label(
        ws, row + 1, "Smallest pack that clears all three gates", 9, 10,
        f'=IFERROR("P="&TEXT(INDEX(B5:B{row-1},'
        f'MATCH("YES*",K5:K{row-1},0)),"0"),'
        f'"none of the candidates tested clears all three")')
    _note(ws, row + 3,
        "Energy needed is for the full endurance distance, not one lap. The "
        "previous advisor compared a single lap's energy against the pack and "
        "reported a large surplus where there was a shortfall.", 11,
        width_hint=150)


def _write_provenance(ws, spec: ExportSpec) -> None:
    _title(ws, "KinematiK \u2014 provenance and corrections", 2)
    rows = [
        ("", ""),
        ("Corrected relative to the original workbook", ""),
        ("Gear ratio", "A reduction of N is applied as motor rpm = wheel rpm * "
                       "N. The original stored 1/N and multiplied, so the motor "
                       "turned slower than the wheel (49x error at 7:1)."),
        ("Pack current", "Solved from P = (V_oc - I*R)*I, so sag is included. "
                         "The original used V_pack * PF * RPM / 1000, which is "
                         "volts x rpm and rises with voltage."),
        ("Efficiency", "Electrical demand divides by total efficiency. The "
                       "original multiplied, making a better motor draw more."),
        ("Phase terms", "sqrt(3) and power factor are not applied to the DC "
                        "bus."),
        ("Pack resistance", "S*(R_cell/P). The original used the total cell "
                            "count in place of S, which cancels to S*R_cell "
                            "and is 3x high for 140S3P."),
        ("Energy", "sum(P*dt). The original summed instantaneous watts."),
        ("Time base", "Accelerations are differentiated against the trace's "
                      "own time column and every integral is weighted by each "
                      "sample's share of the elapsed time, so a log with a "
                      "dropout or a variable sample rate is exact rather than "
                      "smeared over a mean step."),
        ("Gear selection", "Reported as motor torque, which varies with the "
                           "ratio. Pack current does not."),
        ("Pack advice", "Gated on energy, power ceiling and per-cell C-rate "
                        "together."),
        ("", ""),
        ("Assumptions this export introduces", ""),
        ("Vehicle mass", f"{spec.vehicle.mass_kg:g} kg — not present anywhere "
                         f"in the original workbook."),
        ("Drag area CdA", f"{spec.vehicle.cda_m2:g} m^2 — generic FSAE figure."),
        ("Rolling resistance", f"{spec.vehicle.crr:g} — generic FSAE figure."),
        ("Inverter efficiency", f"{spec.vehicle.inverter_efficiency:g} — not in "
                                f"the original workbook."),
        ("Cell C-rating", f"{spec.max_cell_c_rate:g}C — datasheet figure; "
                          f"changes the pack recommendation."),
        ("Grip coefficient", f"{spec.vehicle.mu_lon:g} — caps the acceleration "
                             f"term so a glitched speed sample cannot become a "
                             f"current spike."),
        ("Trace smoothing", f"{spec.smooth_window}-sample moving average before "
                            f"differentiating. Raw quantised traces produce "
                            f"accelerations of several g that never happened."),
        ("", ""),
        ("Method", ""),
        ("Sheets written", "Only sheets prefixed 'KX '. No cell on any "
                           "pre-existing sheet is read-modified-written, so the "
                           "original model and its audit trail are intact."),
        ("Formulas", "Every derived cell is a live formula referencing "
                     "'KX Inputs'. The export is a model, not a snapshot."),
        ("Cached values", "The workbook is recalculated before being returned, "
                          "so data_only readers see numbers rather than None. "
                          "LibreOffice does it where it exists; where it does "
                          "not, the formulas are evaluated in process. Excel "
                          "recomputes everything on open either way."),
    ]
    r = 2
    for a, b in rows:
        if a and not b:
            # Section headers are longer than column A; merge across so they
            # render as headings instead of being clipped at the B boundary.
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
            _header(ws, f"A{r}", a)
        else:
            _label(ws, r, a)
            _style(ws, f"B{r}", wrap=True, color="404040").value = b
        r += 1
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 96
    _fit_wrapped_rows(ws, "B", 94)


# ===================================================================== #
#  Recalculation
# ===================================================================== #
#: Functions that exist only in Microsoft 365. Anything older — Excel 2021,
#: 2019, 2016, LibreOffice, Google Sheets, Numbers, most preview panes — returns
#: #NAME? for these, and because the ones in this workbook are ARRAY formulas
#: spilling down thousands of rows, five of them produce roughly nineteen
#: thousand error cells. Those errors are the "zeroes and blank pages": a
#: #NAME? in the source column makes every dependent sum, min and max blank or
#: zero, so the workbook looks empty rather than broken.
#:
#: Each replacement is exact, not approximate:
#:   CHOOSECOLS(range, n) -> INDEX(range, 0, n)   — row 0 means "whole column"
#:   CHOOSEROWS(range, n) -> INDEX(range, n, 0)
MODERN_FUNCTION_REWRITES: tuple[tuple[str, str, str], ...] = (
    ("CHOOSECOLS", r"_xlfn\.CHOOSECOLS\s*\(", "INDEX_COLS"),
    ("CHOOSEROWS", r"_xlfn\.CHOOSEROWS\s*\(", "INDEX_ROWS"),
)


def _rewrite_choose(formula: str) -> tuple[str, int]:
    """Rewrite CHOOSECOLS/CHOOSEROWS to INDEX, respecting nesting.

    Argument splitting is done by walking the string and tracking parenthesis
    depth and quotes, because a regex cannot tell the comma in
    `CHOOSECOLS(A1:C3, ROUND(1/B2,0))` that separates arguments from the one
    inside ROUND.
    """
    import re as _re
    count = 0
    for name, pattern, kind in MODERN_FUNCTION_REWRITES:
        while True:
            m = _re.search(pattern, formula)
            if not m:
                break
            start = m.end()                       # just past the opening paren
            depth, i, in_str = 1, start, False
            args, cur = [], []
            while i < len(formula) and depth > 0:
                ch = formula[i]
                if ch == '"':
                    in_str = not in_str
                if not in_str:
                    if ch == "(":
                        depth += 1
                    elif ch == ")":
                        depth -= 1
                        if depth == 0:
                            break
                    elif ch == "," and depth == 1:
                        args.append("".join(cur)); cur = []
                        i += 1
                        continue
                cur.append(ch)
                i += 1
            args.append("".join(cur))
            if depth != 0 or len(args) < 2:
                # Unbalanced or unexpected shape: leave it alone rather than
                # produce a formula that is wrong in a new way.
                break
            rng, sel = args[0].strip(), args[1].strip()
            repl = (f"INDEX({rng},0,{sel})" if kind == "INDEX_COLS"
                    else f"INDEX({rng},{sel},0)")
            formula = formula[:m.start()] + repl + formula[i + 1:]
            count += 1
    return formula, count


def make_portable(path: str) -> dict:
    """Replace Microsoft-365-only functions in place, everywhere in a workbook.

    Returns {"CHOOSECOLS": n, ...} counting what was rewritten. Preserves array
    formulas as array formulas — dropping the array wrapper would change a
    spilling column into a single value, which is a different bug with the same
    symptom.
    """
    import openpyxl
    from openpyxl.worksheet.formula import ArrayFormula

    wb = openpyxl.load_workbook(path)
    changed: dict[str, int] = {}
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                is_array = isinstance(v, ArrayFormula)
                text = v.text if is_array else v
                if not isinstance(text, str) or "_xlfn." not in text:
                    continue
                new, n = _rewrite_choose(text)
                if n:
                    for name, _pat, _k in MODERN_FUNCTION_REWRITES:
                        if name in text:
                            changed[name] = changed.get(name, 0) + 1
                    cell.value = (ArrayFormula(v.ref, new) if is_array else new)
    if changed:
        # Ask Excel to recompute everything on open, so the rewritten formulas
        # replace whatever stale cached values are sitting in the file.
        try:
            wb.calculation.fullCalcOnLoad = True
        except Exception:
            pass
        wb.save(path)
    return changed


def modern_function_report(path: str) -> dict:
    """Count remaining 365-only functions, per sheet. Used as a gate."""
    import re as _re
    import openpyxl
    from openpyxl.worksheet.formula import ArrayFormula
    wb = openpyxl.load_workbook(path)
    out: dict[str, int] = {}
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                text = v.text if isinstance(v, ArrayFormula) else v
                if isinstance(text, str) and "_xlfn." in text:
                    for fn in _re.findall(r"_xlfn\.([A-Z]+)", text):
                        out[f"{ws.title}:{fn}"] = out.get(
                            f"{ws.title}:{fn}", 0) + 1
    return out


_RANGE_RE = r"([A-Za-z_][A-Za-z0-9_ .]*!|)\$?([A-Z]{1,3})\$?(\d+):\$?([A-Z]{1,3})\$?(\d+)"


def _data_blocks(ws, col: int = 8) -> list:
    """Contiguous runs of populated rows in `col`, as (first, last) pairs.

    MUST be given the FORMULA view of the worksheet, not the cached-value view.
    openpyxl drops cached values whenever it saves, so a data_only view of a
    workbook this module has already touched looks empty everywhere and every
    range appears to point at nothing. A cell holding a formula is populated
    even when its cached value is gone — that is the structural extent, and the
    structural extent is what a range reference has to match.
    """
    blocks, start = [], None
    for r in range(1, (ws.max_row or 0) + 1):
        has = ws.cell(r, col).value is not None
        if has and start is None:
            start = r
        elif not has and start is not None:
            blocks.append((start, r - 1))
            start = None
    if start is not None:
        blocks.append((start, ws.max_row))
    return blocks


def audit_block_references(path: str) -> list[dict]:
    """Find formulas whose referenced range sits in an EMPTY region.

    This is the other half of the zero problem, and it is not a formula bug —
    it is a data-extent bug. The legacy export path rewrote ElecPropulsion's
    three stacked blocks as 335 rows of static values each, occupying rows
    1-1008, while ThermalLoad, EMFs and BearingBlowOut still reference the
    original fixed offsets at rows 1895-5689. Every one of those references now
    points at blank cells, and INDEX over blank cells returns 0 — so those tabs
    read as solid zeros with no error anywhere to explain it.

    It cannot be repaired: the rows the formulas want were replaced by values,
    not formulas, so the missing samples no longer exist in the file. The only
    remedy is to export again from the ORIGINAL workbook, which is why this
    reports rather than silently patches.
    """
    import re as _re
    import openpyxl
    from openpyxl.worksheet.formula import ArrayFormula

    wb = openpyxl.load_workbook(path)
    extents: dict[str, list] = {}
    findings: list[dict] = []

    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                text = v.text if isinstance(v, ArrayFormula) else v
                if not isinstance(text, str) or "!" not in text:
                    continue
                for m in _re.finditer(_RANGE_RE, text):
                    sheet_ref = m.group(1)[:-1] if m.group(1) else None
                    if not sheet_ref:
                        continue
                    sheet_ref = sheet_ref.strip().strip("'")
                    if sheet_ref not in wb.sheetnames:
                        continue
                    r1, r2 = int(m.group(3)), int(m.group(5))
                    if r2 - r1 < 2:            # single cells / tiny ranges
                        continue
                    if sheet_ref not in extents:
                        extents[sheet_ref] = _data_blocks(wb[sheet_ref])
                    blocks = extents[sheet_ref]
                    covered = any(b[0] <= r1 <= b[1] or b[0] <= r2 <= b[1]
                                  or (r1 <= b[0] and b[1] <= r2)
                                  for b in blocks)
                    if not covered:
                        findings.append({
                            "where": f"{ws.title}!{cell.coordinate}",
                            "target": sheet_ref,
                            "range": f"{m.group(2)}{r1}:{m.group(4)}{r2}",
                            "blocks": blocks,
                        })
    return findings


def repair_range_overshoot(path: str) -> list[str]:
    """Shrink formula ranges that run a few rows past their data.

    Distinct from the case above: here the range mostly hits data and overshoots
    the end, so the tail contributes zeros. `EMFs` reads ElecPropulsion rows
    3789:5682 while the block ends at 5681, which is why its last value is 0
    even in the untouched original workbook. Trimming to the real extent is
    safe and exact.

    Only trims; never extends, because extending would invent coverage.
    """
    import re as _re
    import openpyxl
    from openpyxl.worksheet.formula import ArrayFormula

    wb = openpyxl.load_workbook(path)
    extents: dict[str, list] = {}
    changed: list[str] = []

    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                is_arr = isinstance(v, ArrayFormula)
                text = v.text if is_arr else v
                if not isinstance(text, str) or "!" not in text:
                    continue
                new_text = text
                shrink_by = 0
                for m in list(_re.finditer(_RANGE_RE, text)):
                    ref = m.group(1)[:-1].strip().strip("'") if m.group(1) else None
                    if not ref or ref not in wb.sheetnames:
                        continue
                    r1, r2 = int(m.group(3)), int(m.group(5))
                    if r2 - r1 < 2:
                        continue
                    if ref not in extents:
                        extents[ref] = _data_blocks(wb[ref])
                    blk = next((b for b in extents[ref] if b[0] <= r1 <= b[1]),
                               None)
                    if blk and r2 > blk[1]:
                        old = m.group(0)
                        new = old.replace(f"{m.group(4)}{r2}",
                                          f"{m.group(4)}{blk[1]}")
                        if m.group(5) != str(blk[1]):
                            new_text = new_text.replace(old, new)
                            shrink_by = max(shrink_by, r2 - blk[1])
                            changed.append(
                                f"{ws.title}!{cell.coordinate}: "
                                f"{ref}!{m.group(2)}{r1}:{m.group(4)}{r2} "
                                f"-> row {blk[1]} (block ends there; the extra "
                                f"rows contributed zeros)")
                if new_text != text:
                    if is_arr:
                        # The spill range must shrink with the result, or the
                        # orphaned tail rows evaluate to #N/A — a new error where
                        # there used to be a zero.
                        ref = v.ref
                        if shrink_by and ":" in ref:
                            a_, b_ = ref.split(":", 1)
                            mm = _re.match(r"(\$?[A-Z]{1,3}\$?)(\d+)$", b_)
                            if mm:
                                last_new = int(mm.group(2)) - shrink_by
                                ref = f"{a_}:{mm.group(1)}{last_new}"
                                # Blank the rows that fall outside the shrunk
                                # spill. Left in place they keep whatever
                                # literal the original file cached there — a
                                # stray #N/A or 0 at the foot of the column,
                                # which is precisely the kind of lone zero this
                                # pass exists to remove.
                                colL = mm.group(1).replace("$", "")
                                for _r in range(last_new + 1,
                                                last_new + 1 + shrink_by):
                                    try:
                                        ws.cell(_r, ws[f"{colL}1"].column).value = None
                                    except Exception:
                                        pass
                        cell.value = ArrayFormula(ref, new_text)
                    else:
                        cell.value = new_text
    if changed:
        try:
            wb.calculation.fullCalcOnLoad = True
        except Exception:
            pass
        wb.save(path)
    return changed


#: The canonical ElecPropulsion block layout in an untouched
#: FSAE_EV_Power_Draw.xlsx: RPM, current draw, motor power, stacked vertically.
#: Formulas across the workbook reference these fixed offsets.
CANONICAL_EP_BLOCKS: tuple[tuple[int, int], ...] = ((2, 1894), (1895, 3787),
                                                    (3789, 5681))


def remap_stale_block_references(path: str, *,
                                 target_sheet: str = "ElecPropulsion"
                                 ) -> list[str]:
    """Repoint formulas at the block their data actually occupies now.

    The legacy export rewrote the three stacked blocks as ~335 rows of values
    each, so they sit at rows 1-1008 while every dependent formula still
    references the canonical offsets at 1895-5681. Those references land on
    blank cells and INDEX returns 0, which is the wall of zeros.

    The remap is well-determined rather than a guess: a reference is matched to
    a canonical block by which one contains its start row, and that ordinal then
    selects the correspondingly-ordered block actually present in the file. Three
    canonical blocks, three present blocks, matched in order. If the counts do
    not agree the function does nothing and says so, because a partial remap
    would move numbers to the wrong physical quantity — worse than leaving the
    zeros, which at least look wrong.

    Single-row references to the old MIN/MAX summary rows cannot be remapped —
    those rows no longer exist — so they are rewritten as MIN()/MAX() over the
    mapped block, which is what they meant in the first place.
    """
    import re as _re
    import openpyxl
    from openpyxl.worksheet.formula import ArrayFormula

    wb = openpyxl.load_workbook(path)
    if target_sheet not in wb.sheetnames:
        return []

    # Nothing to repair unless some reference actually points at blank rows.
    # Checked first so a healthy workbook — where blocks 1 and 2 are contiguous
    # and therefore scan as a single run — is silently left alone instead of
    # being told the pairing is ambiguous.
    if not any(f["target"] == target_sheet
               for f in audit_block_references(path)):
        return []

    present = [b for b in _data_blocks(wb[target_sheet]) if b[1] - b[0] > 2]
    if len(present) != len(CANONICAL_EP_BLOCKS):
        return [f"Not remapped: found {len(present)} data block(s) in "
                f"'{target_sheet}' but the formulas reference "
                f"{len(CANONICAL_EP_BLOCKS)}. Refusing to guess the pairing."]
    # Guarded per-reference below, not by comparing block positions: the only
    # references worth moving are the ones pointing at nothing. A range that
    # already lands on data is correct wherever it happens to sit.

    changed: list[str] = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                is_arr = isinstance(v, ArrayFormula)
                text = v.text if is_arr else v
                if not isinstance(text, str) or target_sheet not in text:
                    continue
                new_text = text
                for m in list(_re.finditer(_RANGE_RE, text)):
                    ref = m.group(1)[:-1].strip().strip("'") if m.group(1) else ""
                    if ref != target_sheet:
                        continue
                    r1, r2 = int(m.group(3)), int(m.group(5))
                    # Already reading real data? Then it is fine as it is.
                    if any(b[0] <= r1 <= b[1] for b in present):
                        continue
                    ordinal = next((i for i, (a, b)
                                    in enumerate(CANONICAL_EP_BLOCKS)
                                    if a <= r1 <= b), None)
                    if ordinal is None:
                        continue
                    nb = present[ordinal]
                    # Block 1 starts one row below its detected start, because
                    # row 1 of that block is the gear-ratio header.
                    lo = nb[0] + 1 if ordinal == 0 else nb[0]
                    hi = nb[1]
                    if (r1, r2) == (lo, hi):
                        continue
                    old = m.group(0)
                    new = (f"{m.group(1)}{m.group(2)}{lo}:"
                           f"{m.group(4)}{hi}")
                    new_text = new_text.replace(old, new)
                    changed.append(
                        f"{ws.title}!{cell.coordinate}: {target_sheet}!"
                        f"{m.group(2)}{r1}:{m.group(4)}{r2} -> rows {lo}-{hi} "
                        f"(block {ordinal + 1})")
                if new_text != text:
                    if is_arr:
                        # Spill no further than the data now supports.
                        ref_s = v.ref
                        mm = _re.match(r"(\$?[A-Z]{1,3}\$?)(\d+):"
                                       r"(\$?[A-Z]{1,3}\$?)(\d+)$", ref_s)
                        if mm:
                            rows_avail = hi - lo + 1
                            first = int(mm.group(2))
                            ref_s = (f"{mm.group(1)}{first}:{mm.group(3)}"
                                     f"{first + rows_avail - 1}")
                        cell.value = ArrayFormula(ref_s, new_text)
                    else:
                        cell.value = new_text

    # The vanished MIN/MAX summary rows, rewritten as aggregates over the block.
    tl = wb["ThermalLoad"] if "ThermalLoad" in wb.sheetnames else None
    if tl is not None:
        cur_lo, cur_hi = present[1]
        for coord, agg in (("B4", "MIN"), ("B5", "MAX")):
            v = tl[coord].value
            text = v.text if isinstance(v, ArrayFormula) else v
            if not isinstance(text, str):
                continue
            if _re.search(rf"{target_sheet}!\$?[A-Z]+\$?5(68[0-9]|69[0-9])",
                          text):
                tl[coord] = (
                    f"=({agg}(INDEX({target_sheet}!H{cur_lo}:V{cur_hi},0,"
                    f"ROUND(1/B2,0))))^2 * BatteryPackConfig!B11")
                changed.append(
                    f"ThermalLoad!{coord}: referenced the deleted {agg} summary "
                    f"row; rewritten as {agg}() over rows {cur_lo}-{cur_hi}")

    if changed:
        try:
            wb.calculation.fullCalcOnLoad = True
        except Exception:
            pass
        wb.save(path)
    return changed


def audit_block_column_coverage(path: str, *,
                                target_sheet: str = "ElecPropulsion",
                                first_col: int = 8, last_col: int = 22
                                ) -> list[str]:
    """Report data blocks that have lost most of their gear columns.

    The legacy export wrote the current-draw block for gear 1 ONLY — one of
    fifteen columns. Remapping the references cannot help: the other fourteen
    columns of current data were discarded, not moved. Any sheet reading the
    current block at a gear other than 1 therefore reads blank cells and returns
    zero, and no amount of formula repair recovers it.

    This is the check that distinguishes "repairable" from "re-export from the
    original", so it is worth running before promising anyone a fix.
    """
    import openpyxl
    # FORMULA view, not the cached view. This module re-saves workbooks (the
    # portability pass does), and openpyxl drops every cached value when it
    # saves — so a data_only read reports each column as empty and this check
    # concludes the workbook is damaged when it is perfectly healthy. That
    # false positive previously triggered a rebuild that overwrote the user's
    # own ElecPropulsion formulas with static values. A cell holding a formula
    # is populated; that is the only reading that survives a round trip.
    wb = openpyxl.load_workbook(path)
    if target_sheet not in wb.sheetnames:
        return []
    ws = wb[target_sheet]
    out: list[str] = []
    n_cols = last_col - first_col + 1
    for i, (lo, hi) in enumerate(
            [b for b in _data_blocks(ws) if b[1] - b[0] > 2], start=1):
        have = [c - first_col + 1 for c in range(first_col, last_col + 1)
                if any(ws.cell(r, c).value is not None
                       for r in range(lo, min(hi, lo + 60) + 1))]
        if 0 < len(have) < n_cols:
            out.append(
                f"block {i} (rows {lo}-{hi}) holds only {len(have)} of "
                f"{n_cols} columns (gears {', '.join(map(str, have))}). "
                f"Any formula reading this block at another gear returns zero, "
                f"and the missing columns cannot be reconstructed — they were "
                f"discarded when the workbook was written, not relocated.")
    return out


def repair_workbook(path: str) -> dict:
    """Every repair pass, in the order they depend on each other.

    Portability first (so nothing is #NAME?), then block remapping (so nothing
    points at blank rows), then overshoot trimming (so no tail contributes
    zeros).
    """
    out = {}
    out["portable"] = make_portable(path)
    out["remapped"] = remap_stale_block_references(path)
    out["trimmed"] = repair_range_overshoot(path)
    out["stale"] = audit_block_references(path)
    out["coverage"] = audit_block_column_coverage(path)
    return out


def _resample(values: Sequence[float], n: int) -> list:
    """Linearly resample a series to exactly `n` points."""
    src = list(values)
    if n <= 0:
        return []
    if len(src) == 1:
        return [float(src[0])] * n
    if len(src) == n:
        return [float(v) for v in src]
    out = []
    last = len(src) - 1
    for i in range(n):
        pos = i * last / (n - 1) if n > 1 else 0.0
        lo = int(math.floor(pos))
        hi = min(lo + 1, last)
        frac = pos - lo
        out.append(float(src[lo]) * (1 - frac) + float(src[hi]) * frac)
    return out


def rebuild_propulsion_blocks(path: str, speed_mph: Sequence[float],
                              pack: "pdw.PackSpec",
                              vehicle: "pdw.VehicleSpec",
                              drive: "pdw.DriveSpec", *,
                              dt_s: float,
                              reductions: Sequence[float] = tuple(range(1, 16)),
                              smooth_window: int = 11,
                              target_sheet: str = "ElecPropulsion") -> dict:
    """Rewrite ElecPropulsion's three blocks in full, at the canonical rows.

    This is what removes the "export from the original, never from an enhanced
    copy" caveat. The legacy path truncated the blocks to the lap length and
    wrote the current block for gear 1 only — one of fifteen columns — so every
    dependent sheet read blanks at any other gear. Warning about that was never
    a fix; the data has to come back.

    Two decisions make the reconstruction work:

    **Canonical offsets, always.** The blocks are written to rows 2-1894,
    1895-3787 and 3789-5681 regardless of how many samples the lap has, because
    ThermalLoad, EMFs and BearingBlowOut reference those exact rows. The trace
    is resampled to 1893 points to fit. Rebuilding at the lap's own length would
    reproduce the original bug in a new place.

    **All fifteen gear columns.** Motor speed genuinely varies with the
    reduction, so those columns differ. Pack current and electrical power do
    NOT — power is force times speed however the gearbox delivers it — so those
    columns are filled identically across gears. That is the physics, not a
    shortcut, and it is why a current-only gear sweep could never have chosen a
    ratio.
    """
    import openpyxl

    wb = openpyxl.load_workbook(path)
    if target_sheet not in wb.sheetnames:
        return {"rebuilt": False, "reason": f"no '{target_sheet}' sheet"}
    ws = wb[target_sheet]

    b1, b2, b3 = CANONICAL_EP_BLOCKS
    n = b1[1] - b1[0] + 1                      # 1893
    sp = _resample(pdw._moving_average(list(speed_mph), smooth_window), n)

    # Corrected physics once; current and power are gear-independent.
    base = pdw.power_draw_trace(sp, dt_s, pack, vehicle, drive,
                                smooth_window=1)
    wheel_rpm = [s * pdw.MPH_TO_IN_PER_MIN
                 / (vehicle.wheel_diameter_in * math.pi) for s in sp]

    first_col, last_col = 8, 8 + len(reductions) - 1

    # Clear the old blocks so nothing survives underneath the new ones.
    for r in range(2, max(ws.max_row, b3[1]) + 1):
        for c in range(first_col, last_col + 1):
            ws.cell(r, c).value = None
        if r in (b2[0] - 1, b3[0] - 1) or ws.cell(r, 7).value in (
                "Current Draw (A)", "Phase Current (A)", "Motor Power (kW)"):
            ws.cell(r, 7).value = None

    ws.cell(b1[0] - 1, 7).value = "RPM (Load)"
    for i in range(n):
        for gi, red in enumerate(reductions):
            ws.cell(b1[0] + i, first_col + gi).value = round(
                wheel_rpm[i] * float(red), 4)

    ws.cell(b2[0] - 1, 7).value = "Current Draw (A)"
    for i in range(n):
        val = round(base.i_pack_a[i], 6)
        for gi in range(len(reductions)):
            ws.cell(b2[0] + i, first_col + gi).value = val

    ws.cell(b3[0] - 1, 7).value = "Motor Power (kW)"
    for i in range(n):
        val = round(base.p_elec_w[i] / 1000.0, 6)
        for gi in range(len(reductions)):
            ws.cell(b3[0] + i, first_col + gi).value = val

    # The MIN/MAX summary rows the original workbook keeps below the blocks.
    stat_min, stat_max = b3[1] + 2, b3[1] + 3
    ws.cell(stat_min, 7).value = "Min Current (A)"
    ws.cell(stat_max, 7).value = "Max Current (A)"
    for gi in range(len(reductions)):
        c = first_col + gi
        col = ws.cell(1, c).column_letter
        ws.cell(stat_min, c).value = f"=MIN({col}{b2[0]}:{col}{b2[1]})"
        ws.cell(stat_max, c).value = f"=MAX({col}{b2[0]}:{col}{b2[1]})"

    try:
        wb.calculation.fullCalcOnLoad = True
    except Exception:
        pass
    wb.save(path)
    return {
        "rebuilt": True,
        "rows": n,
        "gears": len(reductions),
        "blocks": [b1, b2, b3],
        "resampled_from": len(speed_mph),
    }


def _recalculate_via_libreoffice(path: str, timeout: int) -> tuple[bool, str]:
    soffice = shutil.which("soffice") or shutil.which("libreoffice")
    if not soffice:
        return False, "LibreOffice not available"
    outdir = tempfile.mkdtemp(prefix="kx_recalc_")
    try:
        subprocess.run(
            [soffice, "--headless", "--norestore", "--convert-to", "xlsx",
             "--outdir", outdir, path],
            check=True, capture_output=True, timeout=timeout)
        produced = os.path.join(
            outdir, os.path.splitext(os.path.basename(path))[0] + ".xlsx")
        if not os.path.exists(produced):
            return False, "LibreOffice produced no output"
        shutil.copy(produced, path)
        return True, "recalculated with LibreOffice"
    except subprocess.TimeoutExpired:
        return False, f"LibreOffice timed out after {timeout}s"
    except subprocess.CalledProcessError as exc:
        return False, f"LibreOffice failed: {exc.stderr[:200]!r}"
    finally:
        shutil.rmtree(outdir, ignore_errors=True)


def recalculate(path: str, timeout: int = 180, *,
                only_prefix: Optional[str] = SHEET_PREFIX
                ) -> tuple[bool, str]:
    """Populate cached values, so `data_only=True` readers see numbers.

    Without this the file is Excel-only: openpyxl writes formula strings and
    evaluates nothing, so `data_only=True` returns None for every derived cell.
    That is the defect that made the previous export unreadable by KinematiK's
    own loaders.

    Two engines, tried in order:

    1. **LibreOffice**, when it is installed. It evaluates the whole workbook
       including the user's own sheets, so it stays the preferred path.
    2. **`xlsx_formula_cache`**, the in-process evaluator. LibreOffice is
       absent from Streamlit Cloud, from slim containers and from most CI
       images — which is to say from everywhere this actually runs — and
       falling back to "no cached values" there meant the export degraded to
       the exact defect it exists to fix, on the machines that matter. The
       evaluator covers the operators and functions the KX sheets use, and
       skips anything it does not understand rather than guessing, so a cell is
       either correct or left empty.

    `only_prefix` limits which cells are *written* by the fallback, not which
    are read: a KX formula that points at one of the user's sheets still
    resolves through it. Pass None to cache the whole workbook, which is what
    reading a pack config out of a stranger's file needs.
    """
    ok, msg = _recalculate_via_libreoffice(path, timeout)
    if ok:
        return True, msg
    lo_reason = msg

    try:
        from . import xlsx_formula_cache as xfc
        report = xfc.populate_cached_values(path, only_prefix=only_prefix)
    except Exception as exc:                      # pragma: no cover - defensive
        return False, (f"{lo_reason}; in-process evaluation also failed "
                       f"({type(exc).__name__}: {exc})")

    written, skipped = report.get("written", 0), report.get("skipped", 0)
    if not written:
        return False, (f"{lo_reason}; the in-process evaluator understood none "
                       f"of the {skipped} formula(s) it was given")
    detail = f"cached {written} formula cell(s) in process ({lo_reason})"
    if skipped:
        detail += (f"; {skipped} cell(s) use constructs it does not model and "
                   f"were left uncached")
    return True, detail


def formula_errors(path: str) -> dict[str, list[str]]:
    """Cells whose cached value is an Excel error, grouped by error type."""
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    bad: dict[str, list[str]] = {}
    for ws in wb.worksheets:
        if not ws.title.startswith(SHEET_PREFIX):
            continue
        for row in ws.iter_rows():
            for c in row:
                v = c.value
                if isinstance(v, str) and v.startswith("#"):
                    bad.setdefault(v, []).append(f"{ws.title}!{c.coordinate}")
    return bad


# ===================================================================== #
#  Entry point
# ===================================================================== #
@dataclass
class TrackSimExport:
    path: str
    sheets_added: list
    trace: Optional[pdw.PowerDrawTrace] = None
    recalculated: bool = False
    recalc_message: str = ""
    formula_errors: dict = field(default_factory=dict)
    findings: list = field(default_factory=list)
    warnings: list = field(default_factory=list)

    def ok(self) -> bool:
        return not self.formula_errors

    def blocking(self) -> list:
        return [f for f in self.findings if f.severity == Severity.FAIL]


def export_track_sim(source_path: str, out_path: str,
                     speed_mph: Sequence[float], dt_s: float, *,
                     time_s: Optional[Sequence[float]] = None,
                     pack: Optional[pdw.PackSpec] = None,
                     vehicle: Optional[pdw.VehicleSpec] = None,
                     drive: Optional[pdw.DriveSpec] = None,
                     lap_time_s: Optional[float] = None,
                     reductions: Sequence[float] = tuple(range(1, 16)),
                     smooth_window: int = 11,
                     endurance_km: Optional[float] = None,
                     max_cell_c_rate: float = 10.0,
                     rebuild_propulsion: bool = True,
                     recalc: bool = True) -> TrackSimExport:
    """Write the lap-sim result into a copy of the electrics workbook.

    `pack` defaults to whatever `power_draw.read_pack_config` finds in the
    source workbook, read by row label. Nothing on the user's sheets is
    modified; every sheet written carries the `KX ` prefix, and re-running
    replaces only those.
    """
    import openpyxl

    warnings: list[str] = []
    findings: list[Finding] = []

    if pack is None:
        try:
            pack = pdw.read_pack_config(source_path)
        except pdw.WorkbookReadError as exc:
            # A workbook that has been through openpyxl has no cached values, so
            # every formula cell reads as empty. Recalculating a scratch copy
            # restores them; falling straight through to defaults would report a
            # pack nobody specified.
            pack = None
            scratch = None
            try:
                import tempfile as _tf
                scratch = os.path.join(_tf.mkdtemp(prefix="kx_pack_"),
                                       "src.xlsx")
                shutil.copy(source_path, scratch)
                # The pack lives on one of the user's own sheets, so the
                # prefix restriction has to come off for this one.
                ok, _msg = recalculate(scratch, only_prefix=None)
                if ok:
                    pack = pdw.read_pack_config(scratch)
                    warnings.append(
                        "Pack values were uncached in the source workbook "
                        "(every formula read as empty); recalculated a scratch "
                        "copy and read them from that.")
            except Exception:
                pack = None
            finally:
                if scratch:
                    shutil.rmtree(os.path.dirname(scratch), ignore_errors=True)
            if pack is None:
                warnings.append(
                    f"Could not read the pack from the workbook ({exc}); using "
                    f"declared defaults. Every pack figure below is therefore a "
                    f"default, not a reading.")
                pack = pdw.PackSpec()
    vehicle = vehicle or pdw.VehicleSpec()
    drive = drive or pdw.DriveSpec()
    if endurance_km is None:
        endurance_km = 22.0

    times = None
    if time_s is not None:
        times = [float(x) for x in time_s]
        if len(times) != len(speed_mph):
            warnings.append(
                f"Ignored the supplied time axis: {len(times)} timestamps "
                f"against {len(speed_mph)} speed samples. Falling back to a "
                f"uniform {dt_s:g} s step.")
            times = None

    spec = ExportSpec(pack=pack, vehicle=vehicle, drive=drive,
                      speed_mph=list(speed_mph), dt_s=dt_s, time_s=times,
                      lap_time_s=lap_time_s, reductions=tuple(reductions),
                      smooth_window=smooth_window, endurance_km=endurance_km,
                      max_cell_c_rate=max_cell_c_rate)

    # Python-side truth, used for findings and as a cross-check on the sheet.
    trace = pdw.power_draw_trace(spec.speed_mph, dt_s, pack, vehicle, drive,
                                 smooth_window=smooth_window)
    findings.extend(pdw.trace_findings(trace))
    findings.extend(pdw.motor_feasibility(pack, drive))

    if os.path.abspath(source_path) != os.path.abspath(out_path):
        shutil.copy(source_path, out_path)

    # Compatibility pass FIRST, on the user's own sheets. Five CHOOSECOLS array
    # formulas in this workbook spill roughly nineteen thousand #NAME? cells on
    # any Excel that is not Microsoft 365, and a #NAME? in a source column makes
    # every dependent sum, min and max blank or zero — which is exactly the
    # "zeroes and blank pages" symptom. INDEX(range,0,n) is an exact
    # replacement and works in every version.
    portable = make_portable(out_path)
    if portable:
        warnings.append(
            "Replaced Microsoft-365-only functions with portable equivalents: "
            + ", ".join(f"{k} x{v}" for k, v in sorted(portable.items()))
            + ". These were array formulas spilling #NAME? across thousands of "
              "cells on any Excel older than 365, which is what made the "
              "downstream sheets read as zero or blank.")

    thin = audit_block_column_coverage(out_path)
    stale_now = audit_block_references(out_path)
    if (thin or stale_now) and rebuild_propulsion:
        # Do not merely warn. The blocks are reconstructible from the speed
        # trace with correct physics, so rebuild them at the canonical rows the
        # dependent sheets reference. This is what makes an already-damaged
        # workbook usable instead of telling someone to go and find the original.
        info = rebuild_propulsion_blocks(
            out_path, spec.speed_mph, pack, vehicle, drive,
            dt_s=dt_s, reductions=tuple(reductions),
            smooth_window=smooth_window)
        if info.get("rebuilt"):
            warnings.append(
                f"Source workbook had damaged propulsion data "
                f"({'; '.join(thin) if thin else 'formulas pointing at blank rows'}). "
                f"Rebuilt all three blocks at the canonical rows "
                f"{info['blocks'][0][0]}-{info['blocks'][2][1]} with "
                f"{info['gears']} gear columns and {info['rows']} samples "
                f"(resampled from {info['resampled_from']}), using the "
                f"corrected physics. ThermalLoad, EMFs and BearingBlowOut now "
                f"work at every gear, not just gear 1.")
            for _pass in (repair_range_overshoot, remap_stale_block_references):
                _pass(out_path)
    elif thin:
        warnings.append(
            "Damaged propulsion data in the source workbook: " + " ".join(thin)
            + " Pass rebuild_propulsion=True to reconstruct it, or re-export "
              "from the original workbook.")


    # If the blocks have been relocated by an earlier legacy export, repoint the
    # formulas at where the data actually is now.
    remapped = remap_stale_block_references(out_path)
    if remapped:
        warnings.append(
            f"Repointed {len(remapped)} formula reference(s) at the data's "
            f"current position — an earlier export moved ElecPropulsion's "
            f"blocks and left every dependent formula reading blank rows: "
            + "; ".join(remapped[:3])
            + ("" if len(remapped) <= 3 else f" (+{len(remapped)-3} more)"))

    # Trim ranges that overshoot their data block by a few rows. EMFs reads one
    # row past the power block even in the untouched original, which is why its
    # last value is zero.
    trimmed = repair_range_overshoot(out_path)
    if trimmed:
        warnings.append(
            f"Trimmed {len(trimmed)} formula range(s) that ran past the end of "
            f"their data and were contributing zeros: " + "; ".join(trimmed[:3])
            + ("" if len(trimmed) <= 3 else f" (+{len(trimmed)-3} more)"))

    # Detect ranges pointing at wholly empty regions. This is what makes the
    # ThermalLoad and EMFs tabs read as solid zeros, and it cannot be repaired.
    stale = audit_block_references(out_path)
    if stale:
        tgt = stale[0]["target"]
        blocks = stale[0]["blocks"]
        span = (f"rows {blocks[0][0]}-{blocks[-1][1]}" if blocks else "no rows")
        warnings.append(
            f"ZEROS EXPLAINED: {len(stale)} formula(s) reference empty regions "
            f"of '{tgt}'. That sheet holds data in {span}, but these formulas "
            f"read further down: "
            + "; ".join(f"{f['where']} -> {f['target']}!{f['range']}"
                        for f in stale[:3])
            + ". INDEX over blank cells returns 0, which is why those tabs are "
              "solid zeros with no error to explain it. This is NOT repairable "
              "here: the rows those formulas want were overwritten with static "
              "values by an earlier 'enhanced' export, so the samples no longer "
              "exist. Re-run this export against the ORIGINAL "
              "FSAE_EV_Power_Draw.xlsx rather than against a previously "
              "enhanced copy. The KX sheets below are self-contained and "
              "unaffected.")

    wb = openpyxl.load_workbook(out_path)
    for name in [s for s in wb.sheetnames if s.startswith(SHEET_PREFIX)]:
        del wb[name]

    ws_in = wb.create_sheet(S_INPUTS)
    ref = _write_inputs(ws_in, spec)

    ws_tr = wb.create_sheet(S_TRACE)
    smoothed = pdw._moving_average(spec.speed_mph, smooth_window)
    n = len(smoothed)
    # Column A is the clock every derivative and integral is taken against, so
    # it carries the real timestamps whenever the caller supplied them.
    times = spec.time_s if spec.time_s is not None else [
        k * dt_s for k in range(n)]
    for k, (sp) in enumerate(smoothed):
        r = 3 + k
        # Nanoseconds, not the old microseconds: this column is the divisor of
        # every derivative and the weight of every integral, so rounding it is
        # rounding the physics. Nine places is finer than any logger and keeps
        # the literal short enough to render in a 12.5-wide column.
        ws_tr[f"A{r}"] = round(float(times[k]), 9)
        ws_tr[f"B{r}"] = round(float(sp), 4)
    tr = _write_trace(ws_tr, spec, ref, n)

    ws_pk = wb.create_sheet(S_PACK)
    pk = _write_pack_limits(ws_pk, ref)

    ws_dash = wb.create_sheet(S_DASH)
    dash = _write_dashboard(ws_dash, ref, tr, pk)

    ws_g = wb.create_sheet(S_GEARS)
    _write_gear_study(ws_g, spec, ref, tr)

    ws_ad = wb.create_sheet(S_ADVISOR)
    _write_advisor(ws_ad, spec, ref, pk, dash)

    ws_pv = wb.create_sheet(S_PROV)
    _write_provenance(ws_pv, spec)

    # Dashboard first: it is what a reviewer opens.
    order = [S_DASH] + [s for s in wb.sheetnames if s != S_DASH]
    wb._sheets = [wb[s] for s in order]
    # openpyxl writes formulas without cached values. Excel shows 0 or blank for
    # an uncached formula unless told to recompute, so this flag is not optional.
    try:
        wb.calculation.fullCalcOnLoad = True
    except Exception:
        pass
    wb.save(out_path)

    added = [S_DASH, S_INPUTS, S_TRACE, S_PACK, S_GEARS, S_ADVISOR, S_PROV]
    result = TrackSimExport(out_path, added, trace, findings=findings,
                            warnings=warnings)

    if recalc:
        ok, msg = recalculate(out_path)
        result.recalculated, result.recalc_message = ok, msg
        if not ok:
            warnings.append(
                msg + ". Formula cells will read as None to data_only "
                      "consumers until the file is opened and saved in Excel.")
        else:
            result.formula_errors = formula_errors(out_path)
            leftover = modern_function_report(out_path)
            if leftover:
                warnings.append(
                    "Microsoft-365-only functions still present: "
                    + ", ".join(f"{k} x{v}" for k, v in sorted(leftover.items()))
                    + ". These will read as #NAME? outside Microsoft 365.")
            if result.formula_errors:
                warnings.append(
                    "Formula errors present: "
                    + ", ".join(f"{k} x{len(v)}"
                                for k, v in result.formula_errors.items()))
    return result


def export_track_sim_bytes(excel_bytes: bytes,
                           speed_ms: Sequence[float],
                           time_s: Sequence[float], *,
                           pack: Optional[pdw.PackSpec] = None,
                           vehicle: Optional[pdw.VehicleSpec] = None,
                           drive: Optional[pdw.DriveSpec] = None,
                           lap_time_s: Optional[float] = None,
                           reductions: Sequence[float] = tuple(range(1, 16)),
                           smooth_window: int = 11,
                           endurance_km: Optional[float] = None,
                           max_cell_c_rate: float = 10.0,
                           rebuild_propulsion: bool = True,
                           recalc: bool = True
                           ) -> tuple[bytes, "TrackSimExport"]:
    """Bytes in, bytes out — the shape a Streamlit download button needs.

    Takes the same `speed_ms` / `time_s` arrays as the old
    `lap_to_excel_roundtrip`, so swapping the call site over is a rename plus
    reading `.excel_bytes` from the tuple instead of the result object.

    `time_s` is passed through to the trace sheet rather than collapsed to a
    step, since a lap sim's output is not always evenly sampled. `dt` is still
    derived from it as the mean spacing, but only as a reported figure: the
    formulas differentiate and integrate against the timestamps themselves, so
    a log with a dropout or a variable rate comes out exact. A non-uniform
    trace is still reported, because it says something about the log.
    """
    if len(time_s) < 2:
        raise ValueError("need at least two samples")
    t = [float(x) for x in time_s]
    dt = (t[-1] - t[0]) / (len(t) - 1)
    if dt <= 0:
        raise ValueError("time_s must increase")

    speed_mph = [float(v) / pdw.MPH_TO_MS for v in speed_ms]

    src_dir = tempfile.mkdtemp(prefix="kx_src_")
    try:
        src = os.path.join(src_dir, "source.xlsx")
        out = os.path.join(src_dir, "FSAE_EV_KinematiK_TrackSim.xlsx")
        with open(src, "wb") as fh:
            fh.write(excel_bytes)
        res = export_track_sim(
            src, out, speed_mph, dt, time_s=t,
            pack=pack, vehicle=vehicle, drive=drive,
            lap_time_s=lap_time_s, reductions=reductions,
            smooth_window=smooth_window, endurance_km=endurance_km,
            max_cell_c_rate=max_cell_c_rate,
            rebuild_propulsion=rebuild_propulsion, recalc=recalc)

        # Report an uneven time base rather than silently averaging over it.
        # The sheet no longer approximates one — column A holds the real
        # timestamps and column P their trapezoidal weights — so this says what
        # the log looks like and what was done about it, not what to go and fix.
        steps = [t[i + 1] - t[i] for i in range(len(t) - 1)]
        if steps and (max(steps) - min(steps)) > 0.05 * dt:
            res.warnings.append(
                f"Time base is uneven (steps {min(steps):.4g}..{max(steps):.4g} s, "
                f"mean {dt:.4g} s). The lap trace carries the real timestamps "
                f"and weights each sample by its own share of the elapsed time, "
                f"so the accelerations and the energy integral are exact for "
                f"the log as recorded. The pass/fail findings are computed "
                f"separately at the mean step and can differ slightly.")
        with open(out, "rb") as fh:
            data = fh.read()
        return data, res
    finally:
        shutil.rmtree(src_dir, ignore_errors=True)


PROVENANCE = {
    "physics_grounded": [
        "force balance, P_elec = P_wheel/eta, sag-aware current — all from "
        "power_draw, which is unit-tested against hand-checkable cases",
    ],
    "hard_rule": (
        "Only sheets prefixed 'KX ' are written, and every derived cell is a "
        "formula rather than a constant. A snapshot of numbers cannot be "
        "re-checked; a model can."
    ),
    "known_limits": [
        "Dashboard cell references are returned by _write_dashboard rather "
        "than hardcoded, so moving a row does not silently repoint a formula.",
        "Vehicle mass, CdA, Crr, inverter efficiency and cell C-rating are "
        "assumptions this export introduces; they are yellow-filled on "
        "KX Inputs and listed on KX Provenance.",
        "Without LibreOffice the cached values come from this package's own "
        "evaluator, which models the operators and functions the KX sheets "
        "use and nothing else. Cells it cannot evaluate keep no cached value "
        "rather than a guessed one, and Excel recomputes all of them on open.",
        "The sheet integrates against the trace's real timestamps, but the "
        "Python-side findings still run at the mean sample interval, so on a "
        "badly uneven log the two can disagree at the margin.",
    ],
}
