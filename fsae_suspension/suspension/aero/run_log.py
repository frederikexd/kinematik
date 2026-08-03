# ============================================================================
#  KinematiK — Formula SAE suspension & vehicle dynamics toolkit
#  Created by Frederik Thio. Copyright (c) 2026 Frederik Thio.
#  Open source. Original author: Frederik Thio, creator of KinematiK.
# ============================================================================
"""
run_log.py — ingest the aero team's ANSYS run log, screen it, average what survives
====================================================================================

THE PROBLEM THIS SOLVES
-----------------------
The wings team keeps a shared spreadsheet: one row per Fluent run, filled in by
whoever drove the solver that night. Mesh settings on the left, coefficients on
the right. It is the real artefact — not a clean `<case>_coeffs.csv` — and it is
mostly noise. Early iterations have inverted mesh sizes, y+ landing in the buffer
layer, a stagnation pressure that says the reference velocity was wrong, or a
contributor called "Khalil - Test". Averaging that sheet as-is produces a
confident number built on runs nobody would defend.

So this module does three things, in order, and shows its work at every step:

    parse_run_log()  ->  screen()  ->  consolidate()  ->  write_workbook()

  1. PARSE tolerantly. The sheet has a two-row header (a banner row above the real
     column names), renamed columns, blank filler rows, and units baked into the
     header text. None of that is an error; it is what the file looks like.
  2. SCREEN each row against explicit, physics-backed acceptance criteria, and
     attach a REASON to every rejection. A row is never silently dropped.
  3. CONSOLIDATE the survivors per operating point — mean, spread, and n — and
     write a workbook whose aggregate cells are live formulas over the accepted
     rows, so the team can audit the average without trusting this code.

THE HONESTY CONTRACT (same discipline as cfd.py / ensemble.py / aeromap.py)
---------------------------------------------------------------------------
  * A rejected run is REPORTED, with the flag, the measured value and the limit it
    broke. There is no quiet filtering.
  * A missing channel stays missing. A blank drag coefficient is derived from the
    drag force ONLY when a reference area is known or inferable, and the derived
    value is labelled `derived`, never presented as reported data.
  * Every threshold lives in `ScreenConfig`, is written into the output workbook's
    Config sheet, and can be overridden. The defaults are documented below with
    their physical basis — they are engineering judgement, not laws.
  * A single run is a sample of one. `ConsolidatedCase` carries `n_accepted` and
    the spread; it never dresses one surviving row up as a converged mean.
  * The statistical outlier pass is DISABLED below `min_n_for_outlier` samples,
    because rejecting an "outlier" out of three points is just picking a favourite.

WHY THESE DEFAULT THRESHOLDS
----------------------------
  y+ band          The one that matters. A wall-function closure (k-epsilon,
                   k-omega with wall functions) needs the first cell centroid in
                   the log layer: y+ >~ 30. Below y+ = 11.06 — where the log law
                   and the linear sublayer profile intersect — the wall function
                   is being asked to model a cell that is physically inside the
                   viscous sublayer, and the wall shear it returns is fiction.
                   A low-Re / sublayer-resolving closure (k-omega SST, Spalart-
                   Allmaras with low-Re treatment) wants the opposite: y+ ~ 1,
                   and degrades once y+ climbs past ~5. The band is therefore
                   chosen FROM the viscous model named in the row, not globally.
  Orthogonal qual. Fluent's own floor for a usable cell is 0.10; below ~0.05 the
                   solve is untrustworthy regardless of what the residuals say.
  Skewness         >0.95 is Fluent's "degenerate cell" line. 0.90-0.95 is a warn.
  Aspect ratio     Deliberately a WARN-only channel by default with a very high
                   reject bar: inflation-layer cells legitimately run AR in the
                   hundreds-to-thousands. A high AR here is a smell, not a fault.
  Stagnation Cp    The best free check on the whole run. Total pressure at a
                   stagnation point is q above static, so max reported gauge
                   pressure / q should sit at Cp ~ 1. If it reads 0.4 or 2.5, the
                   reference velocity, the density or the pressure datum is wrong
                   — and every coefficient in that row is off by that ratio.
  Suction Cp       A wing that produces no suction peak (Cp_min > -0.5) usually
                   means the report summed over the wrong wall zone — the failure
                   mode the Fluent run-sheet warns about first.
  Mass imbalance   Continuity must close. Absolute default because the sheet does
                   not carry inlet mass flow; override `mass_imbalance_*` with a
                   fraction of your inlet flux if you log it.
  Reference area   Inferred per row from |L| / (q * |Cl|) and compared to the
                   group median. A row that disagrees used a different reference
                   area, so its coefficients are not comparable to its neighbours'
                   — the "silent killer" from the Fluent validation run-sheet.

Pure standard library plus a lazy `openpyxl` import (already an optional dep, used
by ev_excel_roundtrip.py and hardpoint_import.py). Importable and unit-testable
with no Streamlit, no solver, no network.
"""

from __future__ import annotations

import csv
import io
import math
import os
import re
import statistics
from dataclasses import dataclass, field, asdict
from typing import Iterable, Optional, Sequence

__all__ = [
    # config / vocabulary
    "ScreenConfig", "Severity", "WallTreatment", "DEFAULT_RHO",
    "LOG_LAW_INTERSECTION_YPLUS",
    # data shapes
    "Flag", "RunRow", "Derived", "Verdict", "CaseKey", "Discretisation",
    "ConsolidatedCase", "ConsolidationReport",
    # pipeline
    "parse_run_log", "parse_rows_from_grid", "screen", "consolidate",
    "process", "to_coeff_results",
    # output
    "write_workbook", "write_csv_bundle", "consolidated_csv",
    # helpers worth reusing / testing
    "wall_treatment_for", "dynamic_pressure", "implied_reference_area",
    "modified_z_scores", "CANONICAL_FIELDS", "SETUP_FIELDS",
    "discretisation_of", "setup_signature",
]


# --------------------------------------------------------------------------- #
#  Physical constants and vocabulary
# --------------------------------------------------------------------------- #
DEFAULT_RHO = 1.225                 # kg/m^3, ISA sea level — Fluent's air default

#: y+ at which the linear (viscous sublayer) and logarithmic profiles intersect.
#: Below this a standard wall function is extrapolating a log law into a cell that
#: physically sits in the sublayer. This is the hard floor for wall-function runs.
LOG_LAW_INTERSECTION_YPLUS = 11.06


class Severity:
    """Flag severity. REJECT removes the row from the average; WARN keeps it."""
    INFO = "info"
    WARN = "warn"
    REJECT = "reject"

    _ORDER = {INFO: 0, WARN: 1, REJECT: 2}

    @classmethod
    def worst(cls, severities: Iterable[str]) -> str:
        best = cls.INFO
        for s in severities:
            if cls._ORDER.get(s, 0) > cls._ORDER[best]:
                best = s
        return best


class WallTreatment:
    """How the named turbulence closure expects the near-wall cell to be sized."""
    WALL_FUNCTION = "wall-function"      # wants y+ in the log layer (~30-300)
    RESOLVED = "resolved"                # wants y+ ~ 1 (viscous sublayer resolved)
    AUTOMATIC = "automatic"              # blends; tolerant across the range
    UNKNOWN = "unknown"


#: Turbulence-model name fragments -> near-wall expectation. Matched case- and
#: punctuation-insensitively against whatever the team typed in "Viscous Model".
_WALL_TREATMENT_PATTERNS: tuple[tuple[str, str], ...] = (
    # Sublayer-resolving / low-Re closures. Checked before the generic k-omega
    # entry so "k-omega SST" resolves to RESOLVED rather than AUTOMATIC.
    (r"sst", WallTreatment.RESOLVED),
    (r"low\s*re", WallTreatment.RESOLVED),
    (r"spalart|s\s*-?\s*a\b", WallTreatment.RESOLVED),
    (r"transition|gamma\s*re|k\s*-?\s*kl", WallTreatment.RESOLVED),
    (r"\bles\b|detached|\bdes\b|\bddes\b", WallTreatment.RESOLVED),
    # Generic k-omega has automatic (blended) wall treatment in Fluent.
    (r"k\s*-?\s*omega|k\s*-?\s*w\b", WallTreatment.AUTOMATIC),
    # Wall-function families.
    (r"k\s*-?\s*epsilon|k\s*-?\s*e\b", WallTreatment.WALL_FUNCTION),
    (r"realizable|rng", WallTreatment.WALL_FUNCTION),
    (r"reynolds\s*stress|\brsm\b", WallTreatment.WALL_FUNCTION),
)


def wall_treatment_for(viscous_model: Optional[str]) -> str:
    """
    Classify a turbulence-model string into its near-wall expectation.

    This decides which y+ band the row is judged against, so it is deliberately
    conservative: an unrecognised model returns UNKNOWN and the y+ gate downgrades
    itself to a warning rather than rejecting a run it cannot reason about.
    """
    if not viscous_model:
        return WallTreatment.UNKNOWN
    text = str(viscous_model).strip().lower()
    if not text:
        return WallTreatment.UNKNOWN
    for pattern, treatment in _WALL_TREATMENT_PATTERNS:
        if re.search(pattern, text):
            return treatment
    return WallTreatment.UNKNOWN


class Discretisation:
    """Spatial discretisation order, as written in the sheet's `Order` column."""
    FIRST = "first-order"
    SECOND = "second-order"
    MIXED = "mixed"
    UNKNOWN = "unknown"


#: `Order` cell text -> discretisation order. Second is checked first so
#: "First to Second Order" (a run that was ramped) reads as MIXED, not FIRST.
_ORDER_PATTERNS: tuple[tuple[str, str], ...] = (
    (r"first.*second|1st.*2nd|ramp|blend", Discretisation.MIXED),
    (r"second|2nd|high.?order|quick|muscl|bounded.?central", Discretisation.SECOND),
    (r"first|1st|upwind\s*$", Discretisation.FIRST),
)


def discretisation_of(order_text: Optional[str]) -> str:
    """
    Classify the `Order` column.

    This is the one solver-setup field with a defensible right answer for a force
    coefficient. First-order upwind is numerically diffusive: it smears the very
    gradients a wing's suction peak is made of, so it under-predicts downforce
    and over-predicts the wake. It is a legitimate way to START a solve and not a
    legitimate way to finish one, which is why the default is a warning rather
    than a rejection — the run may have been ramped to second order without the
    sheet saying so.
    """
    if not order_text:
        return Discretisation.UNKNOWN
    text = str(order_text).strip().lower()
    if not text:
        return Discretisation.UNKNOWN
    for pattern, kind in _ORDER_PATTERNS:
        if re.search(pattern, text):
            return kind
    return Discretisation.UNKNOWN


def _normalise_setup_value(value) -> str:
    """
    Fold a setup cell to a comparable token: lower-cased, punctuation stripped.

    So "SIMPLE", "simple" and "Simple " are one scheme, while genuinely different
    entries stay different. Blank reads as an empty string, which the consistency
    check treats as "not stated" rather than as a distinct choice.
    """
    if value is None:
        return ""
    text = str(value).strip().lower()
    if text in {"", "-", "--", "n/a", "na", "none", "?", "tbd"}:
        return ""
    return re.sub(r"[^a-z0-9]+", "", text)


def setup_signature(row: "RunRow") -> tuple:
    """
    The method a run was solved with, as a comparable tuple.

    Turbulence model, pressure-velocity scheme, discretisation order and
    initialisation. Courant number and pseudo time step are deliberately excluded:
    they affect the PATH to convergence, not the converged answer, so two runs
    that differ only there are still two samples of the same quantity.
    """
    return (
        _normalise_setup_value(row.viscous_model),
        _normalise_setup_value(row.scheme),
        discretisation_of(row.order),
        _normalise_setup_value(row.initialization),
    )


#: Human labels for the parts of a setup signature, for readable flag messages.
_SETUP_PART_LABELS = ("viscous model", "scheme", "discretisation order",
                      "initialization")


def dynamic_pressure(speed_ms: Optional[float], rho: float = DEFAULT_RHO) -> Optional[float]:
    """q = 1/2 rho V^2, in Pa. None if the speed is missing or non-positive."""
    if speed_ms is None or speed_ms <= 0:
        return None
    return 0.5 * rho * speed_ms * speed_ms


def implied_reference_area(force_N: Optional[float], coeff: Optional[float],
                           q_Pa: Optional[float]) -> Optional[float]:
    """
    Back out the reference area a row's own numbers imply: A = |F| / (q * |C|).

    This is how the module detects that two contributors normalised by different
    areas — the failure that makes their coefficients silently incomparable while
    both look perfectly reasonable in isolation.
    """
    if force_N is None or coeff is None or q_Pa is None:
        return None
    if q_Pa <= 0 or abs(coeff) < 1e-9:
        return None
    return abs(force_N) / (q_Pa * abs(coeff))


def modified_z_scores(values: Sequence[float]) -> list[float]:
    """
    Iglewicz-Hoaglin modified z-score, |0.6745 * (x - median) / MAD|.

    Median/MAD rather than mean/stdev because a run log's outliers are exactly the
    values that would poison a mean-based test. When MAD is zero (identical values)
    the scores are all zero — no spread, nothing to reject.
    """
    if not values:
        return []
    med = statistics.median(values)
    deviations = [abs(v - med) for v in values]
    mad = statistics.median(deviations)
    if mad <= 0:
        # Fall back to a mean-absolute-deviation scale so a single wild value in an
        # otherwise identical set is still catchable; if that is zero too, no spread.
        mean_abs = sum(deviations) / len(deviations) if deviations else 0.0
        if mean_abs <= 0:
            return [0.0] * len(values)
        return [abs(0.7979 * (v - med) / mean_abs) for v in values]
    return [abs(0.6745 * (v - med) / mad) for v in values]


# --------------------------------------------------------------------------- #
#  Screening configuration — every threshold, in one auditable place
# --------------------------------------------------------------------------- #
@dataclass
class ScreenConfig:
    """
    Acceptance criteria for a single run row. Defaults are the documented
    engineering judgement in this module's docstring; override any of them.

    Every field is written verbatim into the output workbook's Config sheet, so a
    consolidated number can always be traced back to the rules that produced it.
    """

    # -- fluid / reference conditions ------------------------------------- #
    rho: float = DEFAULT_RHO
    reference_area_m2: Optional[float] = None      # None => infer from the rows

    # -- y+ bands, per near-wall treatment (warn band, then reject band) --- #
    yplus_wf_warn: tuple = (30.0, 300.0)
    yplus_wf_reject: tuple = (LOG_LAW_INTERSECTION_YPLUS, 1000.0)
    yplus_resolved_warn: tuple = (0.0, 5.0)
    yplus_resolved_reject: tuple = (0.0, 30.0)
    yplus_auto_warn: tuple = (0.0, 300.0)
    yplus_auto_reject: tuple = (0.0, 1000.0)
    #: Fractional miss of the mesh's own y+ intent, |avg - desired| / desired.
    yplus_target_miss_warn: float = 0.30

    # -- mesh quality ------------------------------------------------------ #
    ortho_quality_warn: float = 0.20               # warn below
    ortho_quality_reject: float = 0.10             # reject below (Fluent's floor)
    skewness_warn: float = 0.90                    # warn above
    skewness_reject: float = 0.95                  # reject above
    aspect_ratio_warn: float = 5_000.0             # warn above (soft indicator)
    aspect_ratio_reject: float = 100_000.0         # reject above
    #: max/min surface mesh length ratio that looks like a typo rather than grading
    mesh_length_ratio_warn: float = 50.0
    min_layers_warn: int = 5                       # prism layers below this: warn
    min_layers_reject: int = 2

    # -- solution health --------------------------------------------------- #
    mass_imbalance_warn: float = 1e-4              # kg/s, absolute
    mass_imbalance_reject: float = 1e-3

    # -- solver setup ------------------------------------------------------- #
    #: First-order spatial discretisation smears the gradients a suction peak is
    #: made of. Warn by default rather than reject: it is a legitimate way to
    #: start a solve, and a run ramped to second order may not say so in the sheet.
    reject_first_order: bool = False
    #: Pseudo-transient Courant number. Very large values converge the residuals
    #: while the flow field is still moving — a converged-LOOKING answer.
    courant_warn_max: float = 500.0
    courant_warn_min: float = 0.5
    #: Compare each run's method against the rest of its operating point. Two runs
    #: solved differently are not two samples of the same quantity.
    check_setup_consistency: bool = True
    #: A mixed turbulence model within one operating point is the sharpest version
    #: of that problem. Off by default: the tool reports the split rather than
    #: silently picking which half of the team was right.
    reject_mixed_turbulence: bool = False

    # -- pressure-field physics ------------------------------------------- #
    cp_stagnation_warn: tuple = (0.80, 1.30)       # max gauge pressure / q
    cp_stagnation_reject: tuple = (0.30, 2.00)
    #: A downforce surface must show a real suction peak; near-zero means the
    #: force report almost certainly summed over the wrong wall zone.
    cp_suction_min_warn: float = -0.50             # warn if Cp_min is ABOVE this
    cp_suction_floor_reject: float = -100.0        # reject if Cp_min below this

    # -- coefficient / reference-area consistency -------------------------- #
    ref_area_tolerance: float = 0.05               # fractional, vs group median
    ref_area_reject_tolerance: float = 0.25
    #: Wings are expected to make downforce; the sheet's convention is negative
    #: lift = downforce. Set False for a run log of an upforce/loose surface.
    expect_downforce: bool = True

    # -- bookkeeping ------------------------------------------------------- #
    #: Contributor/notes patterns that mark a scratch row rather than real data.
    test_row_patterns: tuple = ("test", "trial", "scratch", "dummy", "ignore",
                                "example", "template", "practice", "wip",
                                "do not use", "delete")
    #: Treat a flagged test row as a rejection (True) or merely a warning (False).
    reject_test_rows: bool = True
    #: Rows explicitly marked not-converged are always rejected.
    reject_unconverged: bool = True

    # -- statistical outlier pass ----------------------------------------- #
    enable_outlier_pass: bool = True
    #: Modified z-score above which a run is an outlier within its own case group.
    outlier_z_threshold: float = 3.5
    #: Below this many accepted rows in a group, the outlier pass is skipped —
    #: with three samples "the odd one out" is noise, not evidence.
    min_n_for_outlier: int = 4

    # -- grouping ---------------------------------------------------------- #
    ride_height_tol_mm: float = 0.5                # bucket width for grouping
    speed_tol_ms: float = 0.5

    #: Optional: keep only the LAST row per (contributor, case). Off by default —
    #: physics screening is the honest filter; this one assumes row order is
    #: chronological, which the sheet does not actually promise.
    prefer_latest_per_contributor: bool = False

    def as_rows(self) -> list[tuple]:
        """Flatten to (setting, value) pairs for the Config sheet."""
        out = []
        for k, v in asdict(self).items():
            if isinstance(v, (tuple, list)):
                v = ", ".join(str(x) for x in v)
            out.append((k, "" if v is None else v))
        return out


# --------------------------------------------------------------------------- #
#  Data shapes
# --------------------------------------------------------------------------- #
@dataclass
class Flag:
    """One screening finding against one row. `value` broke `limit` on `channel`."""
    code: str
    severity: str
    message: str
    channel: str = ""
    value: Optional[float] = None
    limit: Optional[float] = None

    def __str__(self) -> str:
        return f"[{self.severity.upper()}] {self.code}: {self.message}"


#: Canonical field name -> human label used in output sheets. Order matters: it is
#: the column order of the Accepted/Rejected sheets.
#: The 27 columns of the wings-team sheet, IN SHEET ORDER and with the sheet's
#: own labels, so every output table reads as the same document the team filled
#: in rather than a rearrangement of it. `converged` and `iteration` follow as
#: optional extras: they are not on the standard sheet, but teams add them and
#: the parser understands them when they appear.
CANONICAL_FIELDS: tuple[tuple[str, str], ...] = (
    ("contributor",         "Contributor"),
    ("component",           "Front or Rear Wing?"),
    ("ride_height_mm",      "Ride-Height (mm)"),
    ("speed_ms",            "Velocity (m/s)"),
    ("desired_yplus",       "Desired Y+"),
    ("min_surface_mesh",    "Min Surface Mesh Length"),
    ("max_surface_mesh",    "Max Surface Mesh Length"),
    ("first_layer_height_m", "First Layer Height (m)"),
    ("n_layers",            "Number of Layers"),
    ("min_ortho_quality",   "Min Orthogonal Quality"),
    ("max_skewness",        "Max Skewness"),
    ("max_aspect_ratio",    "Max Aspect Ratio"),
    ("viscous_model",       "Viscous Model"),
    ("scheme",              "Scheme"),
    ("order",               "Order"),
    ("pseudo_time_step",    "Pseudo Time Step"),
    ("courant_number",      "Courant Number"),
    ("initialization",      "Initialization"),
    ("lift_force_N",        "Lift Force (N)"),
    ("lift_coeff",          "Lift Coefficient"),
    ("drag_force_N",        "Drag Force (N)"),
    ("drag_coeff",          "Drag Coefficient"),
    ("max_pressure_Pa",     "Max Pressure (Pa)"),
    ("min_pressure_Pa",     "Min. Pressure (Pa)"),
    ("mass_imbalance",      "Mass Imbalance (kg/s)"),
    ("avg_yplus",           "Average Y+"),
    ("notes",               "Notes"),
    # --- optional extras, only populated when the sheet carries them -------- #
    ("converged",           "Converged"),
    ("iteration",           "Iteration"),
)

#: The solver-setup columns. These do not gate a run on their own numbers the way
#: y+ or skewness do — they describe METHOD, and method matters comparatively:
#: two runs at one operating point solved with different turbulence models, or one
#: first-order and one second-order, are not two samples of the same quantity and
#: averaging them is averaging apples and oranges. `_screen_setup_consistency`
#: uses these; `_screen_solver_setup` judges the two that do have a right answer.
SETUP_FIELDS: tuple[str, ...] = (
    "viscous_model", "scheme", "order", "pseudo_time_step",
    "courant_number", "initialization",
)

_NUMERIC_FIELDS = {
    "ride_height_mm", "speed_ms", "desired_yplus", "avg_yplus",
    "min_surface_mesh", "max_surface_mesh", "first_layer_height_m", "n_layers",
    "min_ortho_quality", "max_skewness", "max_aspect_ratio",
    "lift_force_N", "lift_coeff", "drag_force_N", "drag_coeff",
    "max_pressure_Pa", "min_pressure_Pa", "mass_imbalance", "courant_number",
    "iteration",
}


@dataclass
class RunRow:
    """
    One row of the run log, parsed into canonical fields. `raw` keeps every original
    cell keyed by its original header, so nothing the team typed is ever lost — the
    Screening Report can quote a column this module does not model.
    """
    source_row: int = 0
    sheet: str = ""
    contributor: Optional[str] = None
    component: Optional[str] = None
    ride_height_mm: Optional[float] = None
    speed_ms: Optional[float] = None
    desired_yplus: Optional[float] = None
    avg_yplus: Optional[float] = None
    min_surface_mesh: Optional[float] = None
    max_surface_mesh: Optional[float] = None
    first_layer_height_m: Optional[float] = None
    n_layers: Optional[float] = None
    min_ortho_quality: Optional[float] = None
    max_skewness: Optional[float] = None
    max_aspect_ratio: Optional[float] = None
    viscous_model: Optional[str] = None
    scheme: Optional[str] = None
    order: Optional[str] = None
    pseudo_time_step: Optional[str] = None
    courant_number: Optional[float] = None
    initialization: Optional[str] = None
    lift_force_N: Optional[float] = None
    lift_coeff: Optional[float] = None
    drag_force_N: Optional[float] = None
    drag_coeff: Optional[float] = None
    max_pressure_Pa: Optional[float] = None
    min_pressure_Pa: Optional[float] = None
    mass_imbalance: Optional[float] = None
    converged: Optional[bool] = None
    iteration: Optional[float] = None
    notes: Optional[str] = None
    raw: dict = field(default_factory=dict)

    def label(self) -> str:
        who = (self.contributor or "?").strip()
        comp = (self.component or "?").strip()
        rh = "?" if self.ride_height_mm is None else f"{self.ride_height_mm:g}mm"
        return f"row {self.source_row} ({who}, {comp}, {rh})"

    def is_blank(self) -> bool:
        """True if every canonical field is empty — a filler row, not a run."""
        for key, _ in CANONICAL_FIELDS:
            v = getattr(self, key, None)
            if v is not None and not (isinstance(v, str) and not v.strip()):
                return False
        return True


@dataclass
class Derived:
    """Quantities this module computed from a row, kept separate from reported data."""
    q_Pa: Optional[float] = None
    cp_max: Optional[float] = None
    cp_min: Optional[float] = None
    implied_area_lift_m2: Optional[float] = None
    implied_area_drag_m2: Optional[float] = None
    reference_area_used_m2: Optional[float] = None
    lift_coeff_derived: Optional[float] = None
    drag_coeff_derived: Optional[float] = None
    wall_treatment: str = WallTreatment.UNKNOWN
    yplus_target_miss: Optional[float] = None
    outlier_z: Optional[float] = None
    discretisation: str = Discretisation.UNKNOWN
    setup_signature: tuple = ()
    setup_matches_group: Optional[bool] = None     # None = not enough peers to tell

    def effective_lift_coeff(self, row: RunRow) -> Optional[float]:
        """Reported Cl if present, else the derived one, else None."""
        return row.lift_coeff if row.lift_coeff is not None else self.lift_coeff_derived

    def effective_drag_coeff(self, row: RunRow) -> Optional[float]:
        return row.drag_coeff if row.drag_coeff is not None else self.drag_coeff_derived


@dataclass
class Verdict:
    """A row plus everything screening concluded about it."""
    row: RunRow
    flags: list = field(default_factory=list)          # list[Flag]
    derived: Derived = field(default_factory=Derived)
    case: Optional["CaseKey"] = None

    @property
    def severity(self) -> str:
        return Severity.worst(f.severity for f in self.flags)

    @property
    def accepted(self) -> bool:
        return not any(f.severity == Severity.REJECT for f in self.flags)

    @property
    def reject_codes(self) -> list:
        return [f.code for f in self.flags if f.severity == Severity.REJECT]

    @property
    def warn_codes(self) -> list:
        return [f.code for f in self.flags if f.severity == Severity.WARN]

    def reason(self) -> str:
        """One-line human summary of why this row was rejected (or warned)."""
        rejects = [f for f in self.flags if f.severity == Severity.REJECT]
        chosen = rejects or [f for f in self.flags if f.severity == Severity.WARN]
        return "; ".join(f.message for f in chosen)


@dataclass(frozen=True)
class CaseKey:
    """The operating point a run belongs to. Rows sharing a key get averaged."""
    component: str
    ride_height_mm: Optional[float]
    speed_ms: Optional[float]

    def label(self) -> str:
        rh = "?" if self.ride_height_mm is None else f"{self.ride_height_mm:g} mm"
        v = "?" if self.speed_ms is None else f"{self.speed_ms:g} m/s"
        return f"{self.component} @ {rh}, {v}"

    def sort_key(self) -> tuple:
        return (self.component,
                self.ride_height_mm if self.ride_height_mm is not None else 1e18,
                self.speed_ms if self.speed_ms is not None else 1e18)


def _mean_sd(values: Sequence[float]) -> tuple:
    """(mean, sample stdev or None, min, max) for a non-empty sequence."""
    if not values:
        return (None, None, None, None)
    mean = sum(values) / len(values)
    sd = statistics.stdev(values) if len(values) > 1 else None
    return (mean, sd, min(values), max(values))


@dataclass
class ConsolidatedCase:
    """
    The averaged answer at one operating point, with the evidence behind it.

    `spread_pct` is peak-to-peak disagreement across the accepted runs as a
    percentage of the mean — the single number that says whether the team's runs
    actually agree. `n_accepted == 1` means this is one run, not a mean; the
    `confidence` string says so in words.
    """
    case: CaseKey
    n_total: int = 0
    n_accepted: int = 0
    n_rejected: int = 0
    n_warned: int = 0

    lift_coeff_mean: Optional[float] = None
    lift_coeff_sd: Optional[float] = None
    lift_coeff_min: Optional[float] = None
    lift_coeff_max: Optional[float] = None

    drag_coeff_mean: Optional[float] = None
    drag_coeff_sd: Optional[float] = None

    lift_force_mean_N: Optional[float] = None
    lift_force_sd_N: Optional[float] = None
    drag_force_mean_N: Optional[float] = None
    drag_force_sd_N: Optional[float] = None

    reference_area_m2: Optional[float] = None
    reference_area_basis: str = ""
    lift_to_drag: Optional[float] = None
    spread_pct: Optional[float] = None

    #: The METHOD behind the number. A consolidated coefficient is only
    #: meaningful alongside the setup that produced it, so it travels with it.
    viscous_models: list = field(default_factory=list)
    schemes: list = field(default_factory=list)
    discretisations: list = field(default_factory=list)
    initializations: list = field(default_factory=list)
    courant_range: tuple = ()
    setup_consistent: bool = True

    contributors: list = field(default_factory=list)
    reject_reasons: list = field(default_factory=list)     # list[(label, reason)]
    accepted_rows: list = field(default_factory=list)      # list[Verdict]
    notes: list = field(default_factory=list)

    def setup_summary(self) -> str:
        """One line naming the method the accepted runs used."""
        if not self.viscous_models:
            return "method not recorded"
        parts = [" / ".join(self.viscous_models)]
        if self.discretisations:
            parts.append(" / ".join(self.discretisations))
        if self.schemes:
            parts.append(" / ".join(self.schemes))
        if self.initializations:
            parts.append("init " + " / ".join(self.initializations))
        line = ", ".join(parts)
        return line if self.setup_consistent else "MIXED — " + line

    @property
    def confidence(self) -> str:
        if self.n_accepted == 0:
            return "NO DATA — every run at this point was rejected"
        if self.n_accepted == 1:
            return "SINGLE RUN — not a mean; no spread available"
        if not self.setup_consistent:
            return (f"{self.n_accepted} runs, MIXED SETUP — averaged across "
                    f"different solver methods")
        if self.spread_pct is None:
            return f"{self.n_accepted} runs"
        if self.spread_pct <= 5.0:
            return f"{self.n_accepted} runs, tight ({self.spread_pct:.1f}% spread)"
        if self.spread_pct <= 15.0:
            return f"{self.n_accepted} runs, moderate ({self.spread_pct:.1f}% spread)"
        return f"{self.n_accepted} runs, POOR AGREEMENT ({self.spread_pct:.1f}% spread)"

    def summary(self) -> str:
        if self.n_accepted == 0:
            return f"{self.case.label()}: no usable runs ({self.n_rejected} rejected)"
        cl = "n/a" if self.lift_coeff_mean is None else f"{self.lift_coeff_mean:+.4f}"
        cd = "n/a" if self.drag_coeff_mean is None else f"{self.drag_coeff_mean:.4f}"
        return (f"{self.case.label()}: Cl={cl} Cd={cd} "
                f"[{self.n_accepted}/{self.n_total} runs kept] — {self.confidence}")


@dataclass
class ConsolidationReport:
    """Everything the pipeline produced: the answer, the audit trail, the config."""
    cases: list = field(default_factory=list)          # list[ConsolidatedCase]
    verdicts: list = field(default_factory=list)       # list[Verdict], input order
    config: ScreenConfig = field(default_factory=ScreenConfig)
    source: str = ""
    sheet: str = ""
    parse_warnings: list = field(default_factory=list)
    unmapped_headers: list = field(default_factory=list)

    # -- roll-ups ---------------------------------------------------------- #
    @property
    def n_rows(self) -> int:
        return len(self.verdicts)

    @property
    def accepted(self) -> list:
        return [v for v in self.verdicts if v.accepted]

    @property
    def rejected(self) -> list:
        return [v for v in self.verdicts if not v.accepted]

    @property
    def ok(self) -> bool:
        return any(c.n_accepted > 0 for c in self.cases)

    def flag_tally(self) -> dict:
        """code -> count, most common first. What is actually going wrong."""
        tally: dict = {}
        for v in self.verdicts:
            for f in v.flags:
                tally[f.code] = tally.get(f.code, 0) + 1
        return dict(sorted(tally.items(), key=lambda kv: (-kv[1], kv[0])))

    def contributor_stats(self) -> list:
        """
        Per-contributor: runs submitted, accepted, and the flags they hit most.

        The Contributor column was carrying nothing but a name. This is what it
        is actually good for — not ranking people, but showing the team where a
        recurring setup mistake lives, so it gets fixed once at the source
        instead of being screened out of every batch forever.
        """
        by_who: dict = {}
        for v in self.verdicts:
            who = (v.row.contributor or "unattributed").strip() or "unattributed"
            rec = by_who.setdefault(who, {"contributor": who, "runs": 0,
                                          "accepted": 0, "rejected": 0,
                                          "flags": {}})
            rec["runs"] += 1
            rec["accepted" if v.accepted else "rejected"] += 1
            for f in v.flags:
                rec["flags"][f.code] = rec["flags"].get(f.code, 0) + 1
        out = []
        for rec in by_who.values():
            ranked = sorted(rec["flags"].items(), key=lambda kv: (-kv[1], kv[0]))
            rec["top_flags"] = ", ".join(f"{k} x{n}" for k, n in ranked[:4])
            rec["acceptance_pct"] = (100.0 * rec["accepted"] / rec["runs"]
                                     if rec["runs"] else 0.0)
            out.append(rec)
        return sorted(out, key=lambda r: (-r["runs"], r["contributor"]))

    def summary(self) -> str:
        lines = [
            f"{self.n_rows} run(s) parsed from {os.path.basename(self.source) or 'input'}"
            + (f" [{self.sheet}]" if self.sheet else ""),
            f"{len(self.accepted)} accepted, {len(self.rejected)} rejected, "
            f"{len(self.cases)} operating point(s)",
        ]
        for c in self.cases:
            lines.append("  - " + c.summary())
        tally = self.flag_tally()
        if tally:
            top = ", ".join(f"{k} x{v}" for k, v in list(tally.items())[:6])
            lines.append(f"  flags: {top}")
        return "\n".join(lines)


# --------------------------------------------------------------------------- #
#  1) PARSE — find the real header row, map columns, coerce values
# --------------------------------------------------------------------------- #
def _normalise_header(text) -> str:
    """Lowercase, drop parenthetical units and every non-alphanumeric character."""
    if text is None:
        return ""
    s = str(text).strip().lower()
    s = re.sub(r"\([^)]*\)", " ", s)           # drop "(mm)", "(kg/s)", ...
    s = s.replace("+", "plus")                  # keep y+ distinguishable from y
    s = re.sub(r"[^a-z0-9]+", "", s)
    return s


#: normalised header -> canonical field. Longest/most specific aliases first is not
#: needed because matching is exact on the normalised string; overlaps are resolved
#: by the regex fallback in `_match_header`.
_HEADER_ALIASES: dict = {
    # identity / grouping
    "contributor": "contributor", "name": "contributor", "engineer": "contributor",
    "author": "contributor", "who": "contributor", "runby": "contributor",
    "user": "contributor", "member": "contributor",

    "frontorrearwing": "component", "wing": "component", "component": "component",
    "part": "component", "surface": "component", "configuration": "component",
    "config": "component", "geometry": "component", "case": "component",
    "element": "component",

    "rideheight": "ride_height_mm", "rh": "ride_height_mm",
    "rideheightmm": "ride_height_mm", "height": "ride_height_mm",
    "groundclearance": "ride_height_mm",

    "velocity": "speed_ms", "speed": "speed_ms", "freestreamvelocity": "speed_ms",
    "inletvelocity": "speed_ms", "vinf": "speed_ms", "u": "speed_ms",
    "airspeed": "speed_ms",

    # mesh intent / metrics
    "desiredyplus": "desired_yplus", "targetyplus": "desired_yplus",
    "yplustarget": "desired_yplus", "desiredwallyplus": "desired_yplus",
    "averageyplus": "avg_yplus", "avgyplus": "avg_yplus", "meanyplus": "avg_yplus",
    "yplus": "avg_yplus", "actualyplus": "avg_yplus", "wallyplus": "avg_yplus",

    "minsurfacemeshlength": "min_surface_mesh", "minsurfacemesh": "min_surface_mesh",
    "minsurfacesize": "min_surface_mesh", "minelementsize": "min_surface_mesh",
    "maxsurfacemeshlength": "max_surface_mesh", "maxsurfacemesh": "max_surface_mesh",
    "maxsurfacesize": "max_surface_mesh", "maxelementsize": "max_surface_mesh",

    "firstlayerheight": "first_layer_height_m", "firstcellheight": "first_layer_height_m",
    "firstlayerthickness": "first_layer_height_m", "wallspacing": "first_layer_height_m",
    "numberoflayers": "n_layers", "nlayers": "n_layers", "layers": "n_layers",
    "inflationlayers": "n_layers", "prismlayers": "n_layers",
    "numberofinflationlayers": "n_layers",

    "minorthogonalquality": "min_ortho_quality", "minorthoquality": "min_ortho_quality",
    "orthogonalquality": "min_ortho_quality", "minimumorthogonalquality": "min_ortho_quality",
    "maxskewness": "max_skewness", "skewness": "max_skewness",
    "maximumskewness": "max_skewness",
    "maxaspectratio": "max_aspect_ratio", "aspectratio": "max_aspect_ratio",
    "maximumaspectratio": "max_aspect_ratio",

    # solver setup
    "viscousmodel": "viscous_model", "turbulencemodel": "viscous_model",
    "model": "viscous_model", "turbulence": "viscous_model",
    "scheme": "scheme", "pressurevelocitycoupling": "scheme", "coupling": "scheme",
    "solver": "scheme",
    "order": "order", "discretisationorder": "order", "discretizationorder": "order",
    "spatialdiscretization": "order",
    "pseudotimestep": "pseudo_time_step", "pseudotransient": "pseudo_time_step",
    "timestep": "pseudo_time_step",
    "courantnumber": "courant_number", "courant": "courant_number", "cfl": "courant_number",
    "initialization": "initialization", "initialisation": "initialization",
    "init": "initialization",

    # results
    "liftforce": "lift_force_N", "lift": "lift_force_N", "downforce": "lift_force_N",
    "fl": "lift_force_N", "verticalforce": "lift_force_N",
    "liftcoefficient": "lift_coeff", "cl": "lift_coeff", "liftcoeff": "lift_coeff",
    "coefficientoflift": "lift_coeff",
    "dragforce": "drag_force_N", "drag": "drag_force_N", "fd": "drag_force_N",
    "dragcoefficient": "drag_coeff", "cd": "drag_coeff", "dragcoeff": "drag_coeff",
    "coefficientofdrag": "drag_coeff",

    "maxpressure": "max_pressure_Pa", "maximumpressure": "max_pressure_Pa",
    "maxstaticpressure": "max_pressure_Pa",
    "minpressure": "min_pressure_Pa", "minimumpressure": "min_pressure_Pa",
    "minstaticpressure": "min_pressure_Pa",

    "massimbalance": "mass_imbalance", "continuityimbalance": "mass_imbalance",
    "massflowimbalance": "mass_imbalance", "netmassflow": "mass_imbalance",
    "massflowrateimbalance": "mass_imbalance",

    "converged": "converged", "convergence": "converged", "hasconverged": "converged",
    "iteration": "iteration", "iterations": "iteration", "run": "iteration",
    "runid": "iteration", "attempt": "iteration", "runnumber": "iteration",

    "notes": "notes", "note": "notes", "comment": "notes", "comments": "notes",
    "remarks": "notes", "description": "notes",
}

#: Fallback regex probes for headers the alias table misses, tried in order.
_HEADER_PATTERNS: tuple = (
    (r"contributor|engineer|author|runby", "contributor"),
    (r"rideheight|clearance", "ride_height_mm"),
    (r"velocity|speed", "speed_ms"),
    (r"desired.*yplus|target.*yplus", "desired_yplus"),
    (r"(average|avg|mean|actual).*yplus", "avg_yplus"),
    (r"min.*surface.*(mesh|length|size)", "min_surface_mesh"),
    (r"max.*surface.*(mesh|length|size)", "max_surface_mesh"),
    (r"first.*(layer|cell)", "first_layer_height_m"),
    (r"(number|no).*layers|layers", "n_layers"),
    (r"ortho", "min_ortho_quality"),
    (r"skew", "max_skewness"),
    (r"aspect", "max_aspect_ratio"),
    (r"viscous|turbulen", "viscous_model"),
    (r"lift.*coef|^cl$", "lift_coeff"),
    (r"drag.*coef|^cd$", "drag_coeff"),
    (r"lift.*force|downforce", "lift_force_N"),
    (r"drag.*force", "drag_force_N"),
    (r"max.*pressure", "max_pressure_Pa"),
    (r"min.*pressure", "min_pressure_Pa"),
    (r"imbalance", "mass_imbalance"),
    (r"converg", "converged"),
    (r"note|comment|remark", "notes"),
    (r"wing|component|config", "component"),
)


def _match_header(text) -> Optional[str]:
    """Map one header cell to a canonical field name, or None if unrecognised."""
    key = _normalise_header(text)
    if not key:
        return None
    if key in _HEADER_ALIASES:
        return _HEADER_ALIASES[key]
    for pattern, field_name in _HEADER_PATTERNS:
        if re.search(pattern, key):
            return field_name
    return None


_TRUEISH = {"y", "yes", "true", "1", "converged", "ok", "pass", "passed", "done"}
_FALSEISH = {"n", "no", "false", "0", "not converged", "notconverged", "diverged",
             "fail", "failed", "unconverged"}


def _coerce_number(value) -> Optional[float]:
    """
    Best-effort numeric read. Handles the things spreadsheets actually contain:
    scientific notation as text, thousands separators, comma decimals, stray units,
    percent signs, and '-' / 'n/a' placeholders.
    """
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return None if (isinstance(value, float) and math.isnan(value)) else float(value)
    text = str(value).strip()
    if not text or text.lower() in {"-", "--", "n/a", "na", "none", "nan", "?", "tbd"}:
        return None
    percent = text.endswith("%")
    text = text.rstrip("%").strip()
    # Strip a trailing unit token ("26.8 m/s", "50 mm") but keep the number.
    text = re.sub(r"\s*[a-zA-Z/^°µ]+\s*$", "", text).strip()
    text = text.replace("\u2212", "-").replace(" ", "")
    if "," in text and "." in text:
        text = text.replace(",", "")                    # 1,234.5
    elif text.count(",") == 1 and re.search(r",\d{1,3}$", text):
        text = text.replace(",", ".")                   # 26,8 -> 26.8
    else:
        text = text.replace(",", "")
    try:
        num = float(text)
    except ValueError:
        return None
    return num / 100.0 if percent else num


def _coerce_bool(value) -> Optional[bool]:
    if value is None:
        return None
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return bool(value)
    text = str(value).strip().lower()
    if not text:
        return None
    if text in _TRUEISH:
        return True
    if text in _FALSEISH or text.replace(" ", "") in _FALSEISH:
        return False
    return None


def _coerce_text(value) -> Optional[str]:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _load_grid(source, sheet: Optional[str] = None) -> tuple:
    """
    Read any supported input into a rectangular grid of cell values.

    Accepts: a path to .xlsx/.xlsm/.csv/.tsv, raw bytes, a file-like object, or an
    already-built list-of-lists. Returns (grid, sheet_name, source_label).
    """
    # Already a grid.
    if isinstance(source, list):
        return ([list(r) for r in source], sheet or "", "<in-memory>")

    data: Optional[bytes] = None
    label = ""
    if isinstance(source, (bytes, bytearray)):
        data = bytes(source)
        label = "<bytes>"
    elif hasattr(source, "read"):
        data = source.read()
        label = getattr(source, "name", "<stream>")
        if isinstance(data, str):
            data = data.encode("utf-8")
    elif isinstance(source, str):
        label = source
        with open(source, "rb") as fh:
            data = fh.read()
    else:
        raise TypeError(f"unsupported run-log source: {type(source).__name__}")

    is_xlsx = data[:2] == b"PK" or str(label).lower().endswith((".xlsx", ".xlsm", ".xltx"))
    if is_xlsx:
        try:
            from openpyxl import load_workbook            # lazy: optional dependency
        except ImportError as exc:                        # pragma: no cover
            raise ImportError(
                "Reading an .xlsx run log needs openpyxl (`pip install openpyxl`). "
                "Export the sheet as CSV to use this feature without it."
            ) from exc
        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb[wb.sheetnames[0]]
        grid = [list(r) for r in ws.iter_rows(values_only=True)]
        name = ws.title
        wb.close()
        return (grid, name, label)

    # CSV / TSV
    text = data.decode("utf-8-sig", errors="replace")
    try:
        dialect = csv.Sniffer().sniff(text[:4096], delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel
    grid = [row for row in csv.reader(io.StringIO(text), dialect)]
    return (grid, sheet or "", label)


def _find_header_row(grid: Sequence[Sequence]) -> int:
    """
    Locate the real column-header row.

    The wings sheet puts a banner ("Wings Team Simulation Results", "Volume Mesh
    Metrics") above the header, so row 0 is not it. Scoring by how many cells map
    to known fields finds the header wherever it sits, and refuses to be fooled by
    a banner row that happens to contain one recognisable word.
    """
    best_idx, best_score = 0, -1
    for idx, row in enumerate(grid[:25]):          # headers live near the top
        score = sum(1 for cell in row if _match_header(cell))
        # A header row also has several non-empty cells; a 1-word banner does not.
        filled = sum(1 for cell in row if _coerce_text(cell))
        if filled >= 3 and score > best_score:
            best_idx, best_score = idx, score
    return best_idx if best_score >= 3 else 0


def parse_rows_from_grid(grid: Sequence[Sequence], sheet: str = "",
                         header_row: Optional[int] = None) -> tuple:
    """
    Turn a raw grid into RunRows. Returns (rows, warnings, unmapped_headers).

    Split out from `parse_run_log` so tests (and callers holding a grid already)
    can exercise the mapping logic without touching a file.
    """
    warnings: list = []
    if not grid:
        return ([], ["input is empty"], [])

    hdr_idx = _find_header_row(grid) if header_row is None else header_row
    header = list(grid[hdr_idx]) if hdr_idx < len(grid) else []

    mapping: dict = {}                 # column index -> canonical field
    unmapped: list = []
    seen: dict = {}
    for col, cell in enumerate(header):
        canon = _match_header(cell)
        label = _coerce_text(cell)
        if canon is None:
            if label:
                unmapped.append(label)
            continue
        if canon in seen:
            # Two columns claim the same field: keep the first, report the clash
            # rather than letting the later one silently overwrite it.
            warnings.append(
                f"columns {seen[canon]!r} and {label!r} both map to '{canon}'; "
                f"using {seen[canon]!r}")
            if label:
                unmapped.append(label)
            continue
        seen[canon] = label or canon
        mapping[col] = canon

    if not mapping:
        return ([], [f"no recognisable column headers found in row {hdr_idx + 1}"], unmapped)

    missing_core = [f for f in ("lift_force_N", "lift_coeff") if f not in mapping.values()]
    if len(missing_core) == 2:
        warnings.append("no lift force or lift coefficient column found — "
                        "there is nothing to average")

    rows: list = []
    for r_idx in range(hdr_idx + 1, len(grid)):
        raw_row = grid[r_idx]
        row = RunRow(source_row=r_idx + 1, sheet=sheet)
        for col, canon in mapping.items():
            value = raw_row[col] if col < len(raw_row) else None
            if canon in _NUMERIC_FIELDS:
                setattr(row, canon, _coerce_number(value))
            elif canon == "converged":
                row.converged = _coerce_bool(value)
            else:
                setattr(row, canon, _coerce_text(value))
        # Keep every original cell, mapped or not, under its own header.
        for col, cell in enumerate(header):
            label = _coerce_text(cell)
            if label:
                row.raw[label] = raw_row[col] if col < len(raw_row) else None
        if not row.is_blank():
            rows.append(row)

    return (rows, warnings, unmapped)


def parse_run_log(source, sheet: Optional[str] = None,
                  header_row: Optional[int] = None) -> tuple:
    """
    Read an ANSYS run log from a path, bytes, stream or grid.

    Returns (rows, warnings, unmapped_headers, sheet_name, source_label).
    Tolerant by design: a banner row above the header, renamed columns, blank
    filler rows and units in the header text are all normal and are handled.
    """
    grid, sheet_name, label = _load_grid(source, sheet)
    rows, warnings, unmapped = parse_rows_from_grid(grid, sheet_name, header_row)
    return (rows, warnings, unmapped, sheet_name, label)


# --------------------------------------------------------------------------- #
#  2) SCREEN — judge each row, with a reason for every verdict
# --------------------------------------------------------------------------- #
def _looks_like_test_row(row: RunRow, patterns: Sequence[str]) -> Optional[str]:
    """Return the matched marker if the row is scratch data, else None."""
    haystacks = [row.contributor or "", row.notes or "", row.component or ""]
    text = " ".join(haystacks).lower()
    for pattern in patterns:
        # Word-ish match so "Test" hits but "Testarossa" / "latest" do not.
        if re.search(rf"(?<![a-z]){re.escape(pattern.lower())}(?![a-z])", text):
            return pattern
    return None


def _screen_mesh(row: RunRow, cfg: ScreenConfig, flags: list) -> None:
    """Mesh-quality gates: orthogonal quality, skewness, aspect ratio, sizing."""
    q = row.min_ortho_quality
    if q is not None:
        if q < cfg.ortho_quality_reject:
            flags.append(Flag(
                "ORTHO_QUALITY", Severity.REJECT,
                f"min orthogonal quality {q:.3f} is below Fluent's usable floor "
                f"of {cfg.ortho_quality_reject:.2f} — the cells are degenerate",
                "min_ortho_quality", q, cfg.ortho_quality_reject))
        elif q < cfg.ortho_quality_warn:
            flags.append(Flag(
                "ORTHO_QUALITY_LOW", Severity.WARN,
                f"min orthogonal quality {q:.3f} is marginal (want "
                f">{cfg.ortho_quality_warn:.2f})",
                "min_ortho_quality", q, cfg.ortho_quality_warn))

    s = row.max_skewness
    if s is not None:
        if s > cfg.skewness_reject:
            flags.append(Flag(
                "SKEWNESS", Severity.REJECT,
                f"max skewness {s:.3f} exceeds {cfg.skewness_reject:.2f} — "
                f"degenerate cells will corrupt the solution",
                "max_skewness", s, cfg.skewness_reject))
        elif s > cfg.skewness_warn:
            flags.append(Flag(
                "SKEWNESS_HIGH", Severity.WARN,
                f"max skewness {s:.3f} is high (want <{cfg.skewness_warn:.2f})",
                "max_skewness", s, cfg.skewness_warn))

    ar = row.max_aspect_ratio
    if ar is not None:
        if ar > cfg.aspect_ratio_reject:
            flags.append(Flag(
                "ASPECT_RATIO", Severity.REJECT,
                f"max aspect ratio {ar:,.0f} exceeds {cfg.aspect_ratio_reject:,.0f}",
                "max_aspect_ratio", ar, cfg.aspect_ratio_reject))
        elif ar > cfg.aspect_ratio_warn:
            flags.append(Flag(
                "ASPECT_RATIO_HIGH", Severity.WARN,
                f"max aspect ratio {ar:,.0f} is high — expected in inflation "
                f"layers, a concern anywhere else",
                "max_aspect_ratio", ar, cfg.aspect_ratio_warn))

    lo, hi = row.min_surface_mesh, row.max_surface_mesh
    if lo is not None and hi is not None:
        if lo > hi:
            flags.append(Flag(
                "MESH_LENGTH_INVERTED", Severity.REJECT,
                f"min surface mesh length ({lo:g}) is larger than the max ({hi:g}) "
                f"— the two values are swapped or mistyped",
                "min_surface_mesh", lo, hi))
        elif lo > 0 and hi / lo > cfg.mesh_length_ratio_warn:
            flags.append(Flag(
                "MESH_LENGTH_RATIO", Severity.WARN,
                f"surface mesh max/min ratio is {hi / lo:,.0f}:1 — check for a "
                f"decimal-place slip in one of them",
                "max_surface_mesh", hi / lo, cfg.mesh_length_ratio_warn))

    flh = row.first_layer_height_m
    if flh is not None:
        if flh <= 0:
            flags.append(Flag(
                "FIRST_LAYER_HEIGHT", Severity.REJECT,
                f"first layer height {flh:g} is not positive",
                "first_layer_height_m", flh, 0.0))
        elif lo is not None and lo > 0 and flh > lo:
            flags.append(Flag(
                "FIRST_LAYER_TOO_TALL", Severity.WARN,
                f"first layer height ({flh:g}) exceeds the min surface mesh "
                f"length ({lo:g}) — the boundary layer is coarser than the surface",
                "first_layer_height_m", flh, lo))

    n = row.n_layers
    if n is not None:
        if n < cfg.min_layers_reject:
            flags.append(Flag(
                "LAYER_COUNT", Severity.REJECT,
                f"{n:g} inflation layer(s) cannot resolve a boundary layer",
                "n_layers", n, float(cfg.min_layers_reject)))
        elif n < cfg.min_layers_warn:
            flags.append(Flag(
                "LAYER_COUNT_LOW", Severity.WARN,
                f"only {n:g} inflation layers (want >={cfg.min_layers_warn})",
                "n_layers", n, float(cfg.min_layers_warn)))


def _screen_yplus(row: RunRow, cfg: ScreenConfig, derived: Derived,
                  flags: list) -> None:
    """
    The near-wall gate — judged against the band the row's own turbulence model
    requires, because y+ = 25 is excellent for k-omega SST and unusable for
    k-epsilon with standard wall functions.
    """
    treatment = wall_treatment_for(row.viscous_model)
    derived.wall_treatment = treatment

    y = row.avg_yplus
    if y is None:
        return

    if treatment == WallTreatment.WALL_FUNCTION:
        warn_lo, warn_hi = cfg.yplus_wf_warn
        rej_lo, rej_hi = cfg.yplus_wf_reject
        rationale = ("a wall-function closure needs the first cell in the log "
                     "layer")
    elif treatment == WallTreatment.RESOLVED:
        warn_lo, warn_hi = cfg.yplus_resolved_warn
        rej_lo, rej_hi = cfg.yplus_resolved_reject
        rationale = ("a sublayer-resolving closure needs y+ of order 1")
    elif treatment == WallTreatment.AUTOMATIC:
        warn_lo, warn_hi = cfg.yplus_auto_warn
        rej_lo, rej_hi = cfg.yplus_auto_reject
        rationale = "automatic wall treatment blends, but not without limit"
    else:
        # Unknown model: warn on an extreme value, never reject on a rule we
        # cannot justify for a closure we did not recognise.
        if y > cfg.yplus_auto_reject[1] or y <= 0:
            flags.append(Flag(
                "YPLUS_EXTREME", Severity.WARN,
                f"average y+ {y:g} is extreme and the viscous model "
                f"({row.viscous_model or 'unspecified'}) was not recognised, so "
                f"the correct band is unknown",
                "avg_yplus", y, cfg.yplus_auto_reject[1]))
        return

    if y < rej_lo or y > rej_hi:
        if y < rej_lo and treatment == WallTreatment.WALL_FUNCTION:
            detail = (f"below y+ = {rej_lo:g}, where the log law meets the viscous "
                      f"sublayer — the wall function is modelling a cell that sits "
                      f"inside the sublayer, so its wall shear is not physical")
        else:
            detail = f"outside the usable band {rej_lo:g}-{rej_hi:g}"
        flags.append(Flag(
            "YPLUS_BAND", Severity.REJECT,
            f"average y+ {y:g} with {row.viscous_model or 'this model'} "
            f"({treatment}) is {detail}",
            "avg_yplus", y, rej_lo if y < rej_lo else rej_hi))
    elif y < warn_lo or y > warn_hi:
        flags.append(Flag(
            "YPLUS_MARGINAL", Severity.WARN,
            f"average y+ {y:g} sits outside the preferred band "
            f"{warn_lo:g}-{warn_hi:g} for {row.viscous_model or 'this model'} "
            f"({rationale})",
            "avg_yplus", y, warn_lo if y < warn_lo else warn_hi))

    # Did the mesh hit the y+ it was built for?
    if row.desired_yplus:
        miss = abs(y - row.desired_yplus) / abs(row.desired_yplus)
        derived.yplus_target_miss = miss
        if miss > cfg.yplus_target_miss_warn:
            flags.append(Flag(
                "YPLUS_TARGET_MISS", Severity.WARN,
                f"achieved y+ {y:g} misses the {row.desired_yplus:g} the mesh was "
                f"sized for by {miss * 100:.0f}% — the first-layer height needs "
                f"revisiting",
                "avg_yplus", miss, cfg.yplus_target_miss_warn))


def _screen_solution(row: RunRow, cfg: ScreenConfig, flags: list) -> None:
    """Convergence and continuity — an unconverged force is not a force."""
    if row.converged is False and cfg.reject_unconverged:
        flags.append(Flag(
            "NOT_CONVERGED", Severity.REJECT,
            "the run is marked not converged", "converged"))

    mi = row.mass_imbalance
    if mi is not None:
        mag = abs(mi)
        if mag > cfg.mass_imbalance_reject:
            flags.append(Flag(
                "MASS_IMBALANCE", Severity.REJECT,
                f"mass imbalance {mi:.3g} kg/s exceeds "
                f"{cfg.mass_imbalance_reject:.1g} — continuity did not close",
                "mass_imbalance", mag, cfg.mass_imbalance_reject))
        elif mag > cfg.mass_imbalance_warn:
            flags.append(Flag(
                "MASS_IMBALANCE_HIGH", Severity.WARN,
                f"mass imbalance {mi:.3g} kg/s is above "
                f"{cfg.mass_imbalance_warn:.1g}",
                "mass_imbalance", mag, cfg.mass_imbalance_warn))


def _screen_solver_setup(row: RunRow, cfg: ScreenConfig, derived: Derived,
                         flags: list) -> None:
    """
    The solver-setup columns: Scheme, Order, Pseudo Time Step, Courant Number,
    Initialization.

    Only two of these have a defensible right answer on their own. The rest
    describe method, and method is judged comparatively in
    `_screen_setup_consistency` — a SIMPLE run and a Coupled run can both be
    correct, but averaging them together at one operating point is not.
    """
    derived.setup_signature = setup_signature(row)
    derived.discretisation = discretisation_of(row.order)

    if derived.discretisation == Discretisation.FIRST:
        flags.append(Flag(
            "FIRST_ORDER",
            Severity.REJECT if cfg.reject_first_order else Severity.WARN,
            "solved first-order in space, which is numerically diffusive: it "
            "smears the pressure gradients a suction peak is made of, so "
            "downforce comes out low and the wake wide. Fine to start a solve "
            "on, not to report a coefficient from",
            "order"))

    c = row.courant_number
    if c is not None and c > 0:
        if c > cfg.courant_warn_max:
            flags.append(Flag(
                "COURANT_HIGH", Severity.WARN,
                f"pseudo-transient Courant number {c:g} is above "
                f"{cfg.courant_warn_max:g} — large pseudo time steps can settle "
                f"the residuals while the flow field is still moving, which "
                f"looks converged and is not",
                "courant_number", c, cfg.courant_warn_max))
        elif c < cfg.courant_warn_min:
            flags.append(Flag(
                "COURANT_LOW", Severity.WARN,
                f"Courant number {c:g} is very small — the solve advances so "
                f"slowly that the residual history may flatten long before the "
                f"forces have actually stopped moving",
                "courant_number", c, cfg.courant_warn_min))
    elif c is not None and c <= 0:
        flags.append(Flag(
            "COURANT_INVALID", Severity.WARN,
            f"Courant number {c:g} is not positive", "courant_number", c, 0.0))

    # Not a fault, but worth stating: an unrecorded method cannot be compared
    # against the rest of the operating point, so it weakens the group check.
    missing = [label for field_name, label in
               (("viscous_model", "viscous model"), ("scheme", "scheme"),
                ("order", "order"), ("initialization", "initialization"))
               if not _normalise_setup_value(getattr(row, field_name, None))]
    if missing:
        flags.append(Flag(
            "SETUP_UNRECORDED", Severity.WARN,
            "the run does not record its " + ", ".join(missing)
            + " — its method cannot be compared against the other runs at this "
              "operating point",
            "viscous_model"))


def _screen_setup_consistency(verdicts: Sequence[Verdict],
                              cfg: ScreenConfig) -> None:
    """
    Compare each run's method against the rest of its operating point.

    This is the check the solver-setup columns exist for. Every other gate asks
    "is this run valid?"; this one asks "are these runs the same experiment?".
    Two runs at one ride height, one on k-epsilon and one on k-omega SST, can
    both pass every physics gate and still not be two samples of one quantity —
    averaging them produces a number that describes neither.

    The tool reports the split rather than picking a winner: it cannot know which
    model the team meant to standardise on. Set `reject_mixed_turbulence` to make
    the minority a rejection once you have decided.
    """
    if not cfg.check_setup_consistency:
        return

    by_case: dict = {}
    for v in verdicts:
        if v.accepted:
            by_case.setdefault(v.case, []).append(v)

    for case, group in by_case.items():
        if len(group) < 2:
            continue
        # Modal signature across the group, ignoring parts nobody recorded.
        counts: dict = {}
        for v in group:
            counts[v.derived.setup_signature] = \
                counts.get(v.derived.setup_signature, 0) + 1
        modal, modal_n = max(counts.items(), key=lambda kv: (kv[1], kv[0]))
        if len(counts) == 1:
            for v in group:
                v.derived.setup_matches_group = True
            continue

        for v in group:
            sig = v.derived.setup_signature
            if sig == modal:
                v.derived.setup_matches_group = True
                continue
            v.derived.setup_matches_group = False
            differing = [
                f"{label} ({mine or 'unrecorded'} vs {theirs or 'unrecorded'})"
                for label, mine, theirs in zip(_SETUP_PART_LABELS, sig, modal)
                if mine != theirs and (mine or theirs)]
            if not differing:
                continue
            turbulence_differs = sig[0] != modal[0] and (sig[0] or modal[0])
            severity = (Severity.REJECT
                        if turbulence_differs and cfg.reject_mixed_turbulence
                        else Severity.WARN)
            flags_target = v.flags
            flags_target.append(Flag(
                "SETUP_MISMATCH", severity,
                f"solved with a different method from the other "
                f"{modal_n} run(s) at this operating point \u2014 "
                + "; ".join(differing)
                + ". These are not two samples of the same quantity, so the "
                  "mean across them describes neither setup",
                "viscous_model"))


def _screen_pressure(row: RunRow, cfg: ScreenConfig, derived: Derived,
                     flags: list) -> None:
    """
    Pressure-field physics. Assumes the logged pressures are GAUGE (Fluent's
    default), i.e. relative to the operating pressure — the same datum q is in.
    """
    q = derived.q_Pa
    if q is None or q <= 0:
        return

    if row.max_pressure_Pa is not None:
        cp_max = row.max_pressure_Pa / q
        derived.cp_max = cp_max
        lo_r, hi_r = cfg.cp_stagnation_reject
        lo_w, hi_w = cfg.cp_stagnation_warn
        if cp_max < lo_r or cp_max > hi_r:
            flags.append(Flag(
                "CP_STAGNATION", Severity.REJECT,
                f"peak gauge pressure gives Cp_max = {cp_max:.2f}; a stagnation "
                f"point must sit near 1.0, so the reference velocity, density or "
                f"pressure datum in this run is wrong — every coefficient in the "
                f"row is off by that ratio",
                "max_pressure_Pa", cp_max, lo_r if cp_max < lo_r else hi_r))
        elif cp_max < lo_w or cp_max > hi_w:
            flags.append(Flag(
                "CP_STAGNATION_OFF", Severity.WARN,
                f"Cp_max = {cp_max:.2f} deviates from the expected stagnation "
                f"value of 1.0",
                "max_pressure_Pa", cp_max, lo_w if cp_max < lo_w else hi_w))

    if row.min_pressure_Pa is not None:
        cp_min = row.min_pressure_Pa / q
        derived.cp_min = cp_min
        if cp_min < cfg.cp_suction_floor_reject:
            flags.append(Flag(
                "CP_SUCTION_EXTREME", Severity.REJECT,
                f"Cp_min = {cp_min:.1f} is not a physical suction peak — expect a "
                f"single bad cell or a pressure datum error",
                "min_pressure_Pa", cp_min, cfg.cp_suction_floor_reject))
        elif cfg.expect_downforce and cp_min > cfg.cp_suction_min_warn:
            flags.append(Flag(
                "CP_NO_SUCTION", Severity.WARN,
                f"Cp_min = {cp_min:.2f} shows almost no suction on a surface that "
                f"should be making downforce — check the force report is summing "
                f"over the right wall zone",
                "min_pressure_Pa", cp_min, cfg.cp_suction_min_warn))


def _screen_forces(row: RunRow, cfg: ScreenConfig, flags: list) -> None:
    """Is there a result at all, and does its sign make sense?"""
    if row.lift_force_N is None and row.lift_coeff is None:
        flags.append(Flag(
            "NO_RESULT", Severity.REJECT,
            "no lift force and no lift coefficient — nothing to average",
            "lift_force_N"))
        return

    if cfg.expect_downforce:
        # Sheet convention (and KinematiK's): negative lift = downforce.
        for value, name in ((row.lift_force_N, "lift force"),
                            (row.lift_coeff, "lift coefficient")):
            if value is not None and value > 0:
                flags.append(Flag(
                    "LIFT_SIGN", Severity.WARN,
                    f"{name} is positive ({value:g}); in this sheet's convention "
                    f"that is lift, not downforce — either the run is genuinely "
                    f"making lift or Fluent's up-positive value was pasted in "
                    f"without the sign flip",
                    "lift_force_N" if "force" in name else "lift_coeff", value, 0.0))
                break

    if row.lift_coeff is not None and abs(row.lift_coeff) > 20:
        flags.append(Flag(
            "COEFF_MAGNITUDE", Severity.REJECT,
            f"lift coefficient {row.lift_coeff:g} is far outside any physical "
            f"range for a wing — almost certainly a reference-area error",
            "lift_coeff", abs(row.lift_coeff), 20.0))


def _resolve_reference_areas(verdicts: Sequence[Verdict], cfg: ScreenConfig) -> None:
    """
    Establish a reference area per case group, then flag the rows that disagree
    and backfill the coefficients that were left blank.

    Order matters: this must run after the per-row gates (so a row with a broken
    stagnation pressure does not get a vote on the group's area) and before
    consolidation (which needs the backfilled coefficients).
    """
    # Per-row implied areas first.
    for v in verdicts:
        d, row = v.derived, v.row
        d.implied_area_lift_m2 = implied_reference_area(
            row.lift_force_N, row.lift_coeff, d.q_Pa)
        d.implied_area_drag_m2 = implied_reference_area(
            row.drag_force_N, row.drag_coeff, d.q_Pa)

    by_case: dict = {}
    for v in verdicts:
        by_case.setdefault(v.case, []).append(v)

    for case, group in by_case.items():
        if cfg.reference_area_m2 is not None:
            area, basis = cfg.reference_area_m2, "supplied in ScreenConfig"
        else:
            candidates = [v.derived.implied_area_lift_m2 for v in group
                          if v.accepted and v.derived.implied_area_lift_m2]
            if not candidates:
                candidates = [v.derived.implied_area_lift_m2 for v in group
                              if v.derived.implied_area_lift_m2]
            if candidates:
                area = statistics.median(candidates)
                basis = (f"inferred from |L|/(q*|Cl|) across {len(candidates)} "
                         f"run(s) at this point")
            else:
                area, basis = None, "unknown — no row carried both a force and a coefficient"

        for v in group:
            d = v.derived
            d.reference_area_used_m2 = area
            if area:
                # Disagreement check: a different normalisation makes this row's
                # coefficients incomparable to its neighbours' even if both are
                # internally consistent.
                own = d.implied_area_lift_m2
                if own and area > 0:
                    dev = abs(own - area) / area
                    if dev > cfg.ref_area_reject_tolerance:
                        v.flags.append(Flag(
                            "REF_AREA_MISMATCH", Severity.REJECT,
                            f"this row's numbers imply a reference area of "
                            f"{own:.4f} m2 against {area:.4f} m2 for the rest of "
                            f"this operating point ({dev * 100:.0f}% apart) — its "
                            f"coefficients are normalised differently and cannot "
                            f"be averaged with them",
                            "lift_coeff", own, area))
                    elif dev > cfg.ref_area_tolerance:
                        v.flags.append(Flag(
                            "REF_AREA_DRIFT", Severity.WARN,
                            f"implied reference area {own:.4f} m2 differs from the "
                            f"group's {area:.4f} m2 by {dev * 100:.1f}%",
                            "lift_coeff", own, area))

                # Backfill missing coefficients — clearly labelled as derived.
                if v.row.lift_coeff is None and v.row.lift_force_N is not None and d.q_Pa:
                    d.lift_coeff_derived = v.row.lift_force_N / (d.q_Pa * area)
                if v.row.drag_coeff is None and v.row.drag_force_N is not None and d.q_Pa:
                    d.drag_coeff_derived = v.row.drag_force_N / (d.q_Pa * area)


def _bucket(value: Optional[float], tol: float) -> Optional[float]:
    """Round a value onto a tolerance grid so near-identical points group together."""
    if value is None:
        return None
    if tol <= 0:
        return value
    return round(value / tol) * tol


def _case_key(row: RunRow, cfg: ScreenConfig) -> CaseKey:
    component = (row.component or "unspecified").strip()
    return CaseKey(
        component=component,
        ride_height_mm=_bucket(row.ride_height_mm, cfg.ride_height_tol_mm),
        speed_ms=_bucket(row.speed_ms, cfg.speed_tol_ms),
    )


def _screen_outliers(verdicts: Sequence[Verdict], cfg: ScreenConfig) -> None:
    """
    Statistical pass: within each case group, flag runs whose lift disagrees with
    their peers beyond `outlier_z_threshold` modified z-scores.

    Runs last, on rows that already passed every physics gate, so it only ever has
    to explain "these runs were all valid and this one still disagrees". Skipped
    below `min_n_for_outlier` samples.
    """
    if not cfg.enable_outlier_pass:
        return

    by_case: dict = {}
    for v in verdicts:
        if v.accepted:
            by_case.setdefault(v.case, []).append(v)

    for case, group in by_case.items():
        if len(group) < cfg.min_n_for_outlier:
            continue
        # Prefer the coefficient (dimensionless, comparable); fall back to force.
        values, members = [], []
        for v in group:
            cl = v.derived.effective_lift_coeff(v.row)
            if cl is None:
                cl = v.row.lift_force_N
            if cl is not None:
                values.append(float(cl))
                members.append(v)
        if len(values) < cfg.min_n_for_outlier:
            continue
        scores = modified_z_scores(values)
        median = statistics.median(values)
        for v, value, z in zip(members, values, scores):
            v.derived.outlier_z = z
            if z > cfg.outlier_z_threshold:
                v.flags.append(Flag(
                    "STATISTICAL_OUTLIER", Severity.REJECT,
                    f"lift value {value:+.4f} is {z:.1f} modified z-scores from "
                    f"the {median:+.4f} median of the {len(values)} valid runs at "
                    f"this point — it passed every physics check and still "
                    f"disagrees with its peers",
                    "lift_coeff", z, cfg.outlier_z_threshold))


def _screen_supersession(verdicts: Sequence[Verdict], cfg: ScreenConfig) -> None:
    """Optional: keep only the last row per (contributor, case). Off by default."""
    if not cfg.prefer_latest_per_contributor:
        return
    latest: dict = {}
    for v in verdicts:
        if not v.accepted:
            continue
        key = ((v.row.contributor or "").strip().lower(), v.case)
        latest[key] = v                      # input order == sheet order
    keep = {id(v) for v in latest.values()}
    for v in verdicts:
        if v.accepted and id(v) not in keep:
            v.flags.append(Flag(
                "SUPERSEDED", Severity.REJECT,
                f"a later run by {v.row.contributor or 'this contributor'} at the "
                f"same operating point supersedes this one "
                f"(prefer_latest_per_contributor is on)",
                "iteration"))


def screen(rows: Sequence[RunRow], config: Optional[ScreenConfig] = None) -> list:
    """
    Judge every row and return a Verdict for each, in input order.

    The passes run in a deliberate order, cheapest and most certain first:
      bookkeeping -> forces present -> mesh -> y+ -> solution -> solver setup
      -> pressure -> reference-area consistency (needs the group) -> setup
      consistency (needs the group) -> supersession -> statistical outliers
      (needs the survivors).
    """
    cfg = config or ScreenConfig()
    verdicts: list = []

    for row in rows:
        flags: list = []
        derived = Derived(q_Pa=dynamic_pressure(row.speed_ms, cfg.rho))

        marker = _looks_like_test_row(row, cfg.test_row_patterns)
        if marker:
            flags.append(Flag(
                "TEST_ROW",
                Severity.REJECT if cfg.reject_test_rows else Severity.WARN,
                f"marked as scratch data (matched {marker!r} in the contributor, "
                f"component or notes)", "contributor"))

        if row.speed_ms is None:
            flags.append(Flag(
                "NO_VELOCITY", Severity.WARN,
                "no velocity, so dynamic pressure, Cp and the reference-area "
                "cross-check cannot be evaluated for this row", "speed_ms"))

        _screen_forces(row, cfg, flags)
        _screen_mesh(row, cfg, flags)
        _screen_yplus(row, cfg, derived, flags)
        _screen_solution(row, cfg, flags)
        _screen_solver_setup(row, cfg, derived, flags)
        _screen_pressure(row, cfg, derived, flags)

        v = Verdict(row=row, flags=flags, derived=derived)
        v.case = _case_key(row, cfg)
        verdicts.append(v)

    _resolve_reference_areas(verdicts, cfg)
    _screen_setup_consistency(verdicts, cfg)
    _screen_supersession(verdicts, cfg)
    _screen_outliers(verdicts, cfg)
    return verdicts


# --------------------------------------------------------------------------- #
#  3) CONSOLIDATE — average the survivors, per operating point
# --------------------------------------------------------------------------- #
def consolidate(verdicts: Sequence[Verdict],
                config: Optional[ScreenConfig] = None) -> list:
    """
    Group verdicts by operating point and average the accepted rows in each.

    A group with zero survivors is still returned — with `n_accepted == 0` and the
    reasons attached. Silence about a point the team ran is worse than a row that
    says "everything here was rejected, and here is why".
    """
    cfg = config or ScreenConfig()
    by_case: dict = {}
    for v in verdicts:
        by_case.setdefault(v.case, []).append(v)

    cases: list = []
    for case in sorted(by_case, key=lambda c: c.sort_key()):
        group = by_case[case]
        good = [v for v in group if v.accepted]

        cl = [v.derived.effective_lift_coeff(v.row) for v in good]
        cl = [x for x in cl if x is not None]
        cd = [v.derived.effective_drag_coeff(v.row) for v in good]
        cd = [x for x in cd if x is not None]
        lf = [v.row.lift_force_N for v in good if v.row.lift_force_N is not None]
        df = [v.row.drag_force_N for v in good if v.row.drag_force_N is not None]

        cl_mean, cl_sd, cl_lo, cl_hi = _mean_sd(cl)
        cd_mean, cd_sd, _, _ = _mean_sd(cd)
        lf_mean, lf_sd, _, _ = _mean_sd(lf)
        df_mean, df_sd, _, _ = _mean_sd(df)

        spread = None
        if cl_mean is not None and cl_lo is not None and abs(cl_mean) > 1e-12:
            spread = abs(cl_hi - cl_lo) / abs(cl_mean) * 100.0

        l_over_d = None
        if cl_mean is not None and cd_mean not in (None, 0):
            l_over_d = abs(cl_mean) / abs(cd_mean) if cd_mean else None

        areas = [v.derived.reference_area_used_m2 for v in good
                 if v.derived.reference_area_used_m2]
        if not areas:
            areas = [v.derived.reference_area_used_m2 for v in group
                     if v.derived.reference_area_used_m2]
        area = statistics.median(areas) if areas else None
        if cfg.reference_area_m2 is not None:
            basis = "supplied in ScreenConfig"
        elif area:
            basis = "inferred from the accepted runs' own force/coefficient pairs"
        else:
            basis = "unknown"

        notes = []
        derived_cl = sum(1 for v in good if v.row.lift_coeff is None
                         and v.derived.lift_coeff_derived is not None)
        derived_cd = sum(1 for v in good if v.row.drag_coeff is None
                         and v.derived.drag_coeff_derived is not None)
        if derived_cl:
            notes.append(f"{derived_cl} lift coefficient(s) derived from force / (q*A)")
        if derived_cd:
            notes.append(f"{derived_cd} drag coefficient(s) derived from force / (q*A)")
        if good and not cd:
            notes.append("no drag coefficient available — reported or derived")

        def _distinct(attr):
            seen, out = set(), []
            for v in good:
                raw = getattr(v.row, attr, None)
                text = str(raw).strip() if raw is not None else ""
                if text and text.lower() not in seen:
                    seen.add(text.lower())
                    out.append(text)
            return out

        _disc = []
        for v in good:
            d = v.derived.discretisation
            if d != Discretisation.UNKNOWN and d not in _disc:
                _disc.append(d)
        _courants = [v.row.courant_number for v in good
                     if v.row.courant_number is not None]
        _setup_ok = len({v.derived.setup_signature for v in good}) <= 1
        if not _setup_ok:
            notes.append("accepted runs used more than one solver setup \u2014 "
                         "the mean is across different methods")

        cases.append(ConsolidatedCase(
            case=case,
            n_total=len(group),
            n_accepted=len(good),
            n_rejected=len(group) - len(good),
            n_warned=sum(1 for v in good if v.warn_codes),
            lift_coeff_mean=cl_mean, lift_coeff_sd=cl_sd,
            lift_coeff_min=cl_lo, lift_coeff_max=cl_hi,
            drag_coeff_mean=cd_mean, drag_coeff_sd=cd_sd,
            lift_force_mean_N=lf_mean, lift_force_sd_N=lf_sd,
            drag_force_mean_N=df_mean, drag_force_sd_N=df_sd,
            reference_area_m2=area, reference_area_basis=basis,
            lift_to_drag=l_over_d, spread_pct=spread,
            viscous_models=_distinct("viscous_model"),
            schemes=_distinct("scheme"),
            discretisations=_disc,
            initializations=_distinct("initialization"),
            courant_range=((min(_courants), max(_courants))
                           if _courants else ()),
            setup_consistent=_setup_ok,
            contributors=sorted({(v.row.contributor or "?").strip()
                                 for v in good}) or [],
            reject_reasons=[(v.row.label(), v.reason())
                            for v in group if not v.accepted],
            accepted_rows=good,
            notes=notes,
        ))
    return cases


def process(source, config: Optional[ScreenConfig] = None,
            sheet: Optional[str] = None,
            header_row: Optional[int] = None) -> ConsolidationReport:
    """
    The whole pipeline in one call: parse -> screen -> consolidate.

        report = process("wings_runs.xlsx")
        print(report.summary())
        write_workbook(report, "wings_consolidated.xlsx")
    """
    cfg = config or ScreenConfig()
    rows, warnings, unmapped, sheet_name, label = parse_run_log(source, sheet, header_row)
    verdicts = screen(rows, cfg)
    cases = consolidate(verdicts, cfg)
    return ConsolidationReport(
        cases=cases, verdicts=verdicts, config=cfg,
        source=label, sheet=sheet_name,
        parse_warnings=warnings, unmapped_headers=unmapped,
    )


# --------------------------------------------------------------------------- #
#  4) BRIDGE — hand the consolidated points to the rest of KinematiK
# --------------------------------------------------------------------------- #
def to_coeff_results(report: ConsolidationReport, reference_length_m: float = 1.55):
    """
    Turn the consolidated cases into `CoeffResult`s so they flow into `AeroMap`,
    `AeroProvider` and the lap sim — the same objects a solver backend produces.

    Sign convention is preserved: the run log already uses negative = downforce,
    which is KinematiK's convention, so nothing is flipped here. (The Fluent
    run-sheet's `read_fluent_csv` path flips because Fluent reports up-positive;
    this sheet does not.)

    Only cases with at least one accepted run become results, and each carries the
    n / spread in its notes so a single-run point is never mistaken for a mean.
    """
    from .cfd import Attitude, CoeffResult, CFDProvenance, SolverFidelity

    out = []
    for c in report.cases:
        if c.n_accepted == 0 or c.lift_coeff_mean is None:
            continue
        att = Attitude(
            ride_height_mm=c.case.ride_height_mm if c.case.ride_height_mm is not None else 30.0,
            speed_ms=c.case.speed_ms if c.case.speed_ms is not None else 20.0,
        )
        prov = CFDProvenance(
            backend="ansys-run-log",
            fidelity=SolverFidelity.RANS,
            turbulence_model="; ".join(sorted({
                (v.row.viscous_model or "?") for v in c.accepted_rows})),
            notes=(f"consolidated from {c.n_accepted}/{c.n_total} run(s) in "
                   f"{os.path.basename(report.source) or 'run log'}; "
                   f"{c.confidence}"),
        )
        out.append(CoeffResult(
            attitude=att,
            c_lift=c.lift_coeff_mean,
            c_drag=c.drag_coeff_mean,
            converged=True,
            force_monitor_range=c.spread_pct,
            provenance=prov,
            notes=(f"{c.case.component}; ref area "
                   f"{('%.4f m2' % c.reference_area_m2) if c.reference_area_m2 else 'unknown'} "
                   f"({c.reference_area_basis}); " + "; ".join(c.notes)).strip("; "),
        ))
    return out


# --------------------------------------------------------------------------- #
#  5) OUTPUT — the organised workbook, and CSVs for anyone without Excel
# --------------------------------------------------------------------------- #
_CONSOLIDATED_HEADERS = (
    "Case", "Component", "Ride Height (mm)", "Velocity (m/s)",
    "Runs Accepted", "Runs Rejected", "Runs Total",
    "Mean Lift Coefficient", "Lift Coeff SD", "Lift Coeff Min", "Lift Coeff Max",
    "Mean Drag Coefficient", "Drag Coeff SD",
    "Mean Lift Force (N)", "Lift Force SD (N)",
    "Mean Drag Force (N)", "Drag Force SD (N)",
    "L/D", "Spread (%)", "Reference Area (m2)", "Reference Area Basis",
    "Confidence",
    # The solver setup behind the number. A coefficient without its method is
    # not reproducible, and a MIXED entry here is why a mean may be meaningless.
    "Viscous Model(s)", "Scheme(s)", "Discretisation", "Initialization(s)",
    "Courant Range", "Setup Consistent?",
    "Contributors", "Notes",
)

_AUDIT_EXTRA_HEADERS = (
    "Verdict", "Reject Codes", "Warn Codes", "Reason",
    "Dynamic Pressure q (Pa)", "Cp max", "Cp min",
    "Implied Ref Area (m2)", "Ref Area Used (m2)",
    "Wall Treatment", "Discretisation", "Setup Matches Group?",
    "Y+ Target Miss (%)", "Outlier z",
    "Lift Coeff (derived)", "Drag Coeff (derived)",
    "Source Row",
)


def _row_values(v: Verdict) -> list:
    return [getattr(v.row, key, None) for key, _ in CANONICAL_FIELDS]


def _audit_values(v: Verdict) -> list:
    d = v.derived
    return [
        "ACCEPTED" if v.accepted else "REJECTED",
        ", ".join(v.reject_codes),
        ", ".join(v.warn_codes),
        v.reason(),
        d.q_Pa, d.cp_max, d.cp_min,
        d.implied_area_lift_m2, d.reference_area_used_m2,
        d.wall_treatment,
        d.discretisation,
        ("" if d.setup_matches_group is None
         else ("yes" if d.setup_matches_group else "NO")),
        None if d.yplus_target_miss is None else d.yplus_target_miss * 100.0,
        d.outlier_z,
        d.lift_coeff_derived, d.drag_coeff_derived,
        v.row.source_row,
    ]


def write_workbook(report: ConsolidationReport, path: str) -> str:
    """
    Write the organised results workbook. Returns the path written.

    Sheets:
      Consolidated     one row per operating point; the answer. The mean/SD/min/max
                       cells are LIVE FORMULAS over the Accepted Runs sheet, so the
                       team can audit — or re-scope — the average without rerunning
                       this code, and the file recalculates if a row is edited.
      Accepted Runs    the runs behind those means, sorted by case, with everything
                       derived (q, Cp, implied reference area) alongside.
      Rejected Runs    every excluded run with the code and the sentence explaining it.
      Screening Report every row, every flag — the full audit trail.
      Config           the thresholds used, so the numbers are reproducible.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as exc:                            # pragma: no cover
        raise ImportError(
            "Writing an .xlsx report needs openpyxl (`pip install openpyxl`). "
            "Use write_csv_bundle() for a CSV-only output."
        ) from exc

    base = Font(name="Arial", size=10)
    head_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    title_font = Font(name="Arial", size=12, bold=True)
    head_fill = PatternFill("solid", fgColor="1F3864")
    ok_fill = PatternFill("solid", fgColor="E2EFDA")
    bad_fill = PatternFill("solid", fgColor="FCE4E4")
    warn_fill = PatternFill("solid", fgColor="FFF2CC")

    wb = Workbook()

    def style_header(ws, row_idx: int, ncols: int) -> None:
        for col in range(1, ncols + 1):
            cell = ws.cell(row=row_idx, column=col)
            cell.font = head_font
            cell.fill = head_fill
            cell.alignment = Alignment(horizontal="center", vertical="center",
                                       wrap_text=True)
        # Freeze via the string form. `ws.cell(row=row_idx + 1, ...)` would
        # MATERIALISE that row, pushing every later append down by one and
        # silently desynchronising the row cursors the Consolidated formulas are
        # built from — the ranges would then point one row above the real data.
        ws.freeze_panes = f"A{row_idx + 1}"

    def autosize(ws, max_width: int = 42) -> None:
        for col_cells in ws.iter_cols():
            longest = 0
            letter = get_column_letter(col_cells[0].column)
            for cell in col_cells:
                if cell.value is not None:
                    longest = max(longest, len(str(cell.value)))
            ws.column_dimensions[letter].width = min(max(10, longest + 2), max_width)

    # ---------------- Accepted Runs (written first: Consolidated links to it) --
    ws_acc = wb.create_sheet("Accepted Runs")
    acc_headers = [label for _, label in CANONICAL_FIELDS] + list(_AUDIT_EXTRA_HEADERS)
    ws_acc.append(["Runs that passed every screening criterion, grouped by "
                   "operating point. The Consolidated sheet averages these rows."])
    ws_acc.cell(row=1, column=1).font = title_font
    ws_acc.append([])
    ws_acc.append(["Case"] + acc_headers)
    style_header(ws_acc, 3, len(acc_headers) + 1)

    # Sort accepted rows by case so each case owns a CONTIGUOUS block of rows —
    # that is what lets the Consolidated sheet use plain AVERAGE(range) formulas
    # instead of anything array-shaped or version-dependent.
    accepted_by_case: dict = {}
    for v in report.verdicts:
        if v.accepted:
            accepted_by_case.setdefault(v.case, []).append(v)

    case_blocks: dict = {}                      # CaseKey -> (first_row, last_row)
    wrote_any = False
    for c in report.cases:
        group = accepted_by_case.get(c.case, [])
        if not group:
            continue
        start = None
        for v in group:
            ws_acc.append([c.case.label()] + _row_values(v) + _audit_values(v))
            # Read the position back off the sheet rather than tracking it by
            # hand: the Consolidated formulas below reference these exact rows,
            # so an off-by-one here would average the wrong block while still
            # producing a file that opens and recalculates without an error.
            here = ws_acc.max_row
            if start is None:
                start = here
            fill = warn_fill if v.warn_codes else ok_fill
            for col in range(1, len(acc_headers) + 2):
                cell = ws_acc.cell(row=here, column=col)
                cell.font = base
                cell.fill = fill
        case_blocks[c.case] = (start, ws_acc.max_row)
        wrote_any = True
    if not wrote_any:
        ws_acc.append(["(no runs passed screening)"])
        ws_acc.cell(row=ws_acc.max_row, column=1).font = base
    autosize(ws_acc)

    # Column letters on the Accepted sheet, for the formulas below.
    acc_col = {label: get_column_letter(i + 2)
               for i, label in enumerate(acc_headers)}

    # ---------------- Consolidated -------------------------------------------
    ws = wb.create_sheet("Consolidated", 0)
    ws.append(["Consolidated ANSYS results — averaged over screened runs"])
    ws.cell(row=1, column=1).font = title_font
    ws.append([f"Source: {report.source}"
               + (f"  ·  sheet: {report.sheet}" if report.sheet else "")])
    ws.append([f"{report.n_rows} run(s) parsed  ·  {len(report.accepted)} accepted  "
               f"·  {len(report.rejected)} rejected  ·  "
               f"{len(report.cases)} operating point(s)"])
    ws.append(["Mean / SD / min / max cells are formulas over the 'Accepted Runs' "
               "sheet — edit or extend that sheet and these recalculate."])
    for r in (2, 3, 4):
        ws.cell(row=r, column=1).font = base
    ws.append([])
    ws.append(list(_CONSOLIDATED_HEADERS))
    style_header(ws, 6, len(_CONSOLIDATED_HEADERS))

    for c in report.cases:
        block = case_blocks.get(c.case)

        def stat(formula_name: str, header: str):
            """A live formula over this case's block, or the computed value."""
            if not block:
                return None
            lo, hi = block
            letter = acc_col[header]
            ref = f"'Accepted Runs'!{letter}{lo}:{letter}{hi}"
            if formula_name == "STDEV" and lo == hi:
                return None                       # stdev of one sample is undefined
            return f"=IFERROR({formula_name}({ref}),\"\")"

        # Prefer the reported coefficient column; rows where it was derived carry
        # the value in the '(derived)' column, so average across both via the
        # effective values already computed. Where every accepted row reported a
        # coefficient the formula is exact; otherwise fall back to the value.
        all_reported_cl = block and all(
            v.row.lift_coeff is not None for v in accepted_by_case.get(c.case, []))
        all_reported_cd = block and all(
            v.row.drag_coeff is not None for v in accepted_by_case.get(c.case, []))

        cl_mean = stat("AVERAGE", "Lift Coefficient") if all_reported_cl else c.lift_coeff_mean
        cl_sd = stat("STDEV", "Lift Coefficient") if all_reported_cl else c.lift_coeff_sd
        cl_min = stat("MIN", "Lift Coefficient") if all_reported_cl else c.lift_coeff_min
        cl_max = stat("MAX", "Lift Coefficient") if all_reported_cl else c.lift_coeff_max
        cd_mean = stat("AVERAGE", "Drag Coefficient") if all_reported_cd else c.drag_coeff_mean
        cd_sd = stat("STDEV", "Drag Coefficient") if all_reported_cd else c.drag_coeff_sd

        ws.append([
            c.case.label(), c.case.component, c.case.ride_height_mm, c.case.speed_ms,
            c.n_accepted, c.n_rejected, c.n_total,
            cl_mean, cl_sd, cl_min, cl_max,
            cd_mean, cd_sd,
            stat("AVERAGE", "Lift Force (N)") if block else c.lift_force_mean_N,
            stat("STDEV", "Lift Force (N)") if block else c.lift_force_sd_N,
            stat("AVERAGE", "Drag Force (N)") if block else c.drag_force_mean_N,
            stat("STDEV", "Drag Force (N)") if block else c.drag_force_sd_N,
            c.lift_to_drag, c.spread_pct,
            c.reference_area_m2, c.reference_area_basis,
            c.confidence,
            " / ".join(c.viscous_models), " / ".join(c.schemes),
            " / ".join(c.discretisations), " / ".join(c.initializations),
            ("" if not c.courant_range
             else f"{c.courant_range[0]:g}\u2013{c.courant_range[1]:g}"
             if c.courant_range[0] != c.courant_range[1]
             else f"{c.courant_range[0]:g}"),
            ("yes" if c.setup_consistent else "NO \u2014 mixed methods"),
            ", ".join(c.contributors), "; ".join(c.notes),
        ])
        row_idx = ws.max_row
        fill = None
        if c.n_accepted == 0:
            fill = bad_fill
        elif c.n_accepted == 1 or (c.spread_pct or 0) > 15.0:
            fill = warn_fill
        for col in range(1, len(_CONSOLIDATED_HEADERS) + 1):
            cell = ws.cell(row=row_idx, column=col)
            cell.font = base
            if fill:
                cell.fill = fill
            if col in (8, 9, 10, 11, 12, 13):
                cell.number_format = "0.0000"
            elif col in (14, 15, 16, 17, 18, 19):
                cell.number_format = "0.000"
            elif col == 20:
                cell.number_format = "0.00000"

    if not report.cases:
        ws.append(["No operating points were found in the input."])
        ws.cell(row=ws.max_row, column=1).font = base
    autosize(ws)

    # ---------------- Rejected Runs ------------------------------------------
    ws_rej = wb.create_sheet("Rejected Runs")
    ws_rej.append(["Runs excluded from the averages, with the reason for each. "
                   "Nothing is dropped silently — fix the cause and re-run."])
    ws_rej.cell(row=1, column=1).font = title_font
    ws_rej.append([])
    rej_headers = ["Case"] + [label for _, label in CANONICAL_FIELDS] + list(_AUDIT_EXTRA_HEADERS)
    ws_rej.append(rej_headers)
    style_header(ws_rej, 3, len(rej_headers))
    wrote_any = False
    for v in report.verdicts:
        if v.accepted:
            continue
        ws_rej.append([v.case.label() if v.case else ""] + _row_values(v) + _audit_values(v))
        for col in range(1, len(rej_headers) + 1):
            cell = ws_rej.cell(row=ws_rej.max_row, column=col)
            cell.font = base
            cell.fill = bad_fill
        wrote_any = True
    if not wrote_any:
        ws_rej.append(["(no runs were rejected)"])
        ws_rej.cell(row=ws_rej.max_row, column=1).font = base
    autosize(ws_rej)

    # ---------------- Screening Report ---------------------------------------
    ws_log = wb.create_sheet("Screening Report")
    ws_log.append(["Every finding against every row — the full audit trail."])
    ws_log.cell(row=1, column=1).font = title_font
    ws_log.append([])
    log_headers = ["Source Row", "Contributor", "Case", "Verdict", "Severity",
                   "Code", "Channel", "Value", "Limit", "Explanation"]
    ws_log.append(log_headers)
    style_header(ws_log, 3, len(log_headers))
    for v in report.verdicts:
        entries = v.flags or [Flag("CLEAN", Severity.INFO,
                                   "passed every screening criterion")]
        for f in entries:
            ws_log.append([
                v.row.source_row, v.row.contributor,
                v.case.label() if v.case else "",
                "ACCEPTED" if v.accepted else "REJECTED",
                f.severity.upper(), f.code, f.channel, f.value, f.limit, f.message,
            ])
            fill = {Severity.REJECT: bad_fill, Severity.WARN: warn_fill}.get(
                f.severity, ok_fill)
            for col in range(1, len(log_headers) + 1):
                cell = ws_log.cell(row=ws_log.max_row, column=col)
                cell.font = base
                cell.fill = fill
    autosize(ws_log)

    # ---------------- Contributors --------------------------------------- #
    ws_who = wb.create_sheet("Contributors")
    ws_who.append(["Who submitted what, and which checks their runs hit. Not a "
                   "leaderboard \u2014 a map of where a recurring setup mistake "
                   "lives, so it gets fixed at the source."])
    ws_who.cell(row=1, column=1).font = title_font
    ws_who.append([])
    who_headers = ["Contributor", "Runs", "Accepted", "Rejected",
                   "Acceptance (%)", "Most common findings"]
    ws_who.append(who_headers)
    style_header(ws_who, 3, len(who_headers))
    for rec in report.contributor_stats():
        ws_who.append([rec["contributor"], rec["runs"], rec["accepted"],
                       rec["rejected"], round(rec["acceptance_pct"], 1),
                       rec["top_flags"]])
        fill = (ok_fill if rec["rejected"] == 0
                else bad_fill if rec["accepted"] == 0 else warn_fill)
        for col in range(1, len(who_headers) + 1):
            cell = ws_who.cell(row=ws_who.max_row, column=col)
            cell.font = base
            cell.fill = fill
    autosize(ws_who)

    # ---------------- Config --------------------------------------------------
    ws_cfg = wb.create_sheet("Config")
    ws_cfg.append(["Screening criteria used for this run — change these to change "
                   "which runs are kept."])
    ws_cfg.cell(row=1, column=1).font = title_font
    ws_cfg.append([])
    ws_cfg.append(["Setting", "Value"])
    style_header(ws_cfg, 3, 2)
    def cfg_line(label, value):
        ws_cfg.append([label, value])
        for col in (1, 2):
            ws_cfg.cell(row=ws_cfg.max_row, column=col).font = base

    for key, value in report.config.as_rows():
        cfg_line(key, value)
    ws_cfg.append([])
    cfg_line("Parse warnings", "; ".join(report.parse_warnings) or "(none)")
    cfg_line("Unmapped columns (carried through, not screened)",
             ", ".join(report.unmapped_headers) or "(none)")
    ws_cfg.append([])
    ws_cfg.append(["Flag", "Count"])
    style_header(ws_cfg, ws_cfg.max_row, 2)
    for code, count in report.flag_tally().items():
        cfg_line(code, count)
    ws_cfg.freeze_panes = "A4"
    autosize(ws_cfg)

    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    directory = os.path.dirname(os.path.abspath(path))
    if directory:
        os.makedirs(directory, exist_ok=True)
    wb.save(path)
    return path


def consolidated_csv(report: ConsolidationReport) -> str:
    """The Consolidated sheet as CSV text — the deliverable without Excel."""
    buf = io.StringIO()
    w = csv.writer(buf, lineterminator="\n")
    w.writerow(_CONSOLIDATED_HEADERS)
    for c in report.cases:
        w.writerow([
            c.case.label(), c.case.component, c.case.ride_height_mm, c.case.speed_ms,
            c.n_accepted, c.n_rejected, c.n_total,
            c.lift_coeff_mean, c.lift_coeff_sd, c.lift_coeff_min, c.lift_coeff_max,
            c.drag_coeff_mean, c.drag_coeff_sd,
            c.lift_force_mean_N, c.lift_force_sd_N,
            c.drag_force_mean_N, c.drag_force_sd_N,
            c.lift_to_drag, c.spread_pct,
            c.reference_area_m2, c.reference_area_basis,
            c.confidence,
            " / ".join(c.viscous_models), " / ".join(c.schemes),
            " / ".join(c.discretisations), " / ".join(c.initializations),
            ("" if not c.courant_range
             else f"{c.courant_range[0]:g}\u2013{c.courant_range[1]:g}"
             if c.courant_range[0] != c.courant_range[1]
             else f"{c.courant_range[0]:g}"),
            ("yes" if c.setup_consistent else "NO \u2014 mixed methods"),
            ", ".join(c.contributors), "; ".join(c.notes),
        ])
    return buf.getvalue()


def write_csv_bundle(report: ConsolidationReport, directory: str,
                     prefix: str = "aero") -> list:
    """
    Write the same content as four CSVs for teams that do not want a workbook.
    Returns the paths written, consolidated first.
    """
    os.makedirs(directory, exist_ok=True)
    paths: list = []

    p = os.path.join(directory, f"{prefix}_consolidated.csv")
    with open(p, "w", encoding="utf-8", newline="") as fh:
        fh.write(consolidated_csv(report))
    paths.append(p)

    detail_headers = ["Case"] + [label for _, label in CANONICAL_FIELDS] \
        + list(_AUDIT_EXTRA_HEADERS)
    for name, keep in (("accepted", True), ("rejected", False)):
        p = os.path.join(directory, f"{prefix}_{name}_runs.csv")
        with open(p, "w", encoding="utf-8", newline="") as fh:
            w = csv.writer(fh, lineterminator="\n")
            w.writerow(detail_headers)
            for v in report.verdicts:
                if v.accepted is keep:
                    w.writerow([v.case.label() if v.case else ""]
                               + _row_values(v) + _audit_values(v))
        paths.append(p)

    p = os.path.join(directory, f"{prefix}_contributors.csv")
    with open(p, "w", encoding="utf-8", newline="") as fh:
        w = csv.writer(fh, lineterminator="\n")
        w.writerow(["Contributor", "Runs", "Accepted", "Rejected",
                    "Acceptance (%)", "Most common findings"])
        for rec in report.contributor_stats():
            w.writerow([rec["contributor"], rec["runs"], rec["accepted"],
                        rec["rejected"], round(rec["acceptance_pct"], 1),
                        rec["top_flags"]])
    paths.append(p)

    p = os.path.join(directory, f"{prefix}_screening_report.csv")
    with open(p, "w", encoding="utf-8", newline="") as fh:
        w = csv.writer(fh, lineterminator="\n")
        w.writerow(["Source Row", "Contributor", "Case", "Verdict", "Severity",
                    "Code", "Channel", "Value", "Limit", "Explanation"])
        for v in report.verdicts:
            entries = v.flags or [Flag("CLEAN", Severity.INFO,
                                       "passed every screening criterion")]
            for f in entries:
                w.writerow([v.row.source_row, v.row.contributor,
                            v.case.label() if v.case else "",
                            "ACCEPTED" if v.accepted else "REJECTED",
                            f.severity.upper(), f.code, f.channel,
                            f.value, f.limit, f.message])
    paths.append(p)
    return paths
