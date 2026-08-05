# ============================================================================
#  KinematiK — Formula SAE suspension & vehicle dynamics toolkit
#  Created by Frederik Thio. Copyright (c) 2026 Frederik Thio.
#  Open source. Original author: Frederik Thio, creator of KinematiK.
#
#  suspension/xlsx_formula_cache.py — evaluate a workbook's formulas in
#  process and write the results into the file as cached values, so
#  `data_only=True` readers see numbers without LibreOffice or Excel.
# ============================================================================
"""Populate an .xlsx file's formula cache without a spreadsheet application.

WHY THIS EXISTS
---------------
openpyxl writes formula strings and evaluates nothing. A workbook it has
written therefore has no cached values at all, and every formula cell reads
back as ``None`` to `openpyxl.load_workbook(data_only=True)` — which is every
Python consumer, including KinematiK's own loaders. Excel and LibreOffice
recalculate on open and look fine; nothing headless can read the file.

The previous answer was to shell out to LibreOffice. That works on a developer
laptop and fails everywhere the app is actually deployed — Streamlit Cloud,
slim containers, CI — and there the export silently degraded to "correct in
Excel, unreadable by anything else", which is the defect it was written to fix.

So this module evaluates the formulas directly. It is deliberately not a
general spreadsheet engine: it covers the operators and the fourteen functions
the KX sheets actually use, refuses anything it does not understand rather than
guessing, and leaves unevaluated cells exactly as they were. A cell is either
given a value this module is confident in, or given nothing.

WHAT IS AND IS NOT SUPPORTED
----------------------------
Supported: the arithmetic, comparison and concatenation operators; literal
numbers, strings, booleans and errors; relative, absolute and sheet-qualified
references and ranges; and ABS, AND, AVERAGE, COUNT, COUNTA, IF, IFERROR,
INDEX, INT, MATCH, MAX, MIN, NOT, OR, PI, POWER, ROUND, ROUNDDOWN, ROUNDUP,
SIGN, SQRT, SUM, SUMPRODUCT, TEXT.

Not supported, and skipped rather than approximated: array and shared formulas,
defined names, whole-column references, external workbook links, volatile
functions, and every function not in the list above. Skipped cells keep no
cached value, so a reader sees ``None`` — the same honest "I don't know" the
file had before — instead of a plausible wrong number.

Excel remains the authority. `fullCalcOnLoad` is left set by the caller, so the
first time anyone opens the file Excel recomputes every cell from the formulas
and overwrites everything written here. These values exist for headless
readers, not to replace the model.
"""

from __future__ import annotations

import math
import re
import shutil
import tempfile
import zipfile
from dataclasses import dataclass
from functools import lru_cache
from typing import Any
from collections.abc import Sequence
from xml.etree import ElementTree as ET

__all__ = [
    "XlError", "FormulaError", "evaluate_workbook", "populate_cached_values",
    "SUPPORTED_FUNCTIONS",
]

_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
_NS_REL = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"


# ===================================================================== #
#  Values and errors
# ===================================================================== #
class XlError:
    """An Excel error value (#DIV/0!, #VALUE! ...).

    A distinct type rather than a magic string, because a cell can legitimately
    contain the text "#N/A" and the two must not be confused when the result is
    written back into the file with a type attribute.
    """

    __slots__ = ("text",)

    def __init__(self, text: str) -> None:
        self.text = text

    def __repr__(self) -> str:          # pragma: no cover - debug aid
        return f"XlError({self.text!r})"

    def __eq__(self, other: object) -> bool:
        return isinstance(other, XlError) and other.text == self.text

    def __hash__(self) -> int:
        return hash(("XlError", self.text))


DIV0 = XlError("#DIV/0!")
VALUE = XlError("#VALUE!")
REF = XlError("#REF!")
NAME = XlError("#NAME?")
NUM = XlError("#NUM!")
NA = XlError("#N/A")

_ERROR_LITERALS = {e.text: e for e in (DIV0, VALUE, REF, NAME, NUM, NA,
                                       XlError("#NULL!"))}


class FormulaError(Exception):
    """This module cannot evaluate the formula — not a spreadsheet error.

    Raised for unsupported syntax or functions. The caller skips the cell; it
    never becomes a value in the file.
    """


@dataclass
class Matrix:
    """A rectangular block of values — what a range reference evaluates to."""
    rows: list

    def flat(self) -> list:
        return [v for row in self.rows for v in row]

    @property
    def n_rows(self) -> int:
        return len(self.rows)

    @property
    def n_cols(self) -> int:
        return len(self.rows[0]) if self.rows else 0


# ===================================================================== #
#  Tokeniser
# ===================================================================== #
#: One pass, longest-plausible-match first. `func` must precede `ref` because
#: LOG10( tokenises as a three-letter column plus a row number otherwise.
_TOKEN_RE = re.compile(
    r"""
      (?P<ws>\s+)
    | (?P<str>"(?:[^"]|"")*")
    | (?P<err>\#(?:NULL!|DIV/0!|VALUE!|REF!|NAME\?|NUM!|N/A))
    | (?P<bool>(?:TRUE|FALSE)(?![A-Za-z0-9_.(]))
    | (?P<func>[A-Za-z_][A-Za-z0-9_.]*)\s*(?=\()
    | (?P<ref>
          (?:'(?:[^']|'')+'|[A-Za-z_][A-Za-z0-9_.]*)!
          \$?[A-Za-z]{1,3}\$?[0-9]{1,7}
          (?::\$?[A-Za-z]{1,3}\$?[0-9]{1,7})?
        | \$?[A-Za-z]{1,3}\$?[0-9]{1,7}
          (?::\$?[A-Za-z]{1,3}\$?[0-9]{1,7})?
      )
    | (?P<num>[0-9]+(?:\.[0-9]*)?(?:[eE][+-]?[0-9]+)?
             |\.[0-9]+(?:[eE][+-]?[0-9]+)?)
    | (?P<op><>|<=|>=|[-+*/^&=<>%])
    | (?P<lparen>\()
    | (?P<rparen>\))
    | (?P<sep>[,;])
    """,
    re.VERBOSE,
)


def _tokenise(formula: str) -> list:
    out, pos, n = [], 0, len(formula)
    while pos < n:
        m = _TOKEN_RE.match(formula, pos)
        if not m:
            raise FormulaError(f"cannot tokenise at {formula[pos:pos + 20]!r}")
        pos = m.end()
        kind = m.lastgroup
        if kind == "ws":
            continue
        out.append((kind, m.group()))
    return out


# ===================================================================== #
#  Parser  (recursive descent, Excel precedence)
# ===================================================================== #
#  ("num", v) ("str", s) ("bool", b) ("err", XlError)
#  ("ref", text) ("call", name, [args]) ("bin", op, l, r) ("neg", node)
_CMP = {"=", "<>", "<", "<=", ">", ">="}


class _Parser:
    def __init__(self, tokens: list) -> None:
        self.t = tokens
        self.i = 0

    def peek(self) -> tuple | None:
        return self.t[self.i] if self.i < len(self.t) else None

    def take(self) -> tuple:
        if self.i >= len(self.t):
            raise FormulaError("unexpected end of formula")
        tok = self.t[self.i]
        self.i += 1
        return tok

    def expect(self, kind: str, text: str | None = None) -> tuple:
        tok = self.take()
        if tok[0] != kind or (text is not None and tok[1] != text):
            raise FormulaError(f"expected {text or kind}, got {tok[1]!r}")
        return tok

    def parse(self):
        node = self.comparison()
        if self.peek() is not None:
            raise FormulaError(f"trailing input at {self.peek()[1]!r}")
        return node

    def comparison(self):
        node = self.concat()
        while (tok := self.peek()) and tok[0] == "op" and tok[1] in _CMP:
            self.take()
            node = ("bin", tok[1], node, self.concat())
        return node

    def concat(self):
        node = self.additive()
        while (tok := self.peek()) and tok[0] == "op" and tok[1] == "&":
            self.take()
            node = ("bin", "&", node, self.additive())
        return node

    def additive(self):
        node = self.multiplicative()
        while (tok := self.peek()) and tok[0] == "op" and tok[1] in "+-":
            self.take()
            node = ("bin", tok[1], node, self.multiplicative())
        return node

    def multiplicative(self):
        node = self.power()
        while (tok := self.peek()) and tok[0] == "op" and tok[1] in "*/":
            self.take()
            node = ("bin", tok[1], node, self.power())
        return node

    def power(self):
        node = self.unary()
        while (tok := self.peek()) and tok[0] == "op" and tok[1] == "^":
            self.take()
            node = ("bin", "^", node, self.unary())
        return node

    def unary(self):
        tok = self.peek()
        if tok and tok[0] == "op" and tok[1] in "+-":
            self.take()
            node = self.unary()
            return ("neg", node) if tok[1] == "-" else node
        return self.postfix()

    def postfix(self):
        node = self.primary()
        while (tok := self.peek()) and tok[0] == "op" and tok[1] == "%":
            self.take()
            node = ("bin", "/", node, ("num", 100.0))
        return node

    def primary(self):
        tok = self.take()
        kind, text = tok
        if kind == "num":
            return ("num", float(text))
        if kind == "str":
            return ("str", text[1:-1].replace('""', '"'))
        if kind == "bool":
            return ("bool", text.upper() == "TRUE")
        if kind == "err":
            return ("err", _ERROR_LITERALS.get(text, VALUE))
        if kind == "ref":
            return ("ref", text)
        if kind == "func":
            self.expect("lparen")
            args = []
            if (nxt := self.peek()) and nxt[0] == "rparen":
                self.take()
                return ("call", text.upper(), args)
            while True:
                args.append(self.comparison())
                nxt = self.take()
                if nxt[0] == "rparen":
                    break
                if nxt[0] != "sep":
                    raise FormulaError(f"expected , or ) got {nxt[1]!r}")
            return ("call", text.upper(), args)
        if kind == "lparen":
            node = self.comparison()
            self.expect("rparen")
            return node
        raise FormulaError(f"unexpected token {text!r}")


#: ASTs are immutable tuples, so sharing one between cells is safe. Sheets that
#: repeat a formula verbatim (the gear study, the advisor) get it for free; a
#: 5000-row trace, where every row differs, pays the parse once per cell.
@lru_cache(maxsize=16384)
def _parse(formula: str):
    return _Parser(_tokenise(formula)).parse()


# ===================================================================== #
#  Coercion
# ===================================================================== #
def _num(v: Any) -> float:
    """Excel's number coercion. Raises XlError-carrying _Bail on failure."""
    if isinstance(v, XlError):
        raise _Bail(v)
    if v is None or v == "":
        return 0.0
    if isinstance(v, bool):
        return 1.0 if v else 0.0
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        try:
            return float(v.strip())
        except ValueError:
            raise _Bail(VALUE)
    if isinstance(v, Matrix):
        raise _Bail(VALUE)
    raise _Bail(VALUE)


def _text(v: Any) -> str:
    if isinstance(v, XlError):
        raise _Bail(v)
    if v is None:
        return ""
    if isinstance(v, bool):
        return "TRUE" if v else "FALSE"
    if isinstance(v, float):
        if v == int(v) and abs(v) < 1e15:
            return str(int(v))
        return repr(v)
    if isinstance(v, Matrix):
        raise _Bail(VALUE)
    return str(v)


def _bool(v: Any) -> bool:
    if isinstance(v, XlError):
        raise _Bail(v)
    if isinstance(v, bool):
        return v
    if v is None:
        return False
    if isinstance(v, str):
        u = v.strip().upper()
        if u == "TRUE":
            return True
        if u == "FALSE":
            return False
        raise _Bail(VALUE)
    if isinstance(v, Matrix):
        raise _Bail(VALUE)
    return _num(v) != 0.0


class _Bail(Exception):
    """Internal: unwind to the top of the formula carrying an Excel error."""

    def __init__(self, err: XlError) -> None:
        self.err = err


def _finite(x: float) -> float:
    if math.isnan(x) or math.isinf(x):
        raise _Bail(NUM)
    return x


# ===================================================================== #
#  Functions
# ===================================================================== #
def _numbers(args: Sequence[Any]) -> list:
    """Numeric arguments, Excel's aggregate rules: text and blanks ignored.

    Matters for MAX/MIN over a range that includes a header or a blank tail —
    counting those as zero is how a minimum silently becomes 0.
    """
    out = []
    for a in args:
        if isinstance(a, Matrix):
            for v in a.flat():
                if isinstance(v, XlError):
                    raise _Bail(v)
                if isinstance(v, bool):
                    continue                    # booleans in ranges: ignored
                if isinstance(v, (int, float)):
                    out.append(float(v))
        elif isinstance(a, XlError):
            raise _Bail(a)
        elif a is None or a == "":
            continue
        elif isinstance(a, str):
            out.append(_num(a))                 # a literal text number counts
        elif isinstance(a, bool):
            out.append(1.0 if a else 0.0)
        else:
            out.append(float(a))
    return out


def _scalar(v: Any) -> Any:
    """A 1x1 matrix is a scalar; anything larger is not usable as one."""
    if isinstance(v, Matrix):
        if v.n_rows == 1 and v.n_cols == 1:
            return v.rows[0][0]
        raise _Bail(VALUE)
    return v


def _f_sum(args):
    return math.fsum(_numbers(args))


def _f_average(args):
    xs = _numbers(args)
    if not xs:
        raise _Bail(DIV0)
    return math.fsum(xs) / len(xs)


def _f_max(args):
    xs = _numbers(args)
    return max(xs) if xs else 0.0


def _f_min(args):
    xs = _numbers(args)
    return min(xs) if xs else 0.0


def _f_count(args):
    return float(len(_numbers(args)))


def _f_counta(args):
    n = 0
    for a in args:
        vals = a.flat() if isinstance(a, Matrix) else [a]
        n += sum(1 for v in vals if v is not None and v != "")
    return float(n)


def _f_sumproduct(args):
    if not args:
        raise _Bail(VALUE)
    mats = []
    for a in args:
        if isinstance(a, Matrix):
            mats.append(a.flat())
        else:
            mats.append([a])
    size = len(mats[0])
    if any(len(m) != size for m in mats):
        raise _Bail(VALUE)

    def _term(v):
        if isinstance(v, XlError):
            raise _Bail(v)
        if isinstance(v, bool):
            return 0.0                       # SUMPRODUCT ignores booleans
        return float(v) if isinstance(v, (int, float)) else 0.0

    return math.fsum(
        math.prod(_term(m[i]) for m in mats) for i in range(size))


def _f_index(args):
    if len(args) not in (2, 3):
        raise _Bail(VALUE)
    src = args[0]
    if not isinstance(src, Matrix):
        src = Matrix([[src]])
    r = int(_num(_scalar(args[1])))
    c = int(_num(_scalar(args[2]))) if len(args) == 3 else None
    if c is None:
        # One-dimensional addressing over a vector.
        if src.n_rows == 1:
            r, c = 1, r
        else:
            c = 1 if src.n_cols == 1 else 0
    if r == 0 and c == 0:
        return src
    if r == 0:
        if not 1 <= c <= src.n_cols:
            raise _Bail(REF)
        return Matrix([[row[c - 1]] for row in src.rows])
    if c == 0:
        if not 1 <= r <= src.n_rows:
            raise _Bail(REF)
        return Matrix([list(src.rows[r - 1])])
    if not (1 <= r <= src.n_rows and 1 <= c <= src.n_cols):
        raise _Bail(REF)
    return src.rows[r - 1][c - 1]


def _wildcard(pattern: str) -> re.Pattern:
    out = []
    i = 0
    while i < len(pattern):
        ch = pattern[i]
        if ch == "~" and i + 1 < len(pattern) and pattern[i + 1] in "*?~":
            out.append(re.escape(pattern[i + 1]))
            i += 2
            continue
        out.append(".*" if ch == "*" else "." if ch == "?" else re.escape(ch))
        i += 1
    return re.compile("".join(out) + r"\Z", re.IGNORECASE)


def _f_match(args):
    if len(args) not in (2, 3):
        raise _Bail(VALUE)
    needle = _scalar(args[0])
    hay = args[1]
    vals = hay.flat() if isinstance(hay, Matrix) else [hay]
    mode = int(_num(_scalar(args[2]))) if len(args) == 3 else 1

    if mode == 0:
        if isinstance(needle, str):
            rx = (_wildcard(needle) if ("*" in needle or "?" in needle)
                  else None)
            for i, v in enumerate(vals, start=1):
                if isinstance(v, str):
                    if rx.match(v) if rx else v.upper() == needle.upper():
                        return float(i)
            raise _Bail(NA)
        for i, v in enumerate(vals, start=1):
            if isinstance(v, (int, float)) and not isinstance(v, bool) \
                    and float(v) == _num(needle):
                return float(i)
        raise _Bail(NA)

    # Sorted search: ascending for 1, descending for -1.
    target = _num(needle) if not isinstance(needle, str) else None
    if target is None:
        raise _Bail(NA)
    best = None
    for i, v in enumerate(vals, start=1):
        if not isinstance(v, (int, float)) or isinstance(v, bool):
            continue
        if (mode == 1 and v <= target) or (mode == -1 and v >= target):
            best = float(i)
    if best is None:
        raise _Bail(NA)
    return best


#: Number-format subset used by TEXT() in the KX sheets: digit placeholders,
#: an optional thousands separator, an optional percentage. Anything else is
#: refused rather than rendered wrongly — a verdict string carrying a
#: misformatted number is worse than no cached value at all.
_FMT_RE = re.compile(r"\A(?P<int>[#0,]*)(?:\.(?P<dec>[#0]*))?(?P<pct>%?)\Z")


def _f_text(args):
    if len(args) != 2:
        raise _Bail(VALUE)
    v = _scalar(args[0])
    fmt = _text(_scalar(args[1]))
    section = fmt.split(";")[0]
    m = _FMT_RE.match(section)
    if not m:
        raise FormulaError(f"unsupported TEXT format {fmt!r}")
    if isinstance(v, str):
        return v
    x = _num(v)
    if m.group("pct"):
        x *= 100.0
    decimals = len(m.group("dec") or "")
    grouped = "," in (m.group("int") or "")
    body = f"{x:,.{decimals}f}" if grouped else f"{x:.{decimals}f}"
    if body.lstrip("-").startswith("0.") and (m.group("int") or "").endswith("#"):
        body = body.replace("0.", ".", 1)
    return body + m.group("pct")


def _f_round(args, mode="half"):
    if len(args) not in (1, 2):
        raise _Bail(VALUE)
    x = _num(_scalar(args[0]))
    d = int(_num(_scalar(args[1]))) if len(args) == 2 else 0
    f = 10.0 ** d
    y = x * f
    if mode == "up":
        y = math.ceil(abs(y)) * (1 if y >= 0 else -1)
    elif mode == "down":
        y = math.floor(abs(y)) * (1 if y >= 0 else -1)
    else:
        # Excel rounds half away from zero; Python rounds half to even.
        y = math.floor(abs(y) + 0.5) * (1 if y >= 0 else -1)
    return _finite(y / f)


def _f_sqrt(args):
    x = _num(_scalar(args[0]))
    if x < 0:
        raise _Bail(NUM)
    return math.sqrt(x)


def _f_power(args):
    if len(args) != 2:
        raise _Bail(VALUE)
    return _pow(_num(_scalar(args[0])), _num(_scalar(args[1])))


def _pow(a: float, b: float) -> float:
    try:
        r = a ** b
    except (OverflowError, ValueError, ZeroDivisionError):
        raise _Bail(NUM)
    if isinstance(r, complex):
        raise _Bail(NUM)
    return _finite(float(r))


#: name -> (callable, lazy). Lazy functions receive unevaluated AST nodes plus
#: the evaluator, because IF and IFERROR must not evaluate the branch they do
#: not take: IFERROR(1/0, "n/a") has to survive the division.
SUPPORTED_FUNCTIONS = {
    "ABS": (lambda a: abs(_num(_scalar(a[0]))), False),
    "AND": (lambda a: all(_bool(v) for arg in a
                          for v in (arg.flat() if isinstance(arg, Matrix)
                                    else [arg])), False),
    "AVERAGE": (_f_average, False),
    "COUNT": (_f_count, False),
    "COUNTA": (_f_counta, False),
    "INDEX": (_f_index, False),
    "INT": (lambda a: float(math.floor(_num(_scalar(a[0])))), False),
    "MATCH": (_f_match, False),
    "MAX": (_f_max, False),
    "MIN": (_f_min, False),
    "NOT": (lambda a: not _bool(_scalar(a[0])), False),
    "OR": (lambda a: any(_bool(v) for arg in a
                         for v in (arg.flat() if isinstance(arg, Matrix)
                                   else [arg])), False),
    "PI": (lambda a: math.pi, False),
    "POWER": (_f_power, False),
    "ROUND": (lambda a: _f_round(a, "half"), False),
    "ROUNDDOWN": (lambda a: _f_round(a, "down"), False),
    "ROUNDUP": (lambda a: _f_round(a, "up"), False),
    "SIGN": (lambda a: float((_num(_scalar(a[0])) > 0)
                             - (_num(_scalar(a[0])) < 0)), False),
    "SQRT": (_f_sqrt, False),
    "SUM": (_f_sum, False),
    "SUMPRODUCT": (_f_sumproduct, False),
    "TEXT": (_f_text, False),
}


# ===================================================================== #
#  Reference parsing
# ===================================================================== #
_CELL_RE = re.compile(r"\A\$?([A-Za-z]{1,3})\$?([0-9]{1,7})\Z")


def _col_index(letters: str) -> int:
    n = 0
    for ch in letters.upper():
        n = n * 26 + (ord(ch) - 64)
    return n


def _split_ref(text: str) -> tuple:
    """'KX Trace'!C3:C42 -> (sheet or None, 'C3', 'C42')."""
    sheet = None
    body = text
    if "!" in text:
        head, body = text.rsplit("!", 1)
        sheet = head[1:-1].replace("''", "'") if head.startswith("'") else head
    if ":" in body:
        a, b = body.split(":", 1)
    else:
        a = b = body
    if not (_CELL_RE.match(a) and _CELL_RE.match(b)):
        raise FormulaError(f"unsupported reference {text!r}")
    return sheet, a, b


# ===================================================================== #
#  Evaluator
# ===================================================================== #
class _Evaluator:
    """Evaluates one workbook, memoising per cell and detecting cycles."""

    def __init__(self, wb, max_cells: int = 1_000_000) -> None:
        self.wb = wb
        self.max_cells = max_cells
        self.cache: dict = {}
        self.pending: set = set()
        self.unsupported: dict = {}
        self.visited = 0

    # -- cell access -------------------------------------------------- #
    def sheet(self, name: str | None, default: str):
        title = name if name is not None else default
        if title not in self.wb.sheetnames:
            raise _Bail(REF)
        return self.wb[title]

    def cell_value(self, sheet_title: str, coord: str) -> Any:
        key = (sheet_title, coord)
        if key in self.cache:
            return self.cache[key]
        if key in self.pending:
            raise _Bail(REF)                    # circular reference
        ws = self.wb[sheet_title]
        raw = ws[coord].value
        if not (isinstance(raw, str) and raw.startswith("=")):
            self.cache[key] = raw
            return raw
        self.pending.add(key)
        try:
            value = self.evaluate(raw, sheet_title)
        finally:
            self.pending.discard(key)
        self.cache[key] = value
        return value

    def range_value(self, sheet_title: str, a: str, b: str) -> Any:
        ma, mb = _CELL_RE.match(a), _CELL_RE.match(b)
        c1, r1 = _col_index(ma.group(1)), int(ma.group(2))
        c2, r2 = _col_index(mb.group(1)), int(mb.group(2))
        if c1 > c2:
            c1, c2 = c2, c1
        if r1 > r2:
            r1, r2 = r2, r1
        if (r2 - r1 + 1) * (c2 - c1 + 1) > self.max_cells:
            raise FormulaError("range too large")
        ws = self.wb[sheet_title]
        rows = []
        for r in range(r1, r2 + 1):
            rows.append([self.cell_value(sheet_title, f"{_letters(c)}{r}")
                         for c in range(c1, c2 + 1)])
        if r1 == r2 and c1 == c2:
            return rows[0][0]
        return Matrix(rows)

    # -- expression --------------------------------------------------- #
    def evaluate(self, formula: str, sheet_title: str) -> Any:
        try:
            node = _parse(formula[1:] if formula.startswith("=") else formula)
        except FormulaError:
            raise
        try:
            return self.eval_node(node, sheet_title)
        except _Bail as bail:
            return bail.err
        except RecursionError:
            raise FormulaError("formula nests too deeply")

    def eval_node(self, node, sheet: str) -> Any:
        kind = node[0]
        if kind in ("num", "str", "bool", "err"):
            return node[1]
        if kind == "ref":
            name, a, b = _split_ref(node[1])
            title = name if name is not None else sheet
            if title not in self.wb.sheetnames:
                raise _Bail(REF)
            if a == b:
                return self.cell_value(title, a.replace("$", ""))
            return self.range_value(title, a.replace("$", ""),
                                    b.replace("$", ""))
        if kind == "neg":
            return -_num(_scalar(self.eval_node(node[1], sheet)))
        if kind == "bin":
            return self.eval_bin(node[1], node[2], node[3], sheet)
        if kind == "call":
            return self.eval_call(node[1], node[2], sheet)
        raise FormulaError(f"unknown node {kind}")

    def eval_bin(self, op: str, left, right, sheet: str) -> Any:
        lv = _scalar(self.eval_node(left, sheet))
        rv = _scalar(self.eval_node(right, sheet))
        if isinstance(lv, XlError):
            raise _Bail(lv)
        if isinstance(rv, XlError):
            raise _Bail(rv)
        if op == "&":
            return _text(lv) + _text(rv)
        if op in _CMP:
            return _compare(op, lv, rv)
        a, b = _num(lv), _num(rv)
        if op == "+":
            return _finite(a + b)
        if op == "-":
            return _finite(a - b)
        if op == "*":
            return _finite(a * b)
        if op == "/":
            if b == 0:
                raise _Bail(DIV0)
            return _finite(a / b)
        if op == "^":
            return _pow(a, b)
        raise FormulaError(f"unknown operator {op}")

    def eval_call(self, name: str, args, sheet: str) -> Any:
        # Lazy forms first: their whole point is not evaluating every branch.
        if name == "IF":
            if len(args) not in (2, 3):
                raise _Bail(VALUE)
            cond = _bool(_scalar(self.eval_node(args[0], sheet)))
            if cond:
                return self.eval_node(args[1], sheet)
            if len(args) == 3:
                return self.eval_node(args[2], sheet)
            return False
        if name == "IFERROR":
            if len(args) != 2:
                raise _Bail(VALUE)
            try:
                value = self.eval_node(args[0], sheet)
            except _Bail:
                return self.eval_node(args[1], sheet)
            if isinstance(value, XlError):
                return self.eval_node(args[1], sheet)
            return value
        if name == "IFNA":
            if len(args) != 2:
                raise _Bail(VALUE)
            try:
                value = self.eval_node(args[0], sheet)
            except _Bail as bail:
                if bail.err == NA:
                    return self.eval_node(args[1], sheet)
                raise
            return (self.eval_node(args[1], sheet) if value == NA else value)
        if name in ("ISERROR", "ISNA", "ISBLANK", "ISNUMBER", "ISTEXT"):
            return self._eval_is(name, args, sheet)

        entry = SUPPORTED_FUNCTIONS.get(name)
        if entry is None:
            self.unsupported[name] = self.unsupported.get(name, 0) + 1
            raise FormulaError(f"unsupported function {name}")
        fn, _lazy = entry
        values = [self.eval_node(a, sheet) for a in args]
        return fn(values)

    def _eval_is(self, name: str, args, sheet: str) -> bool:
        if len(args) != 1:
            raise _Bail(VALUE)
        try:
            v = self.eval_node(args[0], sheet)
        except _Bail as bail:
            v = bail.err
        if name == "ISERROR":
            return isinstance(v, XlError)
        if name == "ISNA":
            return v == NA
        if isinstance(v, XlError):
            raise _Bail(v)
        v = _scalar(v)
        if name == "ISBLANK":
            return v is None
        if name == "ISNUMBER":
            return isinstance(v, (int, float)) and not isinstance(v, bool)
        return isinstance(v, str)


def _compare(op: str, a: Any, b: Any) -> bool:
    # Excel orders numbers before text; text comparison is case-insensitive.
    a_num = isinstance(a, (int, float)) and not isinstance(a, bool)
    b_num = isinstance(b, (int, float)) and not isinstance(b, bool)
    if a is None and b is None:
        a = b = 0.0
        a_num = b_num = True
    elif a is None:
        a, a_num = (0.0, True) if b_num else ("", False)
    elif b is None:
        b, b_num = (0.0, True) if a_num else ("", False)
    if a_num and b_num:
        x, y = float(a), float(b)
    elif isinstance(a, str) and isinstance(b, str):
        x, y = a.upper(), b.upper()
    elif isinstance(a, bool) or isinstance(b, bool):
        x, y = _num(a), _num(b)
    else:
        # Mixed number/text: any text sorts above any number.
        rank_a, rank_b = (0 if a_num else 1), (0 if b_num else 1)
        if rank_a != rank_b:
            x, y = rank_a, rank_b
        else:                                    # pragma: no cover
            x, y = _text(a), _text(b)
    if op == "=":
        return x == y
    if op == "<>":
        return x != y
    if op == "<":
        return x < y
    if op == "<=":
        return x <= y
    if op == ">":
        return x > y
    return x >= y


def _letters(n: int) -> str:
    out = ""
    while n:
        n, rem = divmod(n - 1, 26)
        out = chr(65 + rem) + out
    return out


# ===================================================================== #
#  Public API
# ===================================================================== #
def evaluate_workbook(path: str, *, only_prefix: str | None = None,
                      max_cells: int = 1_000_000) -> tuple:
    """Evaluate every formula this module understands.

    Returns `(values, report)` where `values` maps `(sheet_title, coord)` to a
    Python value or `XlError`, holding only cells that were evaluated
    confidently. `only_prefix` restricts the work to sheets whose title starts
    with it — the KX sheets, in practice, which is both faster and safer than
    walking a stranger's 5000-row model.
    """
    import openpyxl

    wb = openpyxl.load_workbook(path, data_only=False)
    try:
        ev = _Evaluator(wb, max_cells=max_cells)
        titles = [t for t in wb.sheetnames
                  if only_prefix is None or t.startswith(only_prefix)]
        values: dict = {}
        skipped = 0
        errors: dict = {}
        for title in titles:
            ws = wb[title]
            for row in ws.iter_rows():
                for cell in row:
                    raw = cell.value
                    if not (isinstance(raw, str) and raw.startswith("=")):
                        continue
                    key = (title, cell.coordinate)
                    try:
                        value = ev.cell_value(title, cell.coordinate)
                    except FormulaError:
                        skipped += 1
                        continue
                    except _Bail as bail:
                        value = bail.err
                    except (RecursionError, ValueError, TypeError,
                            OverflowError, ZeroDivisionError, KeyError):
                        skipped += 1
                        continue
                    if isinstance(value, Matrix):
                        skipped += 1
                        continue
                    if isinstance(value, XlError):
                        errors.setdefault(value.text, []).append(
                            f"{title}!{cell.coordinate}")
                    values[key] = value
        report = {
            "evaluated": len(values),
            "skipped": skipped,
            "errors": errors,
            "unsupported_functions": dict(ev.unsupported),
            "sheets": titles,
        }
        return values, report
    finally:
        wb.close()


def _sheet_xml_paths(zf: zipfile.ZipFile) -> dict:
    """Map sheet title -> path of its XML part inside the package."""
    wb_xml = ET.fromstring(zf.read("xl/workbook.xml"))
    rels = ET.fromstring(zf.read("xl/_rels/workbook.xml.rels"))
    target = {}
    for rel in rels:
        tgt = rel.get("Target", "")
        if not tgt.startswith("/"):
            tgt = "xl/" + tgt.lstrip("./")
        else:
            tgt = tgt.lstrip("/")
        target[rel.get("Id")] = tgt
    out = {}
    for sheet in wb_xml.iter(f"{{{_NS}}}sheet"):
        rid = sheet.get(f"{{{_NS_REL}}}id")
        if rid in target:
            out[sheet.get("name")] = target[rid]
    return out


def _cached_repr(value: Any) -> tuple | None:
    """(type attribute, text) for a cached value, or None if unwritable."""
    if isinstance(value, XlError):
        return "e", value.text
    if isinstance(value, bool):
        return "b", "1" if value else "0"
    if isinstance(value, (int, float)):
        x = float(value)
        if math.isnan(x) or math.isinf(x):
            return "e", NUM.text
        if x == int(x) and abs(x) < 1e15:
            return None, str(int(x))
        return None, repr(x)
    if isinstance(value, str):
        return "str", value
    if value is None:
        return None, "0"
    return None


def populate_cached_values(path: str, *, only_prefix: str | None = None,
                           max_cells: int = 1_000_000) -> dict:
    """Evaluate the workbook and write the results in as cached values.

    The file is rewritten in place. Formulas are left exactly as they are —
    only `<v>` elements are added — so the workbook stays a live model and
    Excel recomputes everything on open.
    """
    values, report = evaluate_workbook(path, only_prefix=only_prefix,
                                       max_cells=max_cells)
    if not values:
        report["written"] = 0
        return report

    by_sheet: dict = {}
    for (title, coord), value in values.items():
        by_sheet.setdefault(title, {})[coord] = value

    written = 0
    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    tmp.close()
    with zipfile.ZipFile(path) as src:
        paths = _sheet_xml_paths(src)
        wanted = {paths[t]: cells for t, cells in by_sheet.items()
                  if t in paths}
        with zipfile.ZipFile(tmp.name, "w", zipfile.ZIP_DEFLATED) as dst:
            for item in src.infolist():
                data = src.read(item.filename)
                # calcChain indexes formula cells for incremental recalc. It is
                # an optimisation Excel rebuilds, and a stale one against a file
                # whose cached values just changed is exactly the kind of
                # inconsistency that produces a repair prompt. Drop it.
                if item.filename == "xl/calcChain.xml":
                    continue
                if item.filename in wanted:
                    data, n = _inject(data, wanted[item.filename])
                    written += n
                dst.writestr(item, data)
    shutil.move(tmp.name, path)
    report["written"] = written
    return report


#: Cell and formula elements, matched on the raw bytes. Parsing the sheet with
#: ElementTree and writing it back would be tidier and is not safe: a workbook
#: Excel wrote declares namespace prefixes (mc, x14ac, xr...) that ElementTree
#: renames or drops on serialisation, and `mc:Ignorable` refers to those
#: prefixes by name, so the file comes back needing "repair". Editing the bytes
#: leaves every part of the sheet this module did not compute byte-identical.
#:
#: This is safe against the general XML grammar for the same reason: '>' is
#: escaped in element content by every writer that produces these files, and
#: cell attributes (r, s, t, cm, vm, ph) never contain one.
_C_RE = re.compile(rb"<c(?P<attrs>\s[^>]*?)?(?:/>|>(?P<inner>.*?)</c\s*>)",
                   re.DOTALL)
_F_RE = re.compile(rb"<f(?:\s[^>]*?)?(?:/>|>.*?</f\s*>)", re.DOTALL)
_F_ATTRS_RE = re.compile(rb"<f(?P<attrs>\s[^>]*?)?[/>]")
_R_ATTR_RE = re.compile(rb'\br="([^"]*)"')
_T_ATTR_RE = re.compile(rb'\st="[^"]*"')


def _xml_escape(text: str) -> bytes:
    out = (text.replace("&", "&amp;").replace("<", "&lt;")
               .replace(">", "&gt;"))
    # Control characters are not representable in XML 1.0 at all.
    out = "".join(ch for ch in out
                  if ch >= " " or ch in "\t\n\r")
    return out.encode("utf-8")


def _inject(xml_bytes: bytes, cells: dict) -> tuple:
    """Add <v> to the formula cells we have values for. Returns (xml, count)."""
    written = 0

    def replace(match: re.Match) -> bytes:
        nonlocal written
        attrs = match.group("attrs") or b""
        inner = match.group("inner")
        if inner is None:
            return match.group(0)               # <c .../> has no formula
        rm = _R_ATTR_RE.search(attrs)
        if rm is None:
            return match.group(0)
        coord = rm.group(1).decode()
        if coord not in cells:
            return match.group(0)
        fm = _F_RE.search(inner)
        if fm is None:
            return match.group(0)
        f_attrs = _F_ATTRS_RE.match(fm.group(0))
        f_attr_text = (f_attrs.group("attrs") or b"") if f_attrs else b""
        # Shared, array and data-table formulas have semantics this module does
        # not model — one cell's text governs a block of others.
        if re.search(rb't="(shared|array|dataTable)"', f_attr_text):
            return match.group(0)
        rendered = _cached_repr(cells[coord])
        if rendered is None:
            return match.group(0)
        t_attr, text = rendered

        new_attrs = _T_ATTR_RE.sub(b"", attrs)
        if t_attr is not None:
            new_attrs += b' t="' + t_attr.encode() + b'"'
        body = fm.group(0) + b"<v>" + _xml_escape(text) + b"</v>"
        written += 1
        return b"<c" + new_attrs + b">" + body + b"</c>"

    return _C_RE.sub(replace, xml_bytes), written
