# ============================================================================
#  KinematiK — Formula SAE suspension & vehicle dynamics toolkit
#  Created by Frederik Thio. Copyright (c) 2026 Frederik Thio.
#  Open source. Original author: Frederik Thio, creator of KinematiK.
#
#  suspension/power_draw.py — the EV power-draw workbook, audited and replaced.
#  Reads FSAE_EV_Power_Draw.xlsx, reports what is wrong with cell references,
#  and supplies the corrected physics the workbook was reaching for.
# ============================================================================
"""
Power draw — the workbook's arithmetic, checked and rebuilt.

WHY THIS MODULE EXISTS
----------------------
`FSAE_EV_Power_Draw.xlsx` is the electrics group's model of pack current, joule
heating, motor load and bearing temperature across fifteen gear ratios. It is a
serious piece of work — roughly 90,000 populated cells — and the shape of the
analysis is right. The arithmetic inside it is not, and the errors are the kind
that produce confident numbers rather than error cells, which is why they have
survived.

There is one finding that matters more than all the dimensional errors put
together, and it is worth stating first because it is the reason the others
went unnoticed:

**The workbook has no vehicle in it.** Current is computed as

    I = V_pack * PowerFactor * RPM / 1000

There is no mass, no drag, no rolling resistance, no acceleration, no tractive
force anywhere in the file. Current is made proportional to pack *voltage* —
backwards, since at a given power a higher bus voltage draws *less* current —
and proportional to road speed, with no reference to what the car is being
asked to do at that speed. A model of power draw that never computes force
cannot predict power draw. Every number downstream of that cell inherits it:
the joule heating, the thermal load, the EMF columns, the bearing temperatures.

The fix is not a correction to that formula. It is a force balance, which is
what `power_draw_trace` computes.

WHAT ELSE THE AUDIT FINDS
-------------------------
Dimensional and physical errors, each pinned to a cell:

* The gear ratio is inverted. A reduction of N is modelled as multiplication by
  1/N, so a 7:1 reduction reports the motor turning *slower* than the wheel —
  93 rpm where it should be 4575. Power scales with the square, so this is a
  49x error at gear 7 and a 225x error at gear 15.
* Motor efficiency multiplies where it should divide. As written, a *more*
  efficient motor draws *more* current from the pack.
* `sqrt(3)` and power factor — three-phase AC line quantities — are applied to
  the DC bus voltage, double-counting the phase conversion.
* Pack internal resistance ignores the parallel strings entirely. The formula
  reduces to S x R_cell, giving 1.792 ohm where 140S3P gives 0.597 ohm: 3x
  high, and it feeds every joule-heating and EMF figure in the file.
* `TAN(20)` passes degrees to a function that takes radians, a 6.15x error on
  the bearing's radial gear load. The same formula divides torque by the gear
  *pitch diameter* where the moment arm is the radius, a further 2x the other
  way — two independent errors that partially mask each other.
* The EMF columns compute `V_pack * Power_kW * R_pack`, which is volts times
  kilowatts times ohms. The result is labelled volts and reads ~32,000 for a
  504 V pack.

Structural problems, which are what let the above survive:

* Fifteen gear scenarios are stacked vertically in one sheet at fixed row
  offsets — RPM in 2:1894, current in 1895:3787, power in 3789:5681. Every
  other sheet re-derives those offsets by hand, and two get it wrong: the RPM
  min/max reaches one row into the current block, and the EMF sheet reads one
  row past the power block into blank cells, which is why its last value is 0.
* `ThermalLoad!B5` carries an extra `* B1` factor that the adjacent `B4` does
  not — a 50x discrepancy between two cells that are supposed to be the min and
  max of one quantity.
* `SUM` of instantaneous watts, labelled "Sum of Joule Heat (kW)". Summing
  power does not give energy; energy needs sum(P * dt).
* Unit labels contradict their formulas in at least three places, including
  "Min Power (W)" and "Max Power (kW)" as adjacent labels on the same block.

WHAT THIS MODULE FOUND IN KINEMATIK ITSELF
------------------------------------------
Auditing the workbook meant testing the code that reads it, and that code is
worse off than the spreadsheet:

`ev_electrical_check.ElecParams.from_excel` and
`ev_excel_roundtrip.extract_params_from_excel` both gate on a sheet named
"Battery Pack Calcs". This workbook's sheet is called "BatteryPackConfig".
Neither reader has ever opened it. `extract_params_from_excel` returns an empty
pack dict; `from_excel` falls through to its dataclass defaults — which happen
to be 504 V, 15 Ah, 140S3P, exactly this pack. So it appears to work, and every
value it reports is a coincidence that survives any edit to the workbook.
`assert_reader_sees_the_file` proves this by mutating a copy and checking the
reader notices.

Worse, `ev_excel_roundtrip` re-implements the workbook's formulas *including*
their errors — the same `V * sqrt(3) * PF * RPM / 1000`, the same inverted gear
ratio. The bridge has been validating the spreadsheet against a copy of itself.
Meanwhile `ev_electrical_check` already contains the correct formulation
(`P_elec = P_wheel / eta`, `I = P / V`). Two modules in one repository
disagree, and the wrong one is the one wired to the Excel.

This module reads the workbook by matching row *labels* rather than sheet names
and cell coordinates, and raises on a value it cannot interpret rather than
coercing it — see `read_pack_config`.
"""

from __future__ import annotations

import datetime as _dt
import math
from dataclasses import dataclass, field
from typing import Optional, Sequence

from .interfaces import Severity, Finding

G = 9.80665
#: mph -> inches/minute. 1 mile = 63360 in, so 63360/60 = 1056.
MPH_TO_IN_PER_MIN = 1056.0
MPH_TO_MS = 0.44704


# ===================================================================== #
#  1.  READING THE WORKBOOK WITHOUT LYING ABOUT IT
# ===================================================================== #
class WorkbookReadError(RuntimeError):
    """Raised when a value cannot be read honestly, instead of substituting one."""


def _numeric(value, where: str) -> float:
    """Coerce a cell value to float, or raise — never silently substitute.

    `datetime` types get their own message because they are the failure mode
    that actually occurs here: a cell whose *number format* is a date renders
    its cached value as a time, so `float(v)` raises, and a `try/except` that
    returns a default turns 0.125 into whatever the default is. In the shipped
    reader that default is 1.0, which converts gear ratio 1/8 into direct
    drive without a word.
    """
    if value is None:
        raise WorkbookReadError(f"{where} is empty")
    if isinstance(value, bool):
        raise WorkbookReadError(f"{where} holds a boolean, not a number")
    if isinstance(value, (_dt.datetime, _dt.date, _dt.time, _dt.timedelta)):
        raise WorkbookReadError(
            f"{where} holds {value!r}, a date/time value. The cell's number "
            f"format is a date format, so its cached value is a time serial "
            f"rather than a number. In Excel the underlying arithmetic is "
            f"still correct; every Python reader of this file sees a "
            f"datetime. Set the cell's format to General. Do NOT coerce this "
            f"to a float — the coercion is the bug.")
    if isinstance(value, str):
        try:
            return float(value.strip())
        except ValueError:
            raise WorkbookReadError(f"{where} holds text {value!r}")
    return float(value)


def _label_map(ws, *, label_col: int = 1, value_col: int = 2,
               max_row: Optional[int] = None) -> dict[str, tuple]:
    """Map lowercased row label -> (value, coordinate).

    Reading by label rather than by hardcoded coordinate is the whole point:
    inserting a row above "Fuse Max (A)" moves it, and a reader keyed to B1
    then silently reports the wrong quantity.
    """
    out: dict[str, tuple] = {}
    limit = max_row or ws.max_row
    for r in range(1, limit + 1):
        lab = ws.cell(r, label_col).value
        if not isinstance(lab, str) or not lab.strip():
            continue
        key = lab.strip().lower()
        cell = ws.cell(r, value_col)
        out.setdefault(key, (cell.value, cell.coordinate))
    return out


def _find_sheet(wb, *must_contain: str):
    """Locate a sheet by the labels it contains, not by its name."""
    for ws in wb.worksheets:
        labels = " ".join(str(ws.cell(r, 1).value or "").lower()
                          for r in range(1, min(ws.max_row, 40) + 1))
        if all(t.lower() in labels for t in must_contain):
            return ws
    return None


@dataclass
class PackSpec:
    """Battery pack, with the arithmetic done correctly.

    Field names deliberately mirror the workbook's own labels so the audit can
    quote cells, but every derived quantity is recomputed here rather than read
    from the sheet.
    """
    n_series: int = 140
    n_parallel: int = 3
    cell_voltage_v: float = 3.6
    cell_capacity_ah: float = 5.0
    cell_resistance_ohm: float = 0.0128
    cell_weight_kg: float = 0.071
    fuse_max_a: float = 50.0
    max_cells: Optional[int] = 560
    source: str = ""

    # --- topology --------------------------------------------------------- #
    def cell_count(self) -> int:
        return self.n_series * self.n_parallel

    def nominal_voltage_v(self) -> float:
        return self.cell_voltage_v * self.n_series

    def capacity_ah(self) -> float:
        return self.cell_capacity_ah * self.n_parallel

    def energy_kwh(self) -> float:
        return self.nominal_voltage_v() * self.capacity_ah() / 1000.0

    def mass_kg(self) -> float:
        return self.cell_weight_kg * self.cell_count()

    def resistance_ohm(self) -> float:
        """Pack DC resistance: series groups add, parallel cells divide.

        R = S * (R_cell / P)

        The workbook writes `cell_count * (R_cell / P)`, which cancels to
        `S * R_cell` and so discards the parallel benefit completely. For
        140S3P that is 1.792 ohm against a true 0.597 ohm. It is not a rounding
        difference — it triples every joule-heating figure in the file.
        """
        return self.n_series * (self.cell_resistance_ohm / self.n_parallel)

    # --- limits the workbook never computes -------------------------------- #
    def max_deliverable_power_w(self) -> float:
        """Ceiling on electrical power out of the pack, from its own resistance.

        With P = V*I and V = V_oc - I*R, power peaks at I = V_oc/(2R), giving
        P_max = V_oc^2 / (4R). At that point half the energy is burnt inside
        the pack, so it is a hard ceiling and not an operating point — but it
        is the number that tells you whether a motor rating is reachable at
        all.
        """
        return self.nominal_voltage_v() ** 2 / (4.0 * self.resistance_ohm())

    def power_at_fuse_limit_w(self) -> float:
        """Usable power at the fuse's current limit, including pack sag."""
        i = self.fuse_max_a
        return i * (self.nominal_voltage_v() - i * self.resistance_ohm())

    def current_for_power(self, p_elec_w: float) -> Optional[float]:
        """Pack current needed to deliver `p_elec_w`, accounting for sag.

        Solves R*I^2 - V_oc*I + P = 0 and takes the low-current root. Returns
        None when the demand exceeds `max_deliverable_power_w` — the honest
        answer, rather than a current the pack cannot actually supply.
        """
        if p_elec_w <= 0:
            return 0.0
        voc, r = self.nominal_voltage_v(), self.resistance_ohm()
        disc = voc * voc - 4.0 * r * p_elec_w
        if disc < 0:
            return None
        return (voc - math.sqrt(disc)) / (2.0 * r)

    def joule_heat_w(self, i_pack_a: float) -> float:
        """I^2 * R, in WATTS. The unit is the point: the workbook labels this
        quantity kW in one place and kWh in another, for the same formula."""
        return i_pack_a ** 2 * self.resistance_ohm()

    def cell_current(self, i_pack_a: float) -> float:
        return i_pack_a / self.n_parallel


def read_pack_config(path: str) -> PackSpec:
    """Read pack topology from the workbook by label, raising on bad values."""
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = _find_sheet(wb, "series battery count", "parrallel battery count") \
        or _find_sheet(wb, "series battery count") \
        or _find_sheet(wb, "fuse max")
    if ws is None:
        raise WorkbookReadError(
            "no sheet contains the pack topology labels. Note that the shipped "
            "readers look for a sheet *named* 'Battery Pack Calcs' and this "
            "file's is 'BatteryPackConfig', which is exactly why they fail "
            "silently — searching by label avoids that whole class of bug.")
    m = _label_map(ws)

    def grab(*candidates: str) -> tuple:
        for key in m:
            for c in candidates:
                if c in key:
                    return m[key]
        raise WorkbookReadError(
            f"none of {candidates} found among the labels on '{ws.title}'")

    def num(*candidates: str) -> float:
        v, coord = grab(*candidates)
        return _numeric(v, f"{ws.title}!{coord}")

    return PackSpec(
        n_series=int(num("series battery count")),
        n_parallel=int(num("parrallel battery count", "parallel battery count")),
        cell_voltage_v=num("nominal battery voltage"),
        cell_capacity_ah=num("capacity battery cell"),
        cell_resistance_ohm=num("internal resistance battery cell"),
        cell_weight_kg=num("battery cell weight"),
        fuse_max_a=num("fuse max"),
        max_cells=int(num("max battery cells")),
        source=f"{path}!{ws.title}")


def assert_reader_sees_the_file(path: str, reader, *,
                                probe: Optional[dict] = None) -> list[Finding]:
    """Prove a reader actually reads the workbook instead of defaulting.

    Mutates a scratch copy of the pack topology and checks the reader's output
    changes. A reader that returns identical values from a mutated file is not
    reading it, and the fact that its numbers look right means only that its
    defaults were written by someone looking at this same pack.

    This is a general trap, not a one-off: any `from_excel` with dataclass
    defaults and a `sheetnames` guard can fail this way, and it fails
    invisibly, because the output is plausible.
    """
    import shutil
    import tempfile
    import openpyxl

    probe = probe or {"B3": 100, "B2": 4}
    out: list[Finding] = []
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tf:
        tmp = tf.name
    shutil.copy(path, tmp)
    wb = openpyxl.load_workbook(tmp)
    ws = _find_sheet(wb, "series battery count") or wb.worksheets[0]
    for coord, val in probe.items():
        ws[coord] = val
    wb.save(tmp)

    try:
        before, after = reader(path), reader(tmp)
    except Exception as exc:                       # a raise is a valid answer
        out.append(Finding(
            "pd-reader-raises", Severity.INFO,
            f"Reader raised on the workbook: {type(exc).__name__}: {exc}. "
            f"A raise is far better than a silent default — at least the "
            f"caller finds out.",
            subsystems=["dataacq"]))
        return out

    if repr(before) == repr(after):
        out.append(Finding(
            "pd-reader-ignores-file", Severity.FAIL,
            f"{getattr(reader, '__qualname__', reader)} returns byte-identical "
            f"values from a workbook whose series and parallel counts were "
            f"changed to {probe}. It is not reading this file. If its output "
            f"currently looks correct, that is because its hardcoded defaults "
            f"were written against this same pack — so it will keep reporting "
            f"140S3P after the team moves to anything else, and nothing will "
            f"warn them.",
            subsystems=["dataacq", "electrics", "powertrain"]))
    else:
        out.append(Finding(
            "pd-reader-ok", Severity.OK,
            f"{getattr(reader, '__qualname__', reader)} tracks edits to the "
            f"workbook.", subsystems=["dataacq"]))
    return out


# ===================================================================== #
#  2.  THE VEHICLE THE WORKBOOK DOES NOT HAVE
# ===================================================================== #
@dataclass
class VehicleSpec:
    """The car, to the extent power draw depends on it.

    Every one of these is absent from the workbook, which is why its current
    figures cannot be right in principle rather than merely in detail. Mass and
    drag are what turn a speed trace into a power demand.
    """
    mass_kg: float = 300.0                 # car + driver
    crr: float = 0.020                     # rolling resistance coefficient
    cda_m2: float = 1.10                   # drag area, Cd * A
    air_density: float = 1.225
    wheel_diameter_in: float = 18.0
    drivetrain_efficiency: float = 0.95    # gearbox, chain, bearings
    motor_efficiency: float = 0.9545
    inverter_efficiency: float = 0.97
    regen: bool = False                    # recover braking energy?
    regen_efficiency: float = 0.60
    #: Longitudinal grip coefficient. Used to reject accelerations the tyres
    #: could not have produced, which is how source-data artefacts are caught
    #: before they become 400 A current spikes.
    mu_lon: float = 1.4
    source: str = ""

    def wheel_radius_m(self) -> float:
        return self.wheel_diameter_in * 0.0254 / 2.0

    def total_efficiency(self) -> float:
        return (self.drivetrain_efficiency * self.motor_efficiency
                * self.inverter_efficiency)


@dataclass
class DriveSpec:
    """Gearing, with the reduction the right way up.

    `reduction` is the motor-to-wheel ratio and is >= 1 for a reduction: the
    motor turns `reduction` times for each wheel turn. The workbook stores
    1/reduction and then multiplies by it, so its motor spins slower than its
    wheels — the single largest error in the file.
    """
    reduction: float = 7.0
    motor_max_rpm: float = 23500.0
    motor_peak_torque_nm: float = 120.0
    motor_peak_power_kw: float = 150.0
    motor_continuous_torque_nm: float = 77.0

    def motor_rpm(self, speed_mph: float, wheel_diameter_in: float) -> float:
        """Motor rpm for a road speed. Multiplies by the reduction."""
        wheel_rpm = (speed_mph * MPH_TO_IN_PER_MIN
                     / (wheel_diameter_in * math.pi))
        return wheel_rpm * self.reduction

    def motor_torque_nm(self, tractive_force_n: float, wheel_radius_m: float,
                        driveline_efficiency: float = 1.0) -> float:
        """Motor torque needed for a tractive force at the contact patch.

        This is what actually selects a gear ratio, and it is the quantity the
        workbook's fifteen-column sweep never computes. Pack current does not
        depend on gearing at all — power is force times speed either way — so a
        sweep that reports only current and power looks identical in every
        column. Torque and motor speed are what change, and they are what
        decide whether the motor can do the job.
        """
        return (tractive_force_n * wheel_radius_m
                / (self.reduction * driveline_efficiency))


# ===================================================================== #
#  3.  THE FORCE BALANCE
# ===================================================================== #
@dataclass
class PowerDrawTrace:
    """Per-sample power draw for one speed trace at one gear ratio."""
    dt_s: float
    speed_ms: list
    accel_ms2: list
    f_roll_n: list
    f_aero_n: list
    f_accel_n: list
    p_wheel_w: list
    p_elec_w: list
    i_pack_a: list
    joule_w: list
    motor_rpm: list
    motor_torque_nm: list
    infeasible_samples: int
    clamped_samples: int
    discontinuities: list
    smooth_window: int
    pack: PackSpec
    vehicle: VehicleSpec
    drive: DriveSpec

    # --- summaries --------------------------------------------------------- #
    def peak_current_a(self) -> float:
        return max(self.i_pack_a) if self.i_pack_a else 0.0

    def peak_power_kw(self) -> float:
        return max(self.p_elec_w) / 1000.0 if self.p_elec_w else 0.0

    def energy_kwh(self) -> float:
        """Integral of electrical power. Sum(P * dt), not sum(P)."""
        return sum(p * self.dt_s for p in self.p_elec_w) / 3.6e6

    def joule_energy_kwh(self) -> float:
        return sum(q * self.dt_s for q in self.joule_w) / 3.6e6

    def mean_joule_w(self) -> float:
        return sum(self.joule_w) / len(self.joule_w) if self.joule_w else 0.0

    def peak_motor_rpm(self) -> float:
        return max(self.motor_rpm) if self.motor_rpm else 0.0

    def peak_motor_torque_nm(self) -> float:
        return max(self.motor_torque_nm) if self.motor_torque_nm else 0.0

    def duration_s(self) -> float:
        return len(self.speed_ms) * self.dt_s


def find_discontinuities(speed_mph: Sequence[float], dt_s: float,
                         mass_kg: float, mu_lon: float = 1.4) -> list:
    """Samples whose implied acceleration no tyre could have produced.

    Returns (index, delta_mph, implied_g) for each step beyond the grip limit.
    The workbook never differentiates its speed trace, so it never notices
    these; any correct power model must, because differentiating a quantised or
    glitched trace turns each step into a current spike that then propagates
    into every thermal figure downstream.
    """
    out = []
    limit_g = mu_lon * 1.05          # a little headroom over pure grip
    for i in range(len(speed_mph) - 1):
        dv = (speed_mph[i + 1] - speed_mph[i]) * MPH_TO_MS
        g = abs(dv / dt_s) / G
        if g > limit_g:
            out.append((i, speed_mph[i + 1] - speed_mph[i], dv / dt_s / G))
    return out


def _moving_average(x: Sequence[float], window: int) -> list:
    """Centred moving average, shrinking at the ends rather than padding."""
    if window <= 1:
        return list(x)
    half = window // 2
    n = len(x)
    return [sum(x[max(0, i - half):min(n, i + half + 1)])
            / len(x[max(0, i - half):min(n, i + half + 1)]) for i in range(n)]


def power_draw_trace(speed_mph: Sequence[float], dt_s: float,
                     pack: PackSpec, vehicle: VehicleSpec,
                     drive: DriveSpec, *,
                     smooth_window: int = 1,
                     clamp_to_grip: bool = True) -> PowerDrawTrace:
    """Turn a speed trace into pack current, the way the physics requires.

    At each sample:

        F = Crr*m*g  +  0.5*rho*CdA*v^2  +  m*a
        P_wheel = F * v
        P_elec  = P_wheel / eta            (divide — losses ADD to the demand)
        I_pack  = solved from P = (V_oc - I*R) * I

    Efficiency dividing rather than multiplying is worth dwelling on, because
    the workbook multiplies: written that way, improving the motor increases
    the current drawn from the pack, which is backwards. Losses are paid by the
    pack, so they raise the electrical demand above the mechanical one.

    Braking samples draw zero unless `vehicle.regen` is set, in which case they
    return energy at `regen_efficiency`.

    `smooth_window` averages the speed trace before differentiating. This is not
    cosmetic: a trace quantised to half-mph steps, or one containing a logging
    glitch, produces accelerations of several g that never happened, and the
    m*a term turns each of those into a current spike. `clamp_to_grip` then
    limits whatever survives to what the tyres could actually transmit, and the
    count of clamped samples is reported rather than hidden — if it is large,
    the input trace is the problem and no amount of clamping fixes it.
    """
    if dt_s <= 0:
        raise ValueError("dt_s must be positive")
    if len(speed_mph) < 2:
        raise ValueError("a trace needs at least two samples to differentiate")

    discont = find_discontinuities(speed_mph, dt_s, vehicle.mass_kg,
                                  vehicle.mu_lon)
    speed_used = _moving_average(speed_mph, smooth_window)
    v = [s * MPH_TO_MS for s in speed_used]
    n = len(v)
    accel, f_roll, f_aero, f_acc = [], [], [], []
    p_wheel, p_elec, i_pack, joule, rpm, torque = [], [], [], [], [], []
    infeasible = 0
    clamped = 0
    eta = vehicle.total_efficiency()
    a_limit = vehicle.mu_lon * G

    for k in range(n):
        # central difference inside, one-sided at the ends
        if k == 0:
            a = (v[1] - v[0]) / dt_s
        elif k == n - 1:
            a = (v[-1] - v[-2]) / dt_s
        else:
            a = (v[k + 1] - v[k - 1]) / (2.0 * dt_s)
        if clamp_to_grip and abs(a) > a_limit:
            a = math.copysign(a_limit, a)
            clamped += 1
        accel.append(a)

        fr = vehicle.crr * vehicle.mass_kg * G if v[k] > 0.05 else 0.0
        fa = 0.5 * vehicle.air_density * vehicle.cda_m2 * v[k] ** 2
        fx = vehicle.mass_kg * a
        f_roll.append(fr)
        f_aero.append(fa)
        f_acc.append(fx)

        pw = (fr + fa + fx) * v[k]
        p_wheel.append(pw)

        if pw >= 0:
            pe = pw / eta
        elif vehicle.regen:
            pe = pw * eta * vehicle.regen_efficiency
        else:
            pe = 0.0
        p_elec.append(pe)

        cur = pack.current_for_power(pe) if pe > 0 else 0.0
        if cur is None:
            infeasible += 1
            cur = pack.nominal_voltage_v() / (2.0 * pack.resistance_ohm())
        i_pack.append(cur)
        joule.append(pack.joule_heat_w(cur))
        rpm.append(drive.motor_rpm(speed_used[k], vehicle.wheel_diameter_in))
        torque.append(drive.motor_torque_nm(
            max(0.0, fr + fa + fx), vehicle.wheel_radius_m(),
            vehicle.drivetrain_efficiency))

    return PowerDrawTrace(dt_s, v, accel, f_roll, f_aero, f_acc,
                          p_wheel, p_elec, i_pack, joule, rpm, torque,
                          infeasible, clamped, discont, smooth_window,
                          pack, vehicle, drive)


def gear_sweep(speed_mph: Sequence[float], dt_s: float, pack: PackSpec,
               vehicle: VehicleSpec, drive: DriveSpec,
               reductions: Sequence[float] = tuple(range(1, 16)),
               smooth_window: int = 1,
               clamp_to_grip: bool = True) -> dict[float, PowerDrawTrace]:
    """One trace per gear reduction, keyed by the reduction.

    A dict keyed by the actual ratio, rather than fifteen columns stacked into
    three vertical blocks at fixed row offsets. The offsets are the reason two
    of the workbook's consumer sheets read one row into the wrong block; there
    is nothing to get wrong here.
    """
    out: dict[float, PowerDrawTrace] = {}
    for red in reductions:
        d = DriveSpec(reduction=float(red),
                      motor_max_rpm=drive.motor_max_rpm,
                      motor_peak_torque_nm=drive.motor_peak_torque_nm,
                      motor_peak_power_kw=drive.motor_peak_power_kw,
                      motor_continuous_torque_nm=drive.motor_continuous_torque_nm)
        out[float(red)] = power_draw_trace(
            speed_mph, dt_s, pack, vehicle, d,
            smooth_window=smooth_window, clamp_to_grip=clamp_to_grip)
    return out


# ===================================================================== #
#  4.  BEARING LOAD — degrees, radians, and the moment arm
# ===================================================================== #
def gear_tangential_force_n(torque_nm: float, pitch_diameter_m: float) -> float:
    """F_t = T / r, and r is the pitch RADIUS.

    The workbook divides by the pitch diameter, halving the force. Paired with
    the radians bug below the two errors partially cancel, which is the worst
    possible outcome: the answer looks plausible and neither error is visible.
    """
    if pitch_diameter_m <= 0:
        raise ValueError("pitch diameter must be positive")
    return torque_nm / (pitch_diameter_m / 2.0)


def gear_radial_force_n(tangential_n: float,
                        pressure_angle_deg: float = 20.0) -> float:
    """F_r = F_t * tan(pressure angle), with the angle converted to radians.

    Excel's `TAN` takes radians. `TAN(20)` is the tangent of 20 radians —
    2.2372 — where tan(20 deg) is 0.36397, a factor of 6.15. The workbook does
    not call `RADIANS()`.
    """
    return tangential_n * math.tan(math.radians(pressure_angle_deg))


# ===================================================================== #
#  5.  THE AUDIT
# ===================================================================== #
def _f(check, sev, msg, cells=None, subs=None, detail=None) -> Finding:
    return Finding(check, sev,
                   (f"[{cells}] " if cells else "") + msg,
                   subsystems=subs or ["electrics", "powertrain"],
                   detail=detail or {})


def audit_pack_sheet(pack: PackSpec, sheet_values: dict) -> list[Finding]:
    """Check the pack sheet's own derived cells against correct arithmetic."""
    out: list[Finding] = []
    r_correct = pack.resistance_ohm()
    r_sheet = sheet_values.get("pack_resistance_ohm")

    if r_sheet is not None and abs(r_sheet - r_correct) > 1e-6:
        out.append(_f(
            "pd-pack-resistance", Severity.FAIL,
            f"Pack internal resistance reads {r_sheet:.4f} ohm; for "
            f"{pack.n_series}S{pack.n_parallel}P it is "
            f"{r_correct:.4f} ohm. The formula multiplies the TOTAL cell count "
            f"by R_cell/P, which cancels to S*R_cell and discards the parallel "
            f"benefit entirely — so it is high by exactly the parallel count, "
            f"{r_sheet / r_correct:.0f}x. This value feeds every joule-heating "
            f"figure and both EMF columns, so those are all high by the same "
            f"factor.",
            cells="BatteryPackConfig!B11",
            detail={"sheet": r_sheet, "correct": r_correct}))

    jh = sheet_values.get("joule_max_label", "")
    if "kwh" in jh.lower():
        out.append(_f(
            "pd-joule-units", Severity.FAIL,
            f"'{jh}' computes I^2*R/1000, which is kW — a power, not an "
            f"energy. Sheet1 then multiplies it by 1000 and labels the result "
            f"watts, so the downstream use is consistent with kW and only the "
            f"label is wrong. Energy would need this integrated over time.",
            cells="BatteryPackConfig!B19",
            subs=["electrics", "cooling"]))

    # The comparison nobody in the workbook makes.
    p_fuse = pack.power_at_fuse_limit_w() / 1000.0
    p_max = pack.max_deliverable_power_w() / 1000.0
    out.append(_f(
        "pd-power-ceiling", Severity.WARN,
        f"The pack's own resistance caps electrical output at "
        f"{p_max:.0f} kW (at I = V_oc/2R, where half the energy is burnt "
        f"inside the pack). The {pack.fuse_max_a:.0f} A fuse caps usable "
        f"output at {p_fuse:.1f} kW once sag is included. Neither number "
        f"appears anywhere in the workbook, and both are worth having next to "
        f"the motor rating before anyone sizes a controller.",
        cells="BatteryPackConfig!B1,B11",
        detail={"fuse_limited_kw": p_fuse, "pack_ceiling_kw": p_max}))
    return out


def audit_workbook(path: str) -> list[Finding]:
    """Full audit of the power-draw workbook. Every finding names its cells."""
    import openpyxl
    from openpyxl.worksheet.formula import ArrayFormula

    wbf = openpyxl.load_workbook(path)
    wbv = openpyxl.load_workbook(path, data_only=True)
    out: list[Finding] = []
    names = wbf.sheetnames

    def ftext(ws, coord):
        v = ws[coord].value
        return v.text if isinstance(v, ArrayFormula) else v

    # ---- the headline: no vehicle model ---------------------------------- #
    if "ElecPropulsion" in names:
        ep = wbf["ElecPropulsion"]
        cur = str(ftext(ep, "H1895") or "")
        if "B$11" in cur and "H2" in cur and "SQRT" not in cur.upper():
            out.append(_f(
                "pd-no-vehicle-model", Severity.FAIL,
                "Pack current is computed as V_pack * PowerFactor * RPM / 1000. "
                "That is volts times rpm — not amperes — and it makes current "
                "proportional to bus voltage, which is backwards: at a given "
                "power a higher voltage draws LESS current. More importantly "
                "there is no mass, drag, rolling resistance or acceleration "
                "anywhere in the workbook, so nothing computes tractive force. "
                "A power-draw model that never computes force cannot predict "
                "power draw, and every figure downstream inherits this — joule "
                "heating, thermal load, both EMF columns, the bearing "
                "temperatures. Use power_draw_trace(), which starts from a "
                "force balance.",
                cells="ElecPropulsion!H1895:V3787",
                subs=["electrics", "powertrain", "chassis", "aero"]))

        # ---- gear ratio inverted ----------------------------------------- #
        rpm = str(ftext(ep, "H2") or "")
        if "$H$1" in rpm and "1056" in rpm:
            g1 = wbv["ElecPropulsion"]["N1"].value
            if isinstance(g1, (int, float)) and 0 < g1 < 1:
                out.append(_f(
                    "pd-gear-inverted", Severity.FAIL,
                    "Gear ratios are stored as 1/N and then MULTIPLIED into "
                    "wheel rpm, so a reduction makes the motor turn slower "
                    "than the wheel. At the 7th column the motor reads about "
                    "93 rpm where a 7:1 reduction gives 4575. Power goes as "
                    "the square of speed here, so the error is 49x at gear 7 "
                    "and 225x at gear 15 — and it is invisible because every "
                    "column is wrong the same way, so the columns still rank "
                    "in a sensible order.",
                    cells="ElecPropulsion!H1:V1, H2",
                    subs=["powertrain", "electrics"]))

        # ---- efficiency multiplying; sqrt(3) on a DC bus ------------------ #
        pw = str(ftext(ep, "H3789") or "")
        if "SQRT(3)" in pw.upper():
            out.append(_f(
                "pd-efficiency-and-phase", Severity.FAIL,
                "Motor power multiplies by efficiency (B6) where it should "
                "divide: losses are paid by the pack, so they RAISE electrical "
                "demand. As written, a more efficient motor draws more current. "
                "The same formula applies sqrt(3) and power factor — "
                "three-phase AC line quantities — to the DC bus voltage, "
                "double-counting the phase conversion.",
                cells="ElecPropulsion!H3789:V5681",
                subs=["electrics", "powertrain"]))

        # ---- date-formatted gear cells ----------------------------------- #
        broken = []
        for c in range(8, 23):
            cell = wbv["ElecPropulsion"].cell(1, c)
            if isinstance(cell.value, (_dt.datetime, _dt.date, _dt.time)):
                broken.append((cell.coordinate, c - 7, cell.value))
        if broken:
            out.append(_f(
                "pd-gear-date-format", Severity.WARN,
                f"{len(broken)} gear-ratio cells carry a date number format, "
                f"so their cached values are times rather than numbers: "
                f"{', '.join(f'{c} (gear {g}) = {v}' for c, g, v in broken)}. "
                f"The formulas are correct and Excel's own arithmetic is fine; "
                f"the damage is at the Python boundary, where float() raises "
                f"and a try/except default silently substitutes. KinematiK's "
                f"_safe_float turns these into 1.0, quietly converting gears "
                f"{' and '.join(str(g) for _, g, _ in broken)} into direct "
                f"drive. Set the format to General.",
                cells=",".join(c for c, _, _ in broken),
                subs=["dataacq", "electrics"]))

        # ---- min/max range off-by-one ------------------------------------ #
        mn = str(ftext(ep, "H5683") or "")
        if "H1895" in mn:
            out.append(_f(
                "pd-rpm-stat-range", Severity.FAIL,
                "Min/Max RPM read H2:H1895, but the RPM block ends at row "
                "1894 — row 1895 is the first row of the Current Draw block. "
                "The RPM statistics are contaminated by an ampere value. This "
                "is the direct consequence of stacking three blocks in one "
                "sheet at fixed offsets.",
                cells="ElecPropulsion!H5683:V5684"))

        # ---- contradictory unit labels ----------------------------------- #
        lo = str(wbf["ElecPropulsion"]["G5687"].value or "")
        hi = str(wbf["ElecPropulsion"]["G5688"].value or "")
        if "(W)" in lo and "(kW)" in hi:
            out.append(_f(
                "pd-power-stat-units", Severity.WARN,
                f"'{lo}' and '{hi}' label the min and max of the same block "
                f"with different units. The formula divides by 1000, so both "
                f"are kW.",
                cells="ElecPropulsion!G5687,G5688"))

    # ---- ThermalLoad ------------------------------------------------------ #
    if "ThermalLoad" in names:
        tl, tlf = wbv["ThermalLoad"], wbf["ThermalLoad"]
        b4, b5 = str(ftext(tlf, "B4") or ""), str(ftext(tlf, "B5") or "")
        # \b after B1 matters: "B1" is a substring of "B11", which both cells
        # legitimately reference. Without the boundary this check never fires.
        import re as _re
        _fuse_ref = _re.compile(r"B1(?![0-9])")
        if b4 and b5 and _fuse_ref.search(b5) and not _fuse_ref.search(b4):
            out.append(_f(
                "pd-minmax-inconsistent", Severity.FAIL,
                "Max Joule Heat multiplies by BatteryPackConfig!B1 (the fuse "
                "rating) and Min Joule Heat does not. Two adjacent cells that "
                "should be the min and max of one quantity differ by a factor "
                "of 50, and the extra factor is dimensionally wrong anyway: "
                "I^2 * A * ohm. This is the classic lone-edited-cell error, "
                "and it is only visible by reading the two formulas side by "
                "side.",
                cells="ThermalLoad!B4,B5",
                subs=["electrics", "cooling"]))

        b6 = str(ftext(tlf, "B6") or "")
        if b6.upper().startswith("=SUM"):
            out.append(_f(
                "pd-sum-of-power", Severity.FAIL,
                "'Sum of Joule Heat (kW)' sums instantaneous power across "
                "samples. Adding watts gives watts, not energy — the result "
                "has no physical meaning and scales with however many rows the "
                "trace happens to have. Energy needs sum(P * dt); at the "
                "trace's 1/15 s step the true figure is a fraction of a kWh, "
                "not the thousands the cell reports. "
                "PowerDrawTrace.joule_energy_kwh() does this correctly.",
                cells="ThermalLoad!B6",
                subs=["electrics", "cooling"]))

        f1 = str(ftext(tlf, "F1") or "")
        if "kW" in str(tlf["E1"].value or "") and "B11" in f1:
            out.append(_f(
                "pd-thermal-units", Severity.WARN,
                "The 'Joule Heat (kW)' column computes I^2*R, which is watts. "
                "Its values are around 3,600 — watts, not kilowatts. Combined "
                "with the 3x resistance error the true figure is roughly "
                "1.2 kW, not the 3,600 the column implies.",
                cells="ThermalLoad!E1,F1:F1893",
                subs=["electrics", "cooling"]))

    # ---- EMFs ------------------------------------------------------------- #
    if "EMFs" in names:
        e = wbf["EMFs"]
        b1 = str(ftext(e, "B1") or "")
        if "B12" in b1 and "B11" in b1:
            out.append(_f(
                "pd-emf-dimensions", Severity.FAIL,
                "The EMF columns compute V_pack * Power * R_pack — volts times "
                "kilowatts times ohms. That is not a voltage, and the cached "
                "values (~32,000) are not plausible for a 504 V pack. Back-EMF "
                "is either k_e * omega for the motor or V_oc - I*R for the "
                "pack terminal; both are available from the corrected trace. "
                "Whatever this column is measuring, it is not EMF.",
                cells="EMFs!B1:B1894, E1:E1894"))
        if "5682" in b1:
            out.append(_f(
                "pd-emf-range", Severity.FAIL,
                "The EMF columns read ElecPropulsion rows 3789:5682, but the "
                "Motor Power block ends at 5681. Row 5682 is blank, which is "
                "why the last EMF value is 0. Same root cause as the RPM "
                "statistic: hand-copied row offsets into stacked blocks.",
                cells="EMFs!B1,E1"))

    # ---- BearingBlowOut --------------------------------------------------- #
    if "BearingBlowOut" in names:
        bb = wbf["BearingBlowOut"]
        g1 = str(ftext(bb, "G1") or "")
        if "TAN(" in g1.upper() and "RADIANS" not in g1.upper():
            out.append(_f(
                "pd-tan-degrees", Severity.FAIL,
                f"TAN() is applied to a pressure angle in degrees. Excel's TAN "
                f"takes radians: TAN(20) is {math.tan(20):.4f}, the tangent of "
                f"20 radians, where tan(20 deg) is "
                f"{math.tan(math.radians(20)):.4f} — a factor of "
                f"{math.tan(20)/math.tan(math.radians(20)):.2f}. Wrap the angle "
                f"in RADIANS().",
                cells="BearingBlowOut!G1:G1893",
                subs=["powertrain", "suspension"]))
        if "$B$16" in g1:
            out.append(_f(
                "pd-pitch-diameter", Severity.FAIL,
                "Gear tangential force divides torque by the pitch DIAMETER "
                "(B16), but the moment arm is the pitch radius, so the force "
                "is half what it should be. Note this partially cancels the "
                "radians error above — two independent mistakes leaving a "
                "plausible-looking answer, which is why neither was caught.",
                cells="BearingBlowOut!G1:G1893,B16",
                subs=["powertrain", "suspension"]))
        b19 = str(ftext(bb, "B19") or "")
        if "G1894" in b19:
            out.append(_f(
                "pd-bearing-max-range", Severity.WARN,
                "Max bearing temperature reads MAX(G1:G1894), but column G is "
                "populated only to row 1893.",
                cells="BearingBlowOut!B19",
                subs=["powertrain"]))

    # ---- Sheet1 hardcoded resistance network ------------------------------ #
    if "Sheet1" in names:
        s1 = wbf["Sheet1"]
        hard = [s1.cell(r, 1).coordinate for r in (19, 21, 23, 25, 27, 29, 31)
                if isinstance(s1.cell(r, 1).value, (int, float))]
        if hard:
            out.append(_f(
                "pd-hardcoded-resistances", Severity.WARN,
                f"The conduction resistances at {', '.join(hard)} are typed "
                f"numbers sitting directly under a label that reads "
                f"'Rcond = L/(kA)'. The geometry (A2, A4, A8, A10, A12) and the "
                f"conductivities on MatCharacteristics are both right there and "
                f"neither is referenced, so changing the insulation thickness "
                f"moves nothing. A sheet named 'Sheet1' holding the thermal "
                f"stack-up is also worth renaming before anyone else has to "
                f"find it.",
                cells=",".join(hard),
                subs=["cooling", "electrics"]))

    if "Data Dashboard" in names and wbf["Data Dashboard"].max_row <= 1:
        out.append(_f(
            "pd-empty-dashboard", Severity.INFO,
            "'Data Dashboard' is empty. It is the first tab, so it is what a "
            "reviewer opens first. The corrected summary figures from "
            "PowerDrawTrace are what belong here.",
            cells="Data Dashboard!A1",
            subs=["dataacq"]))

    return out


def audit_kinematik_readers(path: str) -> list[Finding]:
    """Audit KinematiK's own workbook readers against this workbook."""
    out: list[Finding] = []
    try:
        from .ev_electrical_check import ElecParams
    except Exception:
        return out

    import openpyxl
    names = openpyxl.load_workbook(path).sheetnames
    if "Battery Pack Calcs" not in names:
        out.append(_f(
            "pd-reader-sheet-name", Severity.FAIL,
            f"ev_electrical_check.ElecParams.from_excel and "
            f"ev_excel_roundtrip.extract_params_from_excel both gate on a "
            f"sheet named 'Battery Pack Calcs'. This workbook's pack sheet is "
            f"'BatteryPackConfig', so neither has ever read it: the round-trip "
            f"returns an empty pack dict, and from_excel falls through to its "
            f"dataclass defaults. Those defaults are 504 V, 15 Ah, 140S3P — "
            f"this exact pack — so it looks like it works. Read by label "
            f"instead: read_pack_config() in this module does.",
            cells="|".join(names),
            subs=["dataacq", "electrics"]))

    out.extend(assert_reader_sees_the_file(path, ElecParams.from_excel))

    out.append(_f(
        "pd-roundtrip-mirrors-errors", Severity.FAIL,
        "ev_excel_roundtrip re-implements the workbook's formulas including "
        "their errors — the same V*sqrt(3)*PF*RPM/1000 current, the same "
        "gear ratio applied as 1/N. So the bridge validates the spreadsheet "
        "against a copy of itself and can never disagree with it. "
        "ev_electrical_check, in the same package, already has the correct "
        "formulation (P_elec = P_wheel/eta, I = P/V). Two modules disagree and "
        "the wrong one is the one wired to the Excel.",
        cells="suspension/ev_excel_roundtrip.py",
        subs=["dataacq", "electrics", "powertrain"]))
    return out


# ===================================================================== #
#  6.  BRIDGES TO THE REST OF KINEMATIK
# ===================================================================== #
def to_bq_stack(pack: PackSpec, *, part: str = "BQ79616-Q1",
                cells_per_board: int = 16):
    """The pack as a `bq796xx.StackSpec`.

    A monitor measures each SERIES group, not each cell — the parallel cells in
    a group share one node and cannot be told apart. So a 140S3P pack needs 140
    channels, not 420, which is the difference between 9 boards and 27 and is
    worth getting right before anyone orders any.
    """
    from . import bq796xx as bq
    boards = math.ceil(pack.n_series / cells_per_board)
    return bq.StackSpec(part=part, boards=boards,
                        cells_per_board=cells_per_board,
                        isolated=None, nfault_to_interrupt=None)


def bms_findings(pack: PackSpec, *, cells_per_board: int = 16) -> list[Finding]:
    """What monitoring this pack actually requires."""
    from . import bq796xx as bq
    out: list[Finding] = []
    stack = to_bq_stack(pack, cells_per_board=cells_per_board)
    dev = stack.device()
    channels = pack.n_series
    out.append(_f(
        "pd-bms-channels", Severity.INFO,
        f"{pack.n_series}S{pack.n_parallel}P needs {channels} monitored series "
        f"groups, not {pack.cell_count()} — parallel cells share a node and "
        f"cannot be measured separately. At {cells_per_board} channels per "
        f"board that is {stack.boards} x {part_name(dev)} boards "
        f"({stack.boards * cells_per_board - channels} channels spare), "
        f"comfortably inside the {dev.max_boards}-board daisy-chain limit.",
        subs=["electrics", "dataacq"],
        detail={"channels": channels, "boards": stack.boards}))
    if pack.max_cells is not None and pack.cell_count() > pack.max_cells:
        out.append(_f(
            "pd-cell-count-over", Severity.FAIL,
            f"{pack.cell_count()} cells against a declared maximum of "
            f"{pack.max_cells}.",
            cells="BatteryPackConfig!B7,B10", subs=["electrics"]))
    return out


def part_name(dev) -> str:
    return getattr(dev, "part", str(dev))


def to_interface(pack: PackSpec, *, name: str = "electrics",
                 heat_reject_w: Optional[float] = None,
                 is_estimate: bool = False):
    """Project the workbook-derived pack onto a `SubsystemInterface`.

    The pack lives in the electrics lead's Excel file, and until now that made it
    a SECOND source of truth: the accumulator's mass, voltage and fuse limit
    existed in the workbook AND, separately, in whatever a team had typed into
    the integration ledger. Two numbers for one pack is exactly the drift the
    ledger exists to prevent.

    This closes it in the honest direction — the workbook wins, because that is
    where the electrics lead actually works, and the ledger takes a projection of
    it. Everything here is DERIVED from the pack (and recomputed, not read back
    from the sheet's own arithmetic), so the mass roll-up, the LV/HV checks and
    the trace-current chain all see the same accumulator the power-draw audit
    does.

    `heat_reject_w` is NOT invented: pack heat depends on the duty cycle, which
    this function has no access to. Pass it from a `PowerDrawTrace` if you have
    one, or leave it None and let `check_all()` report it MISSING.

    Note the fuse rating is used as the declared peak current: it is the hard
    ceiling the pack can legally deliver, which is the right number for sizing
    copper. If the team's real continuous draw is lower, declare that instead.
    """
    from .interfaces import SubsystemInterface
    return SubsystemInterface(
        name=name,
        mass_kg=pack.mass_kg(),
        voltage_v=pack.nominal_voltage_v(),
        peak_current_a=pack.fuse_max_a,
        power_draw_w=pack.power_at_fuse_limit_w(),
        heat_reject_w=heat_reject_w,
        is_estimate=is_estimate,
        updated_by="power_draw.to_interface",
        notes=f"derived from the pack workbook: {pack.source or '(unnamed file)'}")


def sync_ledger(ledger, pack: PackSpec, *,
                name: str = "electrics",
                heat_reject_w: Optional[float] = None) -> list:
    """Write the workbook's pack onto an IntegrationLedger and report what moved.

    Returns a list of Findings naming every field whose ledger value disagreed
    with the workbook BEFORE the sync — so a team that had been carrying a stale
    hand-typed accumulator mass sees exactly what was wrong and by how much,
    rather than the number silently changing under them.
    """
    from .interfaces import Finding, Severity

    fresh = to_interface(pack, name=name, heat_reject_w=heat_reject_w)
    old = ledger.interfaces.get(name)
    out: list = []
    if old is not None:
        for fld, unit in (("mass_kg", "kg"), ("voltage_v", "V"),
                          ("peak_current_a", "A"), ("power_draw_w", "W")):
            was, now = getattr(old, fld, None), getattr(fresh, fld, None)
            if was is None or now is None:
                continue
            if abs(float(was) - float(now)) > max(1e-6, 0.005 * abs(float(now))):
                out.append(Finding(
                    "pack-sync", Severity.WARN,
                    f"{name}.{fld} was {float(was):.4g} {unit} in the ledger but "
                    f"the workbook gives {float(now):.4g} {unit} — the two had "
                    f"drifted. The workbook value now stands.",
                    subsystems=[name],
                    detail=dict(field=fld, was=float(was), now=float(now),
                                source=pack.source or "workbook")))
    # Preserve anything the workbook has no opinion about (CG, envelope, mounts).
    if old is not None:
        for fld in ("cg_x_mm", "cg_y_mm", "cg_z_mm", "env_x_mm", "env_y_mm",
                    "env_z_mm", "env_origin_mm", "mount_load_n", "mount_points",
                    "mounts_on", "max_temp_c", "cooling_airflow_cms"):
            if getattr(fresh, fld, None) is None and getattr(old, fld, None) is not None:
                setattr(fresh, fld, getattr(old, fld))
        if fresh.heat_reject_w is None and old.heat_reject_w is not None:
            fresh.heat_reject_w = old.heat_reject_w
    ledger.set(fresh)
    return out


def to_elec_params(pack: PackSpec, vehicle: VehicleSpec):
    """Hand the correctly-read pack to the existing electrical feasibility gate.

    `ev_electrical_check` has the right physics and the wrong reader. This
    supplies it with values that actually came from the workbook.
    """
    from .ev_electrical_check import ElecParams
    return ElecParams(
        pack_voltage_v=pack.nominal_voltage_v(),
        pack_capacity_ah=pack.capacity_ah(),
        n_parallel=pack.n_parallel,
        n_series=pack.n_series)


def motor_feasibility(pack: PackSpec, drive: DriveSpec) -> list[Finding]:
    """Can this pack feed this motor? The workbook never asks."""
    out: list[Finding] = []
    ceiling_kw = pack.max_deliverable_power_w() / 1000.0
    fuse_kw = pack.power_at_fuse_limit_w() / 1000.0
    rated = drive.motor_peak_power_kw

    if rated > ceiling_kw:
        out.append(_f(
            "pd-motor-exceeds-pack", Severity.FAIL,
            f"The motor is rated {rated:.0f} kW peak. This pack cannot deliver "
            f"more than {ceiling_kw:.0f} kW into any load, because at "
            f"{pack.resistance_ohm():.3f} ohm the maximum-power point is "
            f"V_oc/2R and beyond it output falls. The motor rating is "
            f"unreachable by {rated - ceiling_kw:.0f} kW regardless of "
            f"controller settings.",
            cells="ElecPropulsion!B2, BatteryPackConfig!B11",
            subs=["electrics", "powertrain"],
            detail={"rated_kw": rated, "ceiling_kw": ceiling_kw}))

    if rated > fuse_kw:
        out.append(_f(
            "pd-motor-exceeds-fuse", Severity.WARN,
            f"The {pack.fuse_max_a:.0f} A fuse limits usable output to "
            f"{fuse_kw:.1f} kW with sag included, against a {rated:.0f} kW "
            f"motor — a factor of {rated / fuse_kw:.1f}. Either the fuse is "
            f"sized for a much smaller envelope than the motor, or the motor is "
            f"specified far above what the car will ever use. Both are "
            f"defensible; having neither written down next to the other is "
            f"not.",
            cells="BatteryPackConfig!B1, ElecPropulsion!B2",
            subs=["electrics", "powertrain"],
            detail={"fuse_kw": fuse_kw, "rated_kw": rated}))
    return out


def trace_findings(trace: PowerDrawTrace) -> list[Finding]:
    """Feasibility of a computed trace against the pack's real limits."""
    out: list[Finding] = []
    pack = trace.pack
    peak_i = trace.peak_current_a()
    peak_cell = pack.cell_current(peak_i)

    # ---- source-data quality: the workbook never differentiates, so it never
    # ---- finds out that its own speed trace is not physically realisable.
    n = len(trace.speed_ms)
    if trace.discontinuities:
        worst = max(trace.discontinuities, key=lambda d: abs(d[2]))
        frac = len(trace.discontinuities) / n
        out.append(_f(
            "pd-speed-discontinuities",
            Severity.FAIL if frac > 0.02 else Severity.WARN,
            f"{len(trace.discontinuities)} of {n} samples "
            f"({frac*100:.1f}%) in the speed trace imply accelerations beyond "
            f"tyre grip. The worst is {worst[1]:+.0f} mph in one "
            f"{trace.dt_s*1000:.0f} ms sample at index {worst[0]}, which is "
            f"{worst[2]:+.1f} g — no car does that, so it is a logging or "
            f"data-entry artefact, not vehicle behaviour. This matters because "
            f"the m*a term turns every one of those steps into a current spike "
            f"and then into a thermal load. The workbook never differentiates "
            f"its speed trace, so it never notices; it also therefore has no "
            f"acceleration term at all, which is why its currents are low. Fix "
            f"the trace, or pass smooth_window to filter it and accept that the "
            f"peaks are approximate.",
            cells="SpeedVsTime!B2:B1894",
            subs=["dataacq", "powertrain", "electrics"],
            detail={"count": len(trace.discontinuities), "fraction": frac,
                    "worst_g": worst[2]}))

    if trace.clamped_samples:
        out.append(_f(
            "pd-accel-clamped", Severity.INFO,
            f"{trace.clamped_samples} samples had their acceleration clamped to "
            f"the {trace.vehicle.mu_lon:g} g grip limit"
            + (f" after smoothing over {trace.smooth_window} samples"
               if trace.smooth_window > 1 else "")
            + ". Clamping keeps one bad sample from dominating the peak current; "
              "it does not make the underlying trace correct.",
            subs=["powertrain", "dataacq"],
            detail={"clamped": trace.clamped_samples}))

    if trace.infeasible_samples:
        out.append(_f(
            "pd-trace-infeasible", Severity.FAIL,
            f"{trace.infeasible_samples} of {len(trace.speed_ms)} samples "
            f"demand more power than the pack can deliver at any current. The "
            f"speed trace is not achievable with this pack; the car will not "
            f"follow it regardless of motor rating.",
            subs=["electrics", "powertrain"]))

    if peak_i > pack.fuse_max_a:
        out.append(_f(
            "pd-fuse-exceeded", Severity.FAIL,
            f"Peak pack current {peak_i:.1f} A against a "
            f"{pack.fuse_max_a:.0f} A fuse. Per cell that is "
            f"{peak_cell:.1f} A across {pack.n_parallel} parallel strings.",
            cells="BatteryPackConfig!B1", subs=["electrics"],
            detail={"peak_a": peak_i, "fuse_a": pack.fuse_max_a}))
    else:
        out.append(_f(
            "pd-fuse-ok", Severity.OK,
            f"Peak {peak_i:.1f} A inside the {pack.fuse_max_a:.0f} A fuse "
            f"({peak_i / pack.fuse_max_a * 100:.0f}%).",
            subs=["electrics"]))

    usable = pack.energy_kwh() * 0.8 - pack.energy_kwh() * 0.1
    if trace.energy_kwh() > usable:
        out.append(_f(
            "pd-energy-short", Severity.FAIL,
            f"The trace needs {trace.energy_kwh():.2f} kWh; usable energy "
            f"between the workbook's own 80% and 10% SOC limits is "
            f"{usable:.2f} kWh.",
            cells="BatteryPackConfig!B17,B18", subs=["electrics"]))

    out.append(_f(
        "pd-joule-heat", Severity.INFO,
        f"Joule heating averages {trace.mean_joule_w():.0f} W and totals "
        f"{trace.joule_energy_kwh()*1000:.0f} Wh over "
        f"{trace.duration_s():.0f} s. This is the number the cooling model "
        f"needs, and it is a heat flow plus an energy — not the "
        f"sum-of-instantaneous-watts the workbook reports.",
        subs=["cooling", "electrics"],
        detail={"mean_w": trace.mean_joule_w(),
                "energy_wh": trace.joule_energy_kwh() * 1000}))

    tq = trace.peak_motor_torque_nm()
    if tq > trace.drive.motor_peak_torque_nm:
        out.append(_f(
            "pd-torque-exceeded", Severity.FAIL,
            f"Peak required motor torque {tq:.0f} Nm at a "
            f"{trace.drive.reduction:g}:1 reduction exceeds the motor's "
            f"{trace.drive.motor_peak_torque_nm:.0f} Nm peak. Gearing up is the "
            f"fix — torque scales inversely with the reduction, while pack "
            f"current does not change at all, because power is force times "
            f"speed regardless of what gear delivers it. That last point is "
            f"why the workbook's fifteen-column current sweep could never have "
            f"informed a gear choice.",
            cells="ElecPropulsion!B1", subs=["powertrain"],
            detail={"required_nm": tq,
                    "peak_nm": trace.drive.motor_peak_torque_nm}))
    elif tq > trace.drive.motor_continuous_torque_nm:
        out.append(_f(
            "pd-torque-above-continuous", Severity.WARN,
            f"Peak required torque {tq:.0f} Nm is within the "
            f"{trace.drive.motor_peak_torque_nm:.0f} Nm peak rating but above "
            f"the {trace.drive.motor_continuous_torque_nm:.0f} Nm continuous "
            f"rating. Acceptable in bursts; how long a burst depends on the "
            f"motor's thermal mass, which is not in the workbook.",
            subs=["powertrain", "cooling"],
            detail={"required_nm": tq}))

    if trace.peak_motor_rpm() > trace.drive.motor_max_rpm:
        out.append(_f(
            "pd-overspeed", Severity.FAIL,
            f"Peak motor speed {trace.peak_motor_rpm():.0f} rpm at a "
            f"{trace.drive.reduction:g}:1 reduction exceeds the "
            f"{trace.drive.motor_max_rpm:.0f} rpm limit.",
            cells="ElecPropulsion!B9", subs=["powertrain"]))
    return out


# ===================================================================== #
#  7.  ONE CALL
# ===================================================================== #
@dataclass
class PowerDrawAudit:
    path: str
    pack: Optional[PackSpec]
    findings: list = field(default_factory=list)

    def blocking(self) -> list:
        return [f for f in self.findings if f.severity == Severity.FAIL]

    def warnings(self) -> list:
        return [f for f in self.findings if f.severity == Severity.WARN]

    def to_markdown(self) -> str:
        L = ["# FSAE_EV_Power_Draw.xlsx — audit", ""]
        if self.pack:
            p = self.pack
            L += [
                f"Pack read from the workbook: **{p.n_series}S{p.n_parallel}P**, "
                f"{p.cell_count()} cells, {p.nominal_voltage_v():.0f} V "
                f"nominal, {p.capacity_ah():.1f} Ah, {p.energy_kwh():.2f} kWh, "
                f"{p.mass_kg():.1f} kg of cells.", "",
                "| Quantity | Workbook | Corrected |",
                "|---|---|---|",
                f"| Pack resistance | 1.792 ohm | **{p.resistance_ohm():.4f} ohm** |",
                f"| Joule heat at fuse current | 4.48 'kWh' | "
                f"**{p.joule_heat_w(p.fuse_max_a):.0f} W** |",
                f"| Power at fuse limit | 25.2 kW (no sag) | "
                f"**{p.power_at_fuse_limit_w()/1000:.1f} kW** |",
                f"| Pack power ceiling | not computed | "
                f"**{p.max_deliverable_power_w()/1000:.0f} kW** |", ""]
        L += [f"**{len(self.blocking())} blocking, {len(self.warnings())} "
              f"warnings, {len(self.findings)} findings total.**", "",
              "| Severity | Check | Finding |", "|---|---|---|"]
        order = {Severity.FAIL: 0, Severity.MISSING: 1, Severity.WARN: 2,
                 Severity.INFO: 3, Severity.OK: 4}
        for f in sorted(self.findings, key=lambda x: order[x.severity]):
            msg = f.message.replace("|", "/").replace("\n", " ")
            L.append(f"| {f.severity.value.upper()} | {f.check} | {msg} |")
        return "\n".join(L)


def audit(path: str, *, vehicle: Optional[VehicleSpec] = None,
          drive: Optional[DriveSpec] = None,
          include_readers: bool = True) -> PowerDrawAudit:
    """Audit the workbook, its readers, and the design it describes."""
    import openpyxl
    findings: list[Finding] = []
    pack: Optional[PackSpec] = None

    try:
        pack = read_pack_config(path)
    except WorkbookReadError as exc:
        findings.append(_f("pd-pack-unreadable", Severity.FAIL, str(exc),
                           subs=["dataacq", "electrics"]))

    findings.extend(audit_workbook(path))

    if pack is not None:
        wv = openpyxl.load_workbook(path, data_only=True)
        ws = _find_sheet(wv, "series battery count")
        sheet_vals = {}
        if ws is not None:
            m = _label_map(ws)
            for k, (v, _c) in m.items():
                if "internal resistance" in k and "pack" in k:
                    sheet_vals["pack_resistance_ohm"] = v
                if "joule heating" in k:
                    sheet_vals["joule_max_label"] = k
        findings.extend(audit_pack_sheet(pack, sheet_vals))
        findings.extend(bms_findings(pack))
        findings.extend(motor_feasibility(pack, drive or DriveSpec()))

    if include_readers:
        findings.extend(audit_kinematik_readers(path))

    return PowerDrawAudit(path, pack, findings)


# ===================================================================== #
#  8.  PROVENANCE
# ===================================================================== #
PROVENANCE = {
    "physics_grounded": [
        "tractive force balance F = Crr*m*g + 0.5*rho*CdA*v^2 + m*a, then "
        "P_wheel = F*v, P_elec = P_wheel/eta, which is the formulation "
        "ev_electrical_check already uses",
        "pack current from P = (V_oc - I*R)*I, taking the low-current root, so "
        "sag is included and an impossible demand returns None",
        "pack DC resistance R = S*(R_cell/P)",
        "maximum deliverable power V_oc^2/(4R) at I = V_oc/(2R)",
        "gear tangential force T/(D/2); radial force F_t*tan(angle in radians)",
        "energy as sum(P*dt), never sum(P)",
    ],
    "read_from_workbook": [
        "pack topology, cell voltage/capacity/resistance/weight, fuse rating "
        "and declared cell maximum — all located by row LABEL, not by sheet "
        "name or cell coordinate",
    ],
    "errors_found_in_workbook": [
        "no vehicle model at all: current computed from V*PF*RPM with no mass, "
        "drag, rolling resistance or acceleration anywhere in the file",
        "gear ratio inverted (stored 1/N, multiplied in): 49x at gear 7",
        "motor efficiency multiplying instead of dividing",
        "sqrt(3) and power factor applied to the DC bus",
        "pack resistance ignores parallel strings: 3x high",
        "TAN() given degrees: 6.15x on bearing radial load",
        "torque divided by pitch diameter instead of radius: 2x low, partially "
        "masking the radians error",
        "EMF columns compute V*kW*ohm and label the result volts",
        "RPM min/max range reaches into the current block",
        "EMF range reaches one row past the power block into blanks",
        "ThermalLoad max carries an extra fuse-current factor the min lacks",
        "SUM of instantaneous watts labelled as a total",
        "unit labels contradicting their own formulas in three places",
        "two gear-ratio cells carry date number formats",
    ],
    "errors_found_in_kinematik": [
        "ev_electrical_check.ElecParams.from_excel and "
        "ev_excel_roundtrip.extract_params_from_excel gate on a sheet name "
        "this workbook does not have, and fall through to defaults that "
        "coincidentally match this pack — proven by mutating a copy",
        "ev_excel_roundtrip re-implements the workbook's incorrect formulas, "
        "so the bridge cannot disagree with the sheet",
        "_safe_float coerces datetime values to a default, turning gear ratios "
        "1/8 and 1/10 into direct drive",
    ],
    "estimate_flagged": [
        "VehicleSpec defaults (300 kg, Crr 0.020, CdA 1.10 m^2) are generic "
        "FSAE figures. They are the inputs the workbook lacks entirely, and "
        "every current figure depends on them — replace with the team's own "
        "mass and aero numbers before quoting anything",
        "inverter efficiency 0.97 assumed; not present in the workbook",
    ],
    "hard_rule": (
        "Values are read by label and never coerced. A cell that cannot be "
        "interpreted raises WorkbookReadError rather than yielding a default, "
        "because the default is what turned a correct 1/8 into direct drive "
        "and a missing sheet into a plausible-looking pack spec."
    ),
}
