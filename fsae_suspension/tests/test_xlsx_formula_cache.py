# ============================================================================
#  KinematiK — Formula SAE suspension & vehicle dynamics toolkit
#  Created by Frederik Thio. Copyright (c) 2026 Frederik Thio.
#  Open source. Original author: Frederik Thio, creator of KinematiK.
# ============================================================================
"""
Tests for suspension.xlsx_formula_cache.

The module's whole claim is "a cell is either right or left empty", so the
tests come in three groups:

* it computes the right answer — checked against hand arithmetic, and, where
  LibreOffice is installed, against LibreOffice cell for cell;
* it refuses what it does not model — unsupported functions, array formulas
  and circular references leave no cached value rather than a plausible one;
* it does not damage the file — formulas survive, the user's sheets are
  untouched, and the result still opens.
"""

import math
import shutil
import subprocess
import tempfile
import os

import openpyxl
import pytest

from suspension import xlsx_formula_cache as fc

_HAS_SOFFICE = bool(shutil.which("soffice") or shutil.which("libreoffice"))
needs_soffice = pytest.mark.skipif(not _HAS_SOFFICE,
                                   reason="LibreOffice not available")


# --------------------------------------------------------------------------- #
#  Helpers
# --------------------------------------------------------------------------- #
def _book(tmp_path, cells, sheet="KX Sheet"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet
    for coord, value in cells.items():
        ws[coord] = value
    path = str(tmp_path / "wb.xlsx")
    wb.save(path)
    return path


def _values(tmp_path, formulas, data=None):
    """Evaluate {coord: formula} with optional {coord: literal} alongside."""
    cells = dict(data or {})
    cells.update(formulas)
    path = _book(tmp_path, cells)
    fc.populate_cached_values(path)
    ws = openpyxl.load_workbook(path, data_only=True)["KX Sheet"]
    return {coord: ws[coord].value for coord in formulas}


def _libreoffice_values(path, sheet):
    outdir = tempfile.mkdtemp(prefix="kx_test_")
    try:
        subprocess.run(
            [shutil.which("soffice") or shutil.which("libreoffice"),
             "--headless", "--norestore", "--convert-to", "xlsx",
             "--outdir", outdir, path],
            check=True, capture_output=True, timeout=180)
        produced = os.path.join(
            outdir, os.path.splitext(os.path.basename(path))[0] + ".xlsx")
        wb = openpyxl.load_workbook(produced, data_only=True)
        ws = wb[sheet]
        return {c.coordinate: c.value for row in ws.iter_rows() for c in row}
    finally:
        shutil.rmtree(outdir, ignore_errors=True)


# --------------------------------------------------------------------------- #
#  1. Arithmetic and operators
# --------------------------------------------------------------------------- #
def test_arithmetic_precedence(tmp_path):
    got = _values(tmp_path, {
        "B1": "=2+3*4",
        "B2": "=(2+3)*4",
        "B3": "=2^3^2",            # Excel is left-associative: (2^3)^2
        "B4": "=-3^2",             # unary minus binds tighter than ^ in Excel
        "B5": "=10/4",
        "B6": "=2*-3",
    })
    assert got["B1"] == 14
    assert got["B2"] == 20
    assert got["B3"] == 64
    assert got["B4"] == 9
    assert got["B5"] == 2.5
    assert got["B6"] == -6


def test_comparison_and_concatenation(tmp_path):
    got = _values(tmp_path, {
        "B1": '="a"&"b"',
        "B2": "=1<2",
        "B3": "=1=2",
        "B4": '="ABC"="abc"',       # Excel compares text case-insensitively
        "B5": '=2&" of "&3',
        "B6": "=3<>3",
    })
    assert got["B1"] == "ab"
    assert got["B2"] is True
    assert got["B3"] is False
    assert got["B4"] is True
    assert got["B5"] == "2 of 3"
    assert got["B6"] is False


def test_percent_and_unary_chains(tmp_path):
    got = _values(tmp_path, {"B1": "=50%", "B2": "=--5", "B3": "=200%*3"})
    assert got["B1"] == 0.5
    assert got["B2"] == 5
    assert got["B3"] == 6


# --------------------------------------------------------------------------- #
#  2. References
# --------------------------------------------------------------------------- #
def test_absolute_and_relative_references_are_the_same_cell(tmp_path):
    got = _values(tmp_path, {"C1": "=A1+$A$1+$A1+A$1"}, data={"A1": 2.5})
    assert got["C1"] == 10.0


def test_cross_sheet_references(tmp_path):
    wb = openpyxl.Workbook()
    a = wb.active
    a.title = "KX Inputs"
    a["B5"] = 7.0
    b = wb.create_sheet("KX Dashboard")
    b["B1"] = "='KX Inputs'!$B$5*2"
    path = str(tmp_path / "x.xlsx")
    wb.save(path)
    fc.populate_cached_values(path, only_prefix="KX ")
    got = openpyxl.load_workbook(path, data_only=True)["KX Dashboard"]["B1"]
    assert got.value == 14.0


def test_a_range_of_one_cell_is_a_scalar(tmp_path):
    got = _values(tmp_path, {"C1": "=SUM(A1:A1)+A1"}, data={"A1": 3.0})
    assert got["C1"] == 6.0


def test_chained_dependencies_resolve_in_any_order(tmp_path):
    """C depends on B depends on A, written in the wrong order on purpose."""
    got = _values(tmp_path, {
        "A3": "=A2*2", "A2": "=A1*2",
    }, data={"A1": 1.5})
    assert got["A2"] == 3.0
    assert got["A3"] == 6.0


# --------------------------------------------------------------------------- #
#  3. Functions
# --------------------------------------------------------------------------- #
def test_aggregates_ignore_text_and_blanks(tmp_path):
    """A header or a blank tail counted as zero is how a MIN becomes 0."""
    data = {"A1": "Speed", "A2": 5.0, "A3": None, "A4": 3.0}
    got = _values(tmp_path, {
        "C1": "=MIN(A1:A4)", "C2": "=MAX(A1:A4)",
        "C3": "=AVERAGE(A1:A4)", "C4": "=SUM(A1:A4)",
        "C5": "=COUNT(A1:A4)",
    }, data=data)
    assert got["C1"] == 3.0
    assert got["C2"] == 5.0
    assert got["C3"] == 4.0
    assert got["C4"] == 8.0
    assert got["C5"] == 2


def test_sumproduct_pairs_two_ranges(tmp_path):
    data = {f"A{i}": float(i) for i in range(1, 5)}
    data.update({f"B{i}": 0.5 for i in range(1, 5)})
    got = _values(tmp_path, {"D1": "=SUMPRODUCT(A1:A4,B1:B4)"}, data=data)
    assert got["D1"] == pytest.approx(5.0)


def test_sumproduct_rejects_mismatched_ranges(tmp_path):
    data = {f"A{i}": 1.0 for i in range(1, 5)}
    data.update({f"B{i}": 1.0 for i in range(1, 4)})
    got = _values(tmp_path, {"D1": "=SUMPRODUCT(A1:A4,B1:B3)"}, data=data)
    assert got["D1"] == "#VALUE!"


def test_index_match_including_a_wildcard(tmp_path):
    data = {"A1": 4.0, "A2": 7.0, "A3": 9.0,
            "B1": "no", "B2": "YES - clears everything", "B3": "no"}
    got = _values(tmp_path, {
        "D1": '=INDEX(A1:A3,MATCH("YES*",B1:B3,0))',
        "D2": '=IFERROR(INDEX(A1:A3,MATCH("NEVER*",B1:B3,0)),"none")',
        "D3": "=INDEX(A1:A3,2)",
    }, data=data)
    assert got["D1"] == 7.0
    assert got["D2"] == "none"
    assert got["D3"] == 7.0


def test_text_formats_the_way_the_verdicts_need(tmp_path):
    got = _values(tmp_path, {
        "B1": '=TEXT(12.345,"0.0")',
        "B2": '=TEXT(12.345,"0")',
        "B3": '=TEXT(0.5,"0.00")',
        "B4": '=TEXT(1234.5,"#,##0")',
        "B5": '=TEXT(-2.5,"0.0")',
    })
    assert got["B1"] == "12.3"
    assert got["B2"] == "12"
    assert got["B3"] == "0.50"
    assert got["B4"] == "1,234"
    assert got["B5"] == "-2.5"


def test_round_goes_half_away_from_zero_like_excel(tmp_path):
    """Python rounds half to even; Excel does not. 2.5 must be 3."""
    got = _values(tmp_path, {
        "B1": "=ROUND(2.5,0)", "B2": "=ROUND(-2.5,0)",
        "B3": "=ROUND(1.2345,2)", "B4": "=ROUNDUP(1.01,0)",
        "B5": "=ROUNDDOWN(1.99,0)",
    })
    assert got["B1"] == 3
    assert got["B2"] == -3
    assert got["B3"] == pytest.approx(1.23)
    assert got["B4"] == 2
    assert got["B5"] == 1


def test_if_does_not_evaluate_the_branch_it_does_not_take(tmp_path):
    got = _values(tmp_path, {"B1": '=IF(1>2,1/0,"safe")'})
    assert got["B1"] == "safe"


def test_iferror_catches_a_division_by_zero(tmp_path):
    got = _values(tmp_path, {"B1": '=IFERROR(1/0,"caught")',
                             "B2": "=IFERROR((3-3)/(3-3),0)"})
    assert got["B1"] == "caught"
    assert got["B2"] == 0


def test_nested_verdict_formula(tmp_path):
    """The shape the dashboard actually uses."""
    data = {"A1": 61.0, "A2": 50.0}
    got = _values(tmp_path, {
        "B1": '=IF(A1<=A2,"PASS - peak "&TEXT(A1,"0.0")&" A of "'
              '&TEXT(A2,"0")&" A","FAIL - peak "&TEXT(A1,"0.0")'
              '&" A exceeds "&TEXT(A2,"0")&" A fuse")',
    }, data=data)
    assert got["B1"] == "FAIL - peak 61.0 A exceeds 50 A fuse"


def test_sqrt_and_pi(tmp_path):
    got = _values(tmp_path, {"B1": "=SQRT(2)", "B2": "=PI()", "B3": "=SQRT(-1)"})
    assert got["B1"] == pytest.approx(math.sqrt(2))
    assert got["B2"] == pytest.approx(math.pi)
    assert got["B3"] == "#NUM!"


def test_and_or_not(tmp_path):
    got = _values(tmp_path, {
        "B1": '=IF(AND(1=1,2>1),"YES","-")',
        "B2": "=OR(FALSE,FALSE)",
        "B3": "=NOT(TRUE)",
    })
    assert got["B1"] == "YES"
    assert got["B2"] is False
    assert got["B3"] is False


# --------------------------------------------------------------------------- #
#  4. Errors and refusals — the half that keeps it honest
# --------------------------------------------------------------------------- #
def test_division_by_zero_is_cached_as_an_error_not_a_number(tmp_path):
    got = _values(tmp_path, {"B1": "=1/0"})
    assert got["B1"] == "#DIV/0!"


def test_an_unsupported_function_leaves_the_cell_uncached(tmp_path):
    path = _book(tmp_path, {"A1": 1.0, "B1": "=XLOOKUP(A1,A1:A1,A1:A1)"})
    report = fc.populate_cached_values(path)
    assert report["written"] == 0
    assert report["skipped"] == 1
    assert "XLOOKUP" in report["unsupported_functions"]
    ws = openpyxl.load_workbook(path, data_only=True)["KX Sheet"]
    assert ws["B1"].value is None, "guessed a value for a function it lacks"


def test_one_bad_formula_does_not_stop_the_others(tmp_path):
    path = _book(tmp_path, {"A1": 2.0, "B1": "=XLOOKUP(1,A1:A1,A1:A1)",
                            "B2": "=A1*3"})
    report = fc.populate_cached_values(path)
    assert report["written"] == 1
    ws = openpyxl.load_workbook(path, data_only=True)["KX Sheet"]
    assert ws["B1"].value is None
    assert ws["B2"].value == 6.0


def test_a_circular_reference_becomes_an_error_not_a_hang(tmp_path):
    got = _values(tmp_path, {"B1": "=B2+1", "B2": "=B1+1"})
    assert got["B1"] == "#REF!"
    assert got["B2"] == "#REF!"


def test_a_reference_to_a_missing_sheet_is_an_error(tmp_path):
    got = _values(tmp_path, {"B1": "='Nowhere'!A1+1"})
    assert got["B1"] == "#REF!"


def test_text_arithmetic_is_a_value_error(tmp_path):
    got = _values(tmp_path, {"B1": "=A1*2"}, data={"A1": "not a number"})
    assert got["B1"] == "#VALUE!"


def test_an_unsupported_text_format_is_skipped_rather_than_guessed(tmp_path):
    path = _book(tmp_path, {"A1": 0.25, "B1": '=TEXT(A1,"yyyy-mm-dd")'})
    report = fc.populate_cached_values(path)
    assert report["skipped"] == 1
    ws = openpyxl.load_workbook(path, data_only=True)["KX Sheet"]
    assert ws["B1"].value is None


def test_only_prefix_limits_what_is_written(tmp_path):
    wb = openpyxl.Workbook()
    theirs = wb.active
    theirs.title = "TheirSheet"
    theirs["A1"] = 2.0
    theirs["B1"] = "=A1*2"
    ours = wb.create_sheet("KX Dashboard")
    ours["B1"] = "=TheirSheet!A1*3"
    path = str(tmp_path / "x.xlsx")
    wb.save(path)

    fc.populate_cached_values(path, only_prefix="KX ")
    wb2 = openpyxl.load_workbook(path, data_only=True)
    assert wb2["KX Dashboard"]["B1"].value == 6.0
    assert wb2["TheirSheet"]["B1"].value is None


# --------------------------------------------------------------------------- #
#  5. The file itself
# --------------------------------------------------------------------------- #
def test_formulas_are_preserved(tmp_path):
    path = _book(tmp_path, {"A1": 2.0, "B1": "=A1*4"})
    fc.populate_cached_values(path)
    ws = openpyxl.load_workbook(path)["KX Sheet"]
    assert ws["B1"].value == "=A1*4", "the model was flattened to a snapshot"


def test_literals_and_styling_survive(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "KX Sheet"
    ws["A1"] = "Label"
    ws["A2"] = 2.0
    ws["A2"].number_format = "0.0000"
    ws["B2"] = "=A2*2"
    ws.column_dimensions["A"].width = 34
    path = str(tmp_path / "x.xlsx")
    wb.save(path)

    fc.populate_cached_values(path)
    ws2 = openpyxl.load_workbook(path)["KX Sheet"]
    assert ws2["A1"].value == "Label"
    assert ws2["A2"].number_format == "0.0000"
    assert ws2.column_dimensions["A"].width == 34


def test_running_it_twice_is_idempotent(tmp_path):
    path = _book(tmp_path, {"A1": 3.0, "B1": "=A1*3"})
    fc.populate_cached_values(path)
    first = openpyxl.load_workbook(path, data_only=True)["KX Sheet"]["B1"].value
    report = fc.populate_cached_values(path)
    second = openpyxl.load_workbook(path, data_only=True)["KX Sheet"]["B1"].value
    assert first == second == 9.0
    assert report["written"] == 1


def test_the_report_counts_what_it_did(tmp_path):
    path = _book(tmp_path, {"A1": 1.0, "B1": "=A1+1", "B2": "=1/0",
                            "B3": "=XLOOKUP(1,A1:A1,A1:A1)"})
    report = fc.populate_cached_values(path)
    assert report["written"] == 2
    assert report["skipped"] == 1
    assert report["errors"]["#DIV/0!"] == ["KX Sheet!B2"]


# --------------------------------------------------------------------------- #
#  6. Agreement with LibreOffice
# --------------------------------------------------------------------------- #
_CROSS_CHECK = {
    "B1": "=2+3*4",
    "B2": "=(2+3)*4",
    "B3": "=2^3^2",
    "B4": "=-3^2",
    "B5": "=A1*A2/A3",
    "B6": "=SQRT(A1^2+A2^2)",
    "B7": "=MAX(A1:A5)",
    "B8": "=MIN(A1:A5)",
    "B9": "=AVERAGE(A1:A5)",
    "B10": "=SUM(A1:A5)",
    "B11": "=SUMPRODUCT(A1:A5,A1:A5)",
    "B12": '=IF(A1>A2,"bigger","smaller")',
    "B13": '=TEXT(A1,"0.00")',
    "B14": '=TEXT(A3,"0")',
    "B15": '=IFERROR(1/(A1-A1),"inf")',
    "B16": "=ROUND(A2,1)",
    "B17": '=INDEX(A1:A5,MATCH(MAX(A1:A5),A1:A5,0))',
    "B18": "=PI()*A1^2",
    "B19": '=IF(AND(A1>0,A2>0),"both","not")',
    "B20": '="peak "&TEXT(A2,"0.0")&" of "&TEXT(A4,"0")',
    "B21": "=A1&\"\"",
    "B22": "=COUNT(A1:A5)",
    "B23": "=ABS(A1-A4)",
    "B24": "=(A2-A1)/(A4-A3)",
}


@needs_soffice
def test_agrees_with_libreoffice_cell_for_cell(tmp_path):
    data = {"A1": 3.0, "A2": 12.345, "A3": 7.0, "A4": 50.0, "A5": -2.0}
    cells = dict(data)
    cells.update(_CROSS_CHECK)
    path = _book(tmp_path, cells)

    reference = _libreoffice_values(path, "KX Sheet")
    fc.populate_cached_values(path)
    ours = openpyxl.load_workbook(path, data_only=True)["KX Sheet"]

    for coord in _CROSS_CHECK:
        want, got = reference.get(coord), ours[coord].value
        if isinstance(want, float):
            assert got == pytest.approx(want, rel=1e-12, abs=1e-12), coord
        else:
            assert got == want, f"{coord}: {got!r} vs {want!r}"
