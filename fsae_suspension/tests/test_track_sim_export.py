# ============================================================================
#  KinematiK — Formula SAE suspension & vehicle dynamics toolkit
#  Created by Frederik Thio. Copyright (c) 2026 Frederik Thio.
#  Open source. Original author: Frederik Thio, creator of KinematiK.
# ============================================================================
"""
Tests for suspension.track_sim_export.

The four defects this module exists to fix each get a test that would have
caught it:

* constants instead of formulas   -> test_derived_cells_are_formulas
* the workbook's physics errors   -> test_peak_power_is_within_the_pack_ceiling,
                                     test_gear_ratio_multiplies
* formulas with no cached values  -> test_recalculated_file_is_readable_by_python
* energy-only pack advice         -> test_advisor_rejects_the_pack_it_used_to_recommend

Recalculation shells out to LibreOffice and is slow, so the tests that need it
use short traces and are marked. Everything else runs on the written formulas
directly.
"""

import io
import math
import shutil

import openpyxl
import pytest

from suspension import power_draw as pdw
from suspension import track_sim_export as tse

_HAS_SOFFICE = bool(shutil.which("soffice") or shutil.which("libreoffice"))
needs_soffice = pytest.mark.skipif(not _HAS_SOFFICE,
                                   reason="LibreOffice not available")


# --------------------------------------------------------------------------- #
#  Fixtures
# --------------------------------------------------------------------------- #
_PACK_ROWS = [
    ("Fuse Max (A)", 50), ("Parrallel Battery Count", 3),
    ("Series Battery Count", 140), ("Nominal Battery Voltage (V)", 3.6),
    ("Capacity Battery Cell (Ah)", 5), ("Endurance Length (km)", 22),
    ("Max Battery Cells", 560),
    ("Internal Resistance Battery Cell (Ohms)", 0.0128),
    ("Battery Cell Weight (kg)", 0.071),
    ("Battery Pack Cell Count", 420),
    ("Battery Pack Internal Resistance (Ohms)", 1.792),
    ("Battery Pack Nominal Voltage (V)", 504),
]


@pytest.fixture
def source(tmp_path):
    """A workbook shaped like the electrics lead's, with their own formulas."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "BatteryPackConfig"
    for i, (a, b) in enumerate(_PACK_ROWS, start=1):
        ws.cell(i, 1, a)
        ws.cell(i, 2, b)
    ws.cell(13, 1, "Power Draw (kW)")
    ws.cell(13, 2, "=(B1*B12)/1000")          # a formula of THEIRS
    ep = wb.create_sheet("ElecPropulsion")
    ep["A1"], ep["B1"] = "Motor Peak Torque (Nm)", 120
    svt = wb.create_sheet("SpeedVsTime")
    svt["A1"], svt["B1"] = "time (s)", "Speed (mph)"
    p = tmp_path / "src.xlsx"
    wb.save(p)
    return str(p)


def _trace(n=40, base=35.0):
    return [base + 8 * math.sin(i / 6) ** 2 for i in range(n)]


def _export(source, tmp_path, *, recalc=False, **kw):
    out = str(tmp_path / "out.xlsx")
    return tse.export_track_sim(source, out, _trace(), 0.0666667,
                                recalc=recalc, **kw)


# --------------------------------------------------------------------------- #
#  1. Nothing of the user's is touched
# --------------------------------------------------------------------------- #
def test_user_sheets_are_untouched(source, tmp_path):
    before = openpyxl.load_workbook(source)
    res = _export(source, tmp_path)
    after = openpyxl.load_workbook(res.path)
    for name in before.sheetnames:
        b, a = before[name], after[name]
        assert b.max_row == a.max_row, f"{name} changed row count"
        for r in range(1, b.max_row + 1):
            for c in range(1, b.max_column + 1):
                assert b.cell(r, c).value == a.cell(r, c).value, \
                    f"{name}!{b.cell(r, c).coordinate} was modified"


def test_the_users_own_formula_survives(source, tmp_path):
    res = _export(source, tmp_path)
    ws = openpyxl.load_workbook(res.path)["BatteryPackConfig"]
    assert ws["B13"].value == "=(B1*B12)/1000"


def test_every_written_sheet_carries_the_prefix(source, tmp_path):
    before = set(openpyxl.load_workbook(source).sheetnames)
    res = _export(source, tmp_path)
    after = set(openpyxl.load_workbook(res.path).sheetnames)
    for name in after - before:
        assert name.startswith(tse.SHEET_PREFIX)


def test_all_expected_sheets_are_created(source, tmp_path):
    res = _export(source, tmp_path)
    names = openpyxl.load_workbook(res.path).sheetnames
    for s in (tse.S_DASH, tse.S_INPUTS, tse.S_TRACE, tse.S_PACK,
              tse.S_GEARS, tse.S_ADVISOR, tse.S_PROV):
        assert s in names


def test_dashboard_is_the_first_sheet(source, tmp_path):
    """It is what a reviewer opens; the original left an empty tab there."""
    res = _export(source, tmp_path)
    assert openpyxl.load_workbook(res.path).sheetnames[0] == tse.S_DASH


def test_rerunning_is_idempotent(source, tmp_path):
    out = str(tmp_path / "out.xlsx")
    tse.export_track_sim(source, out, _trace(), 0.0666667, recalc=False)
    first = openpyxl.load_workbook(out).sheetnames
    tse.export_track_sim(out, out, _trace(), 0.0666667, recalc=False)
    assert openpyxl.load_workbook(out).sheetnames == first


# --------------------------------------------------------------------------- #
#  2. It is a model, not a snapshot
# --------------------------------------------------------------------------- #
def test_derived_cells_are_formulas(source, tmp_path):
    """The previous export wrote constants, which cannot be re-checked."""
    res = _export(source, tmp_path)
    wb = openpyxl.load_workbook(res.path)
    tr = wb[tse.S_TRACE]
    for col in "CDEFGHIJKMNO":
        v = tr[f"{col}3"].value
        assert isinstance(v, str) and v.startswith("="), \
            f"{tse.S_TRACE}!{col}3 is not a formula"
    for cell in ("B5", "B11", "B13", "B14"):
        v = wb[tse.S_DASH][cell].value
        assert isinstance(v, str) and v.startswith("=")


def test_inputs_are_numbers_not_formulas(source, tmp_path):
    """Inputs are typed values; only derived cells are formulas."""
    res = _export(source, tmp_path)
    ws = openpyxl.load_workbook(res.path)[tse.S_INPUTS]
    mass = next(r for r in range(1, 60)
                if str(ws.cell(r, 1).value or "").startswith("Mass"))
    assert isinstance(ws.cell(mass, 2).value, (int, float))


def test_trace_speed_column_is_data(source, tmp_path):
    res = _export(source, tmp_path)
    ws = openpyxl.load_workbook(res.path)[tse.S_TRACE]
    assert isinstance(ws["B3"].value, (int, float))


def test_unverified_assumptions_are_highlighted(source, tmp_path):
    """Mass, CdA and Crr are absent from the original workbook entirely."""
    res = _export(source, tmp_path)
    ws = openpyxl.load_workbook(res.path)[tse.S_INPUTS]
    filled = []
    for r in range(1, 60):
        c = ws.cell(r, 2)
        if c.fill and c.fill.fgColor and c.fill.fgColor.rgb == tse._YELLOW:
            filled.append(str(ws.cell(r, 1).value))
    assert any("Mass" in f for f in filled)
    assert any("Drag area" in f for f in filled)
    assert any("Rolling resistance" in f for f in filled)


# --------------------------------------------------------------------------- #
#  3. The physics is the corrected physics
# --------------------------------------------------------------------------- #
def test_gear_ratio_multiplies(source, tmp_path):
    """motor rpm = wheel rpm * reduction. The original divided."""
    res = _export(source, tmp_path)
    ws = openpyxl.load_workbook(res.path)[tse.S_TRACE]
    f = ws["N3"].value
    assert "reduction" not in f            # it is a cell ref, not a name
    assert "*" in f.split("PI()")[-1] or f.rstrip().endswith(")")
    # the python-side trace is the authority and is already unit-tested
    tr = res.trace
    assert tr.motor_rpm[0] > 0
    faster = pdw.DriveSpec(reduction=14.0).motor_rpm(35.0, 18.0)
    slower = pdw.DriveSpec(reduction=7.0).motor_rpm(35.0, 18.0)
    assert faster == pytest.approx(2 * slower)


def test_peak_power_is_within_the_pack_ceiling(source, tmp_path):
    """The old export reported 294 kW from a pack capped at 106 kW."""
    res = _export(source, tmp_path)
    pack = pdw.PackSpec()
    assert res.trace.peak_power_kw() <= pack.max_deliverable_power_w() / 1000


def test_no_sample_exceeds_the_pack_ceiling_on_a_sane_trace(source, tmp_path):
    res = _export(source, tmp_path)
    assert res.trace.infeasible_samples == 0


def test_pack_resistance_formula_uses_series_not_cell_count(source, tmp_path):
    res = _export(source, tmp_path)
    ws = openpyxl.load_workbook(res.path)[tse.S_PACK]
    row = next(r for r in range(1, 20)
               if str(ws.cell(r, 1).value or "") == "Pack resistance")
    f = ws.cell(row, 2).value
    # n_series * (cell_r / n_parallel) — three refs, no cell-count term
    assert f.count("'" + tse.S_INPUTS + "'!") == 3


def test_energy_is_an_integral(source, tmp_path):
    """SUMPRODUCT(P) * dt, never SUM(P)."""
    res = _export(source, tmp_path)
    ws = openpyxl.load_workbook(res.path)[tse.S_DASH]
    f = ws["B14"].value
    assert "SUMPRODUCT" in f
    assert "3600000" in f


def test_findings_are_attached(source, tmp_path):
    res = _export(source, tmp_path)
    assert res.findings
    assert any(f.check.startswith("pd-") for f in res.findings)


# --------------------------------------------------------------------------- #
#  4. Gear study reports the quantity that selects a ratio
# --------------------------------------------------------------------------- #
def test_gear_study_covers_every_requested_reduction(source, tmp_path):
    res = _export(source, tmp_path, reductions=range(1, 16))
    ws = openpyxl.load_workbook(res.path)[tse.S_GEARS]
    got = [ws.cell(r, 1).value for r in range(5, 20)]
    assert got == [float(i) for i in range(1, 16)]


def test_gear_study_has_a_torque_column_and_a_recommendation(source, tmp_path):
    res = _export(source, tmp_path)
    ws = openpyxl.load_workbook(res.path)[tse.S_GEARS]
    hdr = [ws.cell(4, c).value for c in range(1, 7)]
    assert any("torque" in str(h).lower() for h in hdr)
    assert any("Recommended" in str(h) for h in hdr)
    row = next(r for r in range(18, 26)
               if "Lowest workable" in str(ws.cell(r, 1).value or ""))
    # the label is merged across A:D, so the value sits in E
    assert str(ws.cell(row, 5).value).startswith("=")


def test_torque_scales_inversely_with_reduction():
    veh = pdw.VehicleSpec()
    a = pdw.DriveSpec(reduction=4.0).motor_torque_nm(1000.0,
                                                     veh.wheel_radius_m())
    b = pdw.DriveSpec(reduction=8.0).motor_torque_nm(1000.0,
                                                     veh.wheel_radius_m())
    assert b == pytest.approx(a / 2)


# --------------------------------------------------------------------------- #
#  5. The advisor
# --------------------------------------------------------------------------- #
def test_advisor_rejects_the_pack_it_used_to_recommend(source, tmp_path):
    """140S1P: triples resistance, ~26C per cell. The old advisor advised it."""
    res = _export(source, tmp_path)
    ws = openpyxl.load_workbook(res.path)[tse.S_ADVISOR]
    row = next(r for r in range(5, 15) if ws.cell(r, 2).value == 1)
    verdict = ws.cell(row, 11).value
    assert "c_rate" in verdict or "C rating" in verdict or "rating" in verdict
    assert verdict.startswith("=IF(")


def test_advisor_gates_on_all_three_limits(source, tmp_path):
    res = _export(source, tmp_path)
    ws = openpyxl.load_workbook(res.path)[tse.S_ADVISOR]
    v = ws["K5"].value
    assert "rating" in v          # cell C-rate gate
    assert "ceiling" in v         # power gate
    assert "endurance" in v       # energy gate


def test_advisor_energy_gate_uses_endurance_distance(source, tmp_path):
    """The old sheet compared ONE lap's energy and reported a 643% surplus
    where the endurance requirement was an 8x shortfall."""
    res = _export(source, tmp_path)
    ws = openpyxl.load_workbook(res.path)[tse.S_ADVISOR]
    f = ws["I5"].value
    assert "endurance" in f.lower() or "$B$" in f
    hdr = [ws.cell(4, c).value for c in range(1, 12)]
    assert any("Energy needed" in str(h) for h in hdr)


def test_advisor_lists_the_current_pack_among_candidates(source, tmp_path):
    res = _export(source, tmp_path)
    ws = openpyxl.load_workbook(res.path)[tse.S_ADVISOR]
    pars = [ws.cell(r, 2).value for r in range(5, 15)]
    assert 3 in pars


# --------------------------------------------------------------------------- #
#  6. Cached values — the defect that made the old file unreadable
# --------------------------------------------------------------------------- #
def test_cached_values_do_not_need_libreoffice(source, tmp_path, monkeypatch):
    """The deployment target has no LibreOffice. The file must still read.

    This is the regression that mattered most: on Streamlit Cloud the export
    fell back to "no cached values", which is precisely the defect the module
    exists to fix, on the only machine anyone actually ran it on.
    """
    monkeypatch.setattr(tse.shutil, "which", lambda *_a, **_k: None)
    out = str(tmp_path / "o.xlsx")
    res = tse.export_track_sim(source, out, _trace(30), 0.0666667, recalc=True)
    assert res.recalculated, res.recalc_message
    assert "in process" in res.recalc_message
    ws = openpyxl.load_workbook(out, data_only=True)[tse.S_DASH]
    for cell in ("B11", "B13", "B14"):
        v = ws[cell].value
        assert isinstance(v, (int, float)), f"{cell} read back as {v!r}"


def test_formulas_survive_the_in_process_recalculation(source, tmp_path,
                                                       monkeypatch):
    """Caching values must not turn the model into a snapshot."""
    monkeypatch.setattr(tse.shutil, "which", lambda *_a, **_k: None)
    out = str(tmp_path / "o.xlsx")
    tse.export_track_sim(source, out, _trace(20), 0.0666667, recalc=True)
    ws = openpyxl.load_workbook(out)[tse.S_DASH]
    assert str(ws["B14"].value).startswith("=")


def test_in_process_verdicts_are_strings_not_zeros(source, tmp_path,
                                                   monkeypatch):
    monkeypatch.setattr(tse.shutil, "which", lambda *_a, **_k: None)
    out = str(tmp_path / "o.xlsx")
    tse.export_track_sim(source, out, _trace(20), 0.0666667, recalc=True)
    ws = openpyxl.load_workbook(out, data_only=True)[tse.S_DASH]
    verdicts = [ws.cell(r, 2).value for r in range(1, 40)
                if isinstance(ws.cell(r, 2).value, str)
                and ("PASS" in ws.cell(r, 2).value
                     or "FAIL" in ws.cell(r, 2).value)]
    assert verdicts, "no verdict string was cached"


@needs_soffice
def test_in_process_values_agree_with_libreoffice(source, tmp_path,
                                                  monkeypatch):
    """The fallback is only worth having if it agrees with the real thing."""
    from suspension import xlsx_formula_cache as xfc

    lo = str(tmp_path / "lo.xlsx")
    tse.export_track_sim(source, lo, _trace(25), 0.0666667, recalc=True)

    py = str(tmp_path / "py.xlsx")
    monkeypatch.setattr(tse.shutil, "which", lambda *_a, **_k: None)
    tse.export_track_sim(source, py, _trace(25), 0.0666667, recalc=True)
    monkeypatch.undo()

    a = openpyxl.load_workbook(lo, data_only=True)
    b = openpyxl.load_workbook(py, data_only=True)
    compared = 0
    for name in (tse.S_DASH, tse.S_PACK, tse.S_TRACE, tse.S_GEARS,
                 tse.S_ADVISOR):
        wa, wb_ = a[name], b[name]
        for row in wa.iter_rows():
            for cell in row:
                x, y = cell.value, wb_[cell.coordinate].value
                if isinstance(x, (int, float)) and not isinstance(x, bool):
                    assert isinstance(y, (int, float)), (
                        f"{name}!{cell.coordinate} cached {y!r}, "
                        f"LibreOffice says {x!r}")
                    assert y == pytest.approx(x, rel=1e-9, abs=1e-9), (
                        f"{name}!{cell.coordinate}: {y} vs {x}")
                    compared += 1
                elif isinstance(x, str) and x.strip():
                    assert y == x, f"{name}!{cell.coordinate}: {y!r} vs {x!r}"
                    compared += 1
    assert compared > 200, f"only {compared} cells compared"
    assert xfc.SUPPORTED_FUNCTIONS


@needs_soffice
def test_recalculated_file_is_readable_by_python(source, tmp_path):
    """data_only=True must see numbers, not None."""
    out = str(tmp_path / "o.xlsx")
    res = tse.export_track_sim(source, out, _trace(30), 0.0666667, recalc=True)
    assert res.recalculated, res.recalc_message
    ws = openpyxl.load_workbook(out, data_only=True)[tse.S_DASH]
    for cell in ("B11", "B13", "B14"):
        v = ws[cell].value
        assert isinstance(v, (int, float)), f"{cell} read back as {v!r}"


@needs_soffice
def test_no_formula_errors_after_recalc(source, tmp_path):
    out = str(tmp_path / "o.xlsx")
    res = tse.export_track_sim(source, out, _trace(30), 0.0666667, recalc=True)
    assert res.formula_errors == {}, res.formula_errors
    assert res.ok()


@needs_soffice
def test_it_is_a_live_model(source, tmp_path):
    """Change one input, recalculate, and the conclusions move."""
    out = str(tmp_path / "o.xlsx")
    tse.export_track_sim(source, out, _trace(30), 0.0666667, recalc=True)
    before = openpyxl.load_workbook(out, data_only=True)[tse.S_DASH]["B11"].value

    wb = openpyxl.load_workbook(out)
    ws = wb[tse.S_INPUTS]
    mass = next(r for r in range(1, 60)
                if str(ws.cell(r, 1).value or "").startswith("Mass"))
    ws.cell(mass, 2).value = ws.cell(mass, 2).value * 2
    wb.save(out)
    tse.recalculate(out)

    after = openpyxl.load_workbook(out, data_only=True)[tse.S_DASH]["B11"].value
    assert after > before * 1.5, (
        f"peak current did not respond to doubling the mass "
        f"({before} -> {after}); the workbook is not live")


# --------------------------------------------------------------------------- #
#  7. Provenance
# --------------------------------------------------------------------------- #
def test_provenance_names_the_assumptions_it_introduces(source, tmp_path):
    res = _export(source, tmp_path)
    ws = openpyxl.load_workbook(res.path)[tse.S_PROV]
    text = " ".join(str(ws.cell(r, c).value or "")
                    for r in range(1, ws.max_row + 1) for c in (1, 2))
    for token in ("mass", "CdA", "Rolling resistance", "C-rating",
                  "Inverter efficiency"):
        assert token.lower() in text.lower(), f"{token} not documented"


def test_provenance_records_the_corrections(source, tmp_path):
    res = _export(source, tmp_path)
    ws = openpyxl.load_workbook(res.path)[tse.S_PROV]
    text = " ".join(str(ws.cell(r, c).value or "")
                    for r in range(1, ws.max_row + 1) for c in (1, 2))
    assert "49x" in text or "slower than the wheel" in text
    assert "3x high" in text
    assert "sum(P*dt)" in text or "summed instantaneous" in text


def test_unreadable_pack_falls_back_with_a_warning(tmp_path):
    wb = openpyxl.Workbook()
    wb.active.title = "Nothing"
    wb.active["A1"] = "unrelated"
    src = str(tmp_path / "bare.xlsx")
    wb.save(src)
    out = str(tmp_path / "o.xlsx")
    res = tse.export_track_sim(src, out, _trace(20), 0.0666667, recalc=False)
    assert any("default" in w.lower() for w in res.warnings)


def test_module_provenance_declares_its_limits():
    assert tse.PROVENANCE["known_limits"]
    assert tse.PROVENANCE["hard_rule"]


# --------------------------------------------------------------------------- #
#  8. The bytes wrapper the UI calls
# --------------------------------------------------------------------------- #
def _ms_and_t(n=30, dt=0.0666667):
    mph = _trace(n)
    return ([v * pdw.MPH_TO_MS for v in mph], [i * dt for i in range(n)])


def test_bytes_wrapper_returns_a_workbook_with_kx_sheets(source):
    v_ms, t = _ms_and_t()
    data, res = tse.export_track_sim_bytes(open(source, "rb").read(), v_ms, t,
                                           recalc=False)
    wb = openpyxl.load_workbook(io.BytesIO(data))
    assert [s for s in wb.sheetnames if s.startswith(tse.SHEET_PREFIX)]
    assert res.sheets_added


def test_bytes_wrapper_derives_dt_from_the_time_axis(source):
    v_ms, t = _ms_and_t(n=30, dt=0.05)
    data, res = tse.export_track_sim_bytes(open(source, "rb").read(), v_ms, t,
                                           recalc=False)
    ws = openpyxl.load_workbook(io.BytesIO(data))[tse.S_INPUTS]
    row = next(r for r in range(1, 60)
               if str(ws.cell(r, 1).value or "").startswith("Sample interval"))
    assert ws.cell(row, 2).value == pytest.approx(0.05)


def test_bytes_wrapper_converts_ms_to_mph(source):
    v_ms, t = _ms_and_t()
    data, _ = tse.export_track_sim_bytes(open(source, "rb").read(), v_ms, t,
                                         recalc=False)
    ws = openpyxl.load_workbook(io.BytesIO(data))[tse.S_TRACE]
    assert ws["B3"].value == pytest.approx(v_ms[0] / pdw.MPH_TO_MS, rel=0.2)


def test_uneven_time_base_is_reported(source):
    v_ms, t = _ms_and_t(n=30)
    t[15] += 0.5                      # a gap in the log
    _data, res = tse.export_track_sim_bytes(open(source, "rb").read(), v_ms, t,
                                            recalc=False)
    assert any("uneven" in w.lower() for w in res.warnings)


def test_the_trace_carries_the_real_timestamps(source):
    """A dropped frame must appear in column A, not be averaged away."""
    v_ms, t = _ms_and_t(n=30)
    for i in range(15, 30):
        t[i] += 0.5                   # a half-second gap mid-lap
    data, _res = tse.export_track_sim_bytes(open(source, "rb").read(), v_ms, t,
                                            recalc=False)
    ws = openpyxl.load_workbook(io.BytesIO(data))[tse.S_TRACE]
    written = [ws.cell(r, 1).value for r in range(3, 33)]
    assert written == [pytest.approx(x) for x in t]


def test_sample_weights_sum_to_the_elapsed_time(source, tmp_path, monkeypatch):
    """The trapezoidal weights are what make an uneven integral exact."""
    monkeypatch.setattr(tse.shutil, "which", lambda *_a, **_k: None)
    n = 25
    t = [i * 0.05 for i in range(n)]
    for i in range(12, n):
        t[i] += 0.4
    mph = _trace(n)
    out = str(tmp_path / "o.xlsx")
    tse.export_track_sim(source, out, mph, 0.05, time_s=t, recalc=True)
    ws = openpyxl.load_workbook(out, data_only=True)[tse.S_TRACE]
    weights = [ws.cell(r, 16).value for r in range(3, 3 + n)]
    assert sum(weights) == pytest.approx(t[-1] - t[0])
    dash = openpyxl.load_workbook(out, data_only=True)[tse.S_DASH]
    assert dash["B5"].value == pytest.approx(t[-1] - t[0])


def test_energy_on_an_uneven_log_matches_a_hand_integral(source, tmp_path,
                                                         monkeypatch):
    """Cross-check the sheet's energy against a trapezoid done in Python."""
    monkeypatch.setattr(tse.shutil, "which", lambda *_a, **_k: None)
    n = 25
    t = [i * 0.05 for i in range(n)]
    for i in range(12, n):
        t[i] += 0.4
    out = str(tmp_path / "o.xlsx")
    tse.export_track_sim(source, out, _trace(n), 0.05, time_s=t, recalc=True)
    wb = openpyxl.load_workbook(out, data_only=True)
    tr = wb[tse.S_TRACE]
    p_elec = [tr.cell(r, 10).value for r in range(3, 3 + n)]
    hand = sum((p_elec[i] + p_elec[i + 1]) / 2 * (t[i + 1] - t[i])
               for i in range(n - 1)) / 3600000
    assert wb[tse.S_DASH]["B14"].value == pytest.approx(hand, rel=1e-9)


def test_acceleration_uses_the_real_interval(source, tmp_path):
    """Dividing by a nominal step invents accelerations across a log gap."""
    n = 20
    t = [i * 0.05 for i in range(n)]
    t[10] = t[9] + 0.5                 # one long interval
    for i in range(11, n):
        t[i] = t[i - 1] + 0.05
    out = str(tmp_path / "o.xlsx")
    tse.export_track_sim(source, out, _trace(n), 0.05, time_s=t, recalc=False)
    f = openpyxl.load_workbook(out)[tse.S_TRACE]["D12"].value
    assert "A13-A11" in f.replace(" ", "")
    assert "Inputs'!$B$" not in f.split("MIN(")[-1].split(",")[-1]


def test_even_time_base_is_not_flagged(source):
    v_ms, t = _ms_and_t(n=30)
    _data, res = tse.export_track_sim_bytes(open(source, "rb").read(), v_ms, t,
                                            recalc=False)
    assert not any("uneven" in w.lower() for w in res.warnings)


def test_bytes_wrapper_rejects_a_degenerate_time_axis(source):
    raw = open(source, "rb").read()
    with pytest.raises(ValueError):
        tse.export_track_sim_bytes(raw, [10.0], [0.0], recalc=False)
    with pytest.raises(ValueError):
        tse.export_track_sim_bytes(raw, [10.0, 11.0], [1.0, 1.0], recalc=False)


def test_bytes_wrapper_leaves_the_source_bytes_untouched(source):
    raw = open(source, "rb").read()
    before = bytes(raw)
    v_ms, t = _ms_and_t()
    tse.export_track_sim_bytes(raw, v_ms, t, recalc=False)
    assert raw == before


# --------------------------------------------------------------------------- #
#  9. Rendering
# --------------------------------------------------------------------------- #
#  Long explanatory text was originally written as a plain string in column A.
#  A 240-character sentence in a 34-wide column does not render: it spills over
#  every neighbour, or is clipped by the first non-empty cell to its right.
# --------------------------------------------------------------------------- #
def _merged_rows(ws):
    out = set()
    for rng in ws.merged_cells.ranges:
        for r in range(rng.min_row, rng.max_row + 1):
            out.add(r)
    return out


def test_long_notes_are_merged_and_wrapped(source, tmp_path):
    res = _export(source, tmp_path)
    wb = openpyxl.load_workbook(res.path)
    for name in (tse.S_DASH, tse.S_INPUTS, tse.S_PACK, tse.S_GEARS,
                 tse.S_ADVISOR):
        ws = wb[name]
        c = ws["A2"]
        assert isinstance(c.value, str) and len(c.value) > 60, \
            f"{name}!A2 is not the note"
        assert 2 in _merged_rows(ws), f"{name}!A2 is not merged"
        assert c.alignment.wrap_text, f"{name}!A2 is not wrapped"


def test_merged_notes_have_an_explicit_row_height(source, tmp_path):
    """Excel does not auto-fit the height of a merged wrapped cell."""
    res = _export(source, tmp_path)
    wb = openpyxl.load_workbook(res.path)
    for name in (tse.S_DASH, tse.S_INPUTS, tse.S_PACK, tse.S_GEARS,
                 tse.S_ADVISOR):
        h = wb[name].row_dimensions[2].height
        assert h and h >= tse._LINE_HEIGHT, f"{name} row 2 has no height"


def test_no_unwrapped_cell_badly_overflows_its_column(source, tmp_path):
    res = _export(source, tmp_path)
    wb = openpyxl.load_workbook(res.path)
    problems = []
    for name in [s for s in wb.sheetnames if s.startswith(tse.SHEET_PREFIX)]:
        ws = wb[name]
        merged = _merged_rows(ws)
        for row in ws.iter_rows(max_row=min(ws.max_row, 60)):
            for c in row:
                if c.value is None or c.row in merged:
                    continue
                if isinstance(c.value, str) and c.value.startswith("="):
                    continue
                text = str(c.value)
                if "\n" in text:
                    text = max(text.split("\n"), key=len)
                if len(text) <= 14:
                    continue
                w = ws.column_dimensions[c.column_letter].width
                wrapped = bool(c.alignment and c.alignment.wrap_text)
                if not wrapped and (w is None or len(text) > w * 1.4):
                    problems.append(f"{name}!{c.coordinate} "
                                    f"({len(text)} chars, width {w})")
    assert not problems, "overflowing cells: " + "; ".join(problems[:6])


def test_every_column_with_content_has_a_width(source, tmp_path):
    res = _export(source, tmp_path)
    wb = openpyxl.load_workbook(res.path)
    for name in [s for s in wb.sheetnames if s.startswith(tse.SHEET_PREFIX)]:
        ws = wb[name]
        cols = {c.column_letter for row in ws.iter_rows(max_row=6)
                for c in row if c.value is not None}
        for col in cols:
            assert ws.column_dimensions[col].width, \
                f"{name} column {col} has no width"


def test_trace_header_splits_name_from_unit(source, tmp_path):
    """'Accel (grip-limited) (m/s^2)' is 28 chars in a 12.5-wide column."""
    res = _export(source, tmp_path)
    ws = openpyxl.load_workbook(res.path)[tse.S_TRACE]
    assert "\n" in ws["D2"].value
    assert ws["D2"].alignment.wrap_text
    assert ws.row_dimensions[2].height >= 30


def test_sheets_freeze_their_headers(source, tmp_path):
    res = _export(source, tmp_path)
    wb = openpyxl.load_workbook(res.path)
    for name in (tse.S_DASH, tse.S_TRACE, tse.S_GEARS, tse.S_ADVISOR,
                 tse.S_INPUTS):
        assert wb[name].freeze_panes, f"{name} does not freeze its header"


def test_provenance_section_headers_are_merged(source, tmp_path):
    res = _export(source, tmp_path)
    ws = openpyxl.load_workbook(res.path)[tse.S_PROV]
    merged = _merged_rows(ws)
    headers = [r for r in range(1, ws.max_row + 1)
               if ws.cell(r, 1).value and not ws.cell(r, 2).value]
    assert any(r in merged for r in headers)


def test_autofit_ignores_merged_rows():
    """Including them is what sized column A for a 240-character sentence."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "short"
    ws["A2"] = "x" * 240
    ws.merge_cells("A2:D2")
    tse._autofit(ws, max_w=46.0)
    assert ws.column_dimensions["A"].width < 40


# --------------------------------------------------------------------------- #
#  10. Portability — the actual cause of "zeroes and blank pages"
# --------------------------------------------------------------------------- #
#  The workbook uses CHOOSECOLS, which exists only in Microsoft 365. Everything
#  older — Excel 2021/2019/2016, LibreOffice, Google Sheets, Numbers, preview
#  panes — returns #NAME?. Because the five uses are ARRAY formulas spilling
#  down thousands of rows, they produced 18,949 error cells, and a #NAME? in a
#  source column makes every dependent sum/min/max blank or zero. The sheet
#  therefore looks empty rather than broken.
#
#  INDEX(range, 0, n) is an exact replacement: row index 0 means "whole column".
# --------------------------------------------------------------------------- #
def test_choosecols_rewrites_to_index():
    out, n = tse._rewrite_choose("=_xlfn.CHOOSECOLS(A1:C3,2)")
    assert n == 1
    assert out == "=INDEX(A1:C3,0,2)"


def test_chooserows_rewrites_to_index():
    out, n = tse._rewrite_choose("=_xlfn.CHOOSEROWS(A1:C3,2)")
    assert n == 1
    assert out == "=INDEX(A1:C3,2,0)"


def test_nested_commas_are_not_mistaken_for_argument_separators():
    """A regex would split on the comma inside ROUND()."""
    src = "=_xlfn.CHOOSECOLS(ElecPropulsion!H1895:V3787,ROUND(1/B2,0))"
    out, n = tse._rewrite_choose(src)
    assert n == 1
    assert out == "=INDEX(ElecPropulsion!H1895:V3787,0,ROUND(1/B2,0))"


def test_rewrite_survives_surrounding_arithmetic():
    src = ("=_xlfn.CHOOSECOLS(A1:C3,ROUND(1/B2,0))^2 * "
           "BatteryPackConfig!B11")
    out, n = tse._rewrite_choose(src)
    assert n == 1
    assert out.startswith("=INDEX(A1:C3,0,ROUND(1/B2,0))^2")
    assert out.endswith("BatteryPackConfig!B11")


def test_multiple_occurrences_all_rewritten():
    src = "=_xlfn.CHOOSECOLS(A1:C3,1)+_xlfn.CHOOSECOLS(D1:F3,2)"
    out, n = tse._rewrite_choose(src)
    assert n == 2
    assert "_xlfn" not in out


def test_commas_inside_string_literals_are_respected():
    src = '=_xlfn.CHOOSECOLS(A1:C3,IF(B1>0,1,2))&"a,b"'
    out, n = tse._rewrite_choose(src)
    assert n == 1
    assert out == '=INDEX(A1:C3,0,IF(B1>0,1,2))&"a,b"'


def test_formula_without_modern_functions_is_untouched():
    src = "=SUM(A1:A10)/COUNT(A1:A10)"
    out, n = tse._rewrite_choose(src)
    assert n == 0 and out == src


def test_malformed_call_is_left_alone_rather_than_broken():
    """Better an unchanged #NAME? than a new, differently wrong formula."""
    src = "=_xlfn.CHOOSECOLS(A1:C3"          # unbalanced
    out, n = tse._rewrite_choose(src)
    assert n == 0 and out == src


def _wb_with_choosecols(tmp_path):
    from openpyxl.worksheet.formula import ArrayFormula
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    for r in range(1, 4):
        for c in range(1, 4):
            ws.cell(r, c, r * 10 + c)
    ws["E1"] = ArrayFormula("E1:E3", "=_xlfn.CHOOSECOLS(A1:C3,2)")
    ws["F1"] = "=_xlfn.CHOOSECOLS(A1:C3,1)"
    p = tmp_path / "cc.xlsx"
    wb.save(p)
    return str(p)


def test_make_portable_rewrites_and_counts(tmp_path):
    path = _wb_with_choosecols(tmp_path)
    assert tse.modern_function_report(path)
    changed = tse.make_portable(path)
    assert changed.get("CHOOSECOLS") == 2
    assert tse.modern_function_report(path) == {}


def test_make_portable_keeps_array_formulas_as_arrays(tmp_path):
    """Dropping the array wrapper turns a spilling column into one value —
    a different bug with the same symptom."""
    from openpyxl.worksheet.formula import ArrayFormula
    path = _wb_with_choosecols(tmp_path)
    tse.make_portable(path)
    ws = openpyxl.load_workbook(path)["Data"]
    assert isinstance(ws["E1"].value, ArrayFormula)
    assert ws["E1"].value.ref == "E1:E3"
    assert "INDEX" in ws["E1"].value.text


def test_make_portable_sets_full_recalc_on_load(tmp_path):
    path = _wb_with_choosecols(tmp_path)
    tse.make_portable(path)
    assert openpyxl.load_workbook(path).calculation.fullCalcOnLoad is True


def test_make_portable_is_a_noop_on_a_clean_workbook(tmp_path):
    wb = openpyxl.Workbook()
    wb.active["A1"] = "=SUM(B1:B3)"
    p = str(tmp_path / "clean.xlsx")
    wb.save(p)
    assert tse.make_portable(p) == {}


def test_export_makes_the_source_sheets_portable(source, tmp_path):
    """The user's own sheets are made portable; only formulas change."""
    from openpyxl.worksheet.formula import ArrayFormula
    wb = openpyxl.load_workbook(source)
    wb["ElecPropulsion"]["Z1"] = ArrayFormula(
        "Z1:Z3", "=_xlfn.CHOOSECOLS(A1:C3,2)")
    src2 = str(tmp_path / "with_cc.xlsx")
    wb.save(src2)

    out = str(tmp_path / "out.xlsx")
    res = tse.export_track_sim(src2, out, _trace(), 0.0666667, recalc=False)
    assert tse.modern_function_report(out) == {}
    assert any("365" in w for w in res.warnings)


def test_export_sets_full_recalc_on_load(source, tmp_path):
    """openpyxl writes formulas with no cached value; Excel shows 0 or blank
    for those unless asked to recompute."""
    res = _export(source, tmp_path)
    assert openpyxl.load_workbook(res.path).calculation.fullCalcOnLoad is True


@needs_soffice
def test_the_real_workbook_has_no_error_cells_after_export(tmp_path):
    """End to end on the actual file: 18,949 #NAME? cells before, none after."""
    import os
    src = "/mnt/user-data/uploads/FSAE_EV_Power_Draw.xlsx"
    if not os.path.exists(src):
        pytest.skip("source workbook not present")
    out = str(tmp_path / "real.xlsx")
    tse.export_track_sim(src, out, _trace(40), 0.0666667, recalc=True)
    assert tse.modern_function_report(out) == {}
    assert tse.formula_errors(out) == {}


# --------------------------------------------------------------------------- #
#  11. Zeros from data-extent mismatch
# --------------------------------------------------------------------------- #
#  Second half of the zero problem, and not a formula bug at all. The legacy
#  export rewrote ElecPropulsion's three stacked blocks as ~335 rows of static
#  values in rows 1-1008, while ThermalLoad/EMFs/BearingBlowOut still reference
#  the original offsets at rows 1895-5689. INDEX over blank cells returns 0, so
#  those tabs read as solid zeros with no error to explain it.
# --------------------------------------------------------------------------- #
def _wb_blocks(tmp_path, block_rows, ref_range):
    """A workbook whose dependent formula points at `ref_range`."""
    wb = openpyxl.Workbook()
    svt = wb.active
    svt.title = "SpeedVsTime"
    svt["A1"], svt["B1"] = "time (s)", "Speed (mph)"
    ep = wb.create_sheet("ElecPropulsion")
    ep["A1"], ep["B1"] = "Motor Peak Torque (Nm)", 120
    for i in range(15):
        ep.cell(1, 8 + i, 1.0 / (i + 1))
    for r in range(2, block_rows + 2):
        for c in range(8, 23):
            ep.cell(r, c, 1.0 * r)
    dep = wb.create_sheet("ThermalLoad")
    dep["A1"] = "Joule Heat (kW)"
    dep["F1"] = f"=INDEX(ElecPropulsion!{ref_range},0,7)"
    p = tmp_path / "b.xlsx"
    wb.save(p)
    return str(p)


def test_reference_into_an_empty_region_is_detected(tmp_path):
    path = _wb_blocks(tmp_path, block_rows=300, ref_range="H1895:V3787")
    found = tse.audit_block_references(path)
    assert found, "a reference into blank rows was not detected"
    assert found[0]["target"] == "ElecPropulsion"
    assert "1895" in found[0]["range"]


def test_reference_inside_the_data_is_not_flagged(tmp_path):
    path = _wb_blocks(tmp_path, block_rows=4000, ref_range="H1895:V3787")
    assert tse.audit_block_references(path) == []


def test_extent_detection_uses_the_formula_view(tmp_path):
    """A workbook this module has saved has no cached values. Detecting extents
    from the cached view makes every range look like it points at nothing."""
    path = _wb_blocks(tmp_path, block_rows=4000, ref_range="H1895:V3787")
    tse.make_portable(path)          # re-saves, stripping every cached value
    assert tse.audit_block_references(path) == [], \
        "false positive: extents were read from the cached view"


def test_overshooting_range_is_trimmed(tmp_path):
    """EMFs reads one row past the power block even in the untouched original,
    which is why its last value is zero."""
    path = _wb_blocks(tmp_path, block_rows=100, ref_range="H2:V150")
    changed = tse.repair_range_overshoot(path)
    assert changed
    f = openpyxl.load_workbook(path)["ThermalLoad"]["F1"].value
    assert "V101" in f or "V100" in f
    assert "V150" not in f


def test_trim_never_extends_a_range(tmp_path):
    path = _wb_blocks(tmp_path, block_rows=500, ref_range="H2:V100")
    assert tse.repair_range_overshoot(path) == []
    assert "V100" in openpyxl.load_workbook(path)["ThermalLoad"]["F1"].value


def test_array_spill_shrinks_with_the_trimmed_range(tmp_path):
    """Otherwise the orphaned tail rows become #N/A — a new error replacing an
    old zero."""
    from openpyxl.worksheet.formula import ArrayFormula
    wb = openpyxl.Workbook()
    ep = wb.active
    ep.title = "ElecPropulsion"
    for r in range(2, 52):
        for c in range(8, 23):
            ep.cell(r, c, float(r))
    dep = wb.create_sheet("Dep")
    dep["A1"] = ArrayFormula("A1:A60", "=INDEX(ElecPropulsion!H2:V60,0,1)")
    dep["A60"] = 0.0                       # a stray literal in the orphan row
    p = str(tmp_path / "arr.xlsx")
    wb.save(p)

    tse.repair_range_overshoot(p)
    cell = openpyxl.load_workbook(p)["Dep"]["A1"].value
    assert isinstance(cell, ArrayFormula)
    last = int(cell.ref.split(":")[1].lstrip("$ABCDEFGHIJKLMNOPQRSTUVWXYZ"))
    assert last == 51, f"spill range not shrunk (ref={cell.ref})"


def test_orphan_rows_are_cleared(tmp_path):
    """A literal left below a shrunk spill is exactly the lone stray zero."""
    from openpyxl.worksheet.formula import ArrayFormula
    wb = openpyxl.Workbook()
    ep = wb.active
    ep.title = "ElecPropulsion"
    for r in range(2, 52):
        for c in range(8, 23):
            ep.cell(r, c, float(r))
    dep = wb.create_sheet("Dep")
    dep["A1"] = ArrayFormula("A1:A60", "=INDEX(ElecPropulsion!H2:V60,0,1)")
    dep["A55"] = 0.0
    p = str(tmp_path / "orph.xlsx")
    wb.save(p)
    tse.repair_range_overshoot(p)
    assert openpyxl.load_workbook(p)["Dep"]["A55"].value is None


def test_export_from_a_previously_enhanced_workbook_repairs_it(tmp_path):
    """The user-facing case. This used to only EXPLAIN the zeros and send the
    user off to find the original workbook; it now reconstructs the data."""
    path = _wb_blocks(tmp_path, block_rows=300, ref_range="H1895:V3787")
    out = str(tmp_path / "o.xlsx")
    res = tse.export_track_sim(path, out, _trace(), 0.0666667, recalc=False)
    assert tse.audit_block_references(out) == []
    assert res.trace is not None


def test_declining_the_rebuild_still_explains_the_zeros(tmp_path):
    path, _ = _wb_relocated(tmp_path, rows=300, cols_block2=1)
    out = str(tmp_path / "o.xlsx")
    res = tse.export_track_sim(path, out, _trace(), 0.0666667,
                               rebuild_propulsion=False, recalc=False)
    joined = " ".join(res.warnings)
    assert "1 of 15" in joined
    assert "rebuild_propulsion=True" in joined


def test_export_from_the_original_does_not_cry_wolf(tmp_path):
    path = _wb_blocks(tmp_path, block_rows=4000, ref_range="H1895:V3787")
    out = str(tmp_path / "o.xlsx")
    res = tse.export_track_sim(path, out, _trace(), 0.0666667, recalc=False)
    assert not any("ZEROS EXPLAINED" in w for w in res.warnings)


@needs_soffice
def test_the_real_workbook_ends_with_no_zeros_and_no_errors(tmp_path):
    """End to end: ThermalLoad, EMFs and BearingBlowOut fully populated."""
    import os
    src = "/mnt/user-data/uploads/FSAE_EV_Power_Draw.xlsx"
    if not os.path.exists(src):
        pytest.skip("source workbook not present")
    ws = openpyxl.load_workbook(src, data_only=True)["SpeedVsTime"]
    v = [x for x in (ws.cell(r, 2).value for r in range(2, ws.max_row + 1))
         if isinstance(x, (int, float))]
    out = str(tmp_path / "real.xlsx")
    res = tse.export_track_sim(src, out, v, 0.0666667, recalc=True)
    assert res.formula_errors == {}
    assert not any("ZEROS EXPLAINED" in w for w in res.warnings)


# --------------------------------------------------------------------------- #
#  12. Relocated blocks, and data that is simply gone
# --------------------------------------------------------------------------- #
def _wb_relocated(tmp_path, rows=300, cols_block2=15):
    """Mimics a legacy-exported workbook: three blocks moved up, and the
    current-draw block written for only some gear columns."""
    wb = openpyxl.Workbook()
    svt = wb.active
    svt.title = "SpeedVsTime"
    svt["A1"], svt["B1"] = "time (s)", "Speed (mph)"
    ep = wb.create_sheet("ElecPropulsion")
    ep["A1"], ep["B1"] = "Motor Peak Torque (Nm)", 120
    for i in range(15):
        ep.cell(1, 8 + i, 1.0 / (i + 1))
    b1 = (2, rows + 1)
    b2 = (rows + 3, 2 * rows + 2)
    b3 = (2 * rows + 4, 3 * rows + 3)
    for lo, hi, ncol in ((b1[0], b1[1], 15), (b2[0], b2[1], cols_block2),
                         (b3[0], b3[1], 15)):
        for r in range(lo, hi + 1):
            for c in range(8, 8 + ncol):
                ep.cell(r, c, float(r))
    tl = wb.create_sheet("ThermalLoad")
    tl["A1"], tl["B1"] = "Select Gear Ratio", 7
    tl["A2"], tl["B2"] = "Gear Reduction", "=1/B1"
    tl["F1"] = "=INDEX(ElecPropulsion!H1895:V3787,0,ROUND(1/B2,0))^2"
    tl["A4"] = "Min Joule Heat (kW)"
    tl["B4"] = "=(INDEX(ElecPropulsion!H5685:V5685,1,ROUND(1/B2,0)))^2"
    p = tmp_path / "reloc.xlsx"
    wb.save(p)
    return str(p), (b1, b2, b3)


def test_relocated_blocks_are_repointed(tmp_path):
    path, (b1, b2, b3) = _wb_relocated(tmp_path)
    changed = tse.remap_stale_block_references(path)
    assert changed
    f = openpyxl.load_workbook(path)["ThermalLoad"]["F1"].value
    assert f"H{b2[0]}:V{b2[1]}" in f
    assert "H1895" not in f


def test_deleted_summary_rows_become_aggregates(tmp_path):
    """The MIN/MAX rows no longer exist; MIN() over the block is what they meant."""
    path, (_b1, b2, _b3) = _wb_relocated(tmp_path)
    changed = tse.remap_stale_block_references(path)
    b4 = openpyxl.load_workbook(path)["ThermalLoad"]["B4"].value
    assert b4.startswith("=(MIN(")
    assert f"H{b2[0]}:V{b2[1]}" in b4
    assert any("summary row" in c for c in changed)


def test_remap_refuses_when_the_pairing_is_ambiguous(tmp_path):
    """Two blocks against three references: guessing would put numbers under
    the wrong physical quantity, which is worse than visible zeros."""
    wb = openpyxl.Workbook()
    ep = wb.active
    ep.title = "ElecPropulsion"
    for r in list(range(2, 60)) + list(range(70, 128)):
        for c in range(8, 23):
            ep.cell(r, c, float(r))
    tl = wb.create_sheet("ThermalLoad")
    tl["F1"] = "=INDEX(ElecPropulsion!H1895:V3787,0,7)"
    p = str(tmp_path / "amb.xlsx")
    wb.save(p)
    out = tse.remap_stale_block_references(p)
    assert out and "Refusing to guess" in out[0]
    assert "H1895" in openpyxl.load_workbook(p)["ThermalLoad"]["F1"].value


def test_canonical_workbook_is_not_remapped(tmp_path):
    """Blocks at their canonical rows must be left alone.

    In the real workbook blocks 1 and 2 are CONTIGUOUS (rows 2-1894 then
    1895-3787), so a scan sees them as one run. What makes the canonical case
    safe is not block counting but the per-reference guard: every reference
    already lands on data, so nothing is moved.
    """
    wb = openpyxl.Workbook()
    ep = wb.active
    ep.title = "ElecPropulsion"
    for i in range(15):
        ep.cell(1, 8 + i, 1.0 / (i + 1))
    for r in list(range(2, 3788)) + list(range(3789, 5682)):
        for c in range(8, 23):
            ep.cell(r, c, float(r))
    tl = wb.create_sheet("ThermalLoad")
    tl["B1"], tl["B2"] = 7, "=1/B1"
    tl["F1"] = "=INDEX(ElecPropulsion!H1895:V3787,0,ROUND(1/B2,0))^2"
    p = str(tmp_path / "canon.xlsx")
    wb.save(p)
    assert tse.remap_stale_block_references(p) == []
    assert "H1895:V3787" in openpyxl.load_workbook(p)["ThermalLoad"]["F1"].value


def test_single_block_workbook_refuses_rather_than_guessing(tmp_path):
    """One block against three canonical references is not a pairing.

    The reference must be genuinely stale for this path to be reached — a
    reference that already lands on data needs no repair whatever the block
    count.
    """
    path = _wb_blocks(tmp_path, block_rows=300, ref_range="H1895:V3787")
    out = tse.remap_stale_block_references(path)
    assert out and "Refusing to guess" in out[0]


def test_missing_gear_columns_are_reported_as_unrecoverable(tmp_path):
    """The legacy export kept gear 1 only for the current block — 1 of 15."""
    path, _ = _wb_relocated(tmp_path, cols_block2=1)
    out = tse.audit_block_column_coverage(path)
    assert out
    assert "1 of 15" in out[0]
    assert "cannot be reconstructed" in out[0]


def test_full_column_coverage_is_not_flagged(tmp_path):
    path, _ = _wb_relocated(tmp_path, cols_block2=15)
    assert tse.audit_block_column_coverage(path) == []


def test_repair_workbook_runs_every_pass(tmp_path):
    path, _ = _wb_relocated(tmp_path, cols_block2=1)
    out = tse.repair_workbook(path)
    assert set(out) == {"portable", "remapped", "trimmed", "stale", "coverage"}
    assert out["remapped"]
    assert out["coverage"]
    assert out["stale"] == []        # remapping resolved the empty references


def test_column_loss_is_repaired_rather_than_merely_reported(tmp_path):
    path, _ = _wb_relocated(tmp_path, cols_block2=1)
    out = str(tmp_path / "o.xlsx")
    res = tse.export_track_sim(path, out, _trace(), 0.0666667, recalc=False)
    assert any("Rebuilt all three blocks" in w for w in res.warnings)
    assert tse.audit_block_column_coverage(out) == []


# --------------------------------------------------------------------------- #
#  13. Rebuilding lost propulsion data
# --------------------------------------------------------------------------- #
#  Removes the "export from the original, never from an enhanced copy" caveat.
#  The legacy path wrote the current block for gear 1 only and truncated all
#  three blocks, so dependent sheets read zero at any other gear. Warning about
#  that was never a fix — the data is reconstructible from the speed trace.
# --------------------------------------------------------------------------- #
def test_resample_hits_the_requested_length_and_endpoints():
    out = tse._resample([0.0, 10.0], 5)
    assert len(out) == 5
    assert out[0] == pytest.approx(0.0)
    assert out[-1] == pytest.approx(10.0)
    assert out[2] == pytest.approx(5.0)


def test_resample_handles_degenerate_input():
    assert tse._resample([7.0], 4) == [7.0] * 4
    assert tse._resample([1.0, 2.0, 3.0], 3) == [1.0, 2.0, 3.0]
    assert tse._resample([1.0], 0) == []


def test_rebuild_writes_canonical_rows_and_all_gears(tmp_path):
    """Canonical rows because the dependent sheets reference those exact
    offsets; rebuilding at the lap's own length recreates the original bug."""
    path, _ = _wb_relocated(tmp_path, rows=300, cols_block2=1)
    info = tse.rebuild_propulsion_blocks(
        path, _trace(200), pdw.PackSpec(), pdw.VehicleSpec(), pdw.DriveSpec(),
        dt_s=0.0666667)
    assert info["rebuilt"]
    assert tuple(info["blocks"]) == tse.CANONICAL_EP_BLOCKS
    assert info["gears"] == 15
    assert info["rows"] == 1893

    ws = openpyxl.load_workbook(path)["ElecPropulsion"]
    b1, b2, b3 = tse.CANONICAL_EP_BLOCKS
    for lo, hi in (b1, b2, b3):
        for col in (8, 15, 22):                 # first, middle, last gear
            assert ws.cell(lo, col).value is not None
            assert ws.cell(hi, col).value is not None


def test_rebuild_restores_full_column_coverage(tmp_path):
    path, _ = _wb_relocated(tmp_path, rows=300, cols_block2=1)
    assert tse.audit_block_column_coverage(path)
    tse.rebuild_propulsion_blocks(path, _trace(200), pdw.PackSpec(),
                                  pdw.VehicleSpec(), pdw.DriveSpec(),
                                  dt_s=0.0666667)
    assert tse.audit_block_column_coverage(path) == []


def test_rebuilt_rpm_varies_with_gear_but_current_does_not(tmp_path):
    """Motor speed scales with the reduction. Pack current does not, because
    power is force times speed however the gearbox delivers it."""
    path, _ = _wb_relocated(tmp_path, rows=300, cols_block2=1)
    tse.rebuild_propulsion_blocks(path, _trace(200), pdw.PackSpec(),
                                  pdw.VehicleSpec(), pdw.DriveSpec(),
                                  dt_s=0.0666667)
    ws = openpyxl.load_workbook(path)["ElecPropulsion"]
    b1, b2, _b3 = tse.CANONICAL_EP_BLOCKS
    rpm_g1, rpm_g7 = ws.cell(b1[0] + 5, 8).value, ws.cell(b1[0] + 5, 14).value
    assert rpm_g7 == pytest.approx(rpm_g1 * 7, rel=1e-6)
    cur_g1, cur_g7 = ws.cell(b2[0] + 5, 8).value, ws.cell(b2[0] + 5, 14).value
    assert cur_g1 == pytest.approx(cur_g7)


def test_export_rebuilds_a_damaged_source_by_default(tmp_path):
    path, _ = _wb_relocated(tmp_path, rows=300, cols_block2=1)
    out = str(tmp_path / "o.xlsx")
    res = tse.export_track_sim(path, out, _trace(), 0.0666667, recalc=False)
    assert any("Rebuilt all three blocks" in w for w in res.warnings)
    assert tse.audit_block_column_coverage(out) == []
    assert tse.audit_block_references(out) == []


def test_rebuild_can_be_declined(tmp_path):
    path, _ = _wb_relocated(tmp_path, rows=300, cols_block2=1)
    out = str(tmp_path / "o.xlsx")
    res = tse.export_track_sim(path, out, _trace(), 0.0666667,
                               rebuild_propulsion=False, recalc=False)
    assert not any("Rebuilt all three blocks" in w for w in res.warnings)
    assert any("rebuild_propulsion=True" in w for w in res.warnings)


def test_healthy_source_is_not_rebuilt(tmp_path):
    """A workbook whose references all land on data must be left alone."""
    wb = openpyxl.Workbook()
    ep = wb.active
    ep.title = "ElecPropulsion"
    for i in range(15):
        ep.cell(1, 8 + i, 1.0 / (i + 1))
    for r in list(range(2, 3788)) + list(range(3789, 5682)):
        for c in range(8, 23):
            ep.cell(r, c, float(r))
    svt = wb.create_sheet("SpeedVsTime")
    svt["A1"], svt["B1"] = "time (s)", "Speed (mph)"
    tl = wb.create_sheet("ThermalLoad")
    tl["B1"], tl["B2"] = 7, "=1/B1"
    tl["F1"] = "=INDEX(ElecPropulsion!H1895:V3787,0,ROUND(1/B2,0))^2"
    src = str(tmp_path / "healthy.xlsx")
    wb.save(src)
    out = str(tmp_path / "o.xlsx")
    res = tse.export_track_sim(src, out, _trace(), 0.0666667, recalc=False)
    assert not any("Rebuilt all three blocks" in w for w in res.warnings)
    ws = openpyxl.load_workbook(out)["ElecPropulsion"]
    assert ws.cell(2, 8).value == 2.0        # untouched original values


def test_passes_run_in_dependency_order(tmp_path):
    """Rebuild must precede remapping: remapping first repoints formulas at the
    old positions, which the rebuild then clears."""
    import inspect
    src = inspect.getsource(tse.export_track_sim)
    assert (src.index("audit_block_column_coverage")
            < src.index("rebuild_propulsion_blocks")
            < src.index("remap_stale_block_references"))


def test_coverage_check_reads_uncached_formulas_as_populated(tmp_path):
    """The trap that has now bitten three separate checks in this module.

    NOTE: this test previously shared a name with the one further down this
    file, so Python's module namespace silently discarded it and pytest never
    ran it — a hole in exactly the regression area the docstring warns about.
    Renamed to what it actually asserts: a freshly-written workbook whose cells
    hold formulas with no cached value must not read as empty.

    openpyxl drops every cached value when it saves, and the portability pass
    saves. A data_only read of an already-processed workbook therefore reports
    every column as empty, this check concludes the file is damaged, and the
    rebuild overwrites the user's own formulas with static values. A cell
    holding a formula is populated — that is the only reading that survives a
    round trip.
    """
    wb = openpyxl.Workbook()
    ep = wb.active
    ep.title = "ElecPropulsion"
    for i in range(15):
        ep.cell(1, 8 + i, 1.0 / (i + 1))
    for r in range(2, 40):
        for c in range(8, 23):
            ep.cell(r, c, f"=SpeedVsTime!B{r}*{c}")     # formulas, no cache
    svt = wb.create_sheet("SpeedVsTime")
    for r in range(2, 40):
        svt.cell(r, 2, 30.0)
    path = str(tmp_path / "formulas.xlsx")
    wb.save(path)
    assert tse.audit_block_column_coverage(path) == [], \
        "formula cells were read as empty columns"


def test_a_healthy_workbook_keeps_its_own_formulas(tmp_path):
    """The regression this nearly shipped: the rebuild replacing live formulas
    on a workbook that was never damaged."""
    wb = openpyxl.Workbook()
    ep = wb.active
    ep.title = "ElecPropulsion"
    for i in range(15):
        ep.cell(1, 8 + i, 1.0 / (i + 1))
    for r in list(range(2, 3788)) + list(range(3789, 5682)):
        for c in range(8, 23):
            ep.cell(r, c, float(r))
    ep["H2"] = "=SpeedVsTime!B2*2"
    svt = wb.create_sheet("SpeedVsTime")
    svt["A1"], svt["B1"] = "time (s)", "Speed (mph)"
    tl = wb.create_sheet("ThermalLoad")
    tl["B1"], tl["B2"] = 7, "=1/B1"
    tl["F1"] = "=INDEX(ElecPropulsion!H1895:V3787,0,ROUND(1/B2,0))^2"
    src = str(tmp_path / "healthy2.xlsx")
    wb.save(src)

    out = str(tmp_path / "o.xlsx")
    res = tse.export_track_sim(src, out, _trace(), 0.0666667, recalc=False)
    assert not any("Rebuilt all three blocks" in w for w in res.warnings)
    assert openpyxl.load_workbook(out)["ElecPropulsion"]["H2"].value == \
        "=SpeedVsTime!B2*2"


# --------------------------------------------------------------------------- #
#  14. The cached-value trap, guarded for good
# --------------------------------------------------------------------------- #
#  openpyxl drops every cached value when it saves. Any check that reads the
#  data_only view of a workbook this module has already touched sees an empty
#  sheet. That mistake was made three times here: once in extent detection, once
#  in stale-reference auditing, and once in column-coverage auditing — where it
#  reported the untouched original as damaged and triggered a rebuild that
#  overwrote the user's own formulas with static values.
# --------------------------------------------------------------------------- #
def test_coverage_check_uses_the_formula_view(tmp_path):
    """A workbook whose caches were stripped must not read as damaged."""
    path, _ = _wb_relocated(tmp_path, rows=300, cols_block2=15)
    assert tse.audit_block_column_coverage(path) == []
    tse.make_portable(path)                 # re-saves, stripping every cache
    assert tse.audit_block_column_coverage(path) == [], \
        "false positive: coverage was read from the cached view"


def test_formula_cells_count_as_populated_for_coverage(tmp_path):
    """Formulas with no cached value are still data as far as extent goes."""
    wb = openpyxl.Workbook()
    ep = wb.active
    ep.title = "ElecPropulsion"
    for i in range(15):
        ep.cell(1, 8 + i, 1.0 / (i + 1))
    for r in range(2, 40):
        for c in range(8, 23):
            ep.cell(r, c, "=1+1")           # formula, never evaluated
    p = str(tmp_path / "f.xlsx")
    wb.save(p)
    assert tse.audit_block_column_coverage(p) == []


def test_a_healthy_workbooks_formulas_are_never_overwritten(tmp_path):
    """The regression this guards: rebuilding a healthy sheet replaced live
    formulas with constants, silently turning a model into a snapshot."""
    wb = openpyxl.Workbook()
    ep = wb.active
    ep.title = "ElecPropulsion"
    for i in range(15):
        ep.cell(1, 8 + i, 1.0 / (i + 1))
    for r in list(range(2, 3788)) + list(range(3789, 5682)):
        for c in range(8, 23):
            ep.cell(r, c, "=ROW()")
    svt = wb.create_sheet("SpeedVsTime")
    svt["A1"], svt["B1"] = "time (s)", "Speed (mph)"
    tl = wb.create_sheet("ThermalLoad")
    tl["B1"], tl["B2"] = 7, "=1/B1"
    tl["F1"] = "=INDEX(ElecPropulsion!H1895:V3787,0,ROUND(1/B2,0))^2"
    src = str(tmp_path / "healthy.xlsx")
    wb.save(src)

    out = str(tmp_path / "o.xlsx")
    res = tse.export_track_sim(src, out, _trace(), 0.0666667, recalc=False)
    assert not any("Rebuilt all three blocks" in w for w in res.warnings)
    after = openpyxl.load_workbook(out)["ElecPropulsion"]
    assert after["H2"].value == "=ROW()", "a live formula was replaced"
    assert after["V5681"].value == "=ROW()"
