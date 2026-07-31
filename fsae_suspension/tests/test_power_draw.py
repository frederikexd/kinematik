# ============================================================================
#  KinematiK — Formula SAE suspension & vehicle dynamics toolkit
#  Created by Frederik Thio. Copyright (c) 2026 Frederik Thio.
#  Open source. Original author: Frederik Thio, creator of KinematiK.
# ============================================================================
"""
Tests for suspension.power_draw.

Two halves. The physics tests pin the corrected formulations against
hand-checkable cases — a pack resistance you can do on paper, a constant-speed
cruise whose power is force times velocity, tan(20 degrees) against tan(20
radians). The audit tests pin the *detections*: each one builds a workbook
containing a specific error and asserts the auditor names it, so a future
refactor cannot quietly stop looking.

The reader tests matter most. `_numeric` must raise rather than substitute,
because the substitution is the bug that turned gear ratio 1/8 into direct
drive, and `read_pack_config` must find the pack sheet by label, because
matching on the sheet *name* is what made two shipped readers return their own
defaults from a file they never opened.
"""

import datetime
import math

import pytest

from suspension.interfaces import Severity
from suspension import power_draw as pdw

G = 9.80665


def _sev(findings, check):
    for f in findings:
        if f.check == check:
            return f.severity
    return None


def _checks(findings):
    return {f.check for f in findings}


# --------------------------------------------------------------------------- #
#  Pack arithmetic
# --------------------------------------------------------------------------- #
def test_pack_resistance_divides_by_parallel_strings():
    """140 series groups of 3 parallel cells: R = 140 * (0.0128/3)."""
    p = pdw.PackSpec(n_series=140, n_parallel=3, cell_resistance_ohm=0.0128)
    assert p.resistance_ohm() == pytest.approx(140 * (0.0128 / 3))
    assert p.resistance_ohm() == pytest.approx(0.59733, abs=1e-5)


def test_workbook_resistance_formula_reduces_to_ignoring_parallel():
    """The sheet's cell_count*(R/P) cancels to S*R — the bug, reproduced."""
    p = pdw.PackSpec(n_series=140, n_parallel=3, cell_resistance_ohm=0.0128)
    sheet = p.cell_count() * (p.cell_resistance_ohm / p.n_parallel)
    assert sheet == pytest.approx(p.n_series * p.cell_resistance_ohm)
    assert sheet == pytest.approx(p.resistance_ohm() * p.n_parallel)


def test_parallel_strings_lower_resistance():
    a = pdw.PackSpec(n_series=140, n_parallel=1).resistance_ohm()
    b = pdw.PackSpec(n_series=140, n_parallel=4).resistance_ohm()
    assert b == pytest.approx(a / 4)


def test_pack_topology_basics():
    p = pdw.PackSpec(n_series=140, n_parallel=3, cell_voltage_v=3.6,
                     cell_capacity_ah=5.0, cell_weight_kg=0.071)
    assert p.cell_count() == 420
    assert p.nominal_voltage_v() == pytest.approx(504.0)
    assert p.capacity_ah() == pytest.approx(15.0)
    assert p.energy_kwh() == pytest.approx(7.56)
    assert p.mass_kg() == pytest.approx(29.82)


def test_joule_heat_is_watts():
    p = pdw.PackSpec()
    assert p.joule_heat_w(50.0) == pytest.approx(50.0 ** 2 * p.resistance_ohm())
    assert p.joule_heat_w(50.0) == pytest.approx(1493.3, abs=1.0)


# --------------------------------------------------------------------------- #
#  Current with sag, and the ceiling the workbook never computes
# --------------------------------------------------------------------------- #
def test_current_for_power_accounts_for_sag():
    """P = (V_oc - I*R)*I, so I is above P/V_oc, not equal to it."""
    p = pdw.PackSpec()
    i = p.current_for_power(20_000.0)
    naive = 20_000.0 / p.nominal_voltage_v()
    assert i > naive
    # round-trip: the terminal voltage times the current returns the power
    v_term = p.nominal_voltage_v() - i * p.resistance_ohm()
    assert v_term * i == pytest.approx(20_000.0)


def test_demand_beyond_the_ceiling_returns_none_not_a_number():
    p = pdw.PackSpec()
    assert p.current_for_power(p.max_deliverable_power_w() * 1.01) is None
    assert p.current_for_power(p.max_deliverable_power_w() * 0.99) is not None


def test_max_deliverable_power_is_voc_squared_over_4r():
    p = pdw.PackSpec()
    assert p.max_deliverable_power_w() == pytest.approx(
        p.nominal_voltage_v() ** 2 / (4 * p.resistance_ohm()))
    # and it occurs at I = Voc/2R
    i = p.current_for_power(p.max_deliverable_power_w())
    assert i == pytest.approx(p.nominal_voltage_v() / (2 * p.resistance_ohm()),
                              rel=1e-6)


def test_zero_power_draws_zero_current():
    assert pdw.PackSpec().current_for_power(0.0) == 0.0


def test_fuse_limited_power_is_below_naive_v_times_i():
    """The workbook's 'Power Draw' cell is V_nominal * I with no sag."""
    p = pdw.PackSpec()
    naive = p.fuse_max_a * p.nominal_voltage_v()
    assert p.power_at_fuse_limit_w() < naive
    assert p.power_at_fuse_limit_w() == pytest.approx(
        p.fuse_max_a * (p.nominal_voltage_v()
                        - p.fuse_max_a * p.resistance_ohm()))


def test_cell_current_splits_across_parallel_strings():
    p = pdw.PackSpec(n_parallel=3)
    assert p.cell_current(60.0) == pytest.approx(20.0)


# --------------------------------------------------------------------------- #
#  Gearing — the right way up
# --------------------------------------------------------------------------- #
def test_reduction_multiplies_motor_speed():
    d = pdw.DriveSpec(reduction=7.0)
    wheel_only = pdw.DriveSpec(reduction=1.0).motor_rpm(35.0, 18.0)
    assert d.motor_rpm(35.0, 18.0) == pytest.approx(wheel_only * 7)
    assert d.motor_rpm(35.0, 18.0) == pytest.approx(4575.2, abs=1.0)


def test_workbook_inverted_ratio_is_the_square_wrong_in_power():
    """A 7:1 reduction modelled as 1/7 is 49x out once power is computed."""
    correct = pdw.DriveSpec(reduction=7.0).motor_rpm(35.0, 18.0)
    inverted = pdw.DriveSpec(reduction=1 / 7).motor_rpm(35.0, 18.0)
    assert correct / inverted == pytest.approx(49.0)


def test_wheel_rpm_constant_matches_mph_conversion():
    """1056 = 63360 in/mile / 60 min/hr."""
    assert pdw.MPH_TO_IN_PER_MIN == pytest.approx(63360 / 60)


def test_motor_torque_scales_inversely_with_reduction():
    veh = pdw.VehicleSpec()
    a = pdw.DriveSpec(reduction=4.0).motor_torque_nm(1000.0,
                                                     veh.wheel_radius_m())
    b = pdw.DriveSpec(reduction=8.0).motor_torque_nm(1000.0,
                                                     veh.wheel_radius_m())
    assert b == pytest.approx(a / 2)


# --------------------------------------------------------------------------- #
#  Force balance
# --------------------------------------------------------------------------- #
def _steady(speed_mph, n=40, dt=0.1):
    return [speed_mph] * n, dt


def test_steady_cruise_power_is_force_times_velocity():
    """No acceleration: P_wheel = (F_roll + F_aero) * v, exactly."""
    v, dt = _steady(40.0)
    pack, veh = pdw.PackSpec(), pdw.VehicleSpec(regen=False)
    tr = pdw.power_draw_trace(v, dt, pack, veh, pdw.DriveSpec())
    ms = 40.0 * pdw.MPH_TO_MS
    f_expected = (veh.crr * veh.mass_kg * G
                  + 0.5 * veh.air_density * veh.cda_m2 * ms ** 2)
    mid = len(v) // 2
    assert tr.accel_ms2[mid] == pytest.approx(0.0, abs=1e-9)
    assert tr.p_wheel_w[mid] == pytest.approx(f_expected * ms)


def test_efficiency_divides_so_losses_raise_the_demand():
    """The workbook multiplies; a better motor must not draw MORE current."""
    v, dt = _steady(40.0)
    pack = pdw.PackSpec()
    good = pdw.power_draw_trace(v, dt, pack,
                                pdw.VehicleSpec(motor_efficiency=0.98),
                                pdw.DriveSpec())
    poor = pdw.power_draw_trace(v, dt, pack,
                                pdw.VehicleSpec(motor_efficiency=0.85),
                                pdw.DriveSpec())
    mid = len(v) // 2
    assert poor.p_elec_w[mid] > good.p_elec_w[mid]
    assert poor.i_pack_a[mid] > good.i_pack_a[mid]
    # and electrical demand always exceeds mechanical output
    assert good.p_elec_w[mid] > good.p_wheel_w[mid]


def test_aero_force_scales_with_speed_squared():
    pack, veh = pdw.PackSpec(), pdw.VehicleSpec()
    lo = pdw.power_draw_trace(*_steady(20.0), pack=pack, vehicle=veh,
                              drive=pdw.DriveSpec())
    hi = pdw.power_draw_trace(*_steady(40.0), pack=pack, vehicle=veh,
                              drive=pdw.DriveSpec())
    m = len(lo.speed_ms) // 2
    assert hi.f_aero_n[m] == pytest.approx(4 * lo.f_aero_n[m], rel=1e-6)


def test_pack_current_does_not_depend_on_gearing():
    """Power is force times speed regardless of which gear delivers it.

    This is why the workbook's fifteen-column current sweep could never have
    informed a gear choice, and why the corrected sweep reports torque.
    """
    v, dt = _steady(40.0)
    pack, veh = pdw.PackSpec(), pdw.VehicleSpec()
    sweep = pdw.gear_sweep(v, dt, pack, veh, pdw.DriveSpec(),
                           reductions=[3, 7, 13])
    peaks = {r: t.peak_current_a() for r, t in sweep.items()}
    assert len(set(round(x, 9) for x in peaks.values())) == 1
    # but torque and rpm do change
    torques = {r: t.peak_motor_torque_nm() for r, t in sweep.items()}
    assert len(set(round(x, 6) for x in torques.values())) == 3


def test_energy_is_the_time_integral_not_the_sum():
    v, dt = _steady(40.0, n=100, dt=0.05)
    tr = pdw.power_draw_trace(v, dt, pdw.PackSpec(), pdw.VehicleSpec(),
                              pdw.DriveSpec())
    assert tr.energy_kwh() == pytest.approx(
        sum(tr.p_elec_w) * dt / 3.6e6)
    # halving dt at the same speeds halves the energy
    v2, dt2 = _steady(40.0, n=100, dt=0.025)
    tr2 = pdw.power_draw_trace(v2, dt2, pdw.PackSpec(), pdw.VehicleSpec(),
                               pdw.DriveSpec())
    assert tr2.energy_kwh() == pytest.approx(tr.energy_kwh() / 2)


def test_braking_draws_nothing_without_regen():
    v = [40.0 - i for i in range(20)]
    tr = pdw.power_draw_trace(v, 0.1, pdw.PackSpec(),
                              pdw.VehicleSpec(regen=False), pdw.DriveSpec())
    assert all(p >= 0 for p in tr.p_elec_w)
    assert all(i >= 0 for i in tr.i_pack_a)


def test_regen_returns_energy_when_enabled():
    v = [40.0 - i for i in range(20)]
    off = pdw.power_draw_trace(v, 0.1, pdw.PackSpec(),
                               pdw.VehicleSpec(regen=False), pdw.DriveSpec())
    on = pdw.power_draw_trace(v, 0.1, pdw.PackSpec(),
                              pdw.VehicleSpec(regen=True), pdw.DriveSpec())
    assert on.energy_kwh() < off.energy_kwh()


def test_trace_needs_two_samples_and_positive_dt():
    with pytest.raises(ValueError):
        pdw.power_draw_trace([30.0], 0.1, pdw.PackSpec(), pdw.VehicleSpec(),
                             pdw.DriveSpec())
    with pytest.raises(ValueError):
        pdw.power_draw_trace([30.0, 30.0], 0.0, pdw.PackSpec(),
                             pdw.VehicleSpec(), pdw.DriveSpec())


# --------------------------------------------------------------------------- #
#  Source-data quality
# --------------------------------------------------------------------------- #
def test_discontinuities_are_detected():
    """A 30 -> 14 mph step in one 67 ms sample is about 11 g."""
    v = [30.0] * 10 + [14.0] * 10
    d = pdw.find_discontinuities(v, 0.0666667, 300.0, mu_lon=1.4)
    assert len(d) == 1
    assert abs(d[0][2]) > 10


def test_clean_trace_has_no_discontinuities():
    v = [30.0 + 0.05 * i for i in range(50)]
    assert pdw.find_discontinuities(v, 0.0666667, 300.0) == []


def test_smoothing_reduces_the_peak_from_a_glitch():
    v = [30.0] * 20 + [14.0] * 20
    pack, veh = pdw.PackSpec(), pdw.VehicleSpec()
    raw = pdw.power_draw_trace(v, 0.0666667, pack, veh, pdw.DriveSpec(),
                               smooth_window=1)
    sm = pdw.power_draw_trace(v, 0.0666667, pack, veh, pdw.DriveSpec(),
                              smooth_window=11)
    assert sm.peak_current_a() <= raw.peak_current_a()
    assert sm.smooth_window == 11


def test_grip_clamp_bounds_acceleration_and_counts_itself():
    v = [10.0, 60.0, 10.0, 60.0]
    tr = pdw.power_draw_trace(v, 0.05, pdw.PackSpec(),
                              pdw.VehicleSpec(mu_lon=1.4), pdw.DriveSpec(),
                              clamp_to_grip=True)
    assert tr.clamped_samples > 0
    assert max(abs(a) for a in tr.accel_ms2) <= 1.4 * G + 1e-9


def test_clamp_can_be_disabled():
    v = [10.0, 60.0, 10.0, 60.0]
    tr = pdw.power_draw_trace(v, 0.05, pdw.PackSpec(), pdw.VehicleSpec(),
                              pdw.DriveSpec(), clamp_to_grip=False)
    assert tr.clamped_samples == 0
    assert max(abs(a) for a in tr.accel_ms2) > 1.4 * G


def test_moving_average_preserves_a_constant_and_shrinks_at_ends():
    assert pdw._moving_average([5.0] * 10, 5) == pytest.approx([5.0] * 10)
    assert pdw._moving_average([1.0, 2.0, 3.0], 1) == [1.0, 2.0, 3.0]


def test_discontinuity_finding_escalates_with_frequency():
    pack, veh = pdw.PackSpec(), pdw.VehicleSpec()
    glitchy = [30.0 if i % 2 else 14.0 for i in range(60)]
    tr = pdw.power_draw_trace(glitchy, 0.0666667, pack, veh, pdw.DriveSpec())
    assert _sev(pdw.trace_findings(tr), "pd-speed-discontinuities") == Severity.FAIL


# --------------------------------------------------------------------------- #
#  Gear / bearing physics
# --------------------------------------------------------------------------- #
def test_tangential_force_uses_the_pitch_radius():
    f = pdw.gear_tangential_force_n(77.0, 0.060)
    assert f == pytest.approx(77.0 / 0.030)
    # the workbook divides by the diameter, giving half
    assert f == pytest.approx(2 * (77.0 / 0.060))


def test_radial_force_converts_degrees_to_radians():
    ft = 1000.0
    assert pdw.gear_radial_force_n(ft, 20.0) == pytest.approx(
        ft * math.tan(math.radians(20.0)))
    # the workbook's TAN(20) is the tangent of 20 radians
    assert (ft * math.tan(20.0)) / pdw.gear_radial_force_n(ft, 20.0) == \
        pytest.approx(6.146, abs=0.01)


def test_pitch_diameter_must_be_positive():
    with pytest.raises(ValueError):
        pdw.gear_tangential_force_n(77.0, 0.0)


# --------------------------------------------------------------------------- #
#  Reading without lying
# --------------------------------------------------------------------------- #
def test_numeric_raises_on_a_datetime_instead_of_defaulting():
    """This is the coercion that turned gear ratio 1/8 into direct drive."""
    with pytest.raises(pdw.WorkbookReadError) as e:
        pdw._numeric(datetime.time(3, 0), "ElecPropulsion!O1")
    assert "date" in str(e.value).lower()


def test_numeric_raises_on_empty_text_and_bool():
    for bad in (None, "abc", True):
        with pytest.raises(pdw.WorkbookReadError):
            pdw._numeric(bad, "X!A1")


def test_numeric_accepts_numbers_and_numeric_text():
    assert pdw._numeric(3.5, "X") == 3.5
    assert pdw._numeric(" 7 ", "X") == 7.0


# --------------------------------------------------------------------------- #
#  Audit detections, each against a purpose-built workbook
# --------------------------------------------------------------------------- #
def _wb(tmp_path, build):
    import openpyxl
    wb = openpyxl.Workbook()
    build(wb)
    p = tmp_path / "t.xlsx"
    wb.save(p)
    return str(p)


def _pack_sheet(wb, name="BatteryPackConfig"):
    ws = wb.active
    ws.title = name
    rows = [("Fuse Max (A)", 50), ("Parrallel Battery Count", 3),
            ("Series Battery Count", 140), ("Nominal Battery Voltage (V)", 3.6),
            ("Capacity Battery Cell (Ah)", 5), ("Endurance Length (km)", 22),
            ("Max Battery Cells", 560),
            ("Internal Resistance Battery Cell (Ohms)", 0.0128),
            ("Battery Cell Weight (kg)", 0.071)]
    for i, (a, b) in enumerate(rows, start=1):
        ws.cell(i, 1, a)
        ws.cell(i, 2, b)
    return ws


def test_read_pack_config_finds_the_sheet_by_label(tmp_path):
    """Named anything at all — the labels are what identify it."""
    path = _wb(tmp_path, lambda wb: _pack_sheet(wb, "Something Else Entirely"))
    p = pdw.read_pack_config(path)
    assert p.n_series == 140 and p.n_parallel == 3
    assert p.nominal_voltage_v() == pytest.approx(504.0)


def test_read_pack_config_raises_when_labels_absent(tmp_path):
    def build(wb):
        wb.active.title = "Nothing"
        wb.active["A1"] = "unrelated"
    with pytest.raises(pdw.WorkbookReadError):
        pdw.read_pack_config(_wb(tmp_path, build))


def test_read_pack_config_raises_on_a_date_formatted_cell(tmp_path):
    def build(wb):
        ws = _pack_sheet(wb)
        ws["B3"] = datetime.time(2, 24)
    with pytest.raises(pdw.WorkbookReadError):
        pdw.read_pack_config(_wb(tmp_path, build))


def test_audit_reports_the_resistance_error(tmp_path):
    path = _wb(tmp_path, _pack_sheet)
    pack = pdw.read_pack_config(path)
    f = pdw.audit_pack_sheet(pack, {"pack_resistance_ohm": 1.792})
    assert _sev(f, "pd-pack-resistance") == Severity.FAIL


def test_audit_accepts_a_correct_resistance(tmp_path):
    path = _wb(tmp_path, _pack_sheet)
    pack = pdw.read_pack_config(path)
    f = pdw.audit_pack_sheet(pack, {"pack_resistance_ohm": pack.resistance_ohm()})
    assert "pd-pack-resistance" not in _checks(f)


def test_audit_flags_kwh_label_on_a_power_quantity(tmp_path):
    path = _wb(tmp_path, _pack_sheet)
    pack = pdw.read_pack_config(path)
    f = pdw.audit_pack_sheet(pack, {"joule_max_label":
                                    "joule heating maximum (kwh)"})
    assert _sev(f, "pd-joule-units") == Severity.FAIL


def test_audit_always_reports_the_power_ceiling(tmp_path):
    path = _wb(tmp_path, _pack_sheet)
    pack = pdw.read_pack_config(path)
    assert _sev(pdw.audit_pack_sheet(pack, {}), "pd-power-ceiling") == Severity.WARN


def test_reader_probe_catches_a_reader_that_ignores_the_file(tmp_path):
    path = _wb(tmp_path, _pack_sheet)
    findings = pdw.assert_reader_sees_the_file(path, lambda p: "always the same")
    assert _sev(findings, "pd-reader-ignores-file") == Severity.FAIL


def test_reader_probe_passes_a_reader_that_reads(tmp_path):
    path = _wb(tmp_path, _pack_sheet)
    findings = pdw.assert_reader_sees_the_file(path, pdw.read_pack_config)
    assert _sev(findings, "pd-reader-ok") == Severity.OK


# --------------------------------------------------------------------------- #
#  Feasibility, and the bridges
# --------------------------------------------------------------------------- #
def test_motor_rating_above_the_pack_ceiling_is_a_fail():
    pack = pdw.PackSpec()
    f = pdw.motor_feasibility(pack, pdw.DriveSpec(motor_peak_power_kw=150.0))
    assert _sev(f, "pd-motor-exceeds-pack") == Severity.FAIL


def test_a_modest_motor_clears_the_pack_ceiling():
    pack = pdw.PackSpec()
    f = pdw.motor_feasibility(pack, pdw.DriveSpec(motor_peak_power_kw=20.0))
    assert "pd-motor-exceeds-pack" not in _checks(f)


def test_fuse_limit_is_reported_against_the_motor():
    f = pdw.motor_feasibility(pdw.PackSpec(),
                              pdw.DriveSpec(motor_peak_power_kw=150.0))
    assert _sev(f, "pd-motor-exceeds-fuse") == Severity.WARN


def test_torque_exceeded_at_low_reduction_and_ok_at_high():
    # An accelerating trace, because at steady cruise the roll+aero force alone
    # needs only ~80 Nm at direct drive — enough to exceed the continuous
    # rating but not the peak. The m*a term is what makes gearing matter.
    dt = 0.1
    v = [45.0 + 2.5 * i * dt for i in range(40)]
    pack, veh = pdw.PackSpec(), pdw.VehicleSpec()
    low = pdw.power_draw_trace(v, dt, pack, veh, pdw.DriveSpec(reduction=1.0))
    high = pdw.power_draw_trace(v, dt, pack, veh, pdw.DriveSpec(reduction=15.0))
    assert _sev(pdw.trace_findings(low), "pd-torque-exceeded") == Severity.FAIL
    assert "pd-torque-exceeded" not in _checks(pdw.trace_findings(high))


def test_steady_cruise_at_direct_drive_only_warns():
    """Distinguishes the two torque thresholds: 80 Nm is over the 77 Nm
    continuous rating but well under the 120 Nm peak."""
    v, dt = _steady(45.0)
    tr = pdw.power_draw_trace(v, dt, pdw.PackSpec(), pdw.VehicleSpec(),
                              pdw.DriveSpec(reduction=1.0))
    f = pdw.trace_findings(tr)
    assert _sev(f, "pd-torque-above-continuous") == Severity.WARN
    assert "pd-torque-exceeded" not in _checks(f)


def test_overspeed_is_caught():
    v, dt = _steady(60.0)
    tr = pdw.power_draw_trace(v, dt, pdw.PackSpec(), pdw.VehicleSpec(),
                              pdw.DriveSpec(reduction=15.0,
                                            motor_max_rpm=5000.0))
    assert _sev(pdw.trace_findings(tr), "pd-overspeed") == Severity.FAIL


def test_bms_channels_count_series_groups_not_cells():
    """Parallel cells share a node; a monitor cannot tell them apart."""
    pack = pdw.PackSpec(n_series=140, n_parallel=3)
    f = pdw.bms_findings(pack)
    d = [x for x in f if x.check == "pd-bms-channels"][0].detail
    assert d["channels"] == 140
    assert d["boards"] == math.ceil(140 / 16)      # 9, not 27


def test_bq_stack_bridge_is_within_the_daisy_chain_limit():
    stack = pdw.to_bq_stack(pdw.PackSpec(n_series=140, n_parallel=3))
    assert stack.boards == 9
    assert stack.boards <= stack.device().max_boards


def test_cell_count_over_declared_maximum_fails():
    f = pdw.bms_findings(pdw.PackSpec(n_series=200, n_parallel=3,
                                      max_cells=560))
    assert _sev(f, "pd-cell-count-over") == Severity.FAIL


def test_elec_params_bridge_carries_the_read_values():
    pack = pdw.PackSpec(n_series=100, n_parallel=4, cell_capacity_ah=4.0)
    ep = pdw.to_elec_params(pack, pdw.VehicleSpec())
    assert ep.n_series == 100 and ep.n_parallel == 4
    assert ep.pack_capacity_ah == pytest.approx(16.0)


def test_infeasible_trace_is_reported():
    """A demand the pack cannot meet at any current."""
    pack = pdw.PackSpec(n_series=20, n_parallel=1)      # tiny, high-resistance
    v, dt = _steady(60.0)
    tr = pdw.power_draw_trace(v, dt, pack, pdw.VehicleSpec(), pdw.DriveSpec())
    assert tr.infeasible_samples > 0
    assert _sev(pdw.trace_findings(tr), "pd-trace-infeasible") == Severity.FAIL


def test_fuse_ok_when_current_is_low():
    v, dt = _steady(15.0)
    tr = pdw.power_draw_trace(v, dt, pdw.PackSpec(), pdw.VehicleSpec(),
                              pdw.DriveSpec(reduction=13.0))
    assert _sev(pdw.trace_findings(tr), "pd-fuse-ok") == Severity.OK


def test_markdown_renders(tmp_path):
    path = _wb(tmp_path, _pack_sheet)
    a = pdw.audit(path, include_readers=False)
    md = a.to_markdown()
    assert "audit" in md.lower()
    assert "140S3P" in md


def test_provenance_separates_workbook_errors_from_kinematik_errors():
    assert pdw.PROVENANCE["errors_found_in_workbook"]
    assert pdw.PROVENANCE["errors_found_in_kinematik"]
    joined = " ".join(pdw.PROVENANCE["estimate_flagged"])
    assert "VehicleSpec" in joined
