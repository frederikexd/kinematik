# ============================================================================
#  KinematiK — Formula SAE suspension & vehicle dynamics toolkit
#  Created by Frederik Thio. Copyright (c) 2026 Frederik Thio.
#  Open source. Original author: Frederik Thio, creator of KinematiK.
# ============================================================================
"""
Tests for storage-error diagnosis and gear-ratio recovery.

Both fixes are about the same failure mode: a wrong answer that looks like a
right one.

`diagnose_storage_error` exists because the previous message named the table,
its columns and its row-level-security policy for *every* write failure. For the
most common failure — an expired API key — all three are irrelevant, and the
advice sends someone auditing RLS policies that were never involved.

`gear_ratio_value` exists because two cells in the shipped workbook carry a date
number format, and the old reader defaulted them to 1.0. Not a crash, not a
blank: a plausible direct-drive ratio. The values turn out to be recoverable,
because an Excel time serial is a fraction of a day.
"""

import base64
import datetime as dt
import json
import time

import openpyxl
import pytest

from suspension.project import _jwt_expiry, diagnose_storage_error
from suspension.ev_excel_roundtrip import gear_ratio_value


def _key(exp_offset_days=30, *, with_exp=True, segments=3):
    claims = {"iss": "supabase", "role": "anon"}
    if with_exp:
        claims["exp"] = int(time.time() + exp_offset_days * 86400)
    payload = base64.urlsafe_b64encode(
        json.dumps(claims).encode()).decode().rstrip("=")
    parts = ["eyJhbGciOiJIUzI1NiJ9", payload, "sig"][:segments]
    return ".".join(parts)


class _Backend:
    TABLE = "kinematik_project"

    def __init__(self, key=None):
        self.key = key


# --------------------------------------------------------------------------- #
#  JWT expiry, read locally
# --------------------------------------------------------------------------- #
def test_expiry_is_read_from_the_exp_claim():
    got = _jwt_expiry(_key(exp_offset_days=10))
    expected = dt.datetime.now(dt.UTC) + dt.timedelta(days=10)
    assert abs((got - expected).total_seconds()) < 120


def test_expiry_returns_none_for_a_key_it_cannot_parse():
    assert _jwt_expiry("not-a-jwt") is None
    assert _jwt_expiry(_key(segments=2)) is None
    assert _jwt_expiry(_key(with_exp=False)) is None
    assert _jwt_expiry("") is None
    assert _jwt_expiry(None) is None


def test_expiry_never_raises_on_hostile_input():
    for bad in ("a.b.c", "..", "x." + "!" * 10 + ".y", 12345, object()):
        assert _jwt_expiry(bad) is None


# --------------------------------------------------------------------------- #
#  The reported error
# --------------------------------------------------------------------------- #
_REAL_ERROR = ("{'message': 'JWT expired', 'code': 'PGRST303', "
               "'hint': None, 'details': None}")


def test_expired_key_is_identified_as_a_credential_problem():
    msg = diagnose_storage_error(_REAL_ERROR, backend=_Backend(_key(-9)))
    assert "EXPIRED" in msg
    assert "SUPABASE_KEY" in msg


def test_expired_key_advice_explicitly_rules_out_schema_and_policy():
    """The old message sent people to check exactly these. It must say not to."""
    msg = diagnose_storage_error(_REAL_ERROR, backend=_Backend(_key(-9)))
    low = msg.lower()
    assert "not a schema" in low or "irrelevant" in low
    assert "row-level-security" in low or "row-level security" in low


def test_expired_key_message_names_the_expiry_date():
    msg = diagnose_storage_error(_REAL_ERROR, backend=_Backend(_key(-9)))
    assert "days ago" in msg or "day ago" in msg
    assert "UTC" in msg


def test_expired_key_message_works_without_a_readable_key():
    msg = diagnose_storage_error(_REAL_ERROR, backend=_Backend(None))
    assert "EXPIRED" in msg          # still identifies the cause
    assert "UTC" not in msg          # just cannot date it


def test_hours_are_used_for_a_recent_expiry():
    msg = diagnose_storage_error(_REAL_ERROR,
                                 backend=_Backend(_key(-0.25)))
    assert "h ago" in msg


# --------------------------------------------------------------------------- #
#  Every other cause gets its own advice
# --------------------------------------------------------------------------- #
@pytest.mark.parametrize("err,expect", [
    ('{"code":"PGRST301","message":"JWSInvalid"}', "not a valid token"),
    ('{"code":"42P01","message":"relation does not exist"}', "was not found"),
    ('{"code":"42501","message":"violates row-level security policy"}',
     "refusing the write"),
    ("HTTPSConnectionPool: Read timed out", "network problem"),
    ("413 Payload Too Large", "too large"),
])
def test_each_cause_gets_targeted_advice(err, expect):
    assert expect in diagnose_storage_error(err, backend=_Backend(_key(30)))


def test_unknown_error_gives_an_ordered_checklist():
    msg = diagnose_storage_error("something new", backend=_Backend(_key(30)))
    assert "in this order" in msg
    assert msg.index("key is current") < msg.index("row-level")


def test_diagnosis_uses_the_backend_table_name():
    class Other(_Backend):
        TABLE = "my_custom_table"
    msg = diagnose_storage_error('{"code":"42P01","message":"does not exist"}',
                                 backend=Other())
    assert "my_custom_table" in msg


def test_diagnosis_accepts_an_exception_object_not_just_a_string():
    msg = diagnose_storage_error(RuntimeError(_REAL_ERROR),
                                 backend=_Backend(_key(-3)))
    assert "EXPIRED" in msg


def test_diagnosis_never_raises():
    for bad in (None, 0, [], object(), Exception()):
        assert isinstance(diagnose_storage_error(bad), str)


# --------------------------------------------------------------------------- #
#  Gear ratios: recovered, not defaulted
# --------------------------------------------------------------------------- #
def test_date_formatted_cells_recover_the_intended_fraction():
    """An Excel time serial is a fraction of a day, so 1/8 and 1/10 come back
    exactly rather than becoming 1.0."""
    v, status = gear_ratio_value(dt.time(3, 0))
    assert status == "recovered"
    assert v == pytest.approx(1 / 8)

    v, status = gear_ratio_value(dt.time(2, 24))
    assert status == "recovered"
    assert v == pytest.approx(1 / 10)


def test_plain_numbers_pass_through():
    for raw in (1, 0.5, 0.0666667):
        v, status = gear_ratio_value(raw)
        assert status == "ok"
        assert v == pytest.approx(float(raw))


def test_numeric_text_is_accepted():
    v, status = gear_ratio_value(" 0.25 ")
    assert status == "ok" and v == pytest.approx(0.25)


def test_timedelta_is_also_a_fraction_of_a_day():
    v, status = gear_ratio_value(dt.timedelta(hours=3))
    assert status == "recovered" and v == pytest.approx(0.125)


def test_a_full_datetime_is_not_guessed_at():
    """The fraction is gone once a date epoch is attached; do not invent one."""
    v, status = gear_ratio_value(dt.datetime(2026, 1, 8))
    assert status == "unreadable" and v is None


def test_unreadable_inputs_are_reported_not_defaulted():
    for bad in (None, "", "abc", True, object()):
        v, status = gear_ratio_value(bad)
        assert status == "unreadable", f"{bad!r} was accepted"
        assert v is None


def test_midnight_is_treated_as_unreadable_not_zero():
    """A zero ratio is not a gear; it would divide by zero downstream."""
    v, status = gear_ratio_value(dt.time(0, 0))
    assert status == "unreadable" and v is None


# --------------------------------------------------------------------------- #
#  End to end through the importer
# --------------------------------------------------------------------------- #
def _wb_with_gear_row(values):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SpeedVsTime"
    ws["A1"], ws["B1"] = "time (s)", "Speed (mph)"
    ep = wb.create_sheet("ElecPropulsion")
    ep["A1"], ep["B1"] = "Motor Peak Torque (Nm)", 120
    for i, v in enumerate(values):
        ep.cell(1, 8 + i, v)
    pk = wb.create_sheet("BatteryPackConfig")
    for i, (a, b) in enumerate([("Fuse Max (A)", 50),
                                ("Parrallel Battery Count", 3),
                                ("Series Battery Count", 140),
                                ("Nominal Battery Voltage (V)", 3.6),
                                ("Capacity Battery Cell (Ah)", 5),
                                ("Internal Resistance Battery Cell (Ohms)",
                                 0.0128)], start=1):
        pk.cell(i, 1, a)
        pk.cell(i, 2, b)
    import io
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_importer_recovers_the_two_date_formatted_ratios():
    from suspension.ev_excel_roundtrip import extract_params_from_excel
    vals = [1.0 / i for i in range(1, 16)]
    vals[7] = dt.time(3, 0)          # gear 8
    vals[9] = dt.time(2, 24)         # gear 10
    d = extract_params_from_excel(_wb_with_gear_row(vals))
    assert d["gear_ratios_recovered"] == [8, 10]
    assert d["gear_ratios_unreadable"] == []
    assert d["gear_ratios"][7] == pytest.approx(1 / 8)
    assert d["gear_ratios"][9] == pytest.approx(1 / 10)
    assert len(d["gear_ratio_notes"]) == 2


def test_importer_flags_a_truly_unreadable_ratio():
    from suspension.ev_excel_roundtrip import extract_params_from_excel
    vals = [1.0 / i for i in range(1, 16)]
    vals[4] = "n/a"
    d = extract_params_from_excel(_wb_with_gear_row(vals))
    assert d["gear_ratios_unreadable"] == [5]
    assert d["gear_ratios"][4] == 1.0        # still populated, but declared
    assert any("direct drive" in n for n in d["gear_ratio_notes"])


def test_clean_workbook_produces_no_notes():
    from suspension.ev_excel_roundtrip import extract_params_from_excel
    d = extract_params_from_excel(
        _wb_with_gear_row([1.0 / i for i in range(1, 16)]))
    assert d["gear_ratio_notes"] == []
    assert d["gear_ratios_recovered"] == []


def test_the_real_workbook_yields_fifteen_correct_ratios():
    from suspension.ev_excel_roundtrip import extract_params_from_excel
    path = "/mnt/user-data/uploads/FSAE_EV_Power_Draw.xlsx"
    import os
    if not os.path.exists(path):
        pytest.skip("source workbook not present")
    d = extract_params_from_excel(open(path, "rb").read())
    assert d["gear_ratios_recovered"] == [8, 10]
    for i, r in enumerate(d["gear_ratios"], start=1):
        assert r == pytest.approx(1.0 / i), f"gear {i} is {r}"


# --------------------------------------------------------------------------- #
#  A broken deployment is not a storage problem
# --------------------------------------------------------------------------- #
#  Reported error:
#    Could not write project data: cannot import name 'StaleWriteError' from
#    'suspension.project' (/mount/src/.../suspension/project.py)
#  ...followed by advice to check the Supabase table and its RLS policy. None of
#  that was involved. The repo has TWO project.py files and the package imports
#  suspension.project; overwriting it with the top-level copy strips names the
#  rest of the package needs.
# --------------------------------------------------------------------------- #
_IMPORT_ERROR = ("cannot import name 'StaleWriteError' from "
                 "'suspension.project' (/mount/src/kinematik/"
                 "fsae_suspension/suspension/project.py)")


def test_import_failure_is_not_blamed_on_storage():
    msg = diagnose_storage_error(_IMPORT_ERROR, backend=_Backend(_key(30)))
    assert "NOT a storage problem" in msg
    assert "broken import" in msg


def test_import_failure_advice_names_the_two_project_files():
    msg = diagnose_storage_error(_IMPORT_ERROR, backend=_Backend(_key(30)))
    assert "suspension/project.py" in msg
    assert "suspension.project" in msg


@pytest.mark.parametrize("err", [
    "No module named 'suspension.power_draw'",
    "'ProjectStore' object has no attribute 'save_hint'",
    "cannot import name 'X' from 'y'",
])
def test_all_deployment_errors_take_the_same_branch(err):
    assert "NOT a storage problem" in diagnose_storage_error(err)


def test_stale_write_error_is_still_importable():
    """The regression this whole episode caused: the name must survive."""
    from suspension.project import StaleWriteError
    assert issubclass(StaleWriteError, RuntimeError)


def test_the_live_module_is_the_one_carrying_the_helpers():
    """Both must exist in suspension.project, since that is what the app imports."""
    import suspension.project as sp
    assert callable(sp.diagnose_storage_error)
    assert callable(sp._jwt_expiry)
    assert hasattr(sp, "StaleWriteError")


def test_save_hint_attribute_exists_before_any_save():
    from suspension.project import ProjectStore
    import inspect
    src = inspect.getsource(ProjectStore)
    assert "save_hint" in src
