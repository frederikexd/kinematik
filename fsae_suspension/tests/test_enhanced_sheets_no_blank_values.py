# ============================================================================
#  Nothing on the three added sheets may read as a zero or a dash unless the
#  zero is the answer.
#
#  A dash cannot distinguish "no value exists" from "nobody computed it", and a
#  zero cannot distinguish "the result is zero" from "the reference was blank".
#  Both were the shape of the original defect, so both are pinned here: every
#  placeholder is banned outright, and every zero has to be one of the few the
#  physics actually produces.
# ============================================================================
import io
import os
import re

import numpy as np
import pytest

openpyxl = pytest.importorskip("openpyxl")
from openpyxl.utils import get_column_letter as _CL

from suspension import ev_excel_roundtrip as rt

_SRC = os.path.join(os.path.dirname(__file__), "data",
                    "FSAE_EV_Power_Draw.xlsx")

pytestmark = pytest.mark.skipif(
    not os.path.exists(_SRC),
    reason="stock FSAE_EV_Power_Draw.xlsx fixture not present")

_PLACEHOLDER = re.compile(r"^\s*(-{2,}|—+|–+|n/?a|TBD|\?+|#+|null|None)\s*$",
                          re.IGNORECASE)

#: The only zeros the three sheets are allowed to show, each with the reason.
_JUSTIFIED_ZEROS = {
    "PackHeatmap!A153":  "t = 0, the first sample of the lap",
    "PackHeatmap!G153":  "cumulative heat per cell is zero before the lap starts",
    "LapEnergy!A6":      "t = 0, the first sample of the lap",
    "LapEnergy!D6":      "no energy has been used at t = 0",
    "FeasiblePack!B8":   "no energy shortfall — the pack finishes the lap",
    "FeasiblePack!C8":   "the minimum pack is sized to have no shortfall",
    "FeasiblePack!D8":   "the recommended pack is sized to have no shortfall",
}


@pytest.fixture(scope="module")
def enhanced():
    with open(_SRC, "rb") as fh:
        src = fh.read()
    t = np.arange(335) * 0.04
    v = 18.0 + 14.0 * np.sin(np.linspace(0.0, 7.0, 335)) ** 2
    base = rt.lap_to_excel_roundtrip(speed_ms=v, time_s=t, excel_bytes=src,
                                     lap_time_s=float(t[-1]))
    assert base.ok, base.error
    enh = rt.build_enhanced_excel(
        result=base, excel_bytes=base.excel_bytes,
        thermals=rt.compute_cell_thermals(base.current_draw_a, base.time_s,
                                          base.pack),
        soc_data=rt.compute_soc_and_stop(base.power_kw, base.time_s, base.pack,
                                         base.motor),
        min_pack=rt.compute_minimum_feasible_pack(
            base.power_kw, base.time_s, base.current_draw_a, base.pack,
            base.motor),
        lap_time_s=float(t[-1]))
    return openpyxl.load_workbook(io.BytesIO(enh))


@pytest.mark.parametrize("sheet",
                         ["PackHeatmap", "LapEnergy", "FeasiblePack"])
def test_no_placeholder_text_anywhere(enhanced, sheet):
    """"—" and "?" are what a cell says when a value was never computed."""
    ws = enhanced[sheet]
    found = [f"{ws.title}!{c.coordinate} = {c.value!r}"
             for row in ws.iter_rows() for c in row
             if isinstance(c.value, str) and _PLACEHOLDER.match(c.value)]
    assert not found, f"placeholder cells: {found}"


@pytest.mark.parametrize("sheet",
                         ["PackHeatmap", "LapEnergy", "FeasiblePack"])
def test_every_zero_is_one_we_can_name(enhanced, sheet):
    ws = enhanced[sheet]
    for row in ws.iter_rows():
        for c in row:
            v = c.value
            if isinstance(v, (int, float)) and not isinstance(v, bool) and v == 0:
                key = f"{sheet}!{c.coordinate}"
                assert key in _JUSTIFIED_ZEROS, (
                    f"unexplained zero at {key}; if it is correct, add it to "
                    f"_JUSTIFIED_ZEROS with the reason")


@pytest.mark.parametrize("sheet",
                         ["PackHeatmap", "LapEnergy", "FeasiblePack"])
def test_no_holes_inside_a_populated_row(enhanced, sheet):
    """A gap mid-row is a cell someone forgot, not a layout choice."""
    ws = enhanced[sheet]
    holes = []
    for row in ws.iter_rows():
        filled = [c.column for c in row if c.value is not None]
        if len(filled) < 3:                 # spacer or title row
            continue
        for col in range(min(filled), max(filled) + 1):
            if ws.cell(row=row[0].row, column=col).value is None:
                holes.append(f"{_CL(col)}{row[0].row}")
    # The unit column is legitimately empty for the pass/fail row.
    holes = [h for h in holes if h != "E9"]
    assert not holes, f"{sheet} has interior blanks: {holes[:10]}"


def test_advisor_numbers_are_numbers(enhanced):
    """Stored as text they cannot be charted, summed or reformatted.

    The what-if table below was numeric and the comparison table above was not,
    which is how the inconsistency went unnoticed.
    """
    ws = enhanced["FeasiblePack"]
    for r in range(3, 12):
        label = ws.cell(row=r, column=1).value
        if "Fuse OK" in str(label):          # a verdict, legitimately text
            continue
        for c in (2, 3, 4):
            v = ws.cell(row=r, column=c).value
            assert isinstance(v, (int, float)) and not isinstance(v, bool), \
                f"FeasiblePack!{_CL(c)}{r} ({label}) holds {v!r}, not a number"
            assert ws.cell(row=r, column=c).number_format != "General", \
                f"FeasiblePack!{_CL(c)}{r} has no display format"


def test_pack_independent_quantities_are_stated_in_every_column(enhanced):
    """Peak current and fuse status do not change with pack capacity.

    Showing them under one column and dashing the other two implied they were
    unknown for those packs.
    """
    ws = enhanced["FeasiblePack"]
    rows = {str(ws.cell(row=r, column=1).value): r for r in range(2, 12)}
    for label in ("Peak pack current", "Fuse OK for this lap", "Fuse rating"):
        r = rows[label]
        vals = [ws.cell(row=r, column=c).value for c in (2, 3, 4)]
        assert len(set(vals)) == 1 and vals[0] not in (None, ""), \
            f"{label} differs across columns or is blank: {vals}"


def test_lap_energy_states_what_did_not_happen(enhanced):
    """When the pack finishes there is no stop time — say so, don't dash."""
    ws = enhanced["LapEnergy"]
    header = [ws.cell(row=2, column=c).value for c in range(1, 8)]
    values = [ws.cell(row=3, column=c).value for c in range(1, 8)]
    card = dict(zip(header, values))
    assert card["Pack finishes lap"].endswith("YES")
    assert "finishes" in card["Stop time"]
    assert card["Energy deficit"].startswith("0.000")
    assert card["Energy margin"].endswith("kWh")
