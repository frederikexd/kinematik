# ============================================================================
#  KinematiK — Formula SAE suspension & vehicle dynamics toolkit
#  Created by Frederik Thio. Copyright (c) 2026 Frederik Thio.
#  Open source. Original author: Frederik Thio, creator of KinematiK.
# ============================================================================
"""
Regression tests for the sheet-resolution fix.

The bug: `wb["Battery Pack Calcs"]` on a workbook whose pack sheet is named
`BatteryPackConfig` raised openpyxl's KeyError, which surfaced to the user as

    Lap time section unavailable: 'Worksheet Battery Pack Calcs does not exist.'

— an entire tab lost because a tab had been renamed. Three separate defects sat
behind that one message, and each gets its own test here:

1. The lookup was keyed to one exact spelling.
2. A missing pack sheet was fatal, even though the block it gates is a
   convenience stamp and the lap results the caller asked for were already
   computed.
3. `ElecParams.from_excel` guarded the same lookup with `in wb.sheetnames` and
   fell through to dataclass defaults — which happen to equal this pack, so it
   looked correct while ignoring the file. `test_reader_tracks_edits` is the one
   that would have caught it: it mutates the workbook and asserts the output
   moves.

And one that was found while fixing them: the summary block was stamped at a
hardcoded row 18, which in the current workbook holds the user's own
`Pack Energy 10% SOC` formula.
"""

import io

import numpy as np
import openpyxl
import pytest

from suspension import ev_excel_roundtrip as rt
from suspension.ev_electrical_check import ElecParams


# --------------------------------------------------------------------------- #
#  Fixtures
# --------------------------------------------------------------------------- #
_PACK_ROWS = [
    ("Fuse Max (A)", 50), ("Parrallel Battery Count", 3),
    ("Series Battery Count", 140), ("Nominal Battery Voltage (V)", 3.6),
    ("Capacity Battery Cell (Ah)", 5), ("Endurance Length (km)", 22),
    ("Max Battery Cells", 560),
    ("Internal Resistance Battery Cell (Ohms)", 0.0128),
    ("Battery Cell Weight (kg)", 0.071),
    ("Battery Pack Cell Count", 420),
    ("Battery Pack Internal Resistance (Ohms)", 1.792),
    ("Battery Pack Nominal Voltage (V)", 504),
    ("Current Across a Single Cell (A)", 16.67),
    ("Power Draw (kW)", 25.2), ("Pack Capacity (Ah)", 15),
    ("Pack Energy (kWh)", 7.56), ("Pack Energy 80% SOC", 6.048),
    ("Pack Energy 10% SOC", 0.756),
    ("Joule Heating Maximum (kWh)", 4.48),
]

_EP_ROWS = [
    ("Motor Peak Torque (Nm)", 120), ("Motor Peak Power (kW)", 150),
    ("Motor Frequency (kHz)", 1.96), ("Motor Poles", 4),
    ("Motor Max DC Voltage (V)", 925), ("Motor Efficiency ", 0.9545),
    ("Current From Battery Pack (A)", 50), ("Battery Pack Voltage (V)", 504),
    ("Motor Maximum Speed (rpm)", 23500), ("Wheel Diameter (in)", 18),
    ("Motor Power Factor Assumed", 0.95), ("Resistance (Ohms)", 0.1),
]


def _workbook(pack_sheet_name="BatteryPackConfig", *, with_pack=True):
    wb = openpyxl.Workbook()
    svt = wb.active
    svt.title = "SpeedVsTime"
    svt["A1"], svt["B1"] = "time (s)", "Speed (mph)"
    for i in range(60):
        svt.cell(2 + i, 1, i * 0.0667)
        svt.cell(2 + i, 2, 30.0)

    if with_pack:
        ws = wb.create_sheet(pack_sheet_name)
        for i, (a, b) in enumerate(_PACK_ROWS, start=1):
            ws.cell(i, 1, a)
            ws.cell(i, 2, b)

    ep = wb.create_sheet("ElecPropulsion")
    for i, (a, b) in enumerate(_EP_ROWS, start=1):
        ep.cell(i, 1, a)
        ep.cell(i, 2, b)
    ep["G1"] = "Gear Ratio:"
    for gi in range(15):
        ep.cell(1, 8 + gi, 1.0 / (gi + 1))
    return wb


def _bytes(wb):
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _path(tmp_path, wb, name="wb.xlsx"):
    p = tmp_path / name
    wb.save(p)
    return str(p)


# --------------------------------------------------------------------------- #
#  1. Resolution by alias
# --------------------------------------------------------------------------- #
@pytest.mark.parametrize("name", [
    "Battery Pack Calcs",       # the original spelling
    "BatteryPackConfig",        # the spelling that crashed
    "Battery Pack Config",
    "BatteryPackCalcs",
    "Pack Config",
])
def test_pack_sheet_resolves_under_every_known_alias(name):
    ws = rt.resolve_pack_sheet(_workbook(name))
    assert ws is not None
    assert ws.title == name


def test_alias_matching_ignores_case_and_spacing():
    ws = rt.resolve_pack_sheet(_workbook("batterypackconfig"))
    assert ws is not None


def test_pack_sheet_resolves_by_content_when_renamed_entirely():
    """The durable half: a sheet holding the pack labels IS the pack sheet."""
    ws = rt.resolve_pack_sheet(_workbook("Accumulator v3 FINAL"))
    assert ws is not None
    assert ws.title == "Accumulator v3 FINAL"


def test_speed_and_propulsion_sheets_also_resolve():
    wb = _workbook()
    assert rt.resolve_svt_sheet(wb).title == "SpeedVsTime"
    assert rt.resolve_ep_sheet(wb).title == "ElecPropulsion"


def test_resolver_returns_none_rather_than_raising():
    """A missing sheet is a condition to handle, not an exception to leak.

    openpyxl's KeyError message names a worksheet; it cannot name the feature
    that just died, which is why the user saw a sheet name and no guidance.
    """
    wb = _workbook(with_pack=False)
    assert rt.resolve_pack_sheet(wb) is None
    assert rt.resolve_sheet(wb, ("Nope",), ("nothing at all",)) is None


def test_resolver_prefers_an_exact_alias_over_a_content_match():
    wb = _workbook("Battery Pack Calcs")
    decoy = wb.create_sheet("Old Pack Numbers")
    for i, (a, b) in enumerate(_PACK_ROWS, start=1):
        decoy.cell(i, 1, a)
        decoy.cell(i, 2, b)
    assert rt.resolve_pack_sheet(wb).title == "Battery Pack Calcs"


# --------------------------------------------------------------------------- #
#  2. The crash itself
# --------------------------------------------------------------------------- #
def _lap():
    t = np.linspace(0, 4, 60)
    return 15 + 5 * np.sin(t) ** 2, t


def test_roundtrip_no_longer_raises_on_the_renamed_sheet():
    """The exact reproduction of the reported failure."""
    v, t = _lap()
    r = rt.lap_to_excel_roundtrip(v, t, _bytes(_workbook("BatteryPackConfig")),
                                  lap_time_s=4.0)
    assert r.ok
    assert r.error == ""
    assert r.excel_bytes


def test_roundtrip_works_under_the_original_name_too():
    v, t = _lap()
    r = rt.lap_to_excel_roundtrip(v, t, _bytes(_workbook("Battery Pack Calcs")),
                                  lap_time_s=4.0)
    assert r.ok


def test_missing_pack_sheet_warns_instead_of_killing_the_result():
    """The summary stamp is a convenience. Losing it must not lose the lap."""
    v, t = _lap()
    r = rt.lap_to_excel_roundtrip(v, t, _bytes(_workbook(with_pack=False)),
                                  lap_time_s=4.0)
    assert r.ok
    assert any("pack sheet not found" in w.lower() for w in r.warnings)
    assert r.peak_current_a > 0            # the results still came through
    assert r.excel_bytes


def test_missing_speed_sheet_gives_a_message_naming_the_problem():
    wb = _workbook()
    del wb["SpeedVsTime"]
    v, t = _lap()
    r = rt.lap_to_excel_roundtrip(v, t, _bytes(wb))
    assert not r.ok
    assert "speed" in r.error.lower()
    # and it lists what it did find, so the user can see the mismatch
    assert "ElecPropulsion" in r.error


def test_missing_propulsion_sheet_gives_a_message_naming_the_problem():
    wb = _workbook()
    del wb["ElecPropulsion"]
    v, t = _lap()
    r = rt.lap_to_excel_roundtrip(v, t, _bytes(wb))
    assert not r.ok
    assert "propulsion" in r.error.lower()


# --------------------------------------------------------------------------- #
#  3. The summary stamp must not overwrite the user's own rows
# --------------------------------------------------------------------------- #
def test_summary_block_is_appended_below_existing_data():
    v, t = _lap()
    wb_in = _workbook("BatteryPackConfig")
    original_last = wb_in["BatteryPackConfig"].max_row
    r = rt.lap_to_excel_roundtrip(v, t, _bytes(wb_in), lap_time_s=4.0)
    ws = openpyxl.load_workbook(io.BytesIO(r.excel_bytes))["BatteryPackConfig"]

    # every original label survives
    for row, (label, _val) in enumerate(_PACK_ROWS, start=1):
        assert ws.cell(row, 1).value == label

    # and the stamp begins strictly below them
    header = next(row for row in range(1, ws.max_row + 1)
                  if str(ws.cell(row, 1).value or "").startswith("───"))
    assert header > original_last


def test_summary_block_does_not_land_on_row_18():
    """Row 18 holds 'Pack Energy 10% SOC' in the shipped workbook."""
    v, t = _lap()
    r = rt.lap_to_excel_roundtrip(v, t, _bytes(_workbook()), lap_time_s=4.0)
    ws = openpyxl.load_workbook(io.BytesIO(r.excel_bytes))["BatteryPackConfig"]
    assert ws.cell(18, 1).value == "Pack Energy 10% SOC"
    assert ws.cell(19, 1).value == "Joule Heating Maximum (kWh)"


def test_summary_block_content_is_written():
    v, t = _lap()
    r = rt.lap_to_excel_roundtrip(v, t, _bytes(_workbook()), lap_time_s=4.0)
    ws = openpyxl.load_workbook(io.BytesIO(r.excel_bytes))["BatteryPackConfig"]
    labels = [ws.cell(row, 1).value for row in range(1, ws.max_row + 1)]
    assert "Lap Time (s)" in labels
    assert "Peak Current Draw (A)" in labels


# --------------------------------------------------------------------------- #
#  4. The silent-default bug
# --------------------------------------------------------------------------- #
def test_reader_reads_the_renamed_sheet(tmp_path):
    p = _path(tmp_path, _workbook("BatteryPackConfig"))
    ep = ElecParams.from_excel(p)
    assert ep.n_series == 140
    assert ep.n_parallel == 3


def test_reader_tracks_edits(tmp_path):
    """The test that would have caught the original silent fallback.

    Defaults of 504 V / 15 Ah / 140S3P happen to equal this pack, so comparing
    the output to expected values proves nothing. Mutating the file and checking
    the output moves is the only assertion that distinguishes reading from
    defaulting.
    """
    wb = _workbook("BatteryPackConfig")
    ws = wb["BatteryPackConfig"]
    ws.cell(2, 2, 4)          # parallel
    ws.cell(3, 2, 100)        # series
    ws.cell(5, 2, 4.0)        # cell Ah
    p = _path(tmp_path, wb, "mutated.xlsx")

    ep = ElecParams.from_excel(p)
    assert ep.n_series == 100
    assert ep.n_parallel == 4
    assert (ep.n_series, ep.n_parallel) != (ElecParams().n_series,
                                            ElecParams().n_parallel)


def test_reader_still_defaults_gracefully_when_there_is_no_pack_sheet(tmp_path):
    """No pack sheet is a legitimate reason to fall back — silently defaulting
    when the sheet was RIGHT THERE under another name was not."""
    p = _path(tmp_path, _workbook(with_pack=False))
    ep = ElecParams.from_excel(p)
    assert ep.n_series == ElecParams().n_series


def test_reader_survives_a_content_renamed_sheet(tmp_path):
    p = _path(tmp_path, _workbook("Accumulator v3 FINAL"))
    ep = ElecParams.from_excel(p)
    assert ep.n_series == 140


def test_extract_params_returns_a_populated_pack_dict():
    """This returned {} before the fix."""
    d = rt.extract_params_from_excel(_bytes(_workbook("BatteryPackConfig")))
    assert d["pack"], "pack dict is empty — the sheet was not read"
    assert d["motor"]


# --------------------------------------------------------------------------- #
#  5. The off-by-one row map
# --------------------------------------------------------------------------- #
#  Second reported crash:
#      Lap time section unavailable: The value 1.792 is less than the min_value 100.0.
#  1.792 is BatteryPackConfig!B11, the pack internal resistance. The fixed row
#  map had no entry for that row, so every key from pack_voltage_v down read one
#  row too high and the resistance was handed to a voltage widget.
# --------------------------------------------------------------------------- #
def test_pack_voltage_is_not_the_resistance():
    """The exact reproduction: B11 must never be read as pack voltage."""
    ws = rt.resolve_pack_sheet(_workbook())
    pack = rt.read_pack_values(ws, [])
    assert pack["pack_voltage_v"] == pytest.approx(504.0)
    assert pack["pack_voltage_v"] != pytest.approx(1.792)


def test_every_pack_value_lands_in_its_widget_range():
    """The widget bounds are the real contract — assert against them directly."""
    pack = rt.read_pack_values(rt.resolve_pack_sheet(_workbook()), [])
    bounds = {
        "pack_voltage_v":   (100.0, 1000.0),
        "pack_capacity_ah": (1.0, 100.0),
        "pack_energy_wh":   (100.0, 20000.0),
        "fuse_max_a":       (1.0, 1000.0),
    }
    for key, (lo, hi) in bounds.items():
        assert key in pack, f"{key} missing"
        assert lo <= pack[key] <= hi, f"{key}={pack[key]} outside {lo}..{hi}"


def test_values_are_read_by_label_not_position():
    """Insert a row at the top: coordinates shift, labels don't."""
    wb = _workbook()
    wb["BatteryPackConfig"].insert_rows(1)
    wb["BatteryPackConfig"]["A1"] = "--- notes added by someone ---"
    pack = rt.read_pack_values(rt.resolve_pack_sheet(wb), [])
    assert pack["pack_voltage_v"] == pytest.approx(504.0)
    assert pack["n_series"] == 140
    assert pack["fuse_max_a"] == 50


def test_derived_values_come_from_primary_inputs():
    """Pack voltage/capacity/energy are derived, not read from computed cells.

    Those cells are where the workbook's arithmetic errors live, so reading them
    would import the errors.
    """
    wb = _workbook()
    ws = wb["BatteryPackConfig"]
    ws.cell(12, 2, 999)          # corrupt the sheet's own 'pack nominal voltage'
    pack = rt.read_pack_values(ws, [])
    assert pack["pack_voltage_v"] == pytest.approx(140 * 3.6)


def test_sheet_disagreement_is_reported_not_silently_dropped():
    warns = []
    rt.read_pack_values(rt.resolve_pack_sheet(_workbook()), warns)
    # the fixture carries the workbook's real 1.792 resistance error
    assert any("pack_r_ohm" in w for w in warns)
    assert any("1.792" in w for w in warns)


def test_pack_resistance_is_derived_correctly():
    """S*(R_cell/P), not cell_count*(R_cell/P) which cancels to S*R_cell."""
    pack = rt.read_pack_values(rt.resolve_pack_sheet(_workbook()), [])
    assert pack["pack_r_ohm"] == pytest.approx(140 * (0.0128 / 3))
    assert pack["pack_r_ohm"] != pytest.approx(1.792)


def test_implausible_value_is_dropped_with_an_explanation():
    """Better a missing key than a value that crashes a widget three frames on."""
    wb = _workbook()
    ws = wb["BatteryPackConfig"]
    ws.cell(3, 2, 999999)        # absurd series count
    warns = []
    pack = rt.read_pack_values(ws, warns)
    assert "n_series" not in pack
    assert any("outside the plausible range" in w for w in warns)


def test_abbreviated_labels_still_work():
    """The older sheet vocabulary ('pack V', 'cell I') must keep resolving."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Battery Pack Calcs"
    for i, (lbl, val) in enumerate([
            ("Fuse max (A)", 50), ("n_parallel", 4), ("n_series", 120),
            ("cell V", 3.6), ("cell Ah", 4.2), ("endurance km", 22),
            ("max cells", 600), ("cell R ohm", 0.012), ("cell wt kg", 0.045),
            ("pack cells", 480), ("pack V", 432), ("cell I", 180),
            ("power kW", 78), ("pack Ah", 16.8), ("pack Wh", 7257),
            ("joule kWh", 0.85)], start=1):
        ws.cell(i, 1, lbl)
        ws.cell(i, 2, val)
    pack = rt.read_pack_values(ws, [])
    assert pack["pack_voltage_v"] == pytest.approx(432.0)
    assert pack["cell_current_a"] == 180
    assert pack["joule_heating_kwh"] == pytest.approx(0.85)


def test_roundtrip_exposes_the_resistance_disagreement_as_a_warning():
    v, t = _lap()
    r = rt.lap_to_excel_roundtrip(v, t, _bytes(_workbook()), lap_time_s=4.0)
    assert r.ok
    assert any("pack_r_ohm" in w for w in r.warnings)
